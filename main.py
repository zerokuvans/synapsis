#!/usr/bin/env python3
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file, make_response, Response
import mysql.connector
from mysql.connector import Error
from dotenv import load_dotenv
import os
from functools import wraps
import bcrypt
from datetime import datetime, timedelta
import json
import csv
import io
from collections import defaultdict
import calendar
import pytz
import logging
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.colors import blue
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from io import BytesIO
import jwt
import logging
import base64
import re
import tempfile
from PIL import Image as PILImage
import secrets
import string
import hashlib
import uuid
import pandas as pd
import json
from datetime import datetime, timedelta
import io
import time
import zipfile
import re
import logging
from decimal import Decimal
#from models import Suministro  # Asegúrate de importar el modelo correcto
from flask_login import LoginManager, UserMixin, login_user, logout_user, current_user, login_required

load_dotenv()

# Configuración de zona horaria
TIMEZONE = pytz.timezone('America/Bogota')

# Remove debug print statements for environment variables
# print(f"MYSQL_HOST: {os.getenv('MYSQL_HOST')}")
# print(f"MYSQL_USER: {os.getenv('MYSQL_USER')}")
# print(f"MYSQL_PASSWORD: {os.getenv('MYSQL_PASSWORD')}")
# print(f"MYSQL_DB: {os.getenv('MYSQL_DB')}")
# print(f"MYSQL_PORT: {os.getenv('MYSQL_PORT')}")

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY')
app.config['UPLOAD_FOLDER'] = 'static/uploads'

# Configuración de logging
logging.basicConfig(level=logging.INFO)
handler = logging.StreamHandler()
handler.setFormatter(logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
))
app.logger.addHandler(handler)
app.logger.setLevel(logging.INFO)

# Configuración de Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.unauthorized_handler
def unauthorized():
    """Manejador para usuarios no autorizados"""
    try:
        # Devolver JSON para endpoints de API
        api_endpoints = ['/api/', '/preoperacional', '/login']
        if any(request.path.startswith(endpoint) for endpoint in api_endpoints):
            return jsonify({'error': 'No autorizado', 'message': 'Debe iniciar sesión'}), 401
        else:
            # Redirigir a login para páginas web
            return redirect(url_for('login'))
    except Exception as e:
        app.logger.error(f"Error en unauthorized_handler: {e}")
        return jsonify({'error': 'Error de autorización'}), 401

# Definición de la clase User para Flask-Login
class User(UserMixin):
    def __init__(self, id, nombre, role):
        self.id = id
        self.nombre = nombre
        self.role = role
    
    def has_role(self, role):
        return self.role == role

# Función para cargar usuario para Flask-Login
@login_manager.user_loader
def load_user(user_id):
    connection = get_db_connection()
    if connection is None:
        return None
        
    cursor = connection.cursor(dictionary=True)
    cursor.execute("SELECT id_codigo_consumidor, nombre, id_roles FROM recurso_operativo WHERE id_codigo_consumidor = %s", (user_id,))
    user_data = cursor.fetchone()
    cursor.close()
    connection.close()
    
    if user_data:
        return User(
            id=user_data['id_codigo_consumidor'],
            nombre=user_data['nombre'],
            role=ROLES.get(str(user_data['id_roles']))
        )
    return None

# Importar rutas desde app.py
from app import administrativo_asistencia, obtener_supervisores, obtener_tecnicos_por_supervisor, guardar_asistencia_administrativa

# Importar rutas del módulo analistas desde app.py
from app import analistas_index, analistas_causas, analistas_dashboard, api_grupos_causas_cierre, api_tecnologias_causas_cierre, api_agrupaciones_causas_cierre, api_estadisticas_causas_cierre

# Importar rutas del módulo MPA desde app.py
from app import mpa_dashboard, mpa_dashboard_stats, mpa_vehiculos, mpa_soat, mpa_tecnico_mecanica, mpa_licencias, mpa_licencias_conducir, mpa_inspecciones, mpa_siniestros, mpa_mantenimientos
from app import api_get_vehiculos, api_create_vehiculo, api_get_vehiculo, api_update_vehiculo, api_delete_vehiculo, api_get_tecnicos
from app import api_get_mantenimientos, api_create_mantenimiento, api_get_mantenimiento, api_update_mantenimiento, api_delete_mantenimiento, api_get_placas, api_get_categorias_mantenimiento, api_upload_mantenimiento_image
from app import api_get_soat, api_get_soat_by_id, api_create_soat, api_update_soat, api_delete_soat
from app import api_list_tecnico_mecanica, api_get_tecnico_mecanica, api_create_tecnico_mecanica, api_update_tecnico_mecanica, api_delete_tecnico_mecanica
from app import api_list_licencias_conducir, api_get_licencia_conducir, api_create_licencia_conducir, api_update_licencia_conducir, api_delete_licencia_conducir
from app import api_import_vehiculos_excel, api_import_tecnico_mecanica_excel, api_import_soat_excel, api_import_licencias_excel
from app import mpa_rutas, api_import_rutas_excel, api_rutas_tecnicos, api_rutas_por_tecnico, api_google_directions_route, api_riesgo_motos, api_riesgo_por_localidad, api_localidades

# Importar módulo de dotaciones
from dotaciones_api import registrar_rutas_dotaciones
from encuestas_api import registrar_rutas_encuestas
from avisos_api import registrar_rutas_avisos

# Registrar rutas de app.py
app.route('/administrativo/asistencia')(administrativo_asistencia)
app.route('/api/supervisores', methods=['GET'])(obtener_supervisores)
app.route('/api/tecnicos_por_supervisor', methods=['GET'])(obtener_tecnicos_por_supervisor)
app.route('/api/asistencia/guardar', methods=['POST'])(guardar_asistencia_administrativa)

# Registrar rutas del módulo analistas
app.route('/analistas')(analistas_index)
app.route('/analistas/causas')(analistas_causas)
app.route('/analistas/dashboard')(analistas_dashboard)
app.route('/api/analistas/grupos', methods=['GET'])(api_grupos_causas_cierre)
app.route('/api/analistas/tecnologias', methods=['GET'])(api_tecnologias_causas_cierre)
app.route('/api/analistas/agrupaciones', methods=['GET'])(api_agrupaciones_causas_cierre)
app.route('/api/analistas/estadisticas', methods=['GET'])(api_estadisticas_causas_cierre)

# Registrar rutas del módulo MPA
app.route('/mpa')(mpa_dashboard)
app.route('/api/mpa/dashboard-stats')(mpa_dashboard_stats)
app.route('/mpa/vehiculos')(mpa_vehiculos)
app.route('/mpa/soat')(mpa_soat)
app.route('/mpa/tecnico-mecanica')(mpa_tecnico_mecanica)
app.route('/mpa/licencias')(mpa_licencias)
app.route('/mpa/licencias-conducir')(mpa_licencias_conducir)
app.route('/mpa/inspecciones')(mpa_inspecciones)
app.route('/mpa/siniestros')(mpa_siniestros)
app.route('/mpa/mantenimientos')(mpa_mantenimientos)
app.route('/mpa/rutas')(mpa_rutas)

# Registrar rutas de la API de vehículos MPA
app.route('/api/mpa/vehiculos', methods=['GET'])(api_get_vehiculos)
app.route('/api/mpa/vehiculos', methods=['POST'])(api_create_vehiculo)
app.route('/api/mpa/vehiculos/<int:vehiculo_id>', methods=['GET'])(api_get_vehiculo)
app.route('/api/mpa/vehiculos/<int:vehiculo_id>', methods=['PUT'])(api_update_vehiculo)
app.route('/api/mpa/vehiculos/<int:vehiculo_id>', methods=['DELETE'])(api_delete_vehiculo)
app.route('/api/mpa/tecnicos', methods=['GET'])(api_get_tecnicos)

# Registrar rutas de la API de mantenimientos MPA
app.route('/api/mpa/mantenimientos', methods=['GET'])(api_get_mantenimientos)
app.route('/api/mpa/mantenimientos', methods=['POST'])(api_create_mantenimiento)
app.route('/api/mpa/mantenimientos/<int:mantenimiento_id>', methods=['GET'])(api_get_mantenimiento)
app.route('/api/mpa/mantenimientos/<int:mantenimiento_id>', methods=['PUT'])(api_update_mantenimiento)
app.route('/api/mpa/mantenimientos/<int:mantenimiento_id>', methods=['DELETE'])(api_delete_mantenimiento)
app.route('/api/mpa/vehiculos/placas', methods=['GET'])(api_get_placas)
app.route('/api/mpa/categorias-mantenimiento/<tipo_vehiculo>', methods=['GET'])(api_get_categorias_mantenimiento)
app.route('/api/mpa/mantenimientos/upload-image', methods=['POST'])(api_upload_mantenimiento_image)
app.route('/api/mpa/vehiculos/import-excel', methods=['POST'])(api_import_vehiculos_excel)
app.route('/api/mpa/tecnico_mecanica/import-excel', methods=['POST'])(api_import_tecnico_mecanica_excel)
app.route('/api/mpa/soat/import-excel', methods=['POST'])(api_import_soat_excel)
app.route('/api/mpa/licencias-conducir/import-excel', methods=['POST'])(api_import_licencias_excel)
app.route('/api/mpa/rutas/import-excel', methods=['POST'])(api_import_rutas_excel)
app.route('/api/mpa/rutas/tecnicos', methods=['GET'])(api_rutas_tecnicos)
app.route('/api/mpa/rutas/por-tecnico', methods=['GET'])(api_rutas_por_tecnico)
app.route('/api/mpa/rutas/google-directions', methods=['GET'])(api_google_directions_route)
app.route('/api/mpa/rutas/riesgo-motos', methods=['GET'])(api_riesgo_motos)
app.route('/api/mpa/rutas/riesgo-por-localidad', methods=['GET'])(api_riesgo_por_localidad)
app.route('/api/mpa/localidades', methods=['GET'])(api_localidades)

# Registrar rutas de la API de SOAT MPA
app.route('/api/mpa/soat', methods=['GET'])(api_get_soat)
app.route('/api/mpa/soat', methods=['POST'])(api_create_soat)
app.route('/api/mpa/soat/<int:soat_id>', methods=['GET'])(api_get_soat_by_id)
app.route('/api/mpa/soat/<int:soat_id>', methods=['PUT'])(api_update_soat)
app.route('/api/mpa/soat/<int:soat_id>', methods=['DELETE'])(api_delete_soat)

# Registrar rutas de la API de Técnico Mecánica MPA
app.route('/api/mpa/tecnico_mecanica', methods=['GET'])(api_list_tecnico_mecanica)
app.route('/api/mpa/tecnico_mecanica', methods=['POST'])(api_create_tecnico_mecanica)
app.route('/api/mpa/tecnico_mecanica/<int:tm_id>', methods=['GET'])(api_get_tecnico_mecanica)
app.route('/api/mpa/tecnico_mecanica/<int:tm_id>', methods=['PUT'])(api_update_tecnico_mecanica)
app.route('/api/mpa/tecnico_mecanica/<int:tm_id>', methods=['DELETE'])(api_delete_tecnico_mecanica)

# Registrar rutas de la API de Licencias de Conducir MPA
app.route('/api/mpa/licencias-conducir', methods=['GET'])(api_list_licencias_conducir)
app.route('/api/mpa/licencias-conducir', methods=['POST'])(api_create_licencia_conducir)
app.route('/api/mpa/licencias-conducir/<int:licencia_id>', methods=['GET'])(api_get_licencia_conducir)
app.route('/api/mpa/licencias-conducir/<int:licencia_id>', methods=['PUT'])(api_update_licencia_conducir)
app.route('/api/mpa/licencias-conducir/<int:licencia_id>', methods=['DELETE'])(api_delete_licencia_conducir)

# Registrar rutas del módulo de dotaciones
registrar_rutas_dotaciones(app)

# Registrar rutas del módulo de encuestas
registrar_rutas_encuestas(app)
registrar_rutas_avisos(app)

# Database configuration
db_config = {
    'host': os.getenv('MYSQL_HOST'),
    'user': os.getenv('MYSQL_USER'),
    'password': os.getenv('MYSQL_PASSWORD'),
    'database': os.getenv('MYSQL_DB'),
    'port': int(os.getenv('MYSQL_PORT')),
    'time_zone': '+00:00'  # Configurar MySQL para usar UTC
}

# Función para obtener conexión a la base de datos
def get_db_connection():
    try:
        connection = mysql.connector.connect(**db_config)
        return connection
    except Error as e:
        app.logger.error(f"Error de conexión a MySQL: {str(e)}")
        print(f"Error de conexión a MySQL: {str(e)}")
        return None

# Función para obtener la fecha y hora actual en Bogotá
def get_bogota_datetime():
    return datetime.now(TIMEZONE)

# Función para convertir una fecha UTC a fecha de Bogotá
def convert_to_bogota_time(utc_dt):
    if utc_dt.tzinfo is None:
        utc_dt = pytz.UTC.localize(utc_dt)
    return utc_dt.astimezone(TIMEZONE)

# Probar la conexión a la base de datos
try:
    connection = get_db_connection()
    if connection is None:
        raise Exception('Failed to establish a database connection.')
    connection.close()
except Exception as e:
    pass

# Roles definition
ROLES = {
    '1': 'administrativo',
    '2': 'tecnicos',
    '3': 'operativo',
    '4': 'contabilidad',
    '5': 'logistica',
    '6': 'analista',
    '7': 'lider'
}

# Decorador para requerir rol específico
def role_required(role):
    def role_decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_role' not in session:
                flash('Por favor inicia sesión para acceder a esta página.', 'warning')
                return redirect(url_for('login'))
            if session.get('user_role') != role and session.get('user_role') != 'administrativo':
                flash('No tienes permisos para acceder a esta página.', 'danger')
                return redirect(url_for('dashboard'))
            return f(*args, **kwargs)
        return decorated_function
    return role_decorator

# Login required decorator personalizado para roles
def role_required_login(role=None):
    def role_login_decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if role and not current_user.has_role(role) and current_user.role != 'administrativo':
                flash("No tienes permisos para acceder a esta página.", 'danger')
                return redirect(url_for('login'))
            return f(*args, **kwargs)
        return decorated_function
    return role_login_decorator

# Mantener el decorador original para compatibilidad con código existente
def login_required(role=None):
    """Decorador de autenticación compatible con dos modos:
    - Uso directo: @login_required
    - Uso con argumentos: @login_required() o @login_required(role='...')
    """
    # Modo directo: @login_required
    if callable(role):
        f = role
        role = None

        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('Please log in to access this page.', 'warning')
                return redirect(url_for('login'))

            return f(*args, **kwargs)

        return decorated_function

    # Modo con argumentos: @login_required(...) 
    def login_decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('Please log in to access this page.', 'warning')
                return redirect(url_for('login'))

            user_role = session.get('user_role')

            # Si hay un rol requerido y el usuario no es administrativo
            if role and user_role != 'administrativo':
                if isinstance(role, list):
                    if user_role not in role:
                        flash("No tienes permisos para acceder a esta página.", 'danger')
                        return redirect(url_for('login'))
                elif user_role != role:
                    flash("No tienes permisos para acceder a esta página.", 'danger')
                    return redirect(url_for('login'))

            return f(*args, **kwargs)

        return decorated_function

    return login_decorator

def login_required_api(role=None):
    """Decorador para requerir autenticación en APIs"""
    def api_decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return jsonify({'error': 'Autenticación requerida', 'code': 'AUTH_REQUIRED'}), 401
            
            if role and session.get('user_role') != role and session.get('user_role') != 'administrativo':
                return jsonify({'error': 'Permisos insuficientes', 'code': 'INSUFFICIENT_PERMISSIONS'}), 403
            
            return f(*args, **kwargs)
        return decorated_function
    return api_decorator

def login_required_lider():
    """Decorador específico para el módulo de Líder: administradores, líderes y Sandra Cecilia (52912112)"""
    def lider_decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                flash('Please log in to access this page.', 'warning')
                return redirect(url_for('login'))
            
            user_role = session.get('user_role')
            user_cedula = session.get('user_cedula', '')
            
            # Verificar permisos específicos para el módulo de Líder
            # 1. Administradores tienen acceso completo
            # 2. Usuarios con rol 'lider' tienen acceso
            # 3. Sandra Cecilia Cortes Cuervo (cédula: 52912112) tiene acceso especial
            if (user_role == 'administrativo' or 
                user_role == 'lider' or 
                user_cedula == '52912112'):
                return f(*args, **kwargs)
            else:
                flash("No tienes permisos para acceder al módulo de Líder.", 'danger')
                return redirect(url_for('login'))
        
        return decorated_function
    return lider_decorator

def login_required_lider_api():
    """Decorador específico para APIs del módulo de Líder: administradores, líderes y Sandra Cecilia (52912112)"""
    def lider_api_decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'user_id' not in session:
                return jsonify({'error': 'Autenticación requerida', 'code': 'AUTH_REQUIRED'}), 401
            
            user_role = session.get('user_role')
            user_cedula = session.get('user_cedula', '')
            
            # Verificar permisos específicos para APIs del módulo de Líder
            # 1. Administradores tienen acceso completo
            # 2. Usuarios con rol 'lider' tienen acceso
            # 3. Sandra Cecilia Cortes Cuervo (cédula: 52912112) tiene acceso especial
            if (user_role == 'administrativo' or 
                user_role == 'lider' or 
                user_cedula == '52912112'):
                return f(*args, **kwargs)
            else:
                return jsonify({'error': 'Permisos insuficientes para el módulo de Líder', 'code': 'INSUFFICIENT_PERMISSIONS'}), 403
        
        return decorated_function
    return lider_api_decorator

@app.route('/register', methods=['GET', 'POST'])
@login_required(role='administrativo')  # Solo los administrativos pueden registrar usuarios
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        role_id = request.form['role_id']

        # Generar salt y hash
        salt = bcrypt.gensalt()
        hashed_password = bcrypt.hashpw(password.encode('utf-8'), salt).decode('utf-8')

        try:
            connection = get_db_connection()
            if connection is None:
                flash('Error de conexión a la base de datos.', 'error')
                return redirect(url_for('dashboard'))
                
            cursor = connection.cursor(dictionary=True, buffered=True)
            cursor.execute("INSERT INTO recurso_operativo (recurso_operativo_cedula, recurso_operativo_password, id_roles) VALUES (%s, %s, %s)",
                        (username, hashed_password, role_id))
            connection.commit()
            cursor.close()
            connection.close()
            flash('Usuario registrado exitosamente.', 'success')
            return redirect(url_for('dashboard'))
        except Error as e:
            flash(f'Error al registrar usuario: {str(e)}', 'error')
            return redirect(url_for('dashboard'))

    return redirect(url_for('dashboard'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # Añadir logging para depuración
        app.logger.info("Intento de inicio de sesión recibido")
        
        # Obtener credenciales del formulario
        username = request.form.get('username', '')
        password = request.form.get('password', '').encode('utf-8')
        
        # Validación básica
        if not username or not password:
            app.logger.warning("Intento de inicio de sesión con campos vacíos")
            return jsonify({
                'status': 'error', 
                'message': 'Por favor ingrese usuario y contraseña'
            }), 400

        connection = None
        cursor = None
        try:
            # Intentar conexión a la base de datos
            app.logger.info(f"Intentando conectar a la base de datos para usuario: {username}")
            connection = get_db_connection()
            
            if connection is None:
                app.logger.error("Error al conectar con la base de datos")
                return jsonify({
                    'status': 'error', 
                    'message': 'Error de conexión a la base de datos. Por favor intente más tarde.'
                }), 500
                
            cursor = connection.cursor(dictionary=True, buffered=True)
            
            # Buscar usuario
            app.logger.info(f"Consultando usuario con cedula: {username}")
            cursor.execute("SELECT id_codigo_consumidor, id_roles, recurso_operativo_password, nombre, estado FROM recurso_operativo WHERE recurso_operativo_cedula = %s", (username,))
            user_data = cursor.fetchone()

            # Verificar si el usuario existe
            if not user_data:
                app.logger.warning(f"Usuario no encontrado: {username}")
                return jsonify({
                    'status': 'error', 
                    'message': 'Usuario o contraseña inválidos'
                }), 401
            
            # Verificar si el usuario está activo
            if user_data['estado'] != 'Activo':
                app.logger.warning(f"Intento de acceso de usuario inactivo: {username} - Estado: {user_data['estado']}")
                return jsonify({
                    'status': 'error', 
                    'message': 'Su cuenta se encuentra inactiva. Contacte al administrador para más información.'
                }), 403
            
            # Verificar contraseña
            app.logger.info("Verificando contraseña")
            stored_password = user_data['recurso_operativo_password']
            
            # Asegurar que stored_password es bytes
            if isinstance(stored_password, str):
                stored_password = stored_password.encode('utf-8')
            
            # Verificar si el hash está en formato correcto
            if not stored_password.startswith(b'$2b$') and not stored_password.startswith(b'$2a$'):
                app.logger.error(f"Formato de contraseña inválido para usuario: {username}")
                return jsonify({
                    'status': 'error', 
                    'message': 'Error en el formato de la contraseña almacenada. Contacte al administrador.'
                }), 500
            
            try:
                # Verificar contraseña con bcrypt
                if bcrypt.checkpw(password, stored_password):
                    app.logger.info(f"Inicio de sesión exitoso para: {username}")
                    
                    # Crear objeto User para Flask-Login
                    user = User(
                        id=user_data['id_codigo_consumidor'],
                        nombre=user_data['nombre'],
                        role=ROLES.get(str(user_data['id_roles']))
                    )
                    
                    # Iniciar sesión con Flask-Login
                    login_user(user)
                    
                    # Mantener también las variables de sesión para compatibilidad
                    session['user_id'] = user_data['id_codigo_consumidor']
                    session['id_codigo_consumidor'] = user_data['id_codigo_consumidor']  # Agregar para compatibilidad
                    session['user_cedula'] = username  # Guardar la cédula del usuario
                    session['user_role'] = ROLES.get(str(user_data['id_roles']))
                    session['user_name'] = user_data['nombre']
                    
                    app.logger.info(f"Sesión establecida: ID={user_data['id_codigo_consumidor']}, Rol={ROLES.get(str(user_data['id_roles']))}")

                    # Verificar vencimientos para el usuario
                    try:
                        cursor.execute("""
                            SELECT 
                                fecha_vencimiento_licencia,
                                fecha_vencimiento_soat,
                                fecha_vencimiento_tecnomecanica
                            FROM preoperacional 
                            WHERE id_codigo_consumidor = %s
                            ORDER BY fecha DESC 
                            LIMIT 1
                        """, (user_data['id_codigo_consumidor'],))
                        
                        ultimo_registro = cursor.fetchone()
                        if ultimo_registro:
                            fecha_actual = datetime.now().date()
                            mensajes_vencimiento = []
                            
                            if ultimo_registro['fecha_vencimiento_licencia']:
                                dias_licencia = (ultimo_registro['fecha_vencimiento_licencia'] - fecha_actual).days
                                if 0 <= dias_licencia <= 30:
                                    mensajes_vencimiento.append(f'Tu licencia de conducción vence en {dias_licencia} días')
                            
                            if ultimo_registro['fecha_vencimiento_soat']:
                                dias_soat = (ultimo_registro['fecha_vencimiento_soat'] - fecha_actual).days
                                if 0 <= dias_soat <= 30:
                                    mensajes_vencimiento.append(f'El SOAT vence en {dias_soat} días')
                            
                            if ultimo_registro['fecha_vencimiento_tecnomecanica']:
                                dias_tecno = (ultimo_registro['fecha_vencimiento_tecnomecanica'] - fecha_actual).days
                                if 0 <= dias_tecno <= 30:
                                    mensajes_vencimiento.append(f'La tecnomecánica vence en {dias_tecno} días')
                            
                            if mensajes_vencimiento:
                                flash('¡ATENCIÓN! ' + '. '.join(mensajes_vencimiento), 'warning')
                    except Exception as e:
                        app.logger.error(f"Error al verificar vencimientos: {str(e)}")
                        # No bloqueamos el inicio de sesión por este error
                    
                    # Cerrar conexión a la base de datos
                    if cursor:
                        cursor.close()
                    if connection:
                        connection.close()

                    # Si la solicitud espera JSON, devolver respuesta JSON
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return jsonify({
                            'status': 'success',
                            'message': 'Inicio de sesión exitoso',
                            'user_id': user_data['id_codigo_consumidor'],
                            'user_role': ROLES.get(str(user_data['id_roles'])),
                            'user_name': user_data['nombre'],
                            'redirect_url': url_for('dashboard')
                        })
                    # Si es una solicitud normal, redirigir
                    return redirect(url_for('dashboard'))
                else:
                    app.logger.warning(f"Contraseña incorrecta para usuario: {username}")
                    return jsonify({
                        'status': 'error', 
                        'message': 'Usuario o contraseña inválidos'
                    }), 401
            except Exception as e:
                app.logger.error(f"Error al verificar contraseña: {str(e)}")
                return jsonify({
                    'status': 'error', 
                    'message': 'Error al verificar credenciales. Intente nuevamente.'
                }), 500

        except mysql.connector.Error as e:
            app.logger.error(f"Error de MySQL: {str(e)}")
            return jsonify({
                'status': 'error', 
                'message': f'Error de base de datos: {str(e)}'
            }), 500
        except Exception as e:
            app.logger.error(f"Error inesperado: {str(e)}")
            # Imprimir stack trace para depuración
            import traceback
            app.logger.error(traceback.format_exc())
            return jsonify({
                'status': 'error', 
                'message': 'Error interno del servidor. Por favor contacte al administrador.'
            }), 500
        finally:
            # Asegurarse de cerrar los recursos
            if cursor:
                cursor.close()
            if connection:
                connection.close()
            app.logger.info("Recursos de base de datos liberados")

    # Para solicitudes GET simplemente mostrar la plantilla de login
    return render_template('login.html')

# Redirigir la ruta raíz a la página de login para consistencia y aceptar POST
@app.route('/', methods=['GET', 'POST'])
def root():
    if request.method == 'POST':
        # Delegar a la misma lógica de login para compatibilidad con clientes que postean a '/'
        return login()
    return redirect(url_for('login'))

@app.route('/logout', methods=['GET', 'POST'])
def logout():
    # Obtener el rol del usuario antes de cerrar sesión para personalizar el mensaje
    user_role = session.get('user_role', '')
    
    # Cerrar sesión con Flask-Login
    logout_user()
    
    # Limpiar todas las variables de sesión
    session.clear()
    
    # Asegurarse de que la sesión se elimine
    if session.get('user_id'):
        session.pop('user_id')
    if session.get('user_role'):
        session.pop('user_role')
    
    # Mensaje personalizado según el rol
    if user_role:
        flash(f'Has cerrado sesión exitosamente. ¡Hasta pronto, {user_role.title()}!', 'info')
    else:
        flash('Has cerrado sesión exitosamente.', 'info')
    
    # Redirigir al login
    return redirect(url_for('login'))

# ============================================================================
# MÓDULO ANALISTAS - RUTAS PRINCIPALES
# ============================================================================

@app.route('/analistas')
@login_required()
def main_analistas_index():
    """Renderizar el dashboard del módulo analistas"""
    return render_template('modulos/analistas/dashboard.html')

@app.route('/analistas/causas')
@login_required()
def main_analistas_causas():
    """Renderizar la página de causas de cierre"""
    return render_template('modulos/analistas/index.html')

@app.route('/analistas/dashboard')
@login_required()
def main_analistas_dashboard():
    """Renderizar el dashboard del módulo analistas"""
    return render_template('modulos/analistas/dashboard.html')

@app.route('/analistas/codigos')
@login_required()
def main_analistas_codigos():
    """Renderizar la página de códigos de facturación"""
    return render_template('modulos/analistas/codigos.html')

@app.route('/analistas/inicio-operacion-tecnicos')
@login_required()
def main_inicio_operacion_tecnicos():
    """Renderizar la página de inicio de operación para técnicos"""
    return render_template('inicio_operacion_tecnicos.html')

# ============================================================================
# MÓDULO ANALISTAS - API ENDPOINTS Códigos de Facturación
# ============================================================================

@app.route('/api/analistas/codigos', methods=['GET'])
@login_required()
def api_codigos_facturacion():
    """Lista de códigos de facturación con filtros opcionales"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
        cursor = connection.cursor(dictionary=True)

        texto_busqueda = request.args.get('busqueda', '').strip()
        tecnologia = request.args.get('tecnologia', '').strip()
        agrupacion = request.args.get('agrupacion', '').strip()
        grupo = request.args.get('grupo', '').strip()

        query = """
            SELECT 
                id_base_codigos_facturacion AS idbase_codigos_facturacion,
                codigo AS codigo_codigos_facturacion,
                descripcion AS nombre_codigos_facturacion,
                tecnologia,
                '' AS instrucciones_de_uso_codigos_facturacion,
                categoria,
                nombre,
                0 AS facturable_codigos_facturacion
            FROM base_codigos_facturacion
            WHERE 1=1
        """
        params = []

        if texto_busqueda:
            query += """
                AND (
                    codigo LIKE %s OR 
                    descripcion LIKE %s
                )
            """
            busqueda_param = f"%{texto_busqueda}%"
            params.extend([busqueda_param, busqueda_param])
        if tecnologia:
            query += " AND tecnologia = %s"
            params.append(tecnologia)
        if agrupacion:
            query += " AND categoria = %s"
            params.append(agrupacion)
        if grupo:
            query += " AND nombre = %s"
            params.append(grupo)

        query += " ORDER BY codigo ASC"
        cursor.execute(query, params)
        resultados = cursor.fetchall()
        return jsonify(resultados)

    except mysql.connector.Error as e:
        logging.error(f"Error en API codigos: {str(e)}")
        return jsonify({'error': 'Error al consultar la base de datos'}), 500
    except Exception as e:
        logging.error(f"Error inesperado en API codigos: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

@app.route('/api/analistas/codigos/grupos', methods=['GET'])
@login_required()
def api_grupos_codigos_facturacion():
    """Lista única de grupos en códigos de facturación"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
        cursor = connection.cursor()

        tecnologia = request.args.get('tecnologia', '').strip()
        agrupacion = request.args.get('agrupacion', '').strip()

        query = """
            SELECT DISTINCT nombre
            FROM base_codigos_facturacion
            WHERE nombre IS NOT NULL 
              AND nombre != ''
        """
        params = []
        if tecnologia:
            query += " AND tecnologia = %s"
            params.append(tecnologia)
        if agrupacion:
            query += " AND categoria = %s"
            params.append(agrupacion)
        query += " ORDER BY nombre ASC"

        cursor.execute(query, params) if params else cursor.execute(query)
        resultados = cursor.fetchall()
        grupos = [row[0] for row in resultados]
        return jsonify(grupos)

    except mysql.connector.Error as e:
        logging.error(f"Error en API grupos codigos: {str(e)}")
        return jsonify({'error': 'Error al consultar la base de datos'}), 500
    except Exception as e:
        logging.error(f"Error inesperado en API grupos codigos: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

@app.route('/api/analistas/codigos/tecnologias', methods=['GET'])
@login_required()
def api_tecnologias_codigos_facturacion():
    """Lista única de tecnologías en códigos de facturación"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
        cursor = connection.cursor()

        query = """
            SELECT DISTINCT tecnologia
            FROM base_codigos_facturacion
            WHERE tecnologia IS NOT NULL 
              AND tecnologia != ''
            ORDER BY tecnologia ASC
        """
        cursor.execute(query)
        resultados = cursor.fetchall()
        tecnologias = [row[0] for row in resultados]
        return jsonify(tecnologias)

    except mysql.connector.Error as e:
        logging.error(f"Error en API tecnologias codigos: {str(e)}")
        return jsonify({'error': 'Error al consultar la base de datos'}), 500
    except Exception as e:
        logging.error(f"Error inesperado en API tecnologias codigos: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

@app.route('/api/analistas/codigos/agrupaciones', methods=['GET'])
@login_required()
def api_agrupaciones_codigos_facturacion():
    """Lista única de agrupaciones en códigos de facturación, opcional por tecnología"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
        cursor = connection.cursor()

        tecnologia = request.args.get('tecnologia', '').strip()
        if tecnologia:
            query = """
                SELECT DISTINCT categoria
                FROM base_codigos_facturacion
                WHERE categoria IS NOT NULL 
                  AND categoria != ''
                  AND tecnologia = %s
                ORDER BY categoria ASC
            """
            cursor.execute(query, (tecnologia,))
        else:
            query = """
                SELECT DISTINCT categoria
                FROM base_codigos_facturacion
                WHERE categoria IS NOT NULL 
                  AND categoria != ''
                ORDER BY categoria ASC
            """
            cursor.execute(query)

        resultados = cursor.fetchall()
        agrupaciones = [row[0] for row in resultados]
        return jsonify(agrupaciones)

    except mysql.connector.Error as e:
        logging.error(f"Error en API agrupaciones codigos: {str(e)}")
        return jsonify({'error': 'Error al consultar la base de datos'}), 500
    except Exception as e:
        logging.error(f"Error inesperado en API agrupaciones codigos: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

@app.route('/tecnicos')
@login_required(role='tecnicos')
def tecnicos_dashboard():
    try:
        connection = get_db_connection()
        if connection is None:
            flash('Error de conexión a la base de datos', 'danger')
            resp = make_response(render_template('modulos/tecnicos/dashboard.html', supervisor=None, tiene_asistencia=False))
            resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
            resp.headers['Pragma'] = 'no-cache'
            resp.headers['Expires'] = '0'
            return resp
            
        cursor = connection.cursor(dictionary=True, buffered=True)
        
        # Obtener el supervisor del técnico logueado
        cursor.execute("""
            SELECT super FROM capired.recurso_operativo
            WHERE id_codigo_consumidor = %s
        """, (session['id_codigo_consumidor'],))
        
        supervisor_result = cursor.fetchone()
        supervisor_tecnico = supervisor_result['super'] if supervisor_result and supervisor_result['super'] else None
        
        # Obtener la cédula del usuario logueado para verificar asistencia
        cursor.execute("""
            SELECT recurso_operativo_cedula 
            FROM capired.recurso_operativo 
            WHERE id_codigo_consumidor = %s
        """, (session['id_codigo_consumidor'],))
        
        usuario_actual = cursor.fetchone()
        
        # Verificar si es el usuario especial (52912112) que debe estar exento de la restricción
        if usuario_actual and usuario_actual['recurso_operativo_cedula'] == '52912112':
            # Usuario especial: siempre tiene acceso a todos los botones
            tiene_asistencia = True
        else:
            # Para todos los demás usuarios: verificar asistencia registrada para hoy Y que carpeta_dia no sea 0
            fecha_hoy = datetime.now().strftime('%Y-%m-%d')
            cursor.execute("""
                SELECT COUNT(*) as registros_hoy
                FROM asistencia 
                WHERE id_codigo_consumidor = %s AND DATE(fecha_asistencia) = %s AND carpeta_dia != '0'
            """, (session['id_codigo_consumidor'], fecha_hoy))
            
            registro_existente = cursor.fetchone()
            tiene_asistencia = registro_existente['registros_hoy'] > 0 if registro_existente else False
        
        resp = make_response(render_template('modulos/tecnicos/dashboard.html', supervisor=supervisor_tecnico, tiene_asistencia=tiene_asistencia))
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        resp.headers['Pragma'] = 'no-cache'
        resp.headers['Expires'] = '0'
        return resp
                           
    except mysql.connector.Error as e:
        flash(f'Error al cargar datos del supervisor: {str(e)}', 'warning')
        resp = make_response(render_template('modulos/tecnicos/dashboard.html', supervisor=None, tiene_asistencia=False))
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        resp.headers['Pragma'] = 'no-cache'
        resp.headers['Expires'] = '0'
        return resp
    finally:
        if cursor:
            try:
                # Limpiar cualquier resultado no leído
                while cursor.nextset():
                    pass
            except:
                pass
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/lider/metas-indicadores', methods=['GET'])
@login_required_lider_api()
def api_get_metas_indicadores():
    connection = None
    cursor = None
    try:
        area = request.args.get('area', '').strip()
        periodo_str = request.args.get('periodo')
        from datetime import datetime
        hoy = datetime.now()
        if not periodo_str:
            periodo_str = f"{hoy.year}-{str(hoy.month).zfill(2)}"
        try:
            anio, mes = periodo_str.split('-')
            periodo_fecha = datetime(int(anio), int(mes), 1).strftime('%Y-%m-%d')
        except Exception:
            return jsonify({'error': 'Periodo inválido'}), 400
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS metas_indicadores (
                id INT AUTO_INCREMENT PRIMARY KEY,
                area VARCHAR(50) NOT NULL,
                periodo DATE NOT NULL,
                meta_porcentaje DECIMAL(5,2) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY uniq_area_periodo (area, periodo)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        if not area:
            return jsonify({'error': 'Área requerida'}), 400
        cursor.execute(
            "SELECT meta_porcentaje FROM metas_indicadores WHERE area=%s AND periodo=%s",
            (area, periodo_fecha)
        )
        row = cursor.fetchone()
        meta = row['meta_porcentaje'] if row else None
        try:
            meta = float(meta) if meta is not None else None
        except Exception:
            meta = None
        return jsonify({
            'success': True,
            'area': area,
            'periodo': periodo_str,
            'meta_porcentaje': meta
        })
    except mysql.connector.Error as e:
        return jsonify({'error': f'Error de base de datos: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'error': f'Error interno: {str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/lider/metas-indicadores', methods=['POST'])
@login_required_lider_api()
def api_set_metas_indicadores():
    connection = None
    cursor = None
    try:
        data = request.get_json() or {}
        area = (data.get('area') or '').strip()
        periodo_str = data.get('periodo')
        meta_val = data.get('meta_porcentaje')
        if not area or not periodo_str or meta_val is None:
            return jsonify({'error': 'Datos incompletos'}), 400
        try:
            meta_num = float(meta_val)
        except Exception:
            return jsonify({'error': 'Meta inválida'}), 400
        if meta_num < 0 or meta_num > 100:
            return jsonify({'error': 'Meta fuera de rango'}), 400
        from datetime import datetime
        try:
            anio, mes = periodo_str.split('-')
            periodo_fecha = datetime(int(anio), int(mes), 1).strftime('%Y-%m-%d')
        except Exception:
            return jsonify({'error': 'Periodo inválido'}), 400
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
        cursor = connection.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS metas_indicadores (
                id INT AUTO_INCREMENT PRIMARY KEY,
                area VARCHAR(50) NOT NULL,
                periodo DATE NOT NULL,
                meta_porcentaje DECIMAL(5,2) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE KEY uniq_area_periodo (area, periodo)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        cursor.execute(
            """
            INSERT INTO metas_indicadores (area, periodo, meta_porcentaje)
            VALUES (%s, %s, %s)
            ON DUPLICATE KEY UPDATE meta_porcentaje = VALUES(meta_porcentaje)
            """,
            (area, periodo_fecha, meta_num)
        )
        connection.commit()
        return jsonify({'success': True})
    except mysql.connector.Error as e:
        if connection:
            connection.rollback()
        return jsonify({'error': f'Error de base de datos: {str(e)}'}), 500
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': f'Error interno: {str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()
@app.route('/tecnicos/asignaciones_ferretero')
@login_required(role='tecnicos')
def tecnicos_asignaciones_ferretero():
    try:
        connection = get_db_connection()
        if connection is None:
            flash('Error de conexión a la base de datos', 'danger')
            return render_template('modulos/tecnicos/asignaciones_ferretero.html', asignaciones=[])
            
        cursor = connection.cursor(dictionary=True, buffered=True)
        
        # Obtener las asignaciones de ferretero del técnico logueado
        cursor.execute("""
            SELECT 
                id_ferretero,
                fecha_asignacion,
                silicona,
                amarres_negros,
                amarres_blancos,
                cinta_aislante,
                grapas_negras,
                grapas_blancas,
                id_codigo_consumidor
            FROM ferretero 
            WHERE id_codigo_consumidor = %s
            ORDER BY fecha_asignacion DESC
        """, (session['id_codigo_consumidor'],))
        
        asignaciones = cursor.fetchall()
        
        # Formatear las fechas para mejor visualización
        for asignacion in asignaciones:
            if asignacion['fecha_asignacion']:
                asignacion['fecha_formateada'] = asignacion['fecha_asignacion'].strftime('%d/%m/%Y %H:%M')
            else:
                asignacion['fecha_formateada'] = 'N/A'
        
        return render_template('modulos/tecnicos/asignaciones_ferretero.html', asignaciones=asignaciones)
        
    except mysql.connector.Error as e:
        flash(f'Error al cargar asignaciones: {str(e)}', 'danger')
        return render_template('modulos/tecnicos/asignaciones_ferretero.html', asignaciones=[])
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/tecnicos/ordenes_trabajo')
@login_required(role='tecnicos')
def tecnicos_ordenes_trabajo():
    try:
        return render_template('modulos/tecnicos/ordenes_trabajo.html')
    except Exception as e:
        flash(f'Error al cargar Órdenes de Trabajo: {str(e)}', 'danger')
        return render_template('modulos/tecnicos/ordenes_trabajo.html')

@app.route('/tecnicos/ordenes_trabajo_preview')
@login_required(role='tecnicos')
def tecnicos_ordenes_trabajo_preview():
    try:
        return render_template('modulos/tecnicos/ordenes_trabajo.html')
    except Exception as e:
        flash(f'Error al cargar Órdenes de Trabajo (preview): {str(e)}', 'danger')
        return render_template('modulos/tecnicos/ordenes_trabajo.html')

# ==================== APIs ÓRDENES DE TRABAJO - TÉCNICOS ====================

# Listar órdenes del técnico logueado
@app.route('/tecnicos/ordenes', methods=['GET'])
@app.route('/api/tecnicos/ordenes', methods=['GET'])
@login_required_api(role='tecnicos')
def api_tecnicos_list_ordenes():
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'error': 'Error de conexión a la base de datos'}), 500
        cursor = connection.cursor(dictionary=True)

        user_id = session.get('user_id') or session.get('id_codigo_consumidor')
        if not user_id:
            return jsonify({'success': False, 'error': 'Usuario no autenticado'}), 401

        # Obtener órdenes desde tabla principal
        cursor.execute(
            """
            SELECT 
                ot,
                MIN(cuenta) AS cuenta,
                MIN(servicio) AS servicio,
                MIN(tecnologia) AS tecnologia,
                MIN(agrupacion) AS categoria,
                SUM(COALESCE(valor, 0) * COALESCE(cantidad, 1)) AS total_valor,
                MIN(fecha_creacion) AS fecha_creacion,
                MAX(fecha_creacion) AS fecha_actualizacion
            FROM gestion_ot_tecnicos
            WHERE id_codigo_consumidor = %s
            GROUP BY ot
            ORDER BY fecha_actualizacion DESC
            """,
            (user_id,)
        )
        rows = cursor.fetchall()

        # Obtener cantidad de códigos por OT desde historial
        ots = [row['ot'] for row in rows if row.get('ot') is not None]
        codigos_por_ot = {}
        if ots:
            format_strings = ','.join(['%s'] * len(ots))
            cursor.execute(
                f"""
                SELECT ot, COUNT(*) as cantidad_codigos
                FROM gestion_ot_tecnicos_historial
                WHERE ot IN ({format_strings})
                GROUP BY ot
                """,
                ots
            )
            codigos_por_ot = {row['ot']: row['cantidad_codigos'] for row in cursor.fetchall()}

        ordenes = []
        for r in rows:
            ordenes.append({
                'id': r.get('id'),
                'ot': r.get('ot'),
                'cuenta': r.get('cuenta'),
                'servicio': r.get('servicio'),
                'tecnologia': r.get('tecnologia'),
                'categoria': r.get('categoria'),
                'tecnico_nombre': r.get('tecnico_nombre'),
                'total_valor': float(r.get('total_valor', 0)),
                'cantidad_codigos': codigos_por_ot.get(r.get('ot'), 0),
                'fecha_creacion': r.get('fecha_creacion').strftime('%Y-%m-%d %H:%M:%S') if r.get('fecha_creacion') else None,
                'fecha_actualizacion': r.get('fecha_actualizacion').strftime('%Y-%m-%d %H:%M:%S') if r.get('fecha_actualizacion') else None
            })

        return jsonify({'success': True, 'ordenes': ordenes})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        try:
            if cursor:
                cursor.close()
            if connection and connection.is_connected():
                connection.close()
        except:
            pass

# Ver detalle de una orden por OT desde tabla historial
@app.route('/tecnicos/ordenes/detalle/<ot>', methods=['GET'])
@app.route('/api/tecnicos/ordenes/detalle/<ot>', methods=['GET'])
@login_required_api(role='tecnicos')
def api_tecnicos_get_orden_detalle(ot):
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'error': 'Error de conexión a la base de datos'}), 500
        cursor = connection.cursor(dictionary=True)

        user_id = session.get('user_id') or session.get('id_codigo_consumidor')
        if not user_id:
            return jsonify({'success': False, 'error': 'Usuario no autenticado'}), 401

        # Obtener información general de la OT desde tabla principal
        cursor.execute(
            """
            SELECT 
                MIN(id_gestion_ot_tecnicos) AS id,
                ot,
                MIN(cuenta) AS cuenta,
                MIN(servicio) AS servicio,
                MIN(tecnologia) AS tecnologia,
                MIN(agrupacion) AS categoria,
                SUM(COALESCE(valor, 0) * COALESCE(cantidad, 1)) AS total_valor,
                MIN(fecha_creacion) AS fecha_creacion,
                MAX(fecha_creacion) AS fecha_actualizacion
            FROM gestion_ot_tecnicos
            WHERE ot = %s AND id_codigo_consumidor = %s
            GROUP BY ot
            """,
            (ot, user_id)
        )
        ot_info = cursor.fetchone()
        if not ot_info:
            return jsonify({'success': False, 'error': 'OT no encontrada'}), 404

        # Obtener todos los códigos de la OT desde tabla historial
        cursor.execute(
            """
            SELECT 
                id_gestion_ot_tecnicos_historial AS id,
                codigo,
                nombre,
                descripcion,
                cantidad,
                valor AS valor_unitario,
                (COALESCE(valor, 0) * COALESCE(cantidad, 1)) AS valor_total,
                fecha_creacion
            FROM gestion_ot_tecnicos_historial
            WHERE ot = %s AND id_codigo_consumidor = %s
            ORDER BY fecha_creacion ASC
            """,
            (ot, user_id)
        )
        codigos = cursor.fetchall()

        # Obtener nombre del técnico desde recurso_operativo
        cursor.execute(
            """
            SELECT nombre FROM recurso_operativo 
            WHERE id_codigo_consumidor = %s
            """,
            (user_id,)
        )
        tecnico_res = cursor.fetchone()
        tecnico_nombre = (tecnico_res.get('nombre') if isinstance(tecnico_res, dict) else (tecnico_res[0] if tecnico_res else None)) or 'Técnico'

        orden_detalle = {
            'id': ot_info.get('id'),
            'ot': ot_info['ot'],
            'cuenta': ot_info['cuenta'],
            'servicio': ot_info['servicio'],
            'tecnologia': ot_info['tecnologia'],
            'categoria': ot_info['categoria'],
            'tecnico_nombre': tecnico_nombre,
            'total_valor': float(ot_info['total_valor']),
            'fecha_creacion': ot_info['fecha_creacion'].strftime('%Y-%m-%d %H:%M:%S') if ot_info['fecha_creacion'] else None,
            'fecha_actualizacion': ot_info['fecha_actualizacion'].strftime('%Y-%m-%d %H:%M:%S') if ot_info['fecha_actualizacion'] else None,
            'codigos': [{
                'id': c['id'],
                'codigo': c['codigo'],
                'nombre': c['nombre'],
                'descripcion': c['descripcion'] or '',
                'cantidad': c['cantidad'],
                'valor_unitario': float(c['valor_unitario']),
                'valor_total': float(c['valor_total']),
                'fecha_creacion': c['fecha_creacion'].strftime('%Y-%m-%d %H:%M:%S') if c['fecha_creacion'] else None
            } for c in codigos]
        }

        return jsonify({'success': True, 'data': orden_detalle})
    except Exception as e:
        app.logger.error('[api_tecnicos_get_orden_detalle] error: %s', e, exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        try:
            if cursor:
                cursor.close()
            if connection and connection.is_connected():
                connection.close()
        except:
            pass

# Crear nueva orden (guarda resumen en principal y detalles en historial)
@app.route('/tecnicos/ordenes', methods=['POST'])
@app.route('/api/tecnicos/ordenes', methods=['POST'])
@login_required_api(role='tecnicos')
def api_tecnicos_create_orden():
    import traceback
    try:
        data = request.get_json() or {}
        app.logger.info('[api_tecnicos_create_orden] payload recibido: %s', data)
        required = ['ot', 'cuenta', 'servicio', 'tecnologia', 'codigos']
        for field in required:
            if field not in data or (field == 'codigos' and not isinstance(data['codigos'], list)):
                app.logger.warning('[api_tecnicos_create_orden] faltante/invalid: %s', field)
                return jsonify({'success': False, 'error': f'Campo requerido faltante o inválido: {field}'}), 400

        # Validaciones de tipos básicos
        try:
            ot = int(str(data.get('ot', '')).strip())
            cuenta = int(str(data.get('cuenta', '')).strip())
            servicio = int(str(data.get('servicio', '')).strip())
        except Exception:
            app.logger.error('[api_tecnicos_create_orden] tipos inválidos para ot/cuenta/servicio', exc_info=True)
            return jsonify({'success': False, 'error': 'OT/Cuenta/Servicio deben ser numéricos'}), 400

        tecnologia = data.get('tecnologia')
        if not tecnologia:
            return jsonify({'success': False, 'error': 'Tecnología requerida'}), 400

        connection = get_db_connection()
        if connection is None:
            app.logger.error('[api_tecnicos_create_orden] conexión DB es None')
            return jsonify({'success': False, 'error': 'Error de conexión a la base de datos'}), 500
        cursor = connection.cursor()

        user_id = session.get('user_id') or session.get('id_codigo_consumidor')
        if not user_id:
            app.logger.error('[api_tecnicos_create_orden] usuario no autenticado en sesión')
            return jsonify({'success': False, 'error': 'Usuario no autenticado'}), 401

        categoria = data.get('categoria') or None
        observacion = data.get('observacion') or None
        now = datetime.now()

        # Validar unicidad de OT por usuario: si ya existe y pertenece a otro usuario, bloquear
        try:
            cursor.execute(
                """
                SELECT id_codigo_consumidor, nombre, fecha_creacion
                FROM gestion_ot_tecnicos
                WHERE ot = %s
                ORDER BY fecha_creacion DESC
                LIMIT 1
                """,
                (ot,)
            )
            ot_existente = cursor.fetchone()
        except Exception:
            app.logger.error('[api_tecnicos_create_orden] Error consultando existencia de OT', exc_info=True)
            ot_existente = None

        if ot_existente:
            existing_user_id = ot_existente[0]
            if int(existing_user_id) != int(user_id):
                existing_user_name = None
                try:
                    existing_user_name = ot_existente[1]
                except Exception:
                    existing_user_name = None
                app.logger.warning('[api_tecnicos_create_orden] OT %s ya creada por otro técnico (owner_id=%s)', ot, existing_user_id)
                return jsonify({
                    'success': False,
                    'error': 'OT ya creada por otro técnico',
                    'message': 'La OT ya se encuentra registrada por otro usuario.',
                    'owner_id': int(existing_user_id),
                    'owner_name': existing_user_name
                }), 409
            else:
                app.logger.info('[api_tecnicos_create_orden] La OT %s ya pertenece al usuario actual (%s); se permiten agregar más códigos', ot, user_id)

        # Obtener nombre del técnico
        cursor.execute(
            """
            SELECT nombre FROM recurso_operativo 
            WHERE id_codigo_consumidor = %s
            """,
            (user_id,)
        )
        tecnico_result = cursor.fetchone()
        tecnico_nombre = tecnico_result[0] if tecnico_result else 'Técnico'

        # Calcular total valor de todos los códigos
        total_valor = 0
        for item in data['codigos']:
            cantidad = int(item.get('cantidad') or 1)
            valor_unitario_raw = item.get('valor_unitario') or item.get('valor') or 0
            try:
                valor_unitario = int(float(valor_unitario_raw))
            except Exception:
                valor_unitario = 0
            total_valor += (valor_unitario or 0) * (cantidad or 1)

        # Insertar cada código como fila principal en gestion_ot_tecnicos (estructura real)
        insert_sql_principal = (
            """
            INSERT INTO gestion_ot_tecnicos (
                ot, cuenta, servicio, codigo, tecnologia, agrupacion,
                nombre, observacion, fecha_creacion, id_codigo_consumidor,
                cantidad, valor, descripcion
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
        )
        # Insertar también en historial (usa `ot`, no `ot_id`)
        insert_sql_historial = (
            """
            INSERT INTO gestion_ot_tecnicos_historial (
                id_codigo_consumidor, ot, cuenta, servicio, codigo, tecnologia, agrupacion,
                nombre, observacion, fecha_creacion, cantidad, valor, descripcion
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
        )
        # Fallback con ID explícito si la columna no es AUTO_INCREMENT
        insert_sql_historial_with_id = (
            """
            INSERT INTO gestion_ot_tecnicos_historial (
                id_gestion_ot_tecnicos_historial, id_codigo_consumidor, ot, cuenta, servicio, codigo, tecnologia, agrupacion,
                nombre, observacion, fecha_creacion, cantidad, valor, descripcion
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
        )
        # Intentar asegurar que el ID del historial sea AUTO_INCREMENT
        try:
            cursor.execute("SHOW COLUMNS FROM gestion_ot_tecnicos_historial LIKE 'id_gestion_ot_tecnicos_historial'")
            col = cursor.fetchone()
            # col formato: Field, Type, Null, Key, Default, Extra
            if col and len(col) >= 6 and (not col[5] or 'auto_increment' not in str(col[5]).lower()):
                app.logger.warning('[api_tecnicos_create_orden] id_gestion_ot_tecnicos_historial no es AUTO_INCREMENT, intentando corregir esquema')
                try:
                    cursor.execute("""
                        ALTER TABLE gestion_ot_tecnicos_historial 
                        MODIFY id_gestion_ot_tecnicos_historial INT NOT NULL AUTO_INCREMENT
                    """)
                    app.logger.info('[api_tecnicos_create_orden] Esquema corregido: id_gestion_ot_tecnicos_historial ahora es AUTO_INCREMENT')
                except Exception as alter_err:
                    app.logger.error('[api_tecnicos_create_orden] No se pudo alterar la tabla historial: %s', str(alter_err))
        except Exception:
            app.logger.error('[api_tecnicos_create_orden] Falló verificación de esquema de historial', exc_info=True)

        created_ids = []
        for idx, item in enumerate(data['codigos']):
            app.logger.info('[api_tecnicos_create_orden] procesando item %s: %s', idx, item)

            # Normalización y truncado preventivo
            codigo_raw = item.get('codigo') or ''
            nombre_raw = item.get('nombre') or ''
            descripcion_raw = item.get('descripcion') or ''

            codigo = str(codigo_raw)[:50]
            nombre = str(nombre_raw or codigo_raw)[:45]
            descripcion = str(descripcion_raw or nombre_raw or codigo_raw)[:255]

            cantidad = int(item.get('cantidad') or 1)
            valor_unitario_raw = item.get('valor_unitario') or item.get('valor') or 0
            try:
                valor_unitario = int(float(valor_unitario_raw))
            except Exception:
                app.logger.warning('[api_tecnicos_create_orden] valor_unitario inválido: %s', valor_unitario_raw)
                valor_unitario = 0

            params_principal = (
                ot, cuenta, servicio, codigo, tecnologia, categoria,
                nombre, observacion, now, user_id,
                cantidad, valor_unitario, descripcion
            )
            app.logger.info('[api_tecnicos_create_orden] INSERT principal: ot=%s, codigo=%s, nombre=%s, cantidad=%s, valor=%s',
                            ot, codigo, nombre, cantidad, valor_unitario)
            cursor.execute(insert_sql_principal, params_principal)
            created_ids.append(cursor.lastrowid)

            params_historial = (
                user_id, ot, cuenta, servicio, codigo, tecnologia, categoria,
                nombre, observacion, now, cantidad, valor_unitario, descripcion
            )
            app.logger.info('[api_tecnicos_create_orden] INSERT historial: ot=%s, codigo=%s, nombre=%s, cantidad=%s, valor=%s',
                            ot, codigo, nombre, cantidad, valor_unitario)
            try:
                cursor.execute(insert_sql_historial, params_historial)
            except Exception as hist_err:
                err_msg = str(hist_err)
                if "doesn't have a default value" in err_msg and 'id_gestion_ot_tecnicos_historial' in err_msg:
                    app.logger.warning('[api_tecnicos_create_orden] Fallback: insertando historial con ID explícito por esquema sin default')
                    # Calcular siguiente ID manualmente
                    try:
                        cursor.execute("SELECT COALESCE(MAX(id_gestion_ot_tecnicos_historial), 0) + 1 FROM gestion_ot_tecnicos_historial")
                        next_id_row = cursor.fetchone()
                        next_id = next_id_row[0] if next_id_row else 1
                    except Exception:
                        app.logger.error('[api_tecnicos_create_orden] Error obteniendo next_id para historial', exc_info=True)
                        next_id = None
                    if next_id is None:
                        raise hist_err
                    params_historial_with_id = (
                        next_id, user_id, ot, cuenta, servicio, codigo, tecnologia, categoria,
                        nombre, observacion, now, cantidad, valor_unitario, descripcion
                    )
                    cursor.execute(insert_sql_historial_with_id, params_historial_with_id)
                else:
                    raise

        connection.commit()
        app.logger.info('[api_tecnicos_create_orden] Filas insertadas en gestion_ot_tecnicos=%s y historial=%s, total_valor=%s', 
                        len(created_ids), len(created_ids), total_valor)

        return jsonify({'success': True, 'ids': created_ids, 'valor_total': total_valor})
    except Exception as e:
        app.logger.error('[api_tecnicos_create_orden] excepción: %s\n%s', str(e), traceback.format_exc())
        if 'connection' in locals() and connection:
            try:
                connection.rollback()
            except Exception:
                app.logger.error('[api_tecnicos_create_orden] rollback falló', exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        try:
            if 'cursor' in locals() and cursor:
                cursor.close()
            if 'connection' in locals() and connection and connection.is_connected():
                connection.close()
        except Exception:
            app.logger.error('[api_tecnicos_create_orden] cierre de recursos falló', exc_info=True)
            pass

# Obtener lista de tecnologías disponibles
@app.route('/tecnicos/tecnologias', methods=['GET'])
@app.route('/api/tecnicos/tecnologias', methods=['GET'])
@login_required_api(role='tecnicos')
def api_tecnicos_list_tecnologias():
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'error': 'Error de conexión a la base de datos'}), 500
        cursor = connection.cursor()

        cursor.execute("""
            SELECT DISTINCT tecnologia 
            FROM base_codigos_facturacion 
            WHERE tecnologia IS NOT NULL AND tecnologia <> ''
            ORDER BY tecnologia
        """)
        rows = cursor.fetchall()
        tecnologias = [r[0] for r in rows]
        return jsonify({'success': True, 'tecnologias': tecnologias})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        try:
            if cursor:
                cursor.close()
            if connection and connection.is_connected():
                connection.close()
        except:
            pass

# Obtener categorías por tecnología
@app.route('/tecnicos/categorias/<string:tecnologia>', methods=['GET'])
@app.route('/api/tecnicos/categorias/<string:tecnologia>', methods=['GET'])
@login_required_api(role='tecnicos')
def api_tecnicos_list_categorias(tecnologia):
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'error': 'Error de conexión a la base de datos'}), 500
        cursor = connection.cursor()

        cursor.execute(
            """
            SELECT DISTINCT categoria 
            FROM base_codigos_facturacion 
            WHERE tecnologia = %s AND categoria IS NOT NULL AND categoria <> ''
            ORDER BY categoria
            """,
            (tecnologia,)
        )
        rows = cursor.fetchall()
        categorias = [r[0] for r in rows]
        return jsonify({'success': True, 'categorias': categorias})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        try:
            if cursor:
                cursor.close()
            if connection and connection.is_connected():
                connection.close()
        except:
            pass

# Obtener códigos filtrados por tecnología y categoría
@app.route('/tecnicos/codigos/tecnologia/<string:tecnologia>', methods=['GET'])
@app.route('/tecnicos/codigos/tecnologia/<string:tecnologia>/categoria/<string:categoria>', methods=['GET'])
@app.route('/api/tecnicos/codigos', methods=['GET'])
@login_required_api(role='tecnicos')
def api_tecnicos_list_codigos(tecnologia=None, categoria=None):
    try:
        # Permitir también /api/tecnicos/codigos?tecnologia=...&categoria=...
        if request.path.startswith('/api/tecnicos/codigos'):
            tecnologia = request.args.get('tecnologia', tecnologia)
            categoria = request.args.get('categoria', categoria)

        if not tecnologia:
            return jsonify({'success': False, 'error': 'Parámetro tecnologia requerido'}), 400

        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'error': 'Error de conexión a la base de datos'}), 500
        cursor = connection.cursor(dictionary=True)

        sql = (
            """
            SELECT 
                id_base_codigos_facturacion,
                tecnologia,
                categoria,
                codigo,
                nombre,
                descripcion,
                valor
            FROM base_codigos_facturacion
            WHERE tecnologia = %s
            """
        )
        params = [tecnologia]
        if categoria:
            sql += " AND categoria = %s"
            params.append(categoria)
        sql += " ORDER BY categoria, nombre"

        cursor.execute(sql, params)
        rows = cursor.fetchall()
        codigos = []
        for r in rows:
            codigos.append({
                'id': r.get('id_base_codigos_facturacion'),
                'tecnologia': r.get('tecnologia'),
                'categoria': r.get('categoria'),
                'codigo': r.get('codigo'),
                'nombre': r.get('nombre'),
                'descripcion': r.get('descripcion'),
                'valor': r.get('valor')
            })
        return jsonify({'success': True, 'codigos': codigos})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        try:
            if cursor:
                cursor.close()
            if connection and connection.is_connected():
                connection.close()
        except:
            pass

@app.route('/api/tecnicos/asignacion_ferretero/<int:id_ferretero>')
@login_required(role='tecnicos')
def obtener_detalle_asignacion_ferretero(id_ferretero):
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener el detalle completo de la asignación
        cursor.execute("""
            SELECT 
                f.id_ferretero,
                f.fecha_asignacion,
                f.silicona,
                f.amarres_negros,
                f.amarres_blancos,
                f.cinta_aislante,
                f.grapas_negras,
                f.grapas_blancas,
                f.id_codigo_consumidor,
                ro.nombre as tecnico_nombre,
                ro.recurso_operativo_cedula as tecnico_cedula
            FROM ferretero f
            LEFT JOIN capired.recurso_operativo ro ON f.id_codigo_consumidor = ro.id_codigo_consumidor
            WHERE f.id_ferretero = %s AND f.id_codigo_consumidor = %s
        """, (id_ferretero, session['id_codigo_consumidor']))
        
        asignacion = cursor.fetchone()
        
        if not asignacion:
            return jsonify({'error': 'Asignación no encontrada'}), 404
        
        # Formatear la fecha
        if asignacion['fecha_asignacion']:
            asignacion['fecha_formateada'] = asignacion['fecha_asignacion'].strftime('%d/%m/%Y %H:%M')
        else:
            asignacion['fecha_formateada'] = 'N/A'
        
        return jsonify(asignacion)
        
    except mysql.connector.Error as e:
        return jsonify({'error': f'Error al obtener detalle: {str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/tecnico/obtener_estado_materiales')
@login_required(role='tecnicos')
def obtener_estado_materiales():
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener información del técnico logueado
        cursor.execute("""
            SELECT cargo, carpeta 
            FROM capired.recurso_operativo 
            WHERE id_codigo_consumidor = %s
        """, (session['id_codigo_consumidor'],))
        
        tecnico_info = cursor.fetchone()
        if not tecnico_info:
            return jsonify({'error': 'Técnico no encontrado'}), 404
        
        # Definir límites por área para todos los materiales
        limites_por_area = {
            'INSTALACION': {
                'cinta_aislante': {'cantidad': 5, 'periodo': 15},
                'silicona': {'cantidad': 16, 'periodo': 7},
                'amarres_negros': {'cantidad': 50, 'periodo': 15},
                'amarres_blancos': {'cantidad': 50, 'periodo': 15},
                'grapas_blancas': {'cantidad': 200, 'periodo': 15},
                'grapas_negras': {'cantidad': 200, 'periodo': 15}
            },
            'POSTVENTA': {
                'cinta_aislante': {'cantidad': 3, 'periodo': 15},
                'silicona': {'cantidad': 12, 'periodo': 7},
                'amarres_negros': {'cantidad': 30, 'periodo': 15},
                'amarres_blancos': {'cantidad': 30, 'periodo': 15},
                'grapas_blancas': {'cantidad': 100, 'periodo': 15},
                'grapas_negras': {'cantidad': 100, 'periodo': 15}
            },
            'MANTENIMIENTO': {
                'cinta_aislante': {'cantidad': 4, 'periodo': 15},
                'silicona': {'cantidad': 14, 'periodo': 7},
                'amarres_negros': {'cantidad': 40, 'periodo': 15},
                'amarres_blancos': {'cantidad': 40, 'periodo': 15},
                'grapas_blancas': {'cantidad': 150, 'periodo': 15},
                'grapas_negras': {'cantidad': 150, 'periodo': 15}
            },
            'SUPERVISION': {
                'cinta_aislante': {'cantidad': 2, 'periodo': 15},
                'silicona': {'cantidad': 8, 'periodo': 7},
                'amarres_negros': {'cantidad': 20, 'periodo': 15},
                'amarres_blancos': {'cantidad': 20, 'periodo': 15},
                'grapas_blancas': {'cantidad': 50, 'periodo': 15},
                'grapas_negras': {'cantidad': 50, 'periodo': 15}
            },
            'FTTH INSTALACIONES': {
                'cinta_aislante': {'cantidad': 5, 'periodo': 15},
                'silicona': {'cantidad': 16, 'periodo': 7},
                'amarres_negros': {'cantidad': 50, 'periodo': 15},
                'amarres_blancos': {'cantidad': 50, 'periodo': 15},
                'grapas_blancas': {'cantidad': 200, 'periodo': 15},
                'grapas_negras': {'cantidad': 200, 'periodo': 15}
            },
            'CONDUCTOR': {
                'cinta_aislante': {'cantidad': 99, 'periodo': 15, 'unidad': 'días'},
                'silicona': {'cantidad': 99, 'periodo': 7, 'unidad': 'días'},
                'amarres_negros': {'cantidad': 99, 'periodo': 15, 'unidad': 'días'},
                'amarres_blancos': {'cantidad': 99, 'periodo': 15, 'unidad': 'días'},
                'grapas_blancas': {'cantidad': 99, 'periodo': 7, 'unidad': 'días'},
                'grapas_negras': {'cantidad': 99, 'periodo': 7, 'unidad': 'días'}
            },
            'SUPERVISORES': {
                'cinta_aislante': {'cantidad': 99, 'periodo': 15, 'unidad': 'días'},
                'silicona': {'cantidad': 99, 'periodo': 7, 'unidad': 'días'},
                'amarres_negros': {'cantidad': 99, 'periodo': 15, 'unidad': 'días'},
                'amarres_blancos': {'cantidad': 99, 'periodo': 15, 'unidad': 'días'},
                'grapas_blancas': {'cantidad': 99, 'periodo': 7, 'unidad': 'días'},
                'grapas_negras': {'cantidad': 99, 'periodo': 7, 'unidad': 'días'}
            },
            'BROWNFIELD': {
                'cinta_aislante': {'cantidad': 5, 'periodo': 15, 'unidad': 'días'},
                'silicona': {'cantidad': 16, 'periodo': 7, 'unidad': 'días'},
                'amarres_negros': {'cantidad': 50, 'periodo': 15, 'unidad': 'días'},
                'amarres_blancos': {'cantidad': 50, 'periodo': 15, 'unidad': 'días'},
                'grapas_blancas': {'cantidad': 200, 'periodo': 15, 'unidad': 'días'},
                'grapas_negras': {'cantidad': 200, 'periodo': 15, 'unidad': 'días'}
            },
        }
        
        # Determinar el área de trabajo basado en el cargo y carpeta
        cargo = tecnico_info.get('cargo', '').upper()
        carpeta = tecnico_info.get('carpeta', '').upper()
        
        # Mapeo mejorado considerando cargo y carpeta
        if 'FTTH INSTALACIONES' in cargo:
            area_trabajo = 'FTTH INSTALACIONES'
        elif 'INSTALACION' in cargo or ('FTTH' in cargo and 'INSTALACION' in carpeta):
            area_trabajo = 'INSTALACION'
        elif 'POSTVENTA' in cargo or 'POSTVENTA' in carpeta:
            area_trabajo = 'POSTVENTA'
        elif 'MANTENIMIENTO' in cargo or 'ARREGLOS' in carpeta or 'MANTENIMIENTO' in carpeta:
            area_trabajo = 'MANTENIMIENTO'
        elif 'SUPERVISION' in cargo or 'SUPERVISOR' in cargo:
            area_trabajo = 'SUPERVISION'
        elif 'TECNICO' in cargo and 'ARREGLOS' in carpeta:
            # Caso específico: técnicos con carpeta de arreglos van a mantenimiento
            area_trabajo = 'MANTENIMIENTO'
        else:
            area_trabajo = 'INSTALACION'  # Default
        
        limite_config = limites_por_area[area_trabajo]
        
        # Calcular estado para cada material
        from datetime import datetime, timedelta
        materiales_estado = {}
        
        for material, config in limite_config.items():
            limite_total = config['cantidad']
            periodo_dias = config['periodo']
            
            # Calcular fecha límite para el período
            fecha_limite = datetime.now() - timedelta(days=periodo_dias)
            
            # Obtener material asignado en el período
            cursor.execute(f"""
                SELECT COALESCE(SUM(CAST({material} AS UNSIGNED)), 0) as total_asignadas
                FROM ferretero 
                WHERE id_codigo_consumidor = %s 
                AND fecha_asignacion >= %s
            """, (session['id_codigo_consumidor'], fecha_limite))
            
            resultado = cursor.fetchone()
            asignadas = resultado['total_asignadas'] if resultado else 0
            
            # Calcular disponible
            disponible = max(0, limite_total - asignadas)
            
            materiales_estado[material] = {
                'asignadas': asignadas,
                'disponible': disponible,
                'limite': limite_total,
                'periodo_dias': periodo_dias
            }
        
        # Retornar los datos en el formato que espera el frontend
        response_data = materiales_estado.copy()
        response_data['area_trabajo'] = area_trabajo
        
        return jsonify(response_data)
        
    except mysql.connector.Error as e:
        return jsonify({'error': f'Error al obtener estado de materiales: {str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

# Nuevos endpoints para el sistema preoperacional mejorado
@app.route('/api/tecnicos/datos-preoperacional', methods=['GET'])
@login_required(role='tecnicos')
def obtener_datos_preoperacional():
    """
    Endpoint para obtener todos los datos del vehículo y licencia del técnico logueado
    """
    try:
        print(f"Iniciando datos_preoperacional para sesión: {session.get('id_codigo_consumidor')}")
        
        connection = get_db_connection()
        if connection is None:
            print("Error: No se pudo conectar a la base de datos")
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True, buffered=True)
        
        # Obtener cédula del técnico logueado
        cursor.execute("""
            SELECT recurso_operativo_cedula 
            FROM capired.recurso_operativo 
            WHERE id_codigo_consumidor = %s
        """, (session['id_codigo_consumidor'],))
        
        usuario_actual = cursor.fetchone()
        if not usuario_actual:
            print(f"Usuario no encontrado para id_codigo_consumidor: {session['id_codigo_consumidor']}")
            return jsonify({'success': False, 'message': 'Usuario no encontrado'}), 404
        
        cedula_tecnico = usuario_actual['recurso_operativo_cedula']
        print(f"Cedula del técnico: {cedula_tecnico}")
        
        # Consulta principal para obtener todos los datos usando id_codigo_consumidor
        cursor.execute("""
            SELECT 
                ro.ciudad,
                mv.placa,
                mv.modelo,
                mv.marca,
                mv.tipo_vehiculo,
                mlc.tipo_licencia,
                mlc.fecha_vencimiento as fecha_venc_licencia,
                ms.fecha_vencimiento as fecha_venc_soat,
                mtm.fecha_vencimiento as fecha_venc_tecnico_mecanica,
                0 as ultimo_kilometraje,
                NULL as fecha_ultimo_kilometraje
            FROM capired.recurso_operativo ro
            LEFT JOIN capired.mpa_vehiculos mv ON ro.id_codigo_consumidor = mv.tecnico_asignado
            LEFT JOIN capired.mpa_licencia_conducir mlc ON ro.id_codigo_consumidor = mlc.tecnico
            LEFT JOIN capired.mpa_soat ms ON mv.placa = ms.placa 
                AND ms.estado = 'Activo'
            LEFT JOIN capired.mpa_tecnico_mecanica mtm ON mv.placa = mtm.placa 
                AND mtm.estado = 'Activo'
            WHERE ro.id_codigo_consumidor = %s
        """, (session['id_codigo_consumidor'],))
        
        datos = cursor.fetchone()
        print(f"Datos obtenidos: {datos}")
        
        if not datos:
            print("No se encontraron datos para el técnico")
            return jsonify({'success': False, 'message': 'No se encontraron datos para el técnico'}), 404
        
        # Calcular alertas de documentos próximos a vencer
        alertas = []
        fecha_actual = datetime.now().date()
        
        # Verificar licencia de conducción
        if datos['fecha_venc_licencia']:
            fecha_venc_licencia = datos['fecha_venc_licencia'].date() if hasattr(datos['fecha_venc_licencia'], 'date') else datos['fecha_venc_licencia']
            dias_licencia = (fecha_venc_licencia - fecha_actual).days
            if dias_licencia < 0:
                alertas.append({
                    'tipo': 'error',
                    'mensaje': f'Licencia de conducción vencida desde el {datos["fecha_venc_licencia"].strftime("%d/%m/%Y")}'
                })
            elif dias_licencia <= 30:
                alertas.append({
                    'tipo': 'warning',
                    'mensaje': f'Licencia de conducción próxima a vencer en {dias_licencia} días'
                })
        
        # Verificar SOAT
        if datos['fecha_venc_soat']:
            fecha_venc_soat = datos['fecha_venc_soat'].date() if hasattr(datos['fecha_venc_soat'], 'date') else datos['fecha_venc_soat']
            dias_soat = (fecha_venc_soat - fecha_actual).days
            if dias_soat < 0:
                alertas.append({
                    'tipo': 'error',
                    'mensaje': f'SOAT vencido desde el {datos["fecha_venc_soat"].strftime("%d/%m/%Y")}'
                })
            elif dias_soat <= 30:
                alertas.append({
                    'tipo': 'warning',
                    'mensaje': f'SOAT próximo a vencer en {dias_soat} días'
                })
        
        # Verificar Técnico Mecánica
        if datos['fecha_venc_tecnico_mecanica']:
            fecha_venc_tm = datos['fecha_venc_tecnico_mecanica'].date() if hasattr(datos['fecha_venc_tecnico_mecanica'], 'date') else datos['fecha_venc_tecnico_mecanica']
            dias_tm = (fecha_venc_tm - fecha_actual).days
            if dias_tm < 0:
                alertas.append({
                    'tipo': 'error',
                    'mensaje': f'Técnico mecánica vencida desde el {datos["fecha_venc_tecnico_mecanica"].strftime("%d/%m/%Y")}'
                })
            elif dias_tm <= 30:
                alertas.append({
                    'tipo': 'warning',
                    'mensaje': f'Técnico mecánica próxima a vencer en {dias_tm} días'
                })
        
        # Formatear fechas para el frontend
        response_data = {
            'ciudad': datos['ciudad'] or '',
            'placa': datos['placa'] or '',
            'modelo': datos['modelo'] or '',
            'marca': datos['marca'] or '',
            'tipo_vehiculo': datos['tipo_vehiculo'] or '',
            'tipo_licencia': datos['tipo_licencia'] or '',
            'fecha_venc_licencia': datos['fecha_venc_licencia'].strftime('%Y-%m-%d') if datos['fecha_venc_licencia'] else '',
            'fecha_venc_soat': datos['fecha_venc_soat'].strftime('%Y-%m-%d') if datos['fecha_venc_soat'] else '',
            'fecha_venc_tecnico_mecanica': datos['fecha_venc_tecnico_mecanica'].strftime('%Y-%m-%d') if datos['fecha_venc_tecnico_mecanica'] else '',
            'ultimo_kilometraje': datos['ultimo_kilometraje'],
            'fecha_ultimo_kilometraje': datos['fecha_ultimo_kilometraje'].strftime('%Y-%m-%d') if datos['fecha_ultimo_kilometraje'] else ''
        }
        
        return jsonify({
            'success': True,
            'data': response_data,
            'alertas': alertas
        })
        
    except mysql.connector.Error as e:
        print(f"Error de base de datos: {str(e)}")
        return jsonify({'success': False, 'message': f'Error de base de datos: {str(e)}'}), 500
    except Exception as e:
        print(f"Error interno: {str(e)}")
        return jsonify({'success': False, 'message': f'Error interno: {str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/tecnicos/validar-kilometraje', methods=['POST'])
@login_required(role='tecnicos')
def validar_kilometraje():
    """
    Endpoint para validar kilometraje en tiempo real
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400
        
        placa = data.get('placa')
        kilometraje_propuesto = data.get('kilometraje_propuesto')
        
        if not placa or kilometraje_propuesto is None:
            return jsonify({'success': False, 'message': 'Placa y kilometraje son requeridos'}), 400
        
        try:
            kilometraje_propuesto = int(kilometraje_propuesto)
        except ValueError:
            return jsonify({'success': False, 'message': 'El kilometraje debe ser un número entero'}), 400
        
        if kilometraje_propuesto < 0:
            return jsonify({'success': False, 'message': 'El kilometraje no puede ser negativo'}), 400
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener último kilometraje registrado para la placa (comparación normalizada)
        placa_norm_tecnico = (placa or '').strip().upper().replace('-', '').replace(' ', '')
        cursor.execute("""
            SELECT kilometraje_actual, fecha
            FROM capired.preoperacional 
            WHERE UPPER(REPLACE(REPLACE(TRIM(placa_vehiculo), '-', ''), ' ', '')) = %s 
            ORDER BY fecha DESC 
            LIMIT 1
        """, (placa_norm_tecnico,))
        
        ultimo_registro = cursor.fetchone()
        
        if not ultimo_registro:
            # Primer registro para este vehículo
            return jsonify({
                'success': True,
                'valido': True,
                'ultimo_kilometraje': 0,
                'fecha_ultimo_registro': '',
                'mensaje': 'Primer registro de kilometraje para este vehículo'
            })
        
        ultimo_kilometraje = ultimo_registro['kilometraje_actual']
        fecha_ultimo_registro = ultimo_registro['fecha']
        
        if kilometraje_propuesto < ultimo_kilometraje:
            return jsonify({
                'success': True,
                'valido': False,
                'ultimo_kilometraje': ultimo_kilometraje,
                'fecha_ultimo_registro': fecha_ultimo_registro.strftime('%d/%m/%Y') if fecha_ultimo_registro else '',
                'mensaje': f'El kilometraje no puede ser menor al último registrado: {ultimo_kilometraje} km en fecha {fecha_ultimo_registro.strftime("%d/%m/%Y") if fecha_ultimo_registro else "N/A"}'
            })

        # Nueva validación: límite máximo permitido
        if kilometraje_propuesto > 1000000:
            return jsonify({
                'success': True,
                'valido': False,
                'ultimo_kilometraje': ultimo_kilometraje,
                'fecha_ultimo_registro': fecha_ultimo_registro.strftime('%d/%m/%Y') if fecha_ultimo_registro else '',
                'mensaje': f'El kilometraje no puede superar los 1,000,000 km. Último kilometraje registrado: {ultimo_kilometraje} km'
            })
        
        return jsonify({
            'success': True,
            'valido': True,
            'ultimo_kilometraje': ultimo_kilometraje,
            'fecha_ultimo_registro': fecha_ultimo_registro.strftime('%d/%m/%Y') if fecha_ultimo_registro else '',
            'mensaje': 'Kilometraje válido'
        })
        
    except mysql.connector.Error as e:
        return jsonify({'success': False, 'message': f'Error de base de datos: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error interno: {str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

# ==================== MÓDULO MANTENIMIENTOS - TÉCNICOS ====================

@app.route('/tecnicos/mantenimientos')
@login_required(role='tecnicos')
def tecnicos_mantenimientos():
    """Módulo de mantenimientos para técnicos"""
    try:
        return render_template('modulos/tecnicos/mantenimientos.html')
    except Exception as e:
        flash(f'Error al cargar el módulo de mantenimientos: {str(e)}', 'danger')
        return render_template('modulos/tecnicos/mantenimientos.html')

# ==================== MÓDULO PREOPERACIONAL - OPERATIVOS ====================

@app.route('/api/operativo/datos-preoperacional', methods=['GET'])
@login_required(role=['operativo'])
def obtener_datos_preoperacional_operativo():
    """
    Endpoint para obtener datos del vehículo y licencia del operativo logueado
    """
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500

        cursor = connection.cursor(dictionary=True, buffered=True)

        # Obtener cédula del operativo logueado
        cursor.execute(
            """
            SELECT recurso_operativo_cedula 
            FROM capired.recurso_operativo 
            WHERE id_codigo_consumidor = %s
            """,
            (session['id_codigo_consumidor'],)
        )

        usuario_actual = cursor.fetchone()
        if not usuario_actual:
            return jsonify({'success': False, 'message': 'Usuario no encontrado'}), 404

        # Consulta principal de datos asociado al usuario
        cursor.execute(
            """
            SELECT 
                ro.ciudad,
                ro.super AS supervisor,
                mv.placa,
                mv.modelo,
                mv.marca,
                mv.tipo_vehiculo,
                -- Licencia activa más reciente del operativo
                (
                    SELECT sub.tipo_licencia
                    FROM capired.mpa_licencia_conducir sub
                    WHERE sub.tecnico = ro.id_codigo_consumidor
                      AND sub.fecha_vencimiento IS NOT NULL
                      AND sub.fecha_vencimiento > '1900-01-01'
                      AND (sub.estado IS NULL OR sub.estado = 'Activo')
                    ORDER BY sub.fecha_vencimiento DESC
                    LIMIT 1
                ) AS tipo_licencia,
                (
                    SELECT sub.fecha_vencimiento
                    FROM capired.mpa_licencia_conducir sub
                    WHERE sub.tecnico = ro.id_codigo_consumidor
                      AND sub.fecha_vencimiento IS NOT NULL
                      AND sub.fecha_vencimiento > '1900-01-01'
                      AND (sub.estado IS NULL OR sub.estado = 'Activo')
                    ORDER BY sub.fecha_vencimiento DESC
                    LIMIT 1
                ) AS fecha_venc_licencia,
                -- SOAT activo más reciente por placa del vehículo asignado
                (
                    SELECT s.fecha_vencimiento
                    FROM capired.mpa_soat s
                    WHERE s.placa = mv.placa
                      AND s.estado = 'Activo'
                      AND s.fecha_vencimiento IS NOT NULL
                      AND s.fecha_vencimiento > '1900-01-01'
                    ORDER BY s.fecha_vencimiento DESC
                    LIMIT 1
                ) AS fecha_venc_soat,
                -- Tecnomecánica activa más reciente por placa del vehículo asignado
                (
                    SELECT t.fecha_vencimiento
                    FROM capired.mpa_tecnico_mecanica t
                    WHERE t.placa = mv.placa
                      AND t.estado = 'Activo'
                      AND t.fecha_vencimiento IS NOT NULL
                      AND t.fecha_vencimiento > '1900-01-01'
                    ORDER BY t.fecha_vencimiento DESC
                    LIMIT 1
                ) AS fecha_venc_tecnico_mecanica,
                0 AS ultimo_kilometraje,
                NULL AS fecha_ultimo_kilometraje
            FROM capired.recurso_operativo ro
            LEFT JOIN capired.mpa_vehiculos mv ON ro.id_codigo_consumidor = mv.tecnico_asignado
            -- Se reemplazan joins directos por subconsultas para evitar múltiples resultados y fechas inválidas
            WHERE ro.id_codigo_consumidor = %s
            """,
            (session['id_codigo_consumidor'],)
        )

        datos = cursor.fetchone()
        if not datos:
            return jsonify({'success': False, 'message': 'No se encontraron datos para el usuario'}), 404

        # Calcular alertas de vencimientos
        alertas = []
        fecha_actual = datetime.now().date()

        if datos['fecha_venc_licencia']:
            fecha_venc_licencia = datos['fecha_venc_licencia'].date() if hasattr(datos['fecha_venc_licencia'], 'date') else datos['fecha_venc_licencia']
            dias_licencia = (fecha_venc_licencia - fecha_actual).days
            if dias_licencia < 0:
                alertas.append({'tipo': 'error', 'mensaje': f"Licencia de conducción vencida desde el {datos['fecha_venc_licencia'].strftime('%d/%m/%Y')}"})
            elif dias_licencia <= 30:
                alertas.append({'tipo': 'warning', 'mensaje': f'Licencia de conducción próxima a vencer en {dias_licencia} días'})

        if datos['fecha_venc_soat']:
            fecha_venc_soat = datos['fecha_venc_soat'].date() if hasattr(datos['fecha_venc_soat'], 'date') else datos['fecha_venc_soat']
            dias_soat = (fecha_venc_soat - fecha_actual).days
            if dias_soat < 0:
                alertas.append({'tipo': 'error', 'mensaje': f"SOAT vencido desde el {datos['fecha_venc_soat'].strftime('%d/%m/%Y')}"})
            elif dias_soat <= 30:
                alertas.append({'tipo': 'warning', 'mensaje': f'SOAT próximo a vencer en {dias_soat} días'})

        if datos['fecha_venc_tecnico_mecanica']:
            fecha_venc_tm = datos['fecha_venc_tecnico_mecanica'].date() if hasattr(datos['fecha_venc_tecnico_mecanica'], 'date') else datos['fecha_venc_tecnico_mecanica']
            dias_tm = (fecha_venc_tm - fecha_actual).days
            if dias_tm < 0:
                alertas.append({'tipo': 'error', 'mensaje': f"Técnico mecánica vencida desde el {datos['fecha_venc_tecnico_mecanica'].strftime('%d/%m/%Y')}"})
            elif dias_tm <= 30:
                alertas.append({'tipo': 'warning', 'mensaje': f'Técnico mecánica próxima a vencer en {dias_tm} días'})

        response_data = {
            'ciudad': datos['ciudad'] or '',
            'supervisor': datos.get('supervisor') or '',
            'placa': datos['placa'] or '',
            'modelo': datos['modelo'] or '',
            'marca': datos['marca'] or '',
            'tipo_vehiculo': datos['tipo_vehiculo'] or '',
            'tipo_licencia': datos['tipo_licencia'] or '',
            'fecha_venc_licencia': datos['fecha_venc_licencia'].strftime('%Y-%m-%d') if datos['fecha_venc_licencia'] else '',
            'fecha_venc_soat': datos['fecha_venc_soat'].strftime('%Y-%m-%d') if datos['fecha_venc_soat'] else '',
            'fecha_venc_tecnico_mecanica': datos['fecha_venc_tecnico_mecanica'].strftime('%Y-%m-%d') if datos['fecha_venc_tecnico_mecanica'] else '',
            'ultimo_kilometraje': datos['ultimo_kilometraje'],
            'fecha_ultimo_kilometraje': datos['fecha_ultimo_kilometraje'].strftime('%Y-%m-%d') if datos['fecha_ultimo_kilometraje'] else ''
        }

        return jsonify({'success': True, 'data': response_data, 'alertas': alertas})
    except mysql.connector.Error as e:
        return jsonify({'success': False, 'message': f'Error de base de datos: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error interno: {str(e)}'}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

@app.route('/api/operativo/validar-kilometraje', methods=['POST'])
@login_required(role=['operativo'])
def validar_kilometraje_operativo():
    """
    Endpoint para validar kilometraje en tiempo real (operativos)
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': 'No se recibieron datos'}), 400

        placa = data.get('placa')
        kilometraje_propuesto = data.get('kilometraje_propuesto')
        if not placa or kilometraje_propuesto is None:
            return jsonify({'success': False, 'message': 'Placa y kilometraje son requeridos'}), 400

        try:
            kilometraje_propuesto = int(kilometraje_propuesto)
        except ValueError:
            return jsonify({'success': False, 'message': 'El kilometraje debe ser un número entero'}), 400

        if kilometraje_propuesto < 0:
            return jsonify({'success': False, 'message': 'El kilometraje no puede ser negativo'}), 400

        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500

        # Normalizar placa para evitar falsos "primer registro" por variaciones (espacios/guiones/minúsculas)
        placa_normalizada = (placa or '').strip().upper().replace('-', '').replace(' ', '')
        cursor = connection.cursor(dictionary=True)
        cursor.execute(
            """
            SELECT kilometraje_actual, fecha
            FROM capired.preoperacional 
            WHERE UPPER(REPLACE(REPLACE(TRIM(placa_vehiculo), '-', ''), ' ', '')) = %s
            ORDER BY fecha DESC 
            LIMIT 1
            """,
            (placa_normalizada,)
        )
        ultimo_registro = cursor.fetchone()

        if not ultimo_registro:
            return jsonify({'success': True, 'valido': True, 'ultimo_kilometraje': 0, 'fecha_ultimo_registro': '', 'mensaje': 'Primer registro de kilometraje para este vehículo'})

        ultimo_kilometraje = ultimo_registro['kilometraje_actual']
        fecha_ultimo_registro = ultimo_registro['fecha']

        if kilometraje_propuesto < ultimo_kilometraje:
            return jsonify({
                'success': True,
                'valido': False,
                'ultimo_kilometraje': ultimo_kilometraje,
                'fecha_ultimo_registro': fecha_ultimo_registro.strftime('%d/%m/%Y') if fecha_ultimo_registro else '',
                'mensaje': f'El kilometraje no puede ser menor al último registrado: {ultimo_kilometraje} km en fecha {fecha_ultimo_registro.strftime("%d/%m/%Y") if fecha_ultimo_registro else "N/A"}'
            })

        if kilometraje_propuesto > 1000000:
            return jsonify({
                'success': True,
                'valido': False,
                'ultimo_kilometraje': ultimo_kilometraje,
                'fecha_ultimo_registro': fecha_ultimo_registro.strftime('%d/%m/%Y') if fecha_ultimo_registro else '',
                'mensaje': 'El kilometraje no puede superar los 1,000,000 km.'
            })

        return jsonify({
            'success': True,
            'valido': True,
            'ultimo_kilometraje': ultimo_kilometraje,
            'fecha_ultimo_registro': fecha_ultimo_registro.strftime('%d/%m/%Y') if fecha_ultimo_registro else '',
            'mensaje': 'Kilometraje válido'
        })
    except mysql.connector.Error as e:
        return jsonify({'success': False, 'message': f'Error de base de datos: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error interno: {str(e)}'}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

@app.route('/api/tecnicos/mantenimientos', methods=['GET'])
@login_required(role='tecnicos')
def api_tecnicos_mantenimientos():
    """API para obtener mantenimientos asignados al técnico logueado"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener información del técnico logueado
        cursor.execute("""
            SELECT nombre, recurso_operativo_cedula 
            FROM capired.recurso_operativo 
            WHERE id_codigo_consumidor = %s
        """, (session['id_codigo_consumidor'],))
        
        tecnico_info = cursor.fetchone()
        if not tecnico_info:
            return jsonify({'error': 'Técnico no encontrado'}), 404
        
        tecnico_nombre = tecnico_info['nombre']
        
        # Obtener mantenimientos asignados al técnico
        # Filtrar por el nombre del técnico en el campo 'tecnico'
        query = """
        SELECT 
            m.id_mpa_mantenimientos,
            m.placa,
            m.fecha_mantenimiento,
            m.kilometraje,
            m.observacion,
            m.soporte_foto_geo as foto_taller,
            m.soporte_foto_factura as foto_factura,
            m.tipo_vehiculo,
            m.tecnico as tecnico_nombre,
            m.tipo_mantenimiento
        FROM mpa_mantenimientos m
        WHERE m.tecnico = %s
        ORDER BY m.fecha_mantenimiento DESC
        """
        
        cursor.execute(query, (tecnico_nombre,))
        mantenimientos = cursor.fetchall()
        
        # Formatear fechas para el frontend
        for mantenimiento in mantenimientos:
            if mantenimiento['fecha_mantenimiento']:
                mantenimiento['fecha_mantenimiento'] = mantenimiento['fecha_mantenimiento'].strftime('%Y-%m-%d %H:%M:%S')
        
        return jsonify({
            'success': True,
            'data': mantenimientos,
            'tecnico': tecnico_nombre
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

@app.route('/operativo')
@login_required(role='operativo')
def operativo_dashboard():
    try:
        connection = get_db_connection()
        if connection is None:
            flash('Error de conexión a la base de datos', 'danger')
            resp = make_response(render_template('modulos/operativo/dashboard.html', tiene_asistencia=False))
            resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
            resp.headers['Pragma'] = 'no-cache'
            resp.headers['Expires'] = '0'
            return resp
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener la cédula del usuario logueado
        cursor.execute("""
            SELECT recurso_operativo_cedula 
            FROM capired.recurso_operativo 
            WHERE id_codigo_consumidor = %s
        """, (session['id_codigo_consumidor'],))
        
        usuario_actual = cursor.fetchone()
        
        # Verificar si es el usuario especial (52912112) que debe estar exento de la restricción
        if usuario_actual and usuario_actual['recurso_operativo_cedula'] == '52912112':
            # Usuario especial: siempre tiene acceso a todos los botones
            tiene_asistencia = True
        else:
            # Para todos los demás usuarios: verificar asistencia registrada para hoy
            fecha_hoy = datetime.now().strftime('%Y-%m-%d')
            cursor.execute("""
                SELECT COUNT(*) as registros_hoy
                FROM asistencia 
                WHERE id_codigo_consumidor = %s AND DATE(fecha_asistencia) = %s
            """, (session['id_codigo_consumidor'], fecha_hoy))
            
            registro_existente = cursor.fetchone()
            tiene_asistencia = registro_existente['registros_hoy'] > 0 if registro_existente else False
        
        resp = make_response(render_template('modulos/operativo/dashboard.html', tiene_asistencia=tiene_asistencia))
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        resp.headers['Pragma'] = 'no-cache'
        resp.headers['Expires'] = '0'
        return resp
        
    except mysql.connector.Error as e:
        flash(f'Error al verificar asistencia: {str(e)}', 'danger')
        resp = make_response(render_template('modulos/operativo/dashboard.html', tiene_asistencia=False))
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        resp.headers['Pragma'] = 'no-cache'
        resp.headers['Expires'] = '0'
        return resp
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/lider')
@login_required_lider()
def lider_dashboard():
    """
    Dashboard principal del módulo de Líder
    """
    try:
        return render_template('modulos/lider/dashboard.html')
    except Exception as e:
        flash(f'Error al cargar el panel de líder: {str(e)}', 'danger')
        return redirect(url_for('index'))

@app.route('/lider/turnos-analistas')
@login_required_lider()
def lider_turnos_analistas():
    """
    Página principal del submódulo Turnos Analistas
    """
    try:
        return render_template('modulos/lider/turnos-analistas.html')
    except Exception as e:
        flash(f'Error al cargar turnos de analistas: {str(e)}', 'danger')
        return redirect(url_for('lider_dashboard'))

@app.route('/lider/indicadores')
@login_required_lider()
def lider_indicadores():
    """
    Página principal del submódulo Ver Indicadores
    """
    try:
        return render_template('modulos/lider/indicadores.html')
    except Exception as e:
        flash(f'Error al cargar indicadores: {str(e)}', 'danger')
        return redirect(url_for('lider_dashboard'))

@app.route('/lider/indicadores/operaciones')
@login_required_lider()
def lider_indicadores_operaciones():
    """
    Página del submódulo Indicadores de Operaciones
    """
    try:
        return render_template('modulos/lider/indicadores-operaciones.html')
    except Exception as e:
        flash(f'Error al cargar indicadores de operaciones: {str(e)}', 'danger')
        return redirect(url_for('lider_indicadores'))

@app.route('/lider/indicadores/operaciones/inicio')
@login_required_lider()
def lider_inicio_operacion():
    """
    Página del submódulo Inicio de Operación
    """
    try:
        return render_template('modulos/lider/inicio-operacion.html')
    except Exception as e:
        flash(f'Error al cargar inicio de operación: {str(e)}', 'danger')
        return redirect(url_for('lider_indicadores_operaciones'))

@app.route('/lider/indicadores/operaciones/presupuesto')
@login_required_lider()
def lider_presupuesto():
    """
    Página del submódulo Presupuesto - Muestra tabla de supervisores con valores mensuales
    """
    try:
        return render_template('modulos/lider/presupuesto.html')
    except Exception as e:
        flash(f'Error al cargar presupuesto: {str(e)}', 'danger')
        return redirect(url_for('lider_indicadores_operaciones'))

# APIs para Inicio de Operación
@app.route('/api/lider/inicio-operacion/supervisores', methods=['GET'])
@login_required_lider_api()
def api_inicio_operacion_supervisores():
    """
    API para obtener lista de supervisores desde tabla asistencia
    """
    connection = None
    cursor = None
    try:
        # Conectar a la base de datos capired
        connection = mysql.connector.connect(
            host='localhost',
            port=3306,
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener supervisores únicos de ambas tablas (asistencia y recurso_operativo)
        # Limpiar datos para evitar duplicados por espacios o caracteres especiales
        cursor.execute("""
            SELECT DISTINCT TRIM(REPLACE(REPLACE(supervisor, '\n', ''), '\r', '')) as supervisor
            FROM (
                SELECT DISTINCT super as supervisor 
                FROM asistencia 
                WHERE super IS NOT NULL AND super != ''
                UNION
                SELECT DISTINCT super as supervisor 
                FROM recurso_operativo 
                WHERE super IS NOT NULL AND super != ''
            ) AS combined_supervisors
            WHERE TRIM(REPLACE(REPLACE(supervisor, '\n', ''), '\r', '')) != ''
            ORDER BY supervisor
        """)
        
        supervisores = cursor.fetchall()
        
        return jsonify({
            'success': True,
            'supervisores': [s['supervisor'] for s in supervisores]
        })
        
    except mysql.connector.Error as e:
        print(f"ERROR DB - Supervisores: {str(e)}")
        return jsonify({'error': f'Error de base de datos: {str(e)}'}), 500
    except Exception as e:
        print(f"ERROR - Supervisores: {str(e)}")
        return jsonify({'error': f'Error interno: {str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/lider/inicio-operacion/analistas', methods=['GET'])
@login_required_lider_api()
def api_inicio_operacion_analistas():
    """
    API para obtener lista de analistas desde tabla recurso_operativo
    """
    connection = None
    cursor = None
    try:
        # Conectar a la base de datos capired
        connection = mysql.connector.connect(
            host='localhost',
            port=3306,
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener analistas únicos de la tabla recurso_operativo
        # Limpiar datos para evitar duplicados por espacios o caracteres especiales
        cursor.execute("""
            SELECT DISTINCT TRIM(REPLACE(REPLACE(analista, '\n', ''), '\r', '')) as analista
            FROM recurso_operativo 
            WHERE analista IS NOT NULL AND analista != ''
            AND TRIM(REPLACE(REPLACE(analista, '\n', ''), '\r', '')) != ''
            ORDER BY analista
        """)
        
        analistas = cursor.fetchall()
        
        return jsonify({
            'success': True,
            'analistas': [a['analista'] for a in analistas]
        })
        
    except mysql.connector.Error as e:
        print(f"ERROR DB - Analistas: {str(e)}")
        return jsonify({'error': f'Error de base de datos: {str(e)}'}), 500
    except Exception as e:
        print(f"ERROR - Analistas: {str(e)}")
        return jsonify({'error': f'Error interno: {str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/lider/inicio-operacion/datos', methods=['GET'])
@login_required_lider_api()
def api_inicio_operacion_datos():
    """
    API para obtener datos del dashboard de inicio de operación
    """
    connection = None
    cursor = None
    try:
        # Obtener parámetros de filtro
        fecha = request.args.get('fecha')
        supervisores = request.args.getlist('supervisores[]')
        analistas = request.args.getlist('analistas[]')
        
        # Debug: Imprimir parámetros recibidos
        print(f"DEBUG - Parámetros recibidos:")
        print(f"  - fecha: {fecha}")
        print(f"  - supervisores: {supervisores}")
        print(f"  - analistas: {analistas}")
        
        # Conectar a la base de datos capired
        connection = mysql.connector.connect(
            host='localhost',
            port=3306,
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Construir condiciones WHERE
        where_conditions = []
        params = []
        
        if fecha:
            where_conditions.append("DATE(fecha_asistencia) = %s")
            params.append(fecha)
        
        if supervisores:
            placeholders = ','.join(['%s'] * len(supervisores))
            where_conditions.append(f"a.super IN ({placeholders})")
            params.extend(supervisores)
        
        if analistas:
            # Filtrar por analistas usando JOIN con recurso_operativo
            placeholders = ','.join(['%s'] * len(analistas))
            where_conditions.append(f"ro.analista IN ({placeholders})")
            params.extend(analistas)
        
        where_clause = " AND ".join(where_conditions) if where_conditions else "1=1"
        
        # Definir nombres de tipificación para técnicos (según especificaciones exactas del usuario)
        tipificaciones_tecnicos = [
            'POSTVENTA',
            'POSTVENTA FTTH',
            'FTTH INSTALACIONES',
            'MANTENIMIENTO FTTH',
            'INSTALACIONES DOBLES',
            'ARREGLOS HFC',
            'BROWNFIELD',
            'INSTALACIONES DOBLES BACK'
        ]
        
        # Definir nombres de tipificación para auxiliares (CON ASISTENCIA)
        tipificaciones_auxiliares = [
            'AUXILIAR DE MOTO',
            'AUXILIAR CAMIONETA'
        ]
        
        # Mantener las carpetas originales para validación de asignación
        carpetas_tecnicos = [
            'POSTVENTA',
            'POSTVENTA FTTH',
            'FTTH INSTALACIONES',
            'MANTENIMIENTO FTTH',
            'INSTALACIONES DOBLES',
            'ARREGLOS HFC',
            'BROWNFIELD',
            'INSTALACIONES DOBLES BACK'
        ]
        
        carpetas_auxiliares = [
            'AUXILIAR'
        ]
        
        # Query principal para obtener datos de asistencia con JOIN a tipificacion_asistencia y recurso_operativo
        query = f"""
            SELECT 
                a.cedula,
                a.tecnico,
                a.carpeta,
                a.super,
                a.carpeta_dia,
                t.nombre_tipificacion,
                a.eventos,
                a.valor,
                a.estado,
                a.novedad,
                a.fecha_asistencia,
                a.hora_inicio,
                ro.analista
            FROM asistencia a
            LEFT JOIN tipificacion_asistencia t ON a.carpeta_dia = t.codigo_tipificacion
            LEFT JOIN recurso_operativo ro ON a.cedula = ro.recurso_operativo_cedula
            WHERE {where_clause}
        """
        
        # Debug: Imprimir consulta y parámetros
        print(f"DEBUG - Consulta SQL:")
        print(f"  Query: {query}")
        print(f"  Params: {params}")
        print(f"  Where clause: {where_clause}")
        
        cursor.execute(query, params)
        asistencia_data = cursor.fetchall()
        
        # Debug: Imprimir cantidad de registros obtenidos
        print(f"DEBUG - Registros obtenidos: {len(asistencia_data)}")
        if len(asistencia_data) > 0:
            print(f"DEBUG - Primer registro: {asistencia_data[0]}")
        else:
            print("DEBUG - No se obtuvieron registros")
        
        # Convertir hora_inicio de timedelta a string para serialización JSON
        for registro in asistencia_data:
            if registro.get('hora_inicio') is not None:
                hora_inicio = registro['hora_inicio']
                if hasattr(hora_inicio, 'total_seconds'):  # Es un timedelta
                    # Convertir timedelta a horas y minutos
                    total_seconds = int(hora_inicio.total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    registro['hora_inicio'] = f"{hours:02d}:{minutes:02d}"
                elif isinstance(hora_inicio, str):
                    # Ya es string, mantenerlo como está
                    pass
                else:
                    # Otro tipo, convertir a string vacío
                    registro['hora_inicio'] = ""
            else:
                # Es None, convertir a string vacío
                registro['hora_inicio'] = ""
        
        # Inicializar contadores
        tecnicos_con_asistencia = 0
        tecnicos_sin_asistencia = 0
        apoyo_con_asistencia = 0
        apoyo_sin_asistencia = 0
        auxiliares_con_asistencia = 0
        auxiliares_sin_asistencia = 0
        
        oks_dia = 0
        presupuesto_dia = 0
        cumple = 0
        no_cumple = 0
        
        # Procesar cada registro
        for registro in asistencia_data:
            carpeta_dia = registro.get('carpeta_dia', '')
            nombre_tipificacion = registro.get('nombre_tipificacion', '')
            carpeta = registro.get('carpeta', '')
            eventos = registro.get('eventos', 0) or 0
            valor = registro.get('valor', 0) or 0
            estado = registro.get('estado', '').lower() if registro.get('estado') else ''
            
            # Contar OK's y presupuesto
            oks_dia += eventos
            presupuesto_dia += valor
            
            # Contar cumplimiento
            if estado in ['cumple', 'novedad']:
                cumple += 1
            elif estado in ['no cumple', 'no aplica']:
                no_cumple += 1
            
            # Lógica de asistencia CORREGIDA basada en nombre_tipificacion:
            # La asistencia se determina por el nombre_tipificacion obtenido del JOIN
            # La ausencia es cuando alguien asignado a un rol (carpeta) no tiene el nombre_tipificacion correspondiente
            
            # TÉCNICOS CON ASISTENCIA: carpeta en lista técnicos Y nombre_tipificacion en lista técnicos
            if carpeta and carpeta.strip() and carpeta in carpetas_tecnicos:
                if nombre_tipificacion and nombre_tipificacion.strip() and nombre_tipificacion in tipificaciones_tecnicos:
                    tecnicos_con_asistencia += 1
            
            # APOYO CAMIONETAS CON ASISTENCIA: carpeta = 'APOYO CAMIONETAS' Y nombre_tipificacion = 'APOYO CAMIONETAS'
            elif carpeta and carpeta.strip() and carpeta == 'APOYO CAMIONETAS':
                if nombre_tipificacion and nombre_tipificacion.strip() and nombre_tipificacion == 'APOYO CAMIONETAS':
                    apoyo_con_asistencia += 1
            
            # AUXILIARES CON ASISTENCIA: carpeta = 'AUXILIAR' Y nombre_tipificacion específico
            if carpeta and carpeta.strip() and carpeta == 'AUXILIAR':
                if nombre_tipificacion and nombre_tipificacion.strip() and nombre_tipificacion in tipificaciones_auxiliares:
                    auxiliares_con_asistencia += 1
            
            # SIN ASISTENCIA: Personas asignadas a roles pero sin nombre_tipificacion correspondiente
            if carpeta and carpeta.strip():
                if carpeta in carpetas_tecnicos:
                    # Si está asignado como técnico pero no tiene nombre_tipificacion de técnico
                    if not (nombre_tipificacion and nombre_tipificacion.strip() and nombre_tipificacion in tipificaciones_tecnicos):
                        tecnicos_sin_asistencia += 1
                elif carpeta == 'APOYO CAMIONETAS':
                    # Si está asignado como apoyo pero no tiene nombre_tipificacion de apoyo
                    if not (nombre_tipificacion and nombre_tipificacion.strip() and nombre_tipificacion == 'APOYO CAMIONETAS'):
                        apoyo_sin_asistencia += 1
                elif carpeta in carpetas_auxiliares:
                    # Si está asignado como auxiliar pero no tiene nombre_tipificacion de auxiliar
                    if not (nombre_tipificacion and nombre_tipificacion.strip() and nombre_tipificacion in tipificaciones_auxiliares):
                        auxiliares_sin_asistencia += 1
        
        # Calcular totales
        total_tecnicos = tecnicos_con_asistencia + tecnicos_sin_asistencia
        total_apoyo_camionetas = apoyo_con_asistencia + apoyo_sin_asistencia
        total_auxiliares = auxiliares_con_asistencia + auxiliares_sin_asistencia
        
        # Calcular porcentajes de cumplimiento diario
        total_evaluados = cumple + no_cumple
        cumplimiento_porcentaje = (cumple / total_evaluados * 100) if total_evaluados > 0 else 0
        
        # Presupuesto mes - obtener del primer día del mes con datos
        presupuesto_mes = 0
        if fecha:
            # Extraer año y mes de la fecha seleccionada
            from datetime import datetime
            fecha_obj = datetime.strptime(fecha, '%Y-%m-%d')
            anio = fecha_obj.year
            mes = fecha_obj.month
            
            # Consulta para obtener el presupuesto del primer día del mes que tenga datos
            query_presupuesto_mes = f"""
                SELECT SUM(a.valor) as presupuesto_primer_dia
                FROM asistencia a
                LEFT JOIN recurso_operativo ro ON a.cedula = ro.recurso_operativo_cedula
                WHERE YEAR(a.fecha_asistencia) = %s 
                AND MONTH(a.fecha_asistencia) = %s
                AND a.valor IS NOT NULL
                AND a.valor > 0
                {' AND ro.analista IN (' + ','.join(['%s'] * len(analistas)) + ')' if analistas else ''}
                {' AND a.super IN (' + ','.join(['%s'] * len(supervisores)) + ')' if supervisores else ''}
                AND DATE(a.fecha_asistencia) = (
                    SELECT MIN(DATE(a2.fecha_asistencia))
                    FROM asistencia a2
                    LEFT JOIN recurso_operativo ro2 ON a2.cedula = ro2.recurso_operativo_cedula
                    WHERE YEAR(a2.fecha_asistencia) = %s 
                    AND MONTH(a2.fecha_asistencia) = %s
                    AND a2.valor IS NOT NULL
                    AND a2.valor > 0
                    {' AND ro2.analista IN (' + ','.join(['%s'] * len(analistas)) + ')' if analistas else ''}
                    {' AND a2.super IN (' + ','.join(['%s'] * len(supervisores)) + ')' if supervisores else ''}
                )
            """
            
            # Preparar parámetros para la consulta
            params_presupuesto = [anio, mes]
            if analistas:
                params_presupuesto.extend(analistas)
            if supervisores:
                params_presupuesto.extend(supervisores)
            params_presupuesto.extend([anio, mes])  # Para la subconsulta
            if analistas:
                params_presupuesto.extend(analistas)
            if supervisores:
                params_presupuesto.extend(supervisores)
            
            cursor.execute(query_presupuesto_mes, params_presupuesto)
            resultado_presupuesto = cursor.fetchone()
            
            if resultado_presupuesto and resultado_presupuesto['presupuesto_primer_dia']:
                presupuesto_mes = float(resultado_presupuesto['presupuesto_primer_dia']) * 26
            else:
                # Si no hay datos del mes, usar el presupuesto del día actual * 26
                presupuesto_mes = presupuesto_dia * 26
        else:
            # Si no hay fecha específica, usar el presupuesto del día actual * 26
            presupuesto_mes = presupuesto_dia * 26
        
        # Calcular cumplimiento mensual acumulado
        cumplimiento_mes_porcentaje = 0
        cumple_mes = 0
        novedad_mes = 0
        no_cumple_mes = 0
        no_aplica_mes = 0
        
        if fecha:
            # Extraer año y mes de la fecha seleccionada
            from datetime import datetime
            fecha_obj = datetime.strptime(fecha, '%Y-%m-%d')
            anio = fecha_obj.year
            mes = fecha_obj.month
            
            # Consulta para obtener datos acumulados del mes completo
            query_mes = f"""
                SELECT 
                    a.estado
                FROM asistencia a
                LEFT JOIN recurso_operativo ro ON a.cedula = ro.recurso_operativo_cedula
                WHERE YEAR(a.fecha_asistencia) = %s 
                AND MONTH(a.fecha_asistencia) = %s
                AND a.estado IS NOT NULL 
                AND a.estado != ''
            """
            
            params_mes = [anio, mes]
            
            # Aplicar filtros de supervisores si existen
            if supervisores:
                placeholders = ','.join(['%s'] * len(supervisores))
                query_mes += f" AND a.super IN ({placeholders})"
                params_mes.extend(supervisores)
            
            # Aplicar filtros de analistas si existen
            if analistas:
                placeholders = ','.join(['%s'] * len(analistas))
                query_mes += f" AND ro.analista IN ({placeholders})"
                params_mes.extend(analistas)
            
            cursor.execute(query_mes, params_mes)
            registros_mes = cursor.fetchall()
            
            # Contar estados del mes completo
            for registro in registros_mes:
                estado = registro.get('estado', '').lower() if registro.get('estado') else ''
                if estado == 'cumple':
                    cumple_mes += 1
                elif estado == 'novedad':
                    novedad_mes += 1
                elif estado == 'no cumple':
                    no_cumple_mes += 1
                elif estado == 'no aplica':
                    no_aplica_mes += 1
            
            # Calcular cumplimiento mensual con la nueva fórmula
            # (Cumple_mes + Novedad_mes) / (Cumple_mes + Novedad_mes + No_Cumple_mes + No_Aplica_mes) × 100
            total_mes = cumple_mes + novedad_mes + no_cumple_mes + no_aplica_mes
            if total_mes > 0:
                cumplimiento_mes_porcentaje = ((cumple_mes + novedad_mes) / total_mes) * 100
            
            print(f"DEBUG - Cumplimiento mensual para {mes}/{anio}:")
            print(f"  - Cumple: {cumple_mes}")
            print(f"  - Novedad: {novedad_mes}")
            print(f"  - No Cumple: {no_cumple_mes}")
            print(f"  - No Aplica: {no_aplica_mes}")
            print(f"  - Total evaluados: {total_mes}")
            print(f"  - Porcentaje mensual: {cumplimiento_mes_porcentaje:.1f}%")
        
        # Log para depuración
        print(f"DEBUG - Inicio Operación: Registros procesados: {len(asistencia_data)}")
        print(f"DEBUG - Técnicos: Total={total_tecnicos} (Con asistencia={tecnicos_con_asistencia}, Sin asistencia={tecnicos_sin_asistencia})")
        print(f"DEBUG - Apoyo: Total={total_apoyo_camionetas} (Con asistencia={apoyo_con_asistencia}, Sin asistencia={apoyo_sin_asistencia})")
        print(f"DEBUG - Auxiliares: Total={total_auxiliares} (Con asistencia={auxiliares_con_asistencia}, Sin asistencia={auxiliares_sin_asistencia})")
        print(f"DEBUG - Métricas: OKs={oks_dia}, Presupuesto día=${presupuesto_dia:,}, Presupuesto mes=${presupuesto_mes:,}")
        print(f"DEBUG - Cumplimiento: {cumple} cumple, {no_cumple} no cumple ({cumplimiento_porcentaje:.1f}%)")
        print(f"DEBUG - Lógica asistencia FINAL basada en nombre_tipificacion:")
        print(f"  - TÉCNICOS con asistencia: nombre_tipificacion en {tipificaciones_tecnicos}")
        print(f"  - TÉCNICOS sin asistencia: carpeta en técnicos PERO nombre_tipificacion NO en técnicos")
        print(f"  - APOYO con asistencia: nombre_tipificacion = 'APOYO CAMIONETAS'")
        print(f"  - APOYO sin asistencia: carpeta = 'APOYO CAMIONETAS' PERO nombre_tipificacion ≠ 'APOYO CAMIONETAS'")
        print(f"  - AUXILIARES con asistencia: nombre_tipificacion en {tipificaciones_auxiliares}")
        print(f"  - AUXILIARES sin asistencia: carpeta en auxiliares PERO nombre_tipificacion NO en auxiliares")
        print(f"DEBUG - JOIN con tipificacion_asistencia: carpeta_dia -> codigo_tipificacion -> nombre_tipificacion")
        
        return jsonify({
            'success': True,
            'data': {
                'tecnicos': {
                    'total': total_tecnicos,
                    'con_asistencia': tecnicos_con_asistencia,
                    'sin_asistencia': tecnicos_sin_asistencia
                },
                'apoyo_camionetas': {
                    'total': total_apoyo_camionetas,
                    'con_asistencia': apoyo_con_asistencia,
                    'sin_asistencia': apoyo_sin_asistencia
                },
                'auxiliares': {
                    'total': total_auxiliares,
                    'con_asistencia': auxiliares_con_asistencia,
                    'sin_asistencia': auxiliares_sin_asistencia
                },
                'metricas': {
                    'oks_dia': oks_dia,
                    'presupuesto_dia': presupuesto_dia,
                    'presupuesto_mes': presupuesto_mes
                },
                'cumplimiento': {
                    'cumple': cumple,
                    'no_cumple': no_cumple,
                    'cumplimiento_dia': round(cumplimiento_porcentaje, 1),
                    'cumplimiento_mes': round(cumplimiento_mes_porcentaje, 1),
                    'cumple_mes': cumple_mes,
                    'novedad_mes': novedad_mes,
                    'no_cumple_mes': no_cumple_mes,
                    'no_aplica_mes': no_aplica_mes,
                    'total_mes': cumple_mes + novedad_mes + no_cumple_mes + no_aplica_mes
                },
                'tabla_asistencia': asistencia_data,
                'debug_info': {
                    'total_registros': len(asistencia_data),
                    'filtros_aplicados': {
                        'fecha': fecha,
                        'supervisores': supervisores,
                        'analistas': analistas
                    }
                }
            }
        })
        
    except mysql.connector.Error as e:
        print(f"ERROR DB - Inicio Operación Datos: {str(e)}")
        return jsonify({'error': f'Error de base de datos: {str(e)}'}), 500
    except Exception as e:
        print(f"ERROR - Inicio Operación Datos: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error interno: {str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/lider/inicio-operacion/indicador-diario', methods=['GET'])
@login_required_lider_api()
def api_inicio_operacion_indicador_diario():
    """
    API para indicador diario: tabla por supervisor (cumple, no cumple, % día)
    y gráfica de % mensual por supervisor.
    """
    connection = None
    cursor = None
    try:
        # Parámetros
        fecha = request.args.get('fecha')
        supervisores = request.args.getlist('supervisores[]')
        analistas = request.args.getlist('analistas[]')

        # Conectar DB
        connection = mysql.connector.connect(
            host='localhost',
            port=3306,
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        cursor = connection.cursor(dictionary=True)

        # Si no viene fecha, usar hoy
        if not fecha:
            from datetime import datetime
            fecha = datetime.now().strftime('%Y-%m-%d')

        # Diario por supervisor
        query_diario = """
            SELECT 
                a.super AS supervisor,
                SUM(CASE WHEN LOWER(TRIM(a.estado)) IN ('cumple','novedad') THEN 1 ELSE 0 END) AS cumple,
                SUM(CASE WHEN LOWER(TRIM(a.estado)) IN ('no cumple','no aplica') THEN 1 ELSE 0 END) AS no_cumple,
                COUNT(*) AS total_evaluados
            FROM asistencia a
            LEFT JOIN recurso_operativo ro ON a.cedula = ro.recurso_operativo_cedula
            WHERE DATE(a.fecha_asistencia) = %s
        """
        params_diario = [fecha]
        if supervisores:
            query_diario += f" AND a.super IN ({','.join(['%s']*len(supervisores))})"
            params_diario.extend(supervisores)
        if analistas:
            query_diario += f" AND ro.analista IN ({','.join(['%s']*len(analistas))})"
            params_diario.extend(analistas)
        query_diario += " GROUP BY a.super ORDER BY a.super"

        cursor.execute(query_diario, params_diario)
        diario_rows = cursor.fetchall() or []
        diario = []
        for r in diario_rows:
            cumple = int(r.get('cumple') or 0)
            no_cumple = int(r.get('no_cumple') or 0)
            total = int(r.get('total_evaluados') or 0)
            denominador = cumple + no_cumple  # cumple+novedad + no cumple+no aplica
            porcentaje = (cumple / denominador * 100) if denominador > 0 else 0
            diario.append({
                'supervisor': r.get('supervisor') or '',
                'cumple': cumple,
                'no_cumple': no_cumple,
                'total_evaluados': total,
                'porcentaje': porcentaje
            })

        # Mensual por supervisor
        from datetime import datetime
        dt = datetime.strptime(fecha, '%Y-%m-%d')
        anio, mes = dt.year, dt.month
        query_mensual = """
            SELECT 
                a.super AS supervisor,
                SUM(CASE WHEN LOWER(TRIM(a.estado)) IN ('cumple','novedad') THEN 1 ELSE 0 END) AS cumple,
                SUM(CASE WHEN LOWER(TRIM(a.estado)) IN ('no cumple','no aplica') THEN 1 ELSE 0 END) AS no_cumple
            FROM asistencia a
            LEFT JOIN recurso_operativo ro ON a.cedula = ro.recurso_operativo_cedula
            WHERE YEAR(a.fecha_asistencia) = %s AND MONTH(a.fecha_asistencia) = %s
        """
        params_mensual = [anio, mes]
        if supervisores:
            query_mensual += f" AND a.super IN ({','.join(['%s']*len(supervisores))})"
            params_mensual.extend(supervisores)
        if analistas:
            query_mensual += f" AND ro.analista IN ({','.join(['%s']*len(analistas))})"
            params_mensual.extend(analistas)
        query_mensual += " GROUP BY a.super ORDER BY a.super"

        cursor.execute(query_mensual, params_mensual)
        mensual_rows = cursor.fetchall() or []
        mensual = []
        for r in mensual_rows:
            cumple = int(r.get('cumple') or 0)
            no_cumple = int(r.get('no_cumple') or 0)
            denominador = cumple + no_cumple
            porcentaje = (cumple / denominador * 100) if denominador > 0 else 0
            mensual.append({
                'supervisor': r.get('supervisor') or '',
                'porcentaje': porcentaje
            })

        return jsonify({
            'success': True,
            'fecha': fecha,
            'diario': diario,
            'mensual': mensual
        })

    except mysql.connector.Error as e:
        print(f"ERROR DB - Indicador Diario: {str(e)}")
        return jsonify({'error': f'Error de base de datos: {str(e)}'}), 500
    except Exception as e:
        print(f"ERROR - Indicador Diario: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error interno: {str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/analistas', methods=['GET'])
@login_required_lider_api()
def api_get_analistas():
    """
    API para obtener lista de analistas activos
    """
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener analistas activos
        cursor.execute("""
            SELECT 
                id_codigo_consumidor,
                recurso_operativo_cedula,
                nombre,
                cargo,
                estado
            FROM recurso_operativo 
            WHERE (cargo = 'ANALISTA' OR cargo = 'ANALISTA LOGISTICA') 
            AND estado = 'Activo'
            ORDER BY nombre
        """)
        
        analistas = cursor.fetchall()
        
        return jsonify({
            'success': True,
            'analistas': analistas
        })
        
    except mysql.connector.Error as e:
        return jsonify({'error': f'Error de base de datos: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'error': f'Error interno: {str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/turnos', methods=['GET'])
@login_required_lider_api()
def api_get_turnos():
    """
    API para obtener lista de turnos disponibles
    """
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener turnos disponibles
        cursor.execute("""
            SELECT 
                turnos_horario,
                turnos_nombre,
                turnos_horas_laboradas,
                turnos_breack,
                turnos_almuerzo
            FROM turnos
            ORDER BY turnos_horario
        """)
        
        turnos_data = []
        for row in cursor.fetchall():
            turnos_data.append({
                'horario': row['turnos_horario'],
                'nombre': row['turnos_nombre'],
                'horas_laboradas': row['turnos_horas_laboradas'],
                'break': row['turnos_breack'],
                'almuerzo': row['turnos_almuerzo']
            })
        
        return jsonify({
            'success': True,
            'turnos': turnos_data
        })
        
    except mysql.connector.Error as e:
        return jsonify({'error': f'Error de base de datos: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'error': f'Error interno: {str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/asignar-turno', methods=['POST'])
@login_required_lider_api()
def api_asignar_turno():
    """
    API para asignar turno a un analista
    """
    connection = None
    cursor = None
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'No se recibieron datos'}), 400
            
        # Validar campos requeridos
        required_fields = ['analista', 'fecha', 'turno']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'Campo requerido: {field}'}), 400
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Verificar si ya existe una asignación para esa fecha y analista
        cursor.execute("""
            SELECT id_analistas_turnos FROM analistas_turnos_base 
            WHERE analistas_turnos_fecha = %s AND analistas_turnos_analista = %s
        """, (data['fecha'], data['analista']))
        
        existing = cursor.fetchone()
        
        if existing:
            # Actualizar asignación existente
            cursor.execute("""
                UPDATE analistas_turnos_base 
                SET analistas_turnos_turno = %s,
                    analistas_turnos_almuerzo = %s,
                    analistas_turnos_break = %s,
                    analistas_turnos_horas_trabajadas = %s
                WHERE id_analistas_turnos = %s
            """, (
                data['turno'],
                data.get('almuerzo'),
                data.get('break'),
                data.get('horas_trabajadas'),
                existing['id_analistas_turnos']
            ))
        else:
            # Crear nueva asignación
            cursor.execute("""
                INSERT INTO analistas_turnos_base (
                    analistas_turnos_fecha,
                    analistas_turnos_analista,
                    analistas_turnos_turno,
                    analistas_turnos_almuerzo,
                    analistas_turnos_break,
                    analistas_turnos_horas_trabajadas
                ) VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                data['fecha'],
                data['analista'],
                data['turno'],
                data.get('almuerzo'),
                data.get('break'),
                data.get('horas_trabajadas')
            ))
        
        connection.commit()
        
        return jsonify({
            'success': True,
            'message': 'Turno asignado correctamente'
        })
        
    except mysql.connector.Error as e:
        if connection:
            connection.rollback()
        return jsonify({'error': f'Error de base de datos: {str(e)}'}), 500
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': f'Error interno: {str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/asignar-turnos-semana', methods=['POST'])
@login_required_lider_api()
def api_asignar_turnos_semana():
    """
    API para asignar turnos de una semana completa a un analista
    """
    connection = None
    cursor = None
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'error': 'No se recibieron datos'}), 400
            
        # Validar campos requeridos
        required_fields = ['analista_id', 'turnos']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({'error': f'Campo requerido: {field}'}), 400
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener información de los turnos disponibles
        cursor.execute("SELECT turnos_horario, turnos_nombre, turnos_horas_laboradas, turnos_breack, turnos_almuerzo FROM turnos")
        turnos_info = {}
        for turno in cursor.fetchall():
            turnos_info[turno['turnos_horario']] = {
                'nombre': turno['turnos_nombre'],
                'horas_laboradas': turno['turnos_horas_laboradas'],
                'break': turno['turnos_breack'],
                'almuerzo': turno['turnos_almuerzo']
            }
        
        turnos_asignados = 0
        turnos_actualizados = 0
        errores = []
        
        # Procesar cada día y sus turnos
        for fecha, turnos_dia in data['turnos'].items():
            if not turnos_dia or not isinstance(turnos_dia, list) or len(turnos_dia) == 0:
                continue
                
            # Para cada día, asignar todos los turnos seleccionados
            # Como solo puede haber un turno por día, tomamos el primero
            turno_id = turnos_dia[0]
            
            if not turno_id or turno_id not in turnos_info:
                errores.append(f"Turno inválido para la fecha {fecha}")
                continue
            
            turno_info = turnos_info[turno_id]
            
            # Verificar si ya existe una asignación para esa fecha y analista
            cursor.execute("""
                SELECT id_analistas_turnos FROM analistas_turnos_base 
                WHERE analistas_turnos_fecha = %s AND analistas_turnos_analista = %s
            """, (fecha, data['analista_id']))
            
            existing = cursor.fetchone()
            
            try:
                if existing:
                    # Actualizar asignación existente
                    cursor.execute("""
                        UPDATE analistas_turnos_base 
                        SET analistas_turnos_turno = %s,
                            analistas_turnos_almuerzo = %s,
                            analistas_turnos_break = %s,
                            analistas_turnos_horas_trabajadas = %s
                        WHERE id_analistas_turnos = %s
                    """, (
                        turno_id,
                        turno_info['almuerzo'],
                        turno_info['break'],
                        turno_info['horas_laboradas'],
                        existing['id_analistas_turnos']
                    ))
                    turnos_actualizados += 1
                else:
                    # Crear nueva asignación
                    cursor.execute("""
                        INSERT INTO analistas_turnos_base (
                            analistas_turnos_fecha,
                            analistas_turnos_analista,
                            analistas_turnos_turno,
                            analistas_turnos_almuerzo,
                            analistas_turnos_break,
                            analistas_turnos_horas_trabajadas
                        ) VALUES (%s, %s, %s, %s, %s, %s)
                    """, (
                        fecha,
                        data['analista_id'],
                        turno_id,
                        turno_info['almuerzo'],
                        turno_info['break'],
                        turno_info['horas_laboradas']
                    ))
                    turnos_asignados += 1
            except Exception as e:
                errores.append(f"Error al procesar fecha {fecha}: {str(e)}")
        
        connection.commit()
        
        mensaje = ''
        if turnos_asignados > 0:
            mensaje += f"{turnos_asignados} turno(s) asignado(s). "
        if turnos_actualizados > 0:
            mensaje += f"{turnos_actualizados} turno(s) actualizado(s). "
        if errores:
            mensaje += f"Errores: {', '.join(errores)}"
        
        return jsonify({
            'success': True,
            'message': mensaje or 'Operación completada exitosamente',
            'asignados': turnos_asignados,
            'actualizados': turnos_actualizados,
            'errores': errores
        })
        
    except mysql.connector.Error as e:
        if connection:
            connection.rollback()
        return jsonify({'error': f'Error de base de datos: {str(e)}'}), 500
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({'error': f'Error interno: {str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/turnos-semana', methods=['GET'])
@login_required_lider_api()
def api_get_turnos_semana():
    """
    API para obtener turnos de una semana específica
    """
    connection = None
    cursor = None
    try:
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return jsonify({'error': 'Se requieren fecha_inicio y fecha_fin'}), 400
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener asignaciones de la semana con nombre de analista
        cursor.execute("""
            SELECT 
                atb.analistas_turnos_fecha,
                COALESCE(ro.nombre, atb.analistas_turnos_analista) as analistas_turnos_analista,
                atb.analistas_turnos_turno,
                atb.analistas_turnos_almuerzo,
                atb.analistas_turnos_break,
                atb.analistas_turnos_horas_trabajadas
            FROM analistas_turnos_base atb
            LEFT JOIN recurso_operativo ro ON atb.analistas_turnos_analista = ro.id_codigo_consumidor
            WHERE atb.analistas_turnos_fecha BETWEEN %s AND %s
            ORDER BY atb.analistas_turnos_fecha, ro.nombre
        """, (fecha_inicio, fecha_fin))
        
        turnos_semana = cursor.fetchall()
        
        # Convertir fechas a strings para asegurar compatibilidad con el frontend
        for turno in turnos_semana:
            if 'analistas_turnos_fecha' in turno and turno['analistas_turnos_fecha']:
                turno['analistas_turnos_fecha'] = turno['analistas_turnos_fecha'].strftime('%Y-%m-%d')
        
        return jsonify({
            'success': True,
            'turnos': turnos_semana
        })
        
    except mysql.connector.Error as e:
        return jsonify({'error': f'Error de base de datos: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'error': f'Error interno: {str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/resumen-semanal', methods=['GET'])
@login_required_lider_api()
def api_get_resumen_semanal():
    """
    API para obtener resumen semanal de horas trabajadas por analista
    """
    connection = None
    cursor = None
    try:
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        if not fecha_inicio or not fecha_fin:
            return jsonify({'error': 'Se requieren fecha_inicio y fecha_fin'}), 400
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Consulta para obtener resumen semanal con datos reales de la tabla
        # Solo cuenta horas de almuerzo en días que realmente trabajaron (horas > 0)
        cursor.execute("""
            SELECT 
                ro.nombre as nombre_analista,
                SUM(atb.analistas_turnos_horas_trabajadas) as horas_semanales,
                COUNT(DISTINCT CASE WHEN atb.analistas_turnos_horas_trabajadas > 0 THEN atb.analistas_turnos_fecha END) as dias_trabajados,
                COUNT(DISTINCT CASE WHEN atb.analistas_turnos_horas_trabajadas > 0 THEN atb.analistas_turnos_fecha END) as horas_almuerzo_total,
                (SUM(atb.analistas_turnos_horas_trabajadas) - COUNT(DISTINCT CASE WHEN atb.analistas_turnos_horas_trabajadas > 0 THEN atb.analistas_turnos_fecha END)) as total_horas_netas
            FROM analistas_turnos_base atb
            INNER JOIN recurso_operativo ro ON atb.analistas_turnos_analista = ro.id_codigo_consumidor
            WHERE atb.analistas_turnos_fecha BETWEEN %s AND %s
            GROUP BY atb.analistas_turnos_analista, ro.nombre
            ORDER BY ro.nombre
        """, (fecha_inicio, fecha_fin))
        
        resumen_semanal = cursor.fetchall()
        
        # Asegurar que los valores numéricos sean correctos
        for analista in resumen_semanal:
            analista['horas_semanales'] = float(analista['horas_semanales']) if analista['horas_semanales'] else 0
            analista['dias_trabajados'] = int(analista['dias_trabajados']) if analista['dias_trabajados'] else 0
            analista['horas_almuerzo_total'] = int(analista['horas_almuerzo_total']) if analista['horas_almuerzo_total'] else 0
            analista['total_horas_netas'] = float(analista['total_horas_netas']) if analista['total_horas_netas'] else 0
            
            # Asegurar que el total no sea negativo
            if analista['total_horas_netas'] < 0:
                analista['total_horas_netas'] = 0
        
        return jsonify({
            'success': True,
            'resumen': resumen_semanal
        })
        
    except mysql.connector.Error as e:
        return jsonify({'error': f'Error de base de datos: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'error': f'Error interno: {str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/detalles-dia', methods=['GET'])
@login_required_lider_api()
def api_get_detalles_dia():
    """
    API para obtener detalles de turnos de un día específico
    """
    connection = None
    cursor = None
    try:
        fecha = request.args.get('fecha')
        
        if not fecha:
            return jsonify({'error': 'Se requiere el parámetro fecha'}), 400
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener detalles del día con información completa del turno y nombre de la analista
        cursor.execute("""
            SELECT 
                atb.analistas_turnos_analista,
                ro.nombre as nombre_analista,
                atb.analistas_turnos_turno,
                atb.analistas_turnos_almuerzo,
                atb.analistas_turnos_break,
                atb.analistas_turnos_horas_trabajadas,
                t.turnos_nombre,
                t.turnos_horas_laboradas
            FROM analistas_turnos_base atb
            INNER JOIN recurso_operativo ro ON atb.analistas_turnos_analista = ro.id_codigo_consumidor
            LEFT JOIN turnos t ON atb.analistas_turnos_turno = t.turnos_horario
            WHERE atb.analistas_turnos_fecha = %s
            ORDER BY ro.nombre
        """, (fecha,))
        
        detalles = cursor.fetchall()
        
        return jsonify({
            'success': True,
            'detalles': detalles,
            'fecha': fecha
        })
        
    except mysql.connector.Error as e:
        return jsonify({'error': f'Error de base de datos: {str(e)}'}), 500
    except Exception as e:
        return jsonify({'error': f'Error interno: {str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/operativo/asistencia')
@login_required(role='operativo')
def operativo_asistencia():
    try:
        connection = get_db_connection()
        if connection is None:
            flash('Error de conexión a la base de datos', 'danger')
            return redirect(url_for('operativo_dashboard'))
            
        cursor = connection.cursor(dictionary=True)
        
        # Crear tablas si no existen
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS tipificacion_asistencia (
                id_tipificacion INT AUTO_INCREMENT PRIMARY KEY,
                codigo_tipificacion VARCHAR(50) NOT NULL,
                nombre_tipificacion VARCHAR(200),
                estado CHAR(1) DEFAULT '1',
                fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS asistencia (
                id_asistencia INT AUTO_INCREMENT PRIMARY KEY,
                cedula VARCHAR(20),
                tecnico VARCHAR(100),
                carpeta_dia VARCHAR(50),
                carpeta VARCHAR(50),
                super VARCHAR(100),
                fecha_asistencia TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                id_codigo_consumidor INT
            )
        """)
        connection.commit()
        
        # Obtener el nombre del usuario logueado para usar como filtro de supervisor
        cursor.execute("""
            SELECT nombre FROM capired.recurso_operativo
            WHERE id_codigo_consumidor = %s
        """, (session['id_codigo_consumidor'],))
        
        supervisor_result = cursor.fetchone()
        if not supervisor_result or not supervisor_result['nombre']:
            flash('No se encontró información del usuario logueado', 'warning')
            return redirect(url_for('operativo_dashboard'))
            
        supervisor_usuario = supervisor_result['nombre']
        
        # Verificar si ya existe un registro de asistencia para hoy
        fecha_hoy = datetime.now().strftime('%Y-%m-%d')
        cursor.execute("""
            SELECT COUNT(*) as registros_hoy
            FROM asistencia 
            WHERE super = %s AND DATE(fecha_asistencia) = %s
        """, (supervisor_usuario, fecha_hoy))
        
        registro_existente = cursor.fetchone()
        ya_registrado = registro_existente['registros_hoy'] > 0 if registro_existente else False
        
        # DEBUG: Mostrar todos los valores distintos de carpeta para debug
        cursor.execute("""
            SELECT DISTINCT carpeta, COUNT(*) as cantidad
            FROM capired.recurso_operativo
            WHERE estado = 'Activo'
            GROUP BY carpeta
            ORDER BY carpeta
        """)
        debug_carpetas = cursor.fetchall()
        print("\n=== DEBUG: Valores distintos en columna 'carpeta' ===")
        for carpeta in debug_carpetas:
            print(f"Carpeta: '{carpeta['carpeta']}' - Cantidad: {carpeta['cantidad']}")
        
        # Obtener técnicos del supervisor automáticamente
        # Obtener la cédula del usuario actual para excluirlo de la lista
        cursor.execute("""
            SELECT recurso_operativo_cedula 
            FROM capired.recurso_operativo 
            WHERE id_codigo_consumidor = %s
        """, (session['id_codigo_consumidor'],))
        usuario_actual = cursor.fetchone()
        
        # Filtrar técnicos por supervisor usando la lógica estándar
        cursor.execute("""
            SELECT id_codigo_consumidor, recurso_operativo_cedula, nombre, carpeta
            FROM capired.recurso_operativo
            WHERE super = %s AND estado = 'Activo'
            AND recurso_operativo_cedula != %s
            ORDER BY nombre
        """, (supervisor_usuario, usuario_actual['recurso_operativo_cedula'] if usuario_actual else ''))
        tecnicos = cursor.fetchall()
        print(f"Técnicos encontrados para supervisor {supervisor_usuario} (excluyendo supervisor): {len(tecnicos)}")
        
        # Obtener lista de tipificaciones para carpeta_dia (solo zona OP para operativo)
        cursor.execute("""
            SELECT codigo_tipificacion, nombre_tipificacion
            FROM tipificacion_asistencia
            WHERE estado = '1' AND zona = 'OP'
            ORDER BY codigo_tipificacion
        """)
        carpetas_dia = cursor.fetchall()
        
        return render_template('modulos/operativo/asistencia.html',
                           tecnicos=tecnicos,
                           carpetas_dia=carpetas_dia,
                           supervisor=supervisor_usuario,
                           ya_registrado=ya_registrado,
                           fecha_hoy=fecha_hoy)
                           
    except mysql.connector.Error as e:
        flash(f'Error al cargar datos: {str(e)}', 'danger')
        return redirect(url_for('operativo_dashboard'))
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/logistica')
@login_required(role='logistica')
def logistica_dashboard():
    # Verificar stock bajo para mostrar alertas
    materiales_problematicos = verificar_stock_bajo()
    return render_template('modulos/logistica/dashboard.html', materiales_problematicos=materiales_problematicos)

@app.route('/contabilidad')
@login_required(role='contabilidad')
def contabilidad_dashboard():
    return render_template('modulos/contabilidad/dashboard.html')

@app.route('/admin/buscar_usuarios', methods=['POST'])
@login_required(role='administrativo')
def buscar_usuarios():
    search_query = request.form.get('search_query', '')
    search_type = request.form.get('search_type', 'cedula')
    role = request.form.get('role', '')
    sort = request.form.get('sort', 'cedula')
    
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Construir la consulta SQL base
        query = """
            SELECT 
                id_codigo_consumidor,
                recurso_operativo_cedula,
                id_roles,
                estado
            FROM recurso_operativo
            WHERE 1=1
        """
        params = []
        
        # Aplicar filtros de búsqueda
        if search_query:
            if search_type == 'cedula':
                query += " AND recurso_operativo_cedula LIKE %s"
                params.append(f"%{search_query}%")
            elif search_type == 'codigo':
                query += " AND id_codigo_consumidor LIKE %s"
                params.append(f"%{search_query}%")
            elif search_type == 'rol':
                query += " AND id_roles = %s"
                role_id = next((k for k, v in ROLES.items() if v.lower() == search_query.lower()), None)
                if role_id:
                    params.append(role_id)
        
        if role:
            role_id = next((k for k, v in ROLES.items() if v == role), None)
            if role_id:
                query += " AND id_roles = %s"
                params.append(role_id)
        
        # Aplicar ordenamiento
        if sort == 'cedula':
            query += " ORDER BY recurso_operativo_cedula"
        
        cursor.execute(query, params)
        users = cursor.fetchall()
        
        # Formatear resultados
        formatted_users = []
        for user in users:
            formatted_users.append({
                'id_codigo_consumidor': user['id_codigo_consumidor'],
                'recurso_operativo_cedula': user['recurso_operativo_cedula'],
                'role': ROLES.get(str(user['id_roles']), 'Desconocido'),
                'estado': user.get('estado', 'Activo')
            })
        
        return jsonify(formatted_users)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/admin/exportar_usuarios_csv', methods=['POST'])
@login_required(role='administrativo')
def exportar_usuarios_csv():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener todos los usuarios
        cursor.execute("SELECT id_codigo_consumidor, recurso_operativo_cedula, id_roles, estado FROM recurso_operativo")
        users = cursor.fetchall()
        
        # Crear archivo CSV en memoria
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Escribir encabezados
        writer.writerow(['Código', 'Cédula', 'Rol', 'Estado'])
        
        # Escribir datos
        for user in users:
            writer.writerow([
                user['id_codigo_consumidor'],
                user['recurso_operativo_cedula'],
                ROLES.get(str(user['id_roles']), 'Desconocido'),
                user.get('estado', 'Activo')
            ])
        
        # Preparar respuesta
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'usuarios_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )
        
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/admin/estadisticas_usuarios')
@login_required(role='administrativo')
def estadisticas_usuarios():
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener estadísticas por rol
        cursor.execute("""
            SELECT id_roles, COUNT(*) as count
            FROM recurso_operativo
            GROUP BY id_roles
        """)
        role_stats = cursor.fetchall()
        
        role_labels = []
        role_data = []
        for stat in role_stats:
            role_labels.append(ROLES.get(str(stat['id_roles']), 'Desconocido'))
            role_data.append(stat['count'])
        
        # Obtener estadísticas de actividad (últimos 30 días)
        activity_dates = [(datetime.now() - timedelta(days=x)).strftime('%Y-%m-%d') for x in range(30)]
        activity_counts = []
        
        for date in activity_dates:
            cursor.execute("""
                SELECT COUNT(*) as count
                FROM recurso_operativo
                WHERE DATE(created_at) = %s
            """, (date,))
            result = cursor.fetchone()
            activity_counts.append(result['count'] if result else 0)
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'role_labels': role_labels,
            'role_data': role_data,
            'activity_labels': activity_dates[::-1],
            'activity_data': activity_counts[::-1]
        })
        
    except Error as e:
        return jsonify({'error': str(e)}), 500

@app.route('/dashboard')
@login_required()
def dashboard():
    user_role = session.get('user_role')
    if user_role == 'administrativo':
        try:
            connection = get_db_connection()
            if connection is None:
                flash('Error al cargar usuarios. Por favor, inténtelo de nuevo.', 'error')
                return render_template('modulos/administrativo/dashboard.html')
                
            cursor = connection.cursor(dictionary=True)
            
            # Obtener usuarios
            cursor.execute("""
                SELECT 
                    id_codigo_consumidor, 
                    recurso_operativo_cedula, 
                    id_roles,
                    estado,
                    nombre,
                    cargo
                FROM recurso_operativo
            """)
            users = cursor.fetchall()
            
            # Obtener estadísticas
            total_users = len(users)
            total_roles = len(set(user['id_roles'] for user in users))
            active_users = sum(1 for user in users if user.get('estado') == 'Activo')
            
            cursor.close()
            connection.close()
            
            return render_template('modulos/administrativo/dashboard.html',
                                users=users,
                                ROLES=ROLES,
                                total_users=total_users,
                                total_roles=total_roles,
                                active_users=active_users)
                                
        except Error as e:
            flash(f'Error al cargar usuarios: {str(e)}', 'error')
            return render_template('modulos/administrativo/dashboard.html')
    elif user_role in ROLES.values():
        # Manejar redirección específica para cada rol
        if user_role == 'analista':
            return redirect(url_for('analistas_index'))
        else:
            return redirect(url_for(f'{user_role}_dashboard'))
    else:
        flash('No tienes un rol válido asignado.', 'error')
        return redirect(url_for('logout'))

@app.route('/user/edit/<int:user_id>', methods=['GET', 'POST'])
@login_required(role='administrativo')
def edit_user(user_id):
    try:
        connection = get_db_connection()
        if connection is None:
            flash('Error de conexión a la base de datos.', 'error')
            return redirect(url_for('dashboard'))

        cursor = connection.cursor(dictionary=True)
        
        if request.method == 'POST':
            # Obtener datos del formulario
            new_role = request.form.get('role_id')
            new_password = request.form.get('password')
            
            # Actualizar rol
            if new_role:
                cursor.execute("""
                    UPDATE recurso_operativo 
                    SET id_roles = %s 
                    WHERE id_codigo_consumidor = %s
                """, (new_role, user_id))
            
            # Actualizar contraseña si se proporcionó una nueva
            if new_password:
                hashed_password = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt())
                cursor.execute("""
                    UPDATE recurso_operativo 
                    SET recurso_operativo_password = %s 
                    WHERE id_codigo_consumidor = %s
                """, (hashed_password, user_id))
            
            connection.commit()
            flash('Usuario actualizado exitosamente.', 'success')
            return redirect(url_for('dashboard'))
        
        # Obtener datos del usuario para el formulario
        cursor.execute("""
            SELECT id_codigo_consumidor, recurso_operativo_cedula, id_roles 
            FROM recurso_operativo 
            WHERE id_codigo_consumidor = %s
        """, (user_id,))
        user = cursor.fetchone()
        
        if not user:
            flash('Usuario no encontrado.', 'error')
            return redirect(url_for('dashboard'))
            
        return jsonify({
            'status': 'success',
            'user': user
        })
        
    except Error as e:
        flash(f'Error al editar usuario: {str(e)}', 'error')
        return redirect(url_for('dashboard'))
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

@app.route('/user/delete/<int:user_id>', methods=['POST'])
@login_required(role='administrativo')
def delete_user(user_id):
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'status': 'error', 'message': 'Error de conexión a la base de datos.'})

        cursor = connection.cursor(dictionary=True)
        
        # Verificar que el usuario existe
        cursor.execute("SELECT id_roles FROM recurso_operativo WHERE id_codigo_consumidor = %s", (user_id,))
        user = cursor.fetchone()
        
        if not user:
            return jsonify({'status': 'error', 'message': 'Usuario no encontrado.'})
        
        # No permitir eliminar al último usuario administrativo
        if user['id_roles'] == 1:  # rol administrativo
            cursor.execute("SELECT COUNT(*) as admin_count FROM recurso_operativo WHERE id_roles = 1")
            admin_count = cursor.fetchone()['admin_count']
            if (admin_count <= 1):
                return jsonify({
                    'status': 'error',
                    'message': 'No se puede eliminar el último usuario administrativo.'
                })
        
        # Eliminar usuario
        cursor.execute("DELETE FROM recurso_operativo WHERE id_codigo_consumidor = %s", (user_id,))
        connection.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'Usuario eliminado exitosamente.'
        })
        
    except Error as e:
        return jsonify({
            'status': 'error',
            'message': f'Error al eliminar usuario: {str(e)}'
        })
    finally:
        if connection.is_connected():
            cursor.close()
            connection.close()

@app.route('/create_user', methods=['POST'])
@login_required(role='administrativo')
def create_user():
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500

        cursor = connection.cursor(dictionary=True)
        
        # Obtener datos del formulario
        cedula = request.form.get('recurso_operativo_cedula')  # Nombre corregido según el formulario
        password = request.form.get('password')
        rol = request.form.get('role_id')  # Nombre corregido según el formulario
        estado = request.form.get('estado', 'Activo')  # Por defecto 'Activo'
        nombre = request.form.get('nombre')
        cargo = request.form.get('cargo')
        
        # Obtener los nuevos campos
        carpeta = request.form.get('carpeta')
        cliente = request.form.get('cliente')
        ciudad = request.form.get('ciudad')
        super_valor = request.form.get('super')  # Usando super_valor porque 'super' es palabra reservada

        # Validar datos requeridos
        if not all([cedula, password, rol, nombre]):
            return jsonify({'success': False, 'message': 'Los campos cédula, contraseña, rol y nombre son requeridos'}), 400

        # Verificar si la cédula ya existe
        cursor.execute("SELECT id_codigo_consumidor FROM recurso_operativo WHERE recurso_operativo_cedula = %s", (cedula,))
        if cursor.fetchone():
            return jsonify({'success': False, 'message': 'La cédula ya está registrada'}), 400

        # Encriptar contraseña
        hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())

        # Insertar nuevo usuario con los campos adicionales
        cursor.execute("""
            INSERT INTO recurso_operativo 
            (recurso_operativo_cedula, recurso_operativo_password, id_roles, estado, nombre, cargo, carpeta, cliente, ciudad, super)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (cedula, hashed_password.decode('utf-8'), rol, estado, nombre, cargo, carpeta, cliente, ciudad, super_valor))

        connection.commit()
        cursor.close()
        connection.close()

        return jsonify({'success': True, 'message': 'Usuario creado exitosamente'})

    except Error as e:
        return jsonify({'success': False, 'message': f'Error al crear usuario: {str(e)}'}), 500
    
@app.route('/get_user/<int:user_id>')
@login_required(role='administrativo')
def get_user(user_id):
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500

        cursor = connection.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                id_codigo_consumidor,
                recurso_operativo_cedula,
                id_roles,
                estado
            FROM recurso_operativo 
            WHERE id_codigo_consumidor = %s
        """, (user_id,))
        
        user = cursor.fetchone()
        
        cursor.close()
        connection.close()

        if not user:
            return jsonify({'success': False, 'message': 'Usuario no encontrado'}), 404

        return jsonify(user)

    except Error as e:
        return jsonify({'success': False, 'message': f'Error al obtener usuario: {str(e)}'})
    
@app.route('/update_user', methods=['POST'])
@login_required(role='administrativo')
def update_user():
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500

        cursor = connection.cursor(dictionary=True)
        
        # Obtener datos del formulario
        user_id = request.form.get('id_codigo_consumidor')
        cedula = request.form.get('cedula')
        nombre = request.form.get('nombre')
        password = request.form.get('password')
        rol = request.form.get('rol')
        cargo = request.form.get('cargo')
        estado = request.form.get('estado')

        # Validar datos requeridos
        if not all([user_id, cedula, nombre, rol, cargo, estado]):
            return jsonify({'success': False, 'message': 'Faltan campos requeridos'}), 400

        # Verificar si el usuario existe
        cursor.execute("SELECT id_codigo_consumidor FROM recurso_operativo WHERE id_codigo_consumidor = %s", (user_id,))
        if not cursor.fetchone():
            return jsonify({'success': False, 'message': 'Usuario no encontrado'}), 404

        # Preparar la consulta de actualización
        update_query = """
            UPDATE recurso_operativo 
            SET recurso_operativo_cedula = %s,
                nombre = %s,
                id_roles = %s,
                cargo = %s,
                estado = %s
        """
        params = [cedula, nombre, rol, cargo, estado]

        # Si se proporciona una nueva contraseña, actualizarla
        if password:
            hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
            update_query += ", recurso_operativo_password = %s"
            params.append(hashed_password.decode('utf-8'))

        update_query += " WHERE id_codigo_consumidor = %s"
        params.append(user_id)

        # Ejecutar la actualización
        cursor.execute(update_query, params)
        connection.commit()

        cursor.close()
        connection.close()

        return jsonify({'success': True, 'message': 'Usuario actualizado exitosamente'})

    except Error as e:
        return jsonify({'success': False, 'message': f'Error al actualizar usuario: {str(e)}'}), 500

@app.route('/preoperacional', methods=['POST'])
@login_required(role=['tecnicos','operativo'])
def registrar_preoperacional():
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'status': 'error', 'message': 'Error de conexión a la base de datos.'}), 500

        cursor = connection.cursor(dictionary=True)
        
        # Verificar si ya existe un registro para el día actual
        id_codigo_consumidor = session['id_codigo_consumidor']
        fecha_actual = get_bogota_datetime().date()
        
        cursor.execute("""
            SELECT COUNT(*) as count 
            FROM preoperacional 
            WHERE id_codigo_consumidor = %s 
            AND DATE(CONVERT_TZ(fecha, '+00:00', '-05:00')) = %s
        """, (id_codigo_consumidor, fecha_actual))
        
        resultado = cursor.fetchone()
        app.logger.info(f"Verificación de registro existente: {resultado}")
        
        if resultado['count'] > 0:
            app.logger.warning(f"Ya existe un registro para el usuario {id_codigo_consumidor} en la fecha {fecha_actual}")
            return jsonify({
                'status': 'error',
                'message': 'Ya has registrado un preoperacional para el día de hoy.'
            }), 400

        # Obtener datos del formulario
        data = {
            'centro_de_trabajo': request.form.get('centro_de_trabajo'),
            'ciudad': request.form.get('ciudad'),
            'supervisor': request.form.get('supervisor'),
            'vehiculo_asistio_operacion': request.form.get('vehiculo_asistio_operacion'),
            'tipo_vehiculo': request.form.get('tipo_vehiculo'),
            'placa_vehiculo': request.form.get('placa_vehiculo'),
            'modelo_vehiculo': request.form.get('modelo_vehiculo'),
            'marca_vehiculo': request.form.get('marca_vehiculo'),
            'licencia_conduccion': request.form.get('licencia_conduccion'),
            'fecha_vencimiento_licencia': request.form.get('fecha_vencimiento_licencia'),
            'fecha_vencimiento_soat': request.form.get('fecha_vencimiento_soat'),
            'fecha_vencimiento_tecnomecanica': request.form.get('fecha_vencimiento_tecnomecanica'),
            'estado_espejos': request.form.get('estado_espejos', 0),
            'bocina_pito': request.form.get('bocina_pito', 0),
            'frenos': request.form.get('frenos', 0),
            'encendido': request.form.get('encendido', 0),
            'estado_bateria': request.form.get('estado_bateria', 0),
            'estado_amortiguadores': request.form.get('estado_amortiguadores', 0),
            'estado_llantas': request.form.get('estado_llantas', 0),
            'kilometraje_actual': request.form.get('kilometraje_actual', 0),
            'luces_altas_bajas': request.form.get('luces_altas_bajas', 0),
            'direccionales_delanteras_traseras': request.form.get('direccionales_delanteras_traseras', 0),
            'elementos_prevencion_seguridad_vial_casco': request.form.get('elementos_prevencion_seguridad_vial_casco', 0),
            'casco_certificado': request.form.get('casco_certificado', 0),
            'casco_identificado': request.form.get('casco_identificado', 0),
            'estado_guantes': request.form.get('estado_guantes', 0),
            'estado_rodilleras': request.form.get('estado_rodilleras', 0),
            'impermeable': request.form.get('impermeable', 0),
            'observaciones': request.form.get('observaciones', '')
        }

        # Bloqueo por vencimientos (SOAT / Tecnomecánica / Licencia) dentro de 30 días o vencidos
        try:
            vencimientos_detectados = []
            placa_form = data.get('placa_vehiculo')

            # Fecha actual (Bogotá)
            fecha_hoy = get_bogota_datetime().date()

            # Consultar último SOAT por placa
            if placa_form:
                cursor.execute("""
                    SELECT fecha_vencimiento
                    FROM mpa_soat
                    WHERE placa = %s
                    ORDER BY fecha_vencimiento DESC
                    LIMIT 1
                """, (placa_form,))
                soat_row = cursor.fetchone()
                if soat_row and soat_row.get('fecha_vencimiento'):
                    dias_restantes = (soat_row['fecha_vencimiento'].date() - fecha_hoy).days
                    if dias_restantes <= 0:
                        vencimientos_detectados.append({
                            'tipo': 'SOAT',
                            'fecha': soat_row['fecha_vencimiento'].strftime('%Y-%m-%d'),
                            'dias_restantes': dias_restantes
                        })

                # Consultar Tecnomecánica por placa
                cursor.execute("""
                    SELECT fecha_vencimiento
                    FROM mpa_tecnico_mecanica
                    WHERE placa = %s
                    ORDER BY fecha_vencimiento DESC
                    LIMIT 1
                """, (placa_form,))
                tm_row = cursor.fetchone()
                if tm_row and tm_row.get('fecha_vencimiento'):
                    dias_restantes = (tm_row['fecha_vencimiento'].date() - fecha_hoy).days
                    if dias_restantes <= 0:
                        vencimientos_detectados.append({
                            'tipo': 'Tecnomecánica',
                            'fecha': tm_row['fecha_vencimiento'].strftime('%Y-%m-%d'),
                            'dias_restantes': dias_restantes
                        })

            # Consultar Licencia de Conducir por técnico
            cursor.execute("""
                SELECT fecha_vencimiento
                FROM mpa_licencia_conducir
                WHERE id_codigo_consumidor = %s
                ORDER BY fecha_vencimiento DESC
                LIMIT 1
            """, (id_codigo_consumidor,))
            lic_row = cursor.fetchone()
            if lic_row and lic_row.get('fecha_vencimiento'):
                dias_restantes = (lic_row['fecha_vencimiento'].date() - fecha_hoy).days
                if dias_restantes <= 0:
                    vencimientos_detectados.append({
                        'tipo': 'Licencia de Conducir',
                        'fecha': lic_row['fecha_vencimiento'].strftime('%Y-%m-%d'),
                        'dias_restantes': dias_restantes
                    })

            # Si existen vencimientos, bloquear
            if vencimientos_detectados:
                cursor.close()
                connection.close()
                return jsonify({
                    'status': 'error',
                    'message': 'No puedes registrar el preoperacional: tienes documentos vencidos.',
                    'bloqueo_por_vencimientos': True,
                    'vencimientos': vencimientos_detectados
                }), 400
        except Error as e:
            app.logger.error(f"Error validando vencimientos en preoperacional: {str(e)}")
            # No bloquear por error en validación, continuar

        # Verificar que el id_codigo_consumidor existe en la tabla recurso_operativo
        cursor.execute("SELECT id_codigo_consumidor FROM recurso_operativo WHERE id_codigo_consumidor = %s", (id_codigo_consumidor,))
        usuario_existe = cursor.fetchone()
        app.logger.info(f"Verificación de usuario en recurso_operativo: {usuario_existe}")
        
        if usuario_existe is None:
            app.logger.error(f"Usuario {id_codigo_consumidor} no existe en recurso_operativo")
            return jsonify({
                'status': 'error', 
                'message': 'El id_codigo_consumidor no existe en la tabla recurso_operativo.'
            }), 404

        # VALIDACIÓN CRÍTICA: Verificar si el vehículo tiene mantenimientos abiertos
        placa_vehiculo = data.get('placa_vehiculo')
        if placa_vehiculo:
            app.logger.info(f"Verificando mantenimientos abiertos para vehículo: {placa_vehiculo}")
            cursor.execute("""
                SELECT COUNT(*) FROM mpa_mantenimientos 
                WHERE placa = %s AND estado = 'Abierto'
            """, (placa_vehiculo,))
            
            result = cursor.fetchone()
            mantenimientos_abiertos = result['COUNT(*)'] if result else 0
            app.logger.info(f"Mantenimientos abiertos encontrados: {mantenimientos_abiertos}")
            
            if mantenimientos_abiertos > 0:
                app.logger.warning(f"Bloqueando preoperacional: vehículo {placa_vehiculo} tiene {mantenimientos_abiertos} mantenimiento(s) abierto(s)")
                cursor.close()
                connection.close()
                return jsonify({
                    'status': 'error',
                    'message': f'No se puede completar el preoperacional. El vehículo {placa_vehiculo} tiene {mantenimientos_abiertos} mantenimiento(s) abierto(s) pendiente(s). Contacte al área de MPA para finalizar el mantenimiento antes de continuar.',
                    'tiene_mantenimientos_abiertos': True,
                    'placa': placa_vehiculo,
                    'mantenimientos_abiertos': mantenimientos_abiertos
                }), 400

        # Validación de kilometraje antes de insertar
        placa_vehiculo = data.get('placa_vehiculo')
        try:
            kilometraje_propuesto = int(data.get('kilometraje_actual') or 0)
        except Exception:
            kilometraje_propuesto = 0

        ultimo_kilometraje = 0
        fecha_ultimo_registro = None
        if placa_vehiculo:
            placa_norm_insert = (placa_vehiculo or '').strip().upper().replace('-', '').replace(' ', '')
            cursor.execute("""
                SELECT kilometraje_actual, fecha
                FROM preoperacional
                WHERE UPPER(REPLACE(REPLACE(TRIM(placa_vehiculo), '-', ''), ' ', '')) = %s
                ORDER BY fecha DESC
                LIMIT 1
            """, (placa_norm_insert,))
            ultimo_registro = cursor.fetchone()
            if ultimo_registro:
                ultimo_kilometraje = ultimo_registro['kilometraje_actual'] or 0
                fecha_ultimo_registro = ultimo_registro['fecha']

        # Bloquear si supera 1,000,000 km
        if kilometraje_propuesto > 1000000:
            return jsonify({
                'status': 'error',
                'message': f'El kilometraje no puede superar los 1,000,000 km. Último kilometraje registrado: {ultimo_kilometraje} km'
            }), 400

        # Bloquear si es menor que el último registrado
        if ultimo_kilometraje and kilometraje_propuesto < ultimo_kilometraje:
            return jsonify({
                'status': 'error',
                'message': f'El kilometraje no puede ser menor al último registrado: {ultimo_kilometraje} km en fecha {fecha_ultimo_registro.strftime("%d/%m/%Y") if fecha_ultimo_registro else "N/A"}'
            }), 400

        # Construir la consulta SQL dinámicamente
        columns = list(data.keys()) + ['id_codigo_consumidor']
        values = list(data.values()) + [id_codigo_consumidor]
        placeholders = ['%s'] * len(columns)

        sql = f"""
            INSERT INTO preoperacional (
                {', '.join(columns)}, fecha
            ) VALUES (
                {', '.join(placeholders)}, %s
            )
        """
        
        # Agregar la fecha actual de Bogotá a los valores
        fecha_bogota = get_bogota_datetime()
        values.append(fecha_bogota)
        
        app.logger.info(f"SQL a ejecutar: {sql}")
        app.logger.info(f"Valores a insertar: {values}")
        app.logger.info(f"Fecha de Bogotá: {fecha_bogota}")
        
        cursor.execute(sql, tuple(values))
        app.logger.info("SQL ejecutado exitosamente")
        
        connection.commit()
        app.logger.info("Transacción confirmada (commit)")
        cursor.close()
        connection.close()

        return jsonify({
            'status': 'success',
            'message': 'Preoperacional registrado exitosamente'
        })
        
    except Error as e:
        if connection and connection.is_connected():
            cursor.close()
            connection.close()
        return jsonify({
            'status': 'error',
            'message': f'Error al registrar preoperacional: {str(e)}'
        }), 500

@app.route('/preoperacional_operativo', methods=['POST'])
@login_required(role=['operativo'])
def registrar_preoperacional_operativo():
    try:
        # Debug: Log de inicio de función
        app.logger.info("=== INICIO registrar_preoperacional_operativo ===")
        
        # Debug: Log de todos los datos recibidos
        app.logger.info(f"Datos del formulario recibidos: {dict(request.form)}")
        
        connection = get_db_connection()
        if connection is None:
            app.logger.error("Error: No se pudo establecer conexión a la base de datos")
            return jsonify({'status': 'error', 'message': 'Error de conexión a la base de datos.'}), 500

        cursor = connection.cursor(dictionary=True)
        
        # Verificar si ya existe un registro para el día actual
        id_codigo_consumidor = request.form.get('id_codigo_consumidor')
        app.logger.info(f"ID código consumidor: {id_codigo_consumidor}")
        
        fecha_actual = get_bogota_datetime().date()
        app.logger.info(f"Fecha actual (Bogotá): {fecha_actual}")
        
        cursor.execute("""
            SELECT COUNT(*) as count 
            FROM preoperacional 
            WHERE id_codigo_consumidor = %s 
            AND DATE(CONVERT_TZ(fecha, '+00:00', '-05:00')) = %s
        """, (id_codigo_consumidor, fecha_actual))
        
        resultado = cursor.fetchone()
        if resultado['count'] > 0:
            return jsonify({
                'status': 'error',
                'message': 'Ya has registrado un preoperacional para el día de hoy.'
            }), 400

        # Obtener datos del formulario
        data = {
            'centro_de_trabajo': request.form.get('centro_de_trabajo'),
            'ciudad': request.form.get('ciudad'),
            'supervisor': request.form.get('supervisor'),
            'vehiculo_asistio_operacion': request.form.get('vehiculo_asistio_operacion'),
            'tipo_vehiculo': request.form.get('tipo_vehiculo'),
            'placa_vehiculo': request.form.get('placa_vehiculo'),
            'modelo_vehiculo': request.form.get('modelo_vehiculo'),
            'marca_vehiculo': request.form.get('marca_vehiculo'),
            'licencia_conduccion': request.form.get('licencia_conduccion'),
            'fecha_vencimiento_licencia': request.form.get('fecha_vencimiento_licencia'),
            'fecha_vencimiento_soat': request.form.get('fecha_vencimiento_soat'),
            'fecha_vencimiento_tecnomecanica': request.form.get('fecha_vencimiento_tecnomecanica'),
            'estado_espejos': request.form.get('estado_espejos', 0),
            'bocina_pito': request.form.get('bocina_pito', 0),
            'frenos': request.form.get('frenos', 0),
            'encendido': request.form.get('encendido', 0),
            'estado_bateria': request.form.get('estado_bateria', 0),
            'estado_amortiguadores': request.form.get('estado_amortiguadores', 0),
            'estado_llantas': request.form.get('estado_llantas', 0),
            'kilometraje_actual': request.form.get('kilometraje_actual', 0),
            'luces_altas_bajas': request.form.get('luces_altas_bajas', 0),
            'direccionales_delanteras_traseras': request.form.get('direccionales_delanteras_traseras', 0),
            'elementos_prevencion_seguridad_vial_casco': request.form.get('elementos_prevencion_seguridad_vial_casco', 0),
            'casco_certificado': request.form.get('casco_certificado', 0),
            'casco_identificado': request.form.get('casco_identificado', 0),
            'estado_guantes': request.form.get('estado_guantes', 0),
            'estado_rodilleras': request.form.get('estado_rodilleras', 0),
            'impermeable': request.form.get('impermeable', 0),
            'observaciones': request.form.get('observaciones', '')
        }

        # Bloqueo por vencimientos (SOAT / Tecnomecánica / Licencia) dentro de 30 días o vencidos
        try:
            vencimientos_detectados = []
            placa_form = data.get('placa_vehiculo')

            # Fecha actual (Bogotá)
            fecha_hoy = get_bogota_datetime().date()

            # Consultar último SOAT por placa
            if placa_form:
                cursor.execute("""
                    SELECT fecha_vencimiento
                    FROM mpa_soat
                    WHERE placa = %s
                    ORDER BY fecha_vencimiento DESC
                    LIMIT 1
                """, (placa_form,))
                soat_row = cursor.fetchone()
                if soat_row and soat_row.get('fecha_vencimiento'):
                    dias_restantes = (soat_row['fecha_vencimiento'].date() - fecha_hoy).days
                    if dias_restantes <= 0:
                        vencimientos_detectados.append({
                            'tipo': 'SOAT',
                            'fecha': soat_row['fecha_vencimiento'].strftime('%Y-%m-%d'),
                            'dias_restantes': dias_restantes
                        })

                # Consultar Tecnomecánica por placa
                cursor.execute("""
                    SELECT fecha_vencimiento
                    FROM mpa_tecnico_mecanica
                    WHERE placa = %s
                    ORDER BY fecha_vencimiento DESC
                    LIMIT 1
                """, (placa_form,))
                tm_row = cursor.fetchone()
                if tm_row and tm_row.get('fecha_vencimiento'):
                    dias_restantes = (tm_row['fecha_vencimiento'].date() - fecha_hoy).days
                    if dias_restantes <= 0:
                        vencimientos_detectados.append({
                            'tipo': 'Tecnomecánica',
                            'fecha': tm_row['fecha_vencimiento'].strftime('%Y-%m-%d'),
                            'dias_restantes': dias_restantes
                        })

            # Consultar Licencia de Conducir por operativo (id_codigo_consumidor de la petición)
            cursor.execute("""
                SELECT fecha_vencimiento
                FROM mpa_licencia_conducir
                WHERE id_codigo_consumidor = %s
                ORDER BY fecha_vencimiento DESC
                LIMIT 1
            """, (id_codigo_consumidor,))
            lic_row = cursor.fetchone()
            if lic_row and lic_row.get('fecha_vencimiento'):
                dias_restantes = (lic_row['fecha_vencimiento'].date() - fecha_hoy).days
                if dias_restantes <= 0:
                    vencimientos_detectados.append({
                        'tipo': 'Licencia de Conducir',
                        'fecha': lic_row['fecha_vencimiento'].strftime('%Y-%m-%d'),
                        'dias_restantes': dias_restantes
                    })

            # Si existen vencimientos, bloquear
            if vencimientos_detectados:
                return jsonify({
                    'status': 'error',
                    'message': 'No puedes registrar el preoperacional: tienes documentos vencidos.',
                    'bloqueo_por_vencimientos': True,
                    'vencimientos': vencimientos_detectados
                }), 400
        except Error as e:
            app.logger.error(f"Error validando vencimientos en preoperacional operativo: {str(e)}")
            # No bloquear por error en validación, continuar

        # Verificar que el id_codigo_consumidor existe en la tabla recurso_operativo
        cursor.execute("SELECT id_codigo_consumidor FROM recurso_operativo WHERE id_codigo_consumidor = %s", (id_codigo_consumidor,))
        if cursor.fetchone() is None:
            return jsonify({
                'status': 'error', 
                'message': 'El id_codigo_consumidor no existe en la tabla recurso_operativo.'
            }), 404

        # Validación de kilometraje contra último registro
        placa_form = request.form.get('placa_vehiculo')
        km_actual_str = request.form.get('kilometraje_actual')
        try:
            km_actual = int(km_actual_str) if km_actual_str is not None and km_actual_str != '' else 0
        except ValueError:
            # Cerrar recursos antes de responder
            cursor.close()
            connection.close()
            return jsonify({
                'status': 'error',
                'message': 'El kilometraje debe ser un número entero.'
            }), 400

        if km_actual < 0:
            cursor.close()
            connection.close()
            return jsonify({
                'status': 'error',
                'message': 'El kilometraje no puede ser negativo.'
            }), 400

        if not placa_form:
            cursor.close()
            connection.close()
            return jsonify({
                'status': 'error',
                'message': 'La placa del vehículo es requerida para validar el kilometraje.'
            }), 400

        # Normalizar placa para comparación robusta
        placa_form_normalizada = (placa_form or '').strip().upper().replace('-', '').replace(' ', '')
        cursor.execute(
            """
            SELECT kilometraje_actual, fecha
            FROM capired.preoperacional
            WHERE UPPER(REPLACE(REPLACE(TRIM(placa_vehiculo), '-', ''), ' ', '')) = %s
            ORDER BY fecha DESC
            LIMIT 1
            """,
            (placa_form_normalizada,)
        )
        ultimo_registro_km = cursor.fetchone()

        if ultimo_registro_km:
            ultimo_km = int(ultimo_registro_km.get('kilometraje_actual') or 0)
            fecha_ultimo = ultimo_registro_km.get('fecha')
            if km_actual < ultimo_km:
                cursor.close()
                connection.close()
                return jsonify({
                    'status': 'error',
                    'message': f'El kilometraje no puede ser menor al último registrado: {ultimo_km} km' + (f" en fecha {fecha_ultimo.strftime('%d/%m/%Y')}" if fecha_ultimo else ''),
                    'ultimo_kilometraje': ultimo_km
                }), 400

        if km_actual > 1000000:
            cursor.close()
            connection.close()
            return jsonify({
                'status': 'error',
                'message': 'El kilometraje no puede superar los 1,000,000 km.'
            }), 400

        # Construir la consulta SQL dinámicamente
        columns = list(data.keys()) + ['id_codigo_consumidor']
        values = list(data.values()) + [id_codigo_consumidor]
        placeholders = ['%s'] * len(columns)

        sql = f"""
            INSERT INTO preoperacional (
                {', '.join(columns)}, fecha
            ) VALUES (
                {', '.join(placeholders)}, %s
            )
        """
        
        # Agregar la fecha actual de Bogotá a los valores
        values.append(get_bogota_datetime())
        
        cursor.execute(sql, tuple(values))
        connection.commit()
        cursor.close()
        connection.close()
        
        app.logger.info("=== Preoperacional registrado exitosamente ===")

        # Devolver respuesta JSON con redirección para que el frontend la maneje
        return jsonify({
            'status': 'success',
            'message': 'Preoperacional registrado exitosamente',
            'redirect_url': '/operativo'
        }) 
        
    except Error as e:
        app.logger.error(f"Error de MySQL: {str(e)}")
        app.logger.error(f"Tipo de error: {type(e)}")
        if 'connection' in locals() and connection and connection.is_connected():
            connection.rollback()
            cursor.close()
            connection.close()
        return jsonify({
            'status': 'error',
            'message': f'Error de base de datos: {str(e)}'
        }), 500
    except Exception as e:
        app.logger.error(f"Error general: {str(e)}")
        app.logger.error(f"Tipo de error: {type(e)}")
        if 'connection' in locals() and connection and connection.is_connected():
            connection.rollback()
            cursor.close()
            connection.close()
        return jsonify({
            'status': 'error',
            'message': f'Error interno del servidor: {str(e)}'
        }), 500

@app.route('/check_submission', methods=['GET'])
def check_submission():
    user_id = request.args.get('user_id')
    submission_date = request.args.get('date')
    
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'submitted': False, 'error': 'Database connection failed.'})

        cursor = connection.cursor(dictionary=True)
        query = """
            SELECT 1 FROM preoperacional 
            WHERE id_codigo_consumidor = %s AND DATE(fecha_creacion) = %s
        """
        cursor.execute(query, (user_id, submission_date))
        submission_exists = cursor.fetchone() is not None

        cursor.close()
        connection.close()

        return jsonify({'submitted': submission_exists})
    except Error as e:
        return jsonify({'submitted': False, 'error': str(e)})

@app.route('/preoperacional/listado')
@login_required(role='administrativo')
def listado_preoperacional():
    try:
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        supervisor = request.args.get('supervisor')
        estado_vehiculo = request.args.get('estado_vehiculo')
        
        # Si no se proporcionan filtros de fecha, usar la fecha actual por defecto
        if not fecha_inicio and not fecha_fin:
            fecha_actual = datetime.now().strftime('%Y-%m-%d')
            fecha_inicio = fecha_actual
            fecha_fin = fecha_actual
        
        # Parámetro de paginación
        pagina = request.args.get('pagina', 1, type=int)
        registros_por_pagina = 10

        conn = get_db_connection()
        cur = conn.cursor(dictionary=True)

        # Construir la cláusula WHERE dinámica
        where_clauses = []
        params = []

        if fecha_inicio:
            where_clauses.append("fecha >= %s")
            params.append(fecha_inicio)
        if fecha_fin:
            where_clauses.append("fecha <= %s")
            params.append(fecha_fin + " 23:59:59")
        if supervisor:
            where_clauses.append("supervisor = %s")
            params.append(supervisor)
        
        # Agregar filtro por estado del vehículo
        if estado_vehiculo:
            if estado_vehiculo == 'bueno':
                where_clauses.append("(estado_espejos + bocina_pito + frenos + encendido + estado_bateria + estado_amortiguadores + estado_llantas + luces_altas_bajas + direccionales_delanteras_traseras) = 9")
            elif estado_vehiculo == 'regular':
                where_clauses.append("(estado_espejos + bocina_pito + frenos + encendido + estado_bateria + estado_amortiguadores + estado_llantas + luces_altas_bajas + direccionales_delanteras_traseras) >= 5")
                where_clauses.append("(estado_espejos + bocina_pito + frenos + encendido + estado_bateria + estado_amortiguadores + estado_llantas + luces_altas_bajas + direccionales_delanteras_traseras) < 9")
            elif estado_vehiculo == 'malo':
                where_clauses.append("(estado_espejos + bocina_pito + frenos + encendido + estado_bateria + estado_amortiguadores + estado_llantas + luces_altas_bajas + direccionales_delanteras_traseras) < 5")

        # Construir la consulta SQL completa
        query = """
            SELECT p.*, r.nombre as nombre_tecnico, r.cargo as cargo_tecnico,
                   (estado_espejos + bocina_pito + frenos + encendido + estado_bateria + 
                    estado_amortiguadores + estado_llantas + luces_altas_bajas + 
                    direccionales_delanteras_traseras) as estado_total,
                   CASE 
                       WHEN (estado_espejos + bocina_pito + frenos + encendido + estado_bateria + 
                            estado_amortiguadores + estado_llantas + luces_altas_bajas + 
                            direccionales_delanteras_traseras) = 9 THEN 'Bueno'
                       WHEN (estado_espejos + bocina_pito + frenos + encendido + estado_bateria + 
                            estado_amortiguadores + estado_llantas + luces_altas_bajas + 
                            direccionales_delanteras_traseras) >= 5 THEN 'Regular'
                       ELSE 'Malo'
                   END as estado_vehiculo
            FROM preoperacional p
            LEFT JOIN recurso_operativo r ON p.id_codigo_consumidor = r.id_codigo_consumidor
        """
        
        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)
        
        query += " ORDER BY p.fecha DESC"
        
        cur.execute(query, params)
        registros = cur.fetchall()

        # Calcular paginación
        total_registros = len(registros)
        total_paginas = (total_registros + registros_por_pagina - 1) // registros_por_pagina
        
        # Ajustar página actual si está fuera de rango
        if pagina < 1:
            pagina = 1
        elif pagina > total_paginas and total_paginas > 0:
            pagina = total_paginas
        
        # Obtener registros de la página actual
        inicio = (pagina - 1) * registros_por_pagina
        fin = inicio + registros_por_pagina
        pagina_actual = registros[inicio:fin]

        # Obtener lista de supervisores únicos
        cur.execute("""
            SELECT DISTINCT supervisor 
            FROM preoperacional 
            WHERE supervisor IS NOT NULL 
            AND supervisor != '' 
            ORDER BY supervisor
        """)
        supervisores = [row['supervisor'] for row in cur.fetchall()]

        # Calcular estadísticas básicas
        registros_hoy = sum(1 for r in registros if r['fecha'].date() == datetime.now().date())
        registros_semana = sum(1 for r in registros if r['fecha'].date() >= (datetime.now() - timedelta(days=7)).date())
        registros_mes = sum(1 for r in registros if r['fecha'].date() >= (datetime.now() - timedelta(days=30)).date())

        # Preparar datos para el gráfico de tendencias
        fechas = []
        registros_por_dia = []
        if registros:
            fecha_actual = datetime.now().date()
            for i in range(7):
                fecha = fecha_actual - timedelta(days=i)
                fechas.append(fecha.strftime('%Y-%m-%d'))
                registros_por_dia.append(sum(1 for r in registros if r['fecha'].date() == fecha))
            fechas.reverse()
            registros_por_dia.reverse()

        # ESTADÍSTICAS ADICIONALES

        # 1. Distribución por tipo de vehículo
        tipos_vehiculo = {}
        for r in registros:
            tipo = r['tipo_vehiculo'] or 'No especificado'
            if tipo in tipos_vehiculo:
                tipos_vehiculo[tipo] += 1
            else:
                tipos_vehiculo[tipo] = 1
        
        # Ordenar por cantidad y tomar los 5 principales
        tipos_vehiculo_top = sorted(tipos_vehiculo.items(), key=lambda x: x[1], reverse=True)[:5]
        labels_tipo_vehiculo = [t[0] for t in tipos_vehiculo_top]
        datos_tipo_vehiculo = [t[1] for t in tipos_vehiculo_top]

        # 2. Top 5 técnicos con más registros
        tecnicos = {}
        for r in registros:
            tecnico = r['nombre_tecnico'] or 'No especificado'
            if tecnico in tecnicos:
                tecnicos[tecnico] += 1
            else:
                tecnicos[tecnico] = 1
        
        tecnicos_top = sorted(tecnicos.items(), key=lambda x: x[1], reverse=True)[:5]
        labels_tecnicos = [t[0] for t in tecnicos_top]
        datos_tecnicos = [t[1] for t in tecnicos_top]

        # 3. Top 5 centros de trabajo con más registros
        centros_trabajo = {}
        for r in registros:
            centro = r['centro_de_trabajo'] or 'No especificado'
            if centro in centros_trabajo:
                centros_trabajo[centro] += 1
            else:
                centros_trabajo[centro] = 1
        
        centros_top = sorted(centros_trabajo.items(), key=lambda x: x[1], reverse=True)[:5]
        labels_centros = [c[0] for c in centros_top]
        datos_centros = [c[1] for c in centros_top]

        # 4. Elementos críticos en mal estado (conteo)
        elementos_mal_estado = {
            'Frenos': sum(1 for r in registros if r.get('frenos') == '0'),
            'Luces': sum(1 for r in registros if r.get('luces_altas_bajas') == '0'),
            'Direccionales': sum(1 for r in registros if r.get('direccionales_delanteras_traseras') == '0'),
            'Espejos': sum(1 for r in registros if r.get('estado_espejos') == '0'),
            'Llantas': sum(1 for r in registros if r.get('estado_llantas') == '0')
        }
        
        labels_elementos = list(elementos_mal_estado.keys())
        datos_elementos = list(elementos_mal_estado.values())

        # 5. Distribución por ciudad
        ciudades = {}
        for r in registros:
            ciudad = r['ciudad'] or 'No especificado'
            if ciudad in ciudades:
                ciudades[ciudad] += 1
            else:
                ciudades[ciudad] = 1
        
        ciudades_top = sorted(ciudades.items(), key=lambda x: x[1], reverse=True)[:5]
        labels_ciudades = [c[0] for c in ciudades_top]
        datos_ciudades = [c[1] for c in ciudades_top]

        # 6. Comparativa mes actual vs mes anterior
        fecha_actual = datetime.now().date()
        primer_dia_mes_actual = fecha_actual.replace(day=1)
        ultimo_dia_mes_anterior = primer_dia_mes_actual - timedelta(days=1)
        primer_dia_mes_anterior = ultimo_dia_mes_anterior.replace(day=1)
        
        registros_mes_actual = sum(1 for r in registros if r['fecha'].date() >= primer_dia_mes_actual)
        registros_mes_anterior = sum(1 for r in registros if r['fecha'].date() >= primer_dia_mes_anterior and r['fecha'].date() < primer_dia_mes_actual)
        
        comparativa_meses = {
            'labels': [primer_dia_mes_anterior.strftime('%B %Y'), primer_dia_mes_actual.strftime('%B %Y')],
            'datos': [registros_mes_anterior, registros_mes_actual]
        }

        cur.close()
        conn.close()

        return render_template('modulos/administrativo/listado_preoperacional.html', 
                            registros=registros,
                            pagina_actual=pagina_actual,
                            pagina_actual_num=pagina,
                            total_paginas=total_paginas,
                            total_registros=total_registros,
                            supervisores=supervisores,
                            fecha_inicio=fecha_inicio,
                            fecha_fin=fecha_fin,
                            fecha_actual=datetime.now().strftime('%Y-%m-%d'),
                            supervisor=supervisor,
                            estado_vehiculo=estado_vehiculo,
                            registros_hoy=registros_hoy,
                            registros_semana=registros_semana,
                            registros_mes=registros_mes,
                            fechas=fechas,
                            registros_por_dia=registros_por_dia,
                            # Nuevas estadísticas
                            labels_tipo_vehiculo=labels_tipo_vehiculo,
                            datos_tipo_vehiculo=datos_tipo_vehiculo,
                            labels_tecnicos=labels_tecnicos,
                            datos_tecnicos=datos_tecnicos,
                            labels_centros=labels_centros,
                            datos_centros=datos_centros,
                            labels_elementos=labels_elementos,
                            datos_elementos=datos_elementos,
                            labels_ciudades=labels_ciudades,
                            datos_ciudades=datos_ciudades,
                            comparativa_meses=comparativa_meses)

    except Exception as e:
        flash('Error al cargar el listado preoperacional: ' + str(e), 'error')
        return render_template('modulos/administrativo/listado_preoperacional.html',
                            registros=[],
                            pagina_actual=[],
                            pagina_actual_num=1,
                            total_paginas=0,
                            supervisores=[],
                            fecha_inicio=fecha_inicio,
                            fecha_fin=fecha_fin,
                            supervisor=supervisor,
                            estado_vehiculo=estado_vehiculo,
                            total_registros=0,
                            registros_hoy=0,
                            registros_semana=0,
                            registros_mes=0,
                            fechas=[],
                            registros_por_dia=[],
                            # Valores vacíos para nuevas estadísticas
                            labels_tipo_vehiculo=[],
                            datos_tipo_vehiculo=[],
                            labels_tecnicos=[],
                            datos_tecnicos=[],
                            labels_centros=[],
                            datos_centros=[],
                            labels_elementos=[],
                            datos_elementos=[],
                            labels_ciudades=[],
                            datos_ciudades=[],
                            comparativa_meses={'labels': [], 'datos': []})

@app.route('/preoperacional/exportar_csv')
@login_required(role='administrativo')
def exportar_preoperacional_csv():
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener parámetros de filtro
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        supervisor = request.args.get('supervisor')
        
        # Construir la cláusula WHERE base
        where_clauses = []
        params = []
        
        if fecha_inicio:
            where_clauses.append("DATE(p.fecha) >= %s")
            params.append(fecha_inicio)
        if fecha_fin:
            where_clauses.append("DATE(p.fecha) <= %s")
            params.append(fecha_fin)
        if supervisor:
            where_clauses.append("p.supervisor = %s")
            params.append(supervisor)
            
        where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"
        
        # Obtener registros preoperacionales con información del usuario y filtros aplicados
        query = f"""
            SELECT 
                p.*,
                r.nombre as nombre_tecnico,
                r.cargo as cargo_tecnico,
                r.recurso_operativo_cedula as cedula_tecnico
            FROM preoperacional p
            JOIN recurso_operativo r ON p.id_codigo_consumidor = r.id_codigo_consumidor
            WHERE {where_sql}
            ORDER BY p.fecha DESC
        """
        cursor.execute(query, params)
        registros = cursor.fetchall()
        cursor.close()
        connection.close()
        
        # Crear archivo CSV en memoria con codificación UTF-8-SIG (con BOM)
        output = io.StringIO()
        # Usar csv.excel con delimitador explícito para garantizar correcta separación de columnas
        writer = csv.writer(output, delimiter=',', quotechar='"', quoting=csv.QUOTE_ALL)
        
        # Escribir el BOM de UTF-8
        output.write('\ufeff')
        
        # Escribir encabezados
        writer.writerow([
            'Fecha', 'Técnico', 'Cédula', 'Cargo', 'Centro de Trabajo', 'Ciudad', 'Supervisor',
            'Vehículo Asistió', 'Tipo Vehículo', 'Placa', 'Modelo', 'Marca',
            'Licencia Conducción', 'Vencimiento Licencia', 'Vencimiento SOAT',
            'Vencimiento Tecnomecánica', 'Estado Espejos', 'Bocina/Pito',
            'Frenos', 'Encendido', 'Batería', 'Amortiguadores', 'Llantas',
            'Kilometraje', 'Luces Altas/Bajas', 'Direccionales',
            'Elementos Prevención Casco', 'Casco Certificado', 'Casco Identificado',
            'Estado Guantes', 'Estado Rodilleras', 'Impermeable', 'Observaciones'
        ])
        
        # Escribir datos
        for registro in registros:
            # Convertir fechas a cadena o valor nulo para evitar errores
            fecha_lic = registro['fecha_vencimiento_licencia'].strftime('%Y-%m-%d') if registro['fecha_vencimiento_licencia'] else ''
            fecha_soat = registro['fecha_vencimiento_soat'].strftime('%Y-%m-%d') if registro['fecha_vencimiento_soat'] else ''
            fecha_tec = registro['fecha_vencimiento_tecnomecanica'].strftime('%Y-%m-%d') if registro['fecha_vencimiento_tecnomecanica'] else ''
            
            writer.writerow([
                registro['fecha'].strftime('%Y-%m-%d %H:%M:%S'),
                registro['nombre_tecnico'],
                registro['cedula_tecnico'],
                registro['cargo_tecnico'],
                registro['centro_de_trabajo'] or '',
                registro['ciudad'] or '',
                registro['supervisor'] or '',
                registro['vehiculo_asistio_operacion'] or '',
                registro['tipo_vehiculo'] or '',
                registro['placa_vehiculo'] or '',
                registro['modelo_vehiculo'] or '',
                registro['marca_vehiculo'] or '',
                registro['licencia_conduccion'] or '',
                fecha_lic,
                fecha_soat,
                fecha_tec,
                registro['estado_espejos'] or '',
                registro['bocina_pito'] or '',
                registro['frenos'] or '',
                registro['encendido'] or '',
                registro['estado_bateria'] or '',
                registro['estado_amortiguadores'] or '',
                registro['estado_llantas'] or '',
                registro['kilometraje_actual'] or '',
                registro['luces_altas_bajas'] or '',
                registro['direccionales_delanteras_traseras'] or '',
                registro['elementos_prevencion_seguridad_vial_casco'] or '',
                registro['casco_certificado'] or '',
                registro['casco_identificado'] or '',
                registro['estado_guantes'] or '',
                registro['estado_rodilleras'] or '',
                registro['impermeable'] or '',
                registro['observaciones'] or ''
            ])
        
        # Preparar respuesta
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),  # Usar UTF-8 con BOM
            mimetype='text/csv; charset=utf-8-sig',
            as_attachment=True,
            download_name=f'preoperacional_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )
        
    except Error as e:
        return jsonify({'error': str(e)}), 500

@app.route('/admin/reportes')
@login_required(role='administrativo')
def admin_reportes():
    return render_template('modulos/administrativo/reportes/dashboard_reportes.html')

@app.route('/api/reportes/usuarios')
@login_required(role='administrativo')
def api_reportes_usuarios():
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Estadísticas generales de usuarios
        cursor.execute("""
            SELECT 
                COUNT(*) as total_usuarios,
                SUM(CASE WHEN estado = 'Activo' THEN 1 ELSE 0 END) as usuarios_activos,
                id_roles,
                COUNT(*) as cantidad
            FROM recurso_operativo
            GROUP BY id_roles
        """)
        stats_por_rol = cursor.fetchall()
        
        # Usuarios registrados por mes (últimos 12 meses)
        cursor.execute("""
            SELECT 
                DATE_FORMAT(fecha_creacion, '%Y-%m') as mes,
                COUNT(*) as cantidad
            FROM recurso_operativo
            WHERE fecha_creacion >= DATE_SUB(CURDATE(), INTERVAL 12 MONTH)
            GROUP BY DATE_FORMAT(fecha_creacion, '%Y-%m')
            ORDER BY mes
        """)
        registros_por_mes = cursor.fetchall()
        
        cursor.close()
        connection.close()
        
        # Formatear datos para gráficos
        roles_labels = [ROLES.get(str(stat['id_roles']), 'Desconocido') for stat in stats_por_rol]
        roles_data = [stat['cantidad'] for stat in stats_por_rol]
        
        meses_labels = []
        meses_data = []
        for registro in registros_por_mes:
            meses_labels.append(registro['mes'])
            meses_data.append(registro['cantidad'])
            
        return jsonify({
            'roles': {
                'labels': roles_labels,
                'data': roles_data
            },
            'registros_mensuales': {
                'labels': meses_labels,
                'data': meses_data
            }
        })
        
    except Error as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reportes/filtros')
@login_required(role='administrativo')
def api_reportes_filtros():
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener supervisores únicos
        cursor.execute("""
            SELECT DISTINCT supervisor 
            FROM preoperacional 
            WHERE supervisor IS NOT NULL AND supervisor != ''
            ORDER BY supervisor
        """)
        supervisores = [row['supervisor'] for row in cursor.fetchall()]
        
        # Obtener centros de trabajo únicos
        cursor.execute("""
            SELECT DISTINCT centro_de_trabajo 
            FROM preoperacional 
            WHERE centro_de_trabajo IS NOT NULL AND centro_de_trabajo != ''
            ORDER BY centro_de_trabajo
        """)
        centros_trabajo = [row['centro_de_trabajo'] for row in cursor.fetchall()]
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'supervisores': supervisores,
            'centros_trabajo': centros_trabajo
        })
        
    except Error as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reportes/preoperacional')
@login_required(role='administrativo')
def api_reportes_preoperacional():
    try:
        # Obtener parámetros de filtro
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        supervisor = request.args.get('supervisor')
        centro_trabajo = request.args.get('centro_trabajo')
        
        # Construir la cláusula WHERE base
        where_clauses = []
        params = []
        
        if fecha_inicio:
            where_clauses.append("p.fecha >= %s")
            params.append(fecha_inicio)
        if fecha_fin:
            where_clauses.append("p.fecha <= %s")
            params.append(fecha_fin)
        if supervisor:
            where_clauses.append("p.supervisor = %s")
            params.append(supervisor)
        if centro_trabajo:
            where_clauses.append("p.centro_de_trabajo = %s")
            params.append(centro_trabajo)
            
        where_sql = " AND ".join(where_clauses) if where_clauses else "1=1"
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Estadísticas generales con filtros
        stats_query = f"""
            SELECT 
                COUNT(*) as total_preoperacionales,
                COUNT(DISTINCT id_codigo_consumidor) as total_tecnicos,
                COUNT(DISTINCT DATE(fecha)) as total_dias
            FROM preoperacional p
            WHERE {where_sql}
        """
        cursor.execute(stats_query, params)
        stats_generales = cursor.fetchone()
        
        # Preoperacionales por día con filtros
        registros_query = f"""
            SELECT 
                DATE(fecha) as fecha,
                COUNT(*) as cantidad
            FROM preoperacional p
            WHERE {where_sql}
            GROUP BY DATE(fecha)
            ORDER BY fecha
        """
        cursor.execute(registros_query, params)
        registros_por_dia = cursor.fetchall()
        
        # Estado de vehículos con filtros
        estado_query = f"""
            SELECT 
                COUNT(*) as total,
                SUM(CASE 
                    WHEN (estado_espejos + bocina_pito + frenos + encendido + estado_bateria + 
                         estado_amortiguadores + estado_llantas + luces_altas_bajas + 
                         direccionales_delanteras_traseras) = 9 THEN 1 
                    ELSE 0 
                END) as estado_bueno,
                SUM(CASE 
                    WHEN (estado_espejos + bocina_pito + frenos + encendido + estado_bateria + 
                         estado_amortiguadores + estado_llantas + luces_altas_bajas + 
                         direccionales_delanteras_traseras) >= 5 
                    AND (estado_espejos + bocina_pito + frenos + encendido + estado_bateria + 
                         estado_amortiguadores + estado_llantas + luces_altas_bajas + 
                         direccionales_delanteras_traseras) < 9 THEN 1 
                    ELSE 0 
                END) as estado_regular,
                SUM(CASE 
                    WHEN (estado_espejos + bocina_pito + frenos + encendido + estado_bateria + 
                         estado_amortiguadores + estado_llantas + luces_altas_bajas + 
                         direccionales_delanteras_traseras) < 5 THEN 1 
                    ELSE 0 
                END) as estado_malo
            FROM preoperacional p
            WHERE {where_sql}
        """
        cursor.execute(estado_query, params)
        estado_vehiculos = cursor.fetchone()
        
        # Estadísticas por centro de trabajo con filtros
        centro_trabajo_query = f"""
            SELECT 
                COALESCE(centro_de_trabajo, 'No especificado') as centro_de_trabajo,
                COUNT(*) as cantidad
            FROM preoperacional p
            WHERE {where_sql}
            GROUP BY centro_de_trabajo
            ORDER BY cantidad DESC
            LIMIT 5
        """
        cursor.execute(centro_trabajo_query, params)
        stats_centro_trabajo = cursor.fetchall()
        
        cursor.close()
        connection.close()
        
        # Formatear datos para la respuesta
        dias_labels = [registro['fecha'].strftime('%Y-%m-%d') for registro in registros_por_dia]
        dias_data = [registro['cantidad'] for registro in registros_por_dia]
        
        total = estado_vehiculos['total'] or 1
        estado_labels = ['Bueno', 'Regular', 'Malo']
        estado_data = [
            round((estado_vehiculos['estado_bueno'] or 0) / total * 100, 2),
            round((estado_vehiculos['estado_regular'] or 0) / total * 100, 2),
            round((estado_vehiculos['estado_malo'] or 0) / total * 100, 2)
        ]
        
        return jsonify({
            'stats_generales': stats_generales,
            'registros_diarios': {
                'labels': dias_labels,
                'data': dias_data
            },
            'estado_vehiculos': {
                'labels': estado_labels,
                'data': estado_data,
                'colores': ['#2dce89', '#fb6340', '#f5365c']  # Verde para Bueno, Naranja para Regular, Rojo para Malo
            },
            'centro_trabajo': {
                'labels': [ct['centro_de_trabajo'] for ct in stats_centro_trabajo],
                'data': [ct['cantidad'] for ct in stats_centro_trabajo]
            }
        })
        
    except Error as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reportes/vencimientos')
@login_required(role='administrativo')
def api_reportes_vencimientos():
    try:
        # Obtener parámetros de filtro
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        supervisor = request.args.get('supervisor')
        centro_trabajo = request.args.get('centro_trabajo')
        
        # Construir la cláusula WHERE base
        where_clauses = []
        params = []
        
        if fecha_inicio:
            where_clauses.append("p.fecha >= %s")
            params.append(fecha_inicio)
        if fecha_fin:
            where_clauses.append("p.fecha <= %s")
            params.append(fecha_fin)
        if supervisor:
            where_clauses.append("p.supervisor = %s")
            params.append(supervisor)
        if centro_trabajo:
            where_clauses.append("p.centro_de_trabajo = %s")
            params.append(centro_trabajo)
            
        # Agregar condiciones de vencimiento
        vencimiento_conditions = [
            "p.fecha_vencimiento_licencia IS NOT NULL AND p.fecha_vencimiento_licencia >= CURDATE() AND p.fecha_vencimiento_licencia <= DATE_ADD(CURDATE(), INTERVAL 30 DAY)",
            "p.fecha_vencimiento_soat IS NOT NULL AND p.fecha_vencimiento_soat >= CURDATE() AND p.fecha_vencimiento_soat <= DATE_ADD(CURDATE(), INTERVAL 30 DAY)",
            "p.fecha_vencimiento_tecnomecanica IS NOT NULL AND p.fecha_vencimiento_tecnomecanica >= CURDATE() AND p.fecha_vencimiento_tecnomecanica <= DATE_ADD(CURDATE(), INTERVAL 30 DAY)"
        ]
        
        where_sql = f"""
            ({" OR ".join(vencimiento_conditions)})
            {" AND " + " AND ".join(where_clauses) if where_clauses else ""}
        """
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        query = f"""
            SELECT 
                p.id_preoperacional,
                p.placa_vehiculo,
                p.fecha_vencimiento_licencia,
                p.fecha_vencimiento_soat,
                p.fecha_vencimiento_tecnomecanica,
                r.nombre as nombre_tecnico,
                DATEDIFF(p.fecha_vencimiento_licencia, CURDATE()) as dias_licencia,
                DATEDIFF(p.fecha_vencimiento_soat, CURDATE()) as dias_soat,
                DATEDIFF(p.fecha_vencimiento_tecnomecanica, CURDATE()) as dias_tecnomecanica
            FROM preoperacional p
            JOIN recurso_operativo r ON p.id_codigo_consumidor = r.id_codigo_consumidor
            WHERE {where_sql}
            ORDER BY 
                LEAST(
                    COALESCE(DATEDIFF(p.fecha_vencimiento_licencia, CURDATE()), 999),
                    COALESCE(DATEDIFF(p.fecha_vencimiento_soat, CURDATE()), 999),
                    COALESCE(DATEDIFF(p.fecha_vencimiento_tecnomecanica, CURDATE()), 999)
                )
        """
        
        cursor.execute(query, params)
        vencimientos = cursor.fetchall()
        cursor.close()
        connection.close()
        
        return jsonify({
            'vencimientos': vencimientos
        })
        
    except Error as e:
        return jsonify({'error': str(e)}), 500

@app.route('/verificar_vencimientos')
@login_required()
def verificar_vencimientos():
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener el último registro preoperacional del usuario
        cursor.execute("""
            SELECT 
                fecha_vencimiento_licencia,
                fecha_vencimiento_soat,
                fecha_vencimiento_tecnomecanica
            FROM preoperacional 
            WHERE id_codigo_consumidor = %s
            ORDER BY fecha DESC 
            LIMIT 1
        """, (session.get('user_id'),))
        
        ultimo_registro = cursor.fetchone()
        cursor.close()
        connection.close()
        
        if not ultimo_registro:
            return jsonify({'tiene_vencimientos': False})
            
        vencimientos = []
        # Usar fecha de Bogotá en lugar de la fecha del servidor
        fecha_actual = get_bogota_datetime().date()
        print(f"Verificando vencimientos con fecha de Bogotá: {fecha_actual}")
        
        # Verificar cada tipo de vencimiento
        if ultimo_registro['fecha_vencimiento_licencia']:
            dias_licencia = (ultimo_registro['fecha_vencimiento_licencia'] - fecha_actual).days
            if 0 <= dias_licencia <= 30:
                vencimientos.append({
                    'tipo': 'Licencia de Conducción',
                    'fecha': ultimo_registro['fecha_vencimiento_licencia'].strftime('%Y-%m-%d'),
                    'dias_restantes': dias_licencia
                })
                
        if ultimo_registro['fecha_vencimiento_soat']:
            dias_soat = (ultimo_registro['fecha_vencimiento_soat'] - fecha_actual).days
            if 0 <= dias_soat <= 30:
                vencimientos.append({
                    'tipo': 'SOAT',
                    'fecha': ultimo_registro['fecha_vencimiento_soat'].strftime('%Y-%m-%d'),
                    'dias_restantes': dias_soat
                })
                
        if ultimo_registro['fecha_vencimiento_tecnomecanica']:
            dias_tecno = (ultimo_registro['fecha_vencimiento_tecnomecanica'] - fecha_actual).days
            if 0 <= dias_tecno <= 30:
                vencimientos.append({
                    'tipo': 'Tecnomecánica',
                    'fecha': ultimo_registro['fecha_vencimiento_tecnomecanica'].strftime('%Y-%m-%d'),
                    'dias_restantes': dias_tecno
                })
        
        return jsonify({
            'tiene_vencimientos': len(vencimientos) > 0,
            'vencimientos': vencimientos,
            'fecha_bogota': fecha_actual.strftime('%Y-%m-%d')
        })
        
    except Error as e:
        print("Error en verificar_vencimientos:", str(e))
        return jsonify({'error': str(e)}), 500

@app.route('/verificar_registro_preoperacional')
@login_required(role=['tecnicos', 'operativo'])
def verificar_registro_preoperacional():
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Verificar si existe registro para el día actual en zona horaria de Bogotá
        fecha_actual = get_bogota_datetime().date()
        print(f"Verificando registro preoperacional para fecha Bogotá: {fecha_actual}")
        
        cursor.execute("""
            SELECT COUNT(*) as count, 
                   MAX(fecha) as ultimo_registro
            FROM preoperacional 
            WHERE id_codigo_consumidor = %s 
            AND DATE(CONVERT_TZ(fecha, '+00:00', '-05:00')) = %s
        """, (session.get('id_codigo_consumidor'), fecha_actual))
        
        resultado = cursor.fetchone()
        tiene_registro = resultado['count'] > 0
        ultimo_registro = resultado['ultimo_registro']
        
        # Convertir el último registro a zona horaria de Bogotá si existe
        ultimo_registro_str = None
        if ultimo_registro:
            ultimo_registro_bogota = convert_to_bogota_time(ultimo_registro)
            ultimo_registro_str = ultimo_registro_bogota.strftime('%Y-%m-%d %H:%M:%S')
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'tiene_registro': tiene_registro,
            'ultimo_registro': ultimo_registro_str,
            'fecha_actual': fecha_actual.strftime('%Y-%m-%d'),
            'hora_bogota': get_bogota_datetime().strftime('%H:%M:%S')
        })
        
    except Error as e:
        print(f"Error en verificar_registro_preoperacional: {str(e)}")
        return jsonify({'error': str(e)}), 500

# API Routes para Analistas
@app.route('/api/analistas/causas-cierre', methods=['GET'])
@login_required
def api_causas_cierre():
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify([])
        cursor = connection.cursor(dictionary=True)

        texto_busqueda = request.args.get('busqueda', '').strip()
        tecnologia = request.args.get('tecnologia', '').strip()
        agrupacion = request.args.get('agrupacion', '').strip()
        grupo = request.args.get('grupo', '').strip()

        query = """
            SELECT 
                idbase_causas_cierre,
                codigo_causas_cierre,
                tipo_causas_cierre,
                nombre_causas_cierre,
                tecnologia_causas_cierre,
                instrucciones_de_uso_causas_cierre,
                agrupaciones_causas_cierre,
                todos_los_grupos_causas_cierre,
                facturable_causas_cierre
            FROM base_causas_cierre
            WHERE 1=1
        """

        params = []

        if texto_busqueda:
            query += """
                AND (
                    codigo_causas_cierre LIKE %s OR 
                    nombre_causas_cierre LIKE %s OR 
                    instrucciones_de_uso_causas_cierre LIKE %s
                )
            """
            busqueda_param = f"%{texto_busqueda}%"
            params.extend([busqueda_param, busqueda_param, busqueda_param])

        if tecnologia:
            query += " AND tecnologia_causas_cierre = %s"
            params.append(tecnologia)

        if agrupacion:
            query += " AND agrupaciones_causas_cierre = %s"
            params.append(agrupacion)

        if grupo:
            query += " AND todos_los_grupos_causas_cierre = %s"
            params.append(grupo)

        query += " ORDER BY codigo_causas_cierre ASC"

        cursor.execute(query, params)
        resultados = cursor.fetchall()
        return jsonify(resultados)
    except mysql.connector.Error as e:
        errno = getattr(e, 'errno', None)
        logging.error(f"Error en API causas-cierre: {str(e)} (errno={errno})")
        if errno == 1146:
            return jsonify([])
        return jsonify([])
    except Exception as e:
        logging.error(f"Error inesperado en API causas-cierre: {str(e)}")
        return jsonify([])
    finally:
        if 'cursor' in locals() and cursor:
            try:
                cursor.close()
            except:
                pass
        if 'connection' in locals() and connection:
            try:
                if connection.is_connected():
                    connection.close()
            except:
                pass
def _normalizar_hora(valor):
    """Normaliza distintos formatos de hora a 'HH:MM'."""
    try:
        from datetime import datetime, time as dtime
        import re
        if not valor:
            return None
        # Tipos datetime/time
        if isinstance(valor, dtime):
            return valor.strftime('%H:%M')
        if isinstance(valor, datetime):
            return valor.strftime('%H:%M')

        # Convertir a texto y limpiar variantes
        s = str(valor).strip()
        s = s.replace('.', ':')
        s = (s
             .replace('a. m.', 'AM').replace('p. m.', 'PM')
             .replace('a.m.', 'AM').replace('p.m.', 'PM')
             .replace('A. M.', 'AM').replace('P. M.', 'PM')
             .replace('am', 'AM').replace('pm', 'PM'))
        s = re.sub(r'\s+', ' ', s).strip()

        # Intentar parseo con AM/PM
        for fmt in ('%I:%M %p', '%I:%M%p'):
            try:
                dt = datetime.strptime(s.upper(), fmt)
                return dt.strftime('%H:%M')
            except Exception:
                pass

        # Intentar formato 24h HH:MM
        m = re.search(r'(\d{1,2})[:](\d{2})', s)
        if m:
            h = int(m.group(1)); mnt = int(m.group(2))
            if 0 <= h <= 23 and 0 <= mnt <= 59:
                return f"{h:02d}:{mnt:02d}"

        # Dígitos 3-4 como 830 u 0830
        md = re.search(r'\b(\d{3,4})\b', s)
        if md:
            d = md.group(1)
            if len(d) == 3:
                h = int(d[0]); mnt = int(d[1:])
            else:
                h = int(d[:2]); mnt = int(d[2:])
            if 0 <= h <= 23 and 0 <= mnt <= 59:
                return f"{h:02d}:{mnt:02d}"

        return None
    except Exception:
        # Si ocurre cualquier error inesperado al normalizar, retornar None
        return None

@app.route('/api/analistas/tecnicos-asignados', methods=['GET'])
@login_required()
def main_api_tecnicos_asignados():
    """API endpoint para obtener técnicos asignados al analista actual con su información de asistencia"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener el nombre del analista actual desde la sesión
        analista_nombre = session.get('user_name') if session.get('user_name') else None
        
        if not analista_nombre:
            return jsonify({'error': 'No se pudo identificar al analista actual'}), 401
            
        # Obtener parámetro de fecha (por defecto fecha actual)
        fecha_filtro = request.args.get('fecha')
        if not fecha_filtro:
            from datetime import date
            fecha_filtro = date.today().strftime('%Y-%m-%d')
        
        # Consulta para obtener técnicos asignados al analista actual
        query_tecnicos = """
            SELECT DISTINCT
                ro.id_codigo_consumidor,
                ro.recurso_operativo_cedula as cedula,
                ro.nombre as tecnico,
                ro.carpeta,
                ro.cliente,
                ro.ciudad,
                ro.super as supervisor,
                ro.analista
            FROM recurso_operativo ro
            WHERE ro.analista = %s
            AND ro.estado = 'Activo'
            ORDER BY ro.nombre
        """
        
        cursor.execute(query_tecnicos, (analista_nombre,))
        tecnicos = cursor.fetchall()
        
        # Para cada técnico, obtener información de asistencia del día actual
        tecnicos_con_asistencia = []
        
        for tecnico in tecnicos:
            # Buscar asistencia del día filtrado
            query_asistencia = """
                SELECT 
                    a.carpeta_dia,
                    a.fecha_asistencia,
                    a.hora_inicio,
                    a.estado,
                    a.novedad,
                    ta.nombre_tipificacion,
                    ta.codigo_tipificacion,
                    ta.grupo as grupo_tipificacion,
                    pc.presupuesto_eventos,
                    pc.presupuesto_diario,
                    pc.presupuesto_carpeta as nombre_presupuesto_carpeta
                FROM asistencia a
                LEFT JOIN tipificacion_asistencia ta ON a.carpeta_dia = ta.codigo_tipificacion
                LEFT JOIN presupuesto_carpeta pc ON ta.nombre_tipificacion = pc.presupuesto_carpeta
                WHERE a.cedula = %s
                AND DATE(a.fecha_asistencia) = %s
                ORDER BY a.fecha_asistencia DESC
                LIMIT 1
            """
            
            cursor.execute(query_asistencia, (tecnico['cedula'], fecha_filtro))
            asistencia = cursor.fetchone()
            
            # Agregar información de asistencia al técnico
            tecnico_info = {
                'id_codigo_consumidor': tecnico['id_codigo_consumidor'],
                'cedula': tecnico['cedula'],
                'tecnico': tecnico['tecnico'],
                'carpeta': tecnico['carpeta'],
                'cliente': tecnico['cliente'],
                'ciudad': tecnico['ciudad'],
                'supervisor': tecnico['supervisor'],
                'analista': tecnico['analista'],
                'asistencia_hoy': {
                    'carpeta_dia': asistencia['carpeta_dia'] if asistencia else None,
                    'fecha_asistencia': asistencia['fecha_asistencia'].isoformat() if asistencia and asistencia['fecha_asistencia'] else None,
                    'hora_inicio': _normalizar_hora(asistencia['hora_inicio']) if asistencia and asistencia['hora_inicio'] else None,
                    'estado': asistencia['estado'] if asistencia else None,
                    'novedad': asistencia['novedad'] if asistencia else None,
                    'tipificacion': asistencia['nombre_tipificacion'] if asistencia else 'Sin registro',
                    'codigo_tipificacion': asistencia['codigo_tipificacion'] if asistencia else None,
                    'grupo_tipificacion': asistencia['grupo_tipificacion'] if asistencia else None,
                    'presupuesto_eventos': asistencia['presupuesto_eventos'] if asistencia else None,
                    'presupuesto_diario': asistencia['presupuesto_diario'] if asistencia else None,
                    'nombre_presupuesto_carpeta': asistencia['nombre_presupuesto_carpeta'] if asistencia else None
                }
            }
            
            tecnicos_con_asistencia.append(tecnico_info)
        
        return jsonify({
            'success': True,
            'analista': analista_nombre,
            'total_tecnicos': len(tecnicos_con_asistencia),
            'tecnicos': tecnicos_con_asistencia
        })
        
    except mysql.connector.Error as e:
        logging.error(f"Error en API técnicos asignados: {str(e)}")
        return jsonify({'error': 'Error al consultar la base de datos'}), 500
    except Exception as e:
        logging.error(f"Error inesperado en API técnicos asignados: {str(e)}")
        return jsonify({'error': 'Error interno del servidor'}), 500
    finally:
        try:
            if 'cursor' in locals() and cursor:
                cursor.close()
        except:
            pass
        try:
            if 'connection' in locals() and connection and connection.is_connected():
                connection.close()
        except:
            pass

@app.route('/logistica/asignaciones')
@login_required()
@role_required('logistica')
def ver_asignaciones():
    try:
        # Verificar stock bajo para mostrar alertas
        materiales_problematicos = verificar_stock_bajo()
        
        connection = get_db_connection()
        if connection is None:
            return render_template('error.html', 
                               mensaje='Error de conexión a la base de datos',
                               error='No se pudo establecer conexión con la base de datos')
                               
        cursor = connection.cursor(dictionary=True)
        
        # Obtener todas las asignaciones con información del técnico
        cursor.execute("""
            SELECT a.*, r.nombre, r.recurso_operativo_cedula 
            FROM asignacion a 
            LEFT JOIN recurso_operativo r ON a.id_codigo_consumidor = r.id_codigo_consumidor 
            ORDER BY a.asignacion_fecha DESC
        """)
        asignaciones = cursor.fetchall()

        # Obtener lista de técnicos disponibles
        cursor.execute("""
            SELECT id_codigo_consumidor, nombre, recurso_operativo_cedula, cargo 
            FROM recurso_operativo 
            WHERE cargo LIKE '%TECNICO%' OR cargo LIKE '%TÉCNICO%'
            ORDER BY nombre
        """)
        tecnicos = cursor.fetchall()
        
        cursor.close()
        connection.close()

        return render_template('modulos/logistica/asignaciones.html', 
                            asignaciones=asignaciones,
                            tecnicos=tecnicos,
                            materiales_problematicos=materiales_problematicos)

    except Exception as e:
        print(f"Error al obtener asignaciones: {str(e)}")
        return render_template('error.html', 
                           mensaje='Error al cargar las asignaciones',
                           error=str(e))

@app.route('/logistica/registrar_asignacion', methods=['POST'])
@login_required()
@role_required('logistica')
def registrar_asignacion():
    connection = None
    cursor = None
    try:
        # Obtener datos básicos
        id_codigo_consumidor = request.form.get('id_codigo_consumidor')
        fecha = request.form.get('fecha')
        cargo = request.form.get('cargo')

        # Validar campos requeridos
        if not all([id_codigo_consumidor, fecha, cargo]):
            return jsonify({
                'status': 'error',
                'message': 'Los campos ID, fecha y cargo son requeridos'
            }), 400

        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'status': 'error',
                'message': 'Error de conexión a la base de datos'
            }), 500

        cursor = connection.cursor(dictionary=True)

        # Verificar que el técnico existe
        cursor.execute('SELECT nombre FROM recurso_operativo WHERE id_codigo_consumidor = %s', (id_codigo_consumidor,))
        tecnico = cursor.fetchone()
        
        if not tecnico:
            return jsonify({
                'status': 'error',
                'message': 'El técnico seleccionado no existe'
            }), 404

        # Obtener la estructura de la tabla
        cursor.execute("""
            SELECT COLUMN_NAME 
            FROM INFORMATION_SCHEMA.COLUMNS 
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME = 'asignacion'
        """)
        columnas_existentes = {row['COLUMN_NAME'].lower() for row in cursor.fetchall()}
        
        print("Columnas existentes:", columnas_existentes)  # Debug log

        # Campos base que siempre deben existir
        campos = ['id_codigo_consumidor', 'asignacion_fecha', 'asignacion_cargo']
        valores = [id_codigo_consumidor, fecha, cargo]

        # Lista de todas las herramientas posibles
        herramientas = [
            'adaptador_mandril', 'alicate', 'barra_45cm', 'bisturi_metalico',
            'broca_3/8', 'broca_3/8/6_ran', 'broca_1/2/6_ran',
            'broca_metal/madera_1/4', 'broca_metal/madera_3/8', 'broca_metal/madera_5/16',
            'caja_de_herramientas', 'cajon_rojo', 'cinta_de_senal', 'cono_retractil',
            'cortafrio', 'destor_de_estrella', 'destor_de_pala', 'destor_tester',
            'espatula', 'exten_de_corr_10_mts', 'llave_locking_male', 'llave_reliance',
            'llave_torque_rg_11', 'llave_torque_rg_6', 'llaves_mandril',
            'mandril_para_taladro', 'martillo_de_una', 'multimetro',
            'pelacable_rg_6y_rg_11', 'pinza_de_punta', 'pistola_de_silicona',
            'planillero', 'ponchadora_rg_6_y_rg_11', 'ponchadora_rj_45_y_rj11',
            'probador_de_tonos', 'probador_de_tonos_utp', 'puntas_para_multimetro',
            'sonda_metalica', 'tacos_de_madera', 'taladro_percutor',
            'telefono_de_pruebas', 'power_miter', 'bfl_laser', 'cortadora',
            'stripper_fibra', 'pelachaqueta'
        ]

        # Procesar cada herramienta
        for herramienta in herramientas:
            nombre_campo = f'asignacion_{herramienta}'
            if nombre_campo.lower() in columnas_existentes:
                campos.append(nombre_campo)
                valor = request.form.get(herramienta, '0')
                valores.append(valor if valor == '1' else '0')
            else:
                print(f"Campo ignorado (no existe en la tabla): {nombre_campo}")

        # Agregar estado por defecto si existe la columna
        if 'asignacion_estado' in columnas_existentes:
            campos.append('asignacion_estado')
            valores.append('1')

        # Construir y ejecutar la consulta SQL
        sql = f"""
            INSERT INTO asignacion ({', '.join(campos)})
            VALUES ({', '.join(['%s'] * len(valores))})
        """
        
        print("SQL Query:", sql)  # Debug log
        print("Valores:", valores)  # Debug log

        cursor.execute(sql, tuple(valores))
        connection.commit()

        return jsonify({
            'status': 'success',
            'message': 'Asignación registrada exitosamente'
        }), 201

    except mysql.connector.Error as e:
        print(f"Error MySQL: {str(e)}")  # Debug log
        return jsonify({
            'status': 'error',
            'message': f'Error en la base de datos: {str(e)}'
        }), 500
    except Exception as e:
        print(f"Error general: {str(e)}")  # Debug log
        return jsonify({
            'status': 'error',
            'message': f'Error al registrar la asignación: {str(e)}'
        }), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/logistica/asignacion/<int:id_asignacion>')
@login_required()
@role_required('logistica')
def ver_detalle_asignacion(id_asignacion):
    connection = None
    cursor = None
    
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'status': 'error',
                'message': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener los detalles de la asignación
        cursor.execute("""
            SELECT a.*, r.nombre, r.recurso_operativo_cedula, r.cargo
            FROM asignacion a 
            LEFT JOIN recurso_operativo r ON a.id_codigo_consumidor = r.id_codigo_consumidor 
            WHERE a.id_asignacion = %s
        """, (id_asignacion,))
        
        asignacion = cursor.fetchone()
        
        if not asignacion:
            return jsonify({
                'status': 'error',
                'message': 'No se encontró la asignación'
            }), 404

        # Función auxiliar para formatear valores
        def formatear_valor(valor):
            if valor is None:
                return '0'
            return str(valor)

        # Organizar los datos
        detalles = {
            'info_basica': {
                'tecnico': asignacion['nombre'],
                'cedula': asignacion['recurso_operativo_cedula'],
                'cargo': asignacion.get('asignacion_cargo', 'No especificado'),
                'fecha': asignacion['asignacion_fecha'].strftime('%Y-%m-%d %H:%M:%S'),
                'estado': 'Activo' if str(asignacion.get('asignacion_estado', '0')) == '1' else 'Inactivo'
            },
            'herramientas_basicas': {
                'adaptador_mandril': formatear_valor(asignacion.get('asignacion_adaptador_mandril')),
                'alicate': formatear_valor(asignacion.get('asignacion_alicate')),
                'barra_45cm': formatear_valor(asignacion.get('asignacion_barra_45cm')),
                'bisturi_metalico': formatear_valor(asignacion.get('asignacion_bisturi_metalico')),
                'caja_de_herramientas': formatear_valor(asignacion.get('asignacion_caja_de_herramientas')),
                'cortafrio': formatear_valor(asignacion.get('asignacion_cortafrio')),
                'destor_de_estrella': formatear_valor(asignacion.get('asignacion_destor_de_estrella')),
                'destor_de_pala': formatear_valor(asignacion.get('asignacion_destor_de_pala')),
                'destor_tester': formatear_valor(asignacion.get('asignacion_destor_tester')),
                'espatula': formatear_valor(asignacion.get('asignacion_espatula')),
                'martillo_de_una': formatear_valor(asignacion.get('asignacion_martillo_de_una')),
                'pinza_de_punta': formatear_valor(asignacion.get('asignacion_pinza_de_punta'))
            },
            'brocas': {
                'broca_3/8': formatear_valor(asignacion.get('asignacion_broca_3/8')),
                'broca_3/8/6_ran': formatear_valor(asignacion.get('asignacion_broca_3/8/6_ran')),
                'broca_1/2/6_ran': formatear_valor(asignacion.get('asignacion_broca_1/2/6_ran')),
                'broca_metal/madera_1/4': formatear_valor(asignacion.get('asignacion_broca_metal/madera_1/4')),
                'broca_metal/madera_3/8': formatear_valor(asignacion.get('asignacion_broca_metal/madera_3/8')),
                'broca_metal/madera_5/16': formatear_valor(asignacion.get('asignacion_broca_metal/madera_5/16'))
            },
            'herramientas_red': {
                'cajon_rojo': formatear_valor(asignacion.get('asignacion_cajon_rojo')),
                'cinta_de_senal': formatear_valor(asignacion.get('asignacion_cinta_de_senal')),
                'cono_retractil': formatear_valor(asignacion.get('asignacion_cono_retractil')),
                'exten_de_corr_10_mts': formatear_valor(asignacion.get('asignacion_exten_de_corr_10_mts')),
                'llave_locking_male': formatear_valor(asignacion.get('asignacion_llave_locking_male')),
                'llave_reliance': formatear_valor(asignacion.get('asignacion_llave_reliance')),
                'llave_torque_rg_11': formatear_valor(asignacion.get('asignacion_llave_torque_rg_11')),
                'llave_torque_rg_6': formatear_valor(asignacion.get('asignacion_llave_torque_rg_6')),
                'llaves_mandril': formatear_valor(asignacion.get('asignacion_llaves_mandril')),
                'mandril_para_taladro': formatear_valor(asignacion.get('asignacion_mandril_para_taladro')),
                'pelacable_rg_6y_rg_11': formatear_valor(asignacion.get('asignacion_pelacable_rg_6y_rg_11')),
                'pistola_de_silicona': formatear_valor(asignacion.get('asignacion_pistola_de_silicona')),
                'planillero': formatear_valor(asignacion.get('asignacion_planillero')),
                'ponchadora_rg_6_y_rg_11': formatear_valor(asignacion.get('asignacion_ponchadora_rg_6_y_rg_11')),
                'ponchadora_rj_45_y_rj11': formatear_valor(asignacion.get('asignacion_ponchadora_rj_45_y_rj11')),
                'probador_de_tonos': formatear_valor(asignacion.get('asignacion_probador_de_tonos')),
                'probador_de_tonos_utp': formatear_valor(asignacion.get('asignacion_probador_de_tonos_utp')),
                'puntas_para_multimetro': formatear_valor(asignacion.get('asignacion_puntas_para_multimetro')),
                'sonda_metalica': formatear_valor(asignacion.get('asignacion_sonda_metalica')),
                'tacos_de_madera': formatear_valor(asignacion.get('asignacion_tacos_de_madera')),
                'telefono_de_pruebas': formatear_valor(asignacion.get('asignacion_telefono_de_pruebas')),
                'power_miter': formatear_valor(asignacion.get('asignacion_power_miter')),
                'bfl_laser': formatear_valor(asignacion.get('asignacion_bfl_laser')),
                'cortadora': formatear_valor(asignacion.get('asignacion_cortadora')),
                'stripper_fibra': formatear_valor(asignacion.get('asignacion_stripper_fibra')),
                'pelachaqueta': formatear_valor(asignacion.get('asignacion_pelachaqueta'))
            }
        }

        print("Detalles a enviar:", detalles)  # Debug log
        
        return jsonify({
            'status': 'success',
            'data': detalles
        })

    except Exception as e:
        print(f"Error al obtener detalles de asignación: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error al obtener detalles de la asignación: {str(e)}'
        }), 500
        
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/logistica/exportar_asignaciones_csv')
@login_required()
@role_required('logistica')
def exportar_asignaciones_csv():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener todas las asignaciones con información del técnico
        cursor.execute("""
            SELECT 
                a.*,
                r.nombre as nombre_tecnico,
                r.recurso_operativo_cedula as cedula_tecnico,
                r.cargo as cargo_tecnico
            FROM asignacion a 
            LEFT JOIN recurso_operativo r ON a.id_codigo_consumidor = r.id_codigo_consumidor 
            ORDER BY a.asignacion_fecha DESC
        """)
        asignaciones = cursor.fetchall()
        
        # Crear archivo CSV en memoria
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Escribir encabezados
        writer.writerow([
            'ID Asignación',
            'Fecha',
            'Técnico',
            'Cédula',
            'Cargo',
            'Estado',
            'Adaptador Mandril',
            'Alicate',
            'Barra 45cm',
            'Bisturí Metálico',
            'Caja de Herramientas',
            'Cortafrío',
            'Destornillador Estrella',
            'Destornillador Pala',
            'Destornillador Tester',
            'Martillo',
            'Pinza de Punta',
            'Multímetro',
            'Taladro Percutor',
            'Ponchadora RG6/RG11',
            'Ponchadora RJ45/RJ11',
            'Power Meter',
            'BFL Laser',
            'Cortadora',
            'Stripper Fibra',
            'Arnés',
            'Eslinga',
            'Casco Tipo II',
            'Araña Casco',
            'Barbuquejo',
            'Guantes de Vaqueta',
            'Gafas',
            'Línea de Vida'
        ])
        
        # Escribir datos
        for asignacion in asignaciones:
            writer.writerow([
                asignacion['id_asignacion'],
                asignacion['asignacion_fecha'].strftime('%Y-%m-%d %H:%M:%S'),
                asignacion.get('nombre_tecnico', 'No asignado'),
                asignacion.get('cedula_tecnico', 'No disponible'),
                asignacion.get('asignacion_cargo', 'No especificado'),
                'Activo' if str(asignacion.get('asignacion_estado', '0')) == '1' else 'Inactivo',
                asignacion.get('asignacion_adaptador_mandril', '0'),
                asignacion.get('asignacion_alicate', '0'),
                asignacion.get('asignacion_barra_45cm', '0'),
                asignacion.get('asignacion_bisturi_metalico', '0'),
                asignacion.get('asignacion_caja_de_herramientas', '0'),
                asignacion.get('asignacion_cortafrio', '0'),
                asignacion.get('asignacion_destor_de_estrella', '0'),
                asignacion.get('asignacion_destor_de_pala', '0'),
                asignacion.get('asignacion_destor_tester', '0'),
                asignacion.get('asignacion_martillo_de_una', '0'),
                asignacion.get('asignacion_pinza_de_punta', '0'),
                asignacion.get('asignacion_multimetro', '0'),
                asignacion.get('asignacion_taladro_percutor', '0'),
                asignacion.get('asignacion_ponchadora_rg_6_y_rg_11', '0'),
                asignacion.get('asignacion_ponchadora_rj_45_y_rj11', '0'),
                asignacion.get('asignacion_power_miter', '0'),
                asignacion.get('asignacion_bfl_laser', '0'),
                asignacion.get('asignacion_cortadora', '0'),
                asignacion.get('asignacion_stripper_fibra', '0'),
                
            ])
        
        # Preparar respuesta
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),  # Usar UTF-8 con BOM para soporte de caracteres especiales
            mimetype='text/csv; charset=utf-8-sig',
            as_attachment=True,
            download_name=f'asignaciones_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )
        
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/logistica/inventario')
@login_required()
@role_required('logistica')
def ver_inventario():
    try:
        # Verificar stock bajo para mostrar alertas
        materiales_problematicos = verificar_stock_bajo()
        
        connection = get_db_connection()
        if connection is None:
            return render_template('error.html', 
                               mensaje='Error de conexión a la base de datos',
                               error='No se pudo establecer conexión con la base de datos')
                               
        cursor = connection.cursor(dictionary=True)
        
        # Obtener conteo de herramientas básicas asignadas
        herramientas_basicas_query = """
            SELECT 
                SUM(asignacion_adaptador_mandril) as adaptador_mandril,
                SUM(asignacion_alicate) as alicate,
                SUM(asignacion_barra_45cm) as barra_45cm,
                SUM(asignacion_bisturi_metalico) as bisturi_metalico,
                SUM(asignacion_caja_de_herramientas) as caja_herramientas,
                SUM(asignacion_cortafrio) as cortafrio,
                SUM(asignacion_destor_de_estrella) as destornillador_estrella,
                SUM(asignacion_destor_de_pala) as destornillador_pala,
                SUM(asignacion_destor_tester) as destornillador_tester,
                SUM(asignacion_martillo_de_una) as martillo,
                SUM(asignacion_pinza_de_punta) as pinza_punta
            FROM asignacion 
            WHERE asignacion_estado = '1'
        """
        cursor.execute(herramientas_basicas_query)
        herramientas_basicas = cursor.fetchone()
        
        # Obtener conteo de herramientas especializadas asignadas
        herramientas_especializadas_query = """
            SELECT 
                SUM(asignacion_multimetro) as multimetro,
                SUM(asignacion_taladro_percutor) as taladro_percutor,
                SUM(asignacion_ponchadora_rg_6_y_rg_11) as ponchadora_rg6_rg11,
                SUM(asignacion_ponchadora_rj_45_y_rj11) as ponchadora_rj45_rj11,
                SUM(asignacion_power_miter) as power_meter,
                SUM(asignacion_bfl_laser) as bfl_laser,
                SUM(asignacion_cortadora) as cortadora,
                SUM(asignacion_stripper_fibra) as stripper_fibra
            FROM asignacion 
            WHERE asignacion_estado = '1'
        """
        cursor.execute(herramientas_especializadas_query)
        herramientas_especializadas = cursor.fetchone()
        
       

        # Obtener conteo de material ferretero asignado
        material_ferretero_query = """
            SELECT 
                SUM(silicona) as silicona,
                SUM(amarres_negros) as amarres_negros,
                SUM(amarres_blancos) as amarres_blancos,
                SUM(cinta_aislante) as cinta_aislante,
                SUM(grapas_blancas) as grapas_blancas,
                SUM(grapas_negras) as grapas_negras
            FROM ferretero
        """
        cursor.execute(material_ferretero_query)
        material_ferretero = cursor.fetchone()
        
        # Obtener historial de movimientos de inventario (para gráficos de tendencia)
        historial_movimientos_query = """
            SELECT 
                DATE(fecha_movimiento) as fecha,
                tipo_movimiento,
                item_afectado,
                cantidad,
                usuario,
                ubicacion
            FROM movimientos_inventario
            ORDER BY fecha_movimiento DESC
            LIMIT 100
        """
        cursor.execute(historial_movimientos_query)
        historial_movimientos = cursor.fetchall()
        
        # Generar datos para gráficos de tendencias
        ultimos_30_dias = []
        hoy = datetime.now().date()
        for i in range(30, -1, -1):
            fecha = hoy - timedelta(days=i)
            ultimos_30_dias.append(fecha.strftime('%Y-%m-%d'))
        
        # Conteo de movimientos por fecha para gráfico de tendencia
        movimientos_por_dia = {}
        for fecha in ultimos_30_dias:
            movimientos_por_dia[fecha] = 0
            
        for movimiento in historial_movimientos:
            fecha_str = movimiento['fecha'].strftime('%Y-%m-%d')
            if fecha_str in movimientos_por_dia:
                movimientos_por_dia[fecha_str] += 1
                
        # Conteo por categoría para gráficos de distribución
        categorias = {
            'Herramientas Básicas': sum(value for key, value in herramientas_basicas.items() if value is not None),
            'Herramientas Especializadas': sum(value for key, value in herramientas_especializadas.items() if value is not None),
            'Material Ferretero': sum(value for key, value in material_ferretero.items() if value is not None)
        }
        
        # Información para alertas (items con nivel crítico)
        items_criticos = []
        
        # Definir umbrales para cada tipo de elemento
        umbrales_minimos = {
            'adaptador_mandril': 5, 'alicate': 10, 'barra_45cm': 5,
            'bisturi_metalico': 15, 'caja_herramientas': 10, 'cortafrio': 10,
            'destornillador_estrella': 15, 'destornillador_pala': 15,
            'destornillador_tester': 10, 'martillo': 10, 'pinza_punta': 10,
            'multimetro': 5, 'taladro_percutor': 3, 'ponchadora_rg6_rg11': 5,
            'ponchadora_rj45_rj11': 5, 'power_meter': 3, 'bfl_laser': 3,
            'cortadora': 3, 'stripper_fibra': 3, 'arnes': 10, 'eslinga': 10,
            'casco_tipo_ii': 20, 'arana_casco': 20, 'barbuquejo': 20,
            'guantes_vaqueta': 30, 'gafas': 30, 'linea_vida': 5,
            'silicona': 10, 'amarres_negros': 50, 'amarres_blancos': 50,
            'cinta_aislante': 20, 'grapas_blancas': 100, 'grapas_negras': 100
        }
        
        # Verificar herramientas básicas bajo umbral
        for key, value in herramientas_basicas.items():
            if key in umbrales_minimos and value is not None:
                if value <= umbrales_minimos[key]:
                    nombre_item = key.replace('_', ' ').title()
                    items_criticos.append({
                        'item': nombre_item, 
                        'actual': value, 
                        'minimo': umbrales_minimos[key],
                        'categoria': 'Herramientas Básicas'
                    })
        
        # Verificar herramientas especializadas bajo umbral
        for key, value in herramientas_especializadas.items():
            if key in umbrales_minimos and value is not None:
                if value <= umbrales_minimos[key]:
                    nombre_item = key.replace('_', ' ').title()
                    items_criticos.append({
                        'item': nombre_item, 
                        'actual': value, 
                        'minimo': umbrales_minimos[key],
                        'categoria': 'Herramientas Especializadas'
                    })
        
      
        
        # Verificar material ferretero bajo umbral
        for key, value in material_ferretero.items():
            if key in umbrales_minimos and value is not None:
                if value <= umbrales_minimos[key]:
                    nombre_item = key.replace('_', ' ').title()
                    items_criticos.append({
                        'item': nombre_item, 
                        'actual': value, 
                        'minimo': umbrales_minimos[key],
                        'categoria': 'Material Ferretero'
                    })
        
        # Obtener datos para comparativa periódica (mes actual vs mes anterior)
        # Esta es una simulación - en un caso real se obtendría de los registros históricos
        mes_actual = datetime.now().month
        año_actual = datetime.now().year
        
        comparativa_meses = {
            'labels': ['Mes Anterior', 'Mes Actual'],
            'datos': [215, 243]  # Datos de ejemplo
        }
        
        # Simulación de datos para ubicaciones
        ubicaciones = [
            {'nombre': 'Almacén Principal', 'items': 452},
            {'nombre': 'Bodega Norte', 'items': 230},
            {'nombre': 'Bodega Sur', 'items': 195},
            {'nombre': 'Vehículos', 'items': 123},
            {'nombre': 'Personal', 'items': 87}
        ]
        
        cursor.close()
        connection.close()

        return render_template('modulos/logistica/inventario.html',
                           herramientas_basicas=herramientas_basicas,
                           herramientas_especializadas=herramientas_especializadas,
                           
                           material_ferretero=material_ferretero,
                           historial_movimientos=historial_movimientos,
                           movimientos_por_dia=movimientos_por_dia,
                           ultimos_30_dias=ultimos_30_dias,
                           categorias=categorias,
                           items_criticos=items_criticos,
                           comparativa_meses=comparativa_meses,
                           ubicaciones=ubicaciones,
                           materiales_problematicos=materiales_problematicos)

    except Exception as e:
        print(f"Error al obtener inventario: {str(e)}")
        return render_template('error.html',
                           mensaje='Error al cargar el inventario',
                           error=str(e))

@app.route('/logistica/ferretero')
@login_required()
@role_required('logistica')
def ver_asignaciones_ferretero():
    try:
        # Verificar stock bajo para mostrar alertas
        materiales_problematicos = verificar_stock_bajo()
        
        connection = get_db_connection()
        if connection is None:
            return render_template('error.html', 
                               mensaje='Error de conexión a la base de datos',
                               error='No se pudo establecer conexión con la base de datos')
                               
        cursor = connection.cursor(dictionary=True)
        
        # Filtro de fecha: mostrar solo registros del día seleccionado (por defecto, hoy)
        fecha_seleccionada = request.args.get('fecha') or datetime.now().strftime('%Y-%m-%d')
        cursor.execute("""
            SELECT f.*, r.nombre, r.recurso_operativo_cedula, r.cargo
            FROM ferretero f 
            LEFT JOIN recurso_operativo r ON f.id_codigo_consumidor = r.id_codigo_consumidor 
            WHERE DATE(f.fecha_asignacion) = %s
            ORDER BY f.fecha_asignacion DESC
        """, (fecha_seleccionada,))
        asignaciones = cursor.fetchall()

        # Obtener lista de técnicos disponibles
        cursor.execute("""
            SELECT id_codigo_consumidor, nombre, recurso_operativo_cedula, cargo, carpeta 
            FROM recurso_operativo 
            WHERE cargo LIKE '%TECNICO%' OR cargo LIKE '%TÉCNICO%'
            ORDER BY nombre
        """)
        tecnicos = cursor.fetchall()
        
        # Preparar información de límites para cada técnico
        fecha_actual = datetime.now()
        tecnicos_con_limites = []
        
        # Función para obtener límites desde la base de datos
        def obtener_limites_desde_db(cursor):
            """Obtiene los límites configurables desde la base de datos"""
            try:
                cursor.execute("""
                    SELECT 
                        area_trabajo,
                        material_tipo,
                        cantidad_limite,
                        periodo_dias,
                        unidad_medida
                    FROM limites_ferretero 
                    WHERE activo = TRUE
                    ORDER BY area_trabajo, material_tipo
                """)
                limites_db = cursor.fetchall()
                
                # Convertir a estructura de diccionario
                limites = {}
                for limite in limites_db:
                    area = limite['area_trabajo']
                    material = limite['material_tipo']
                    
                    if area not in limites:
                        limites[area] = {}
                    
                    limites[area][material] = {
                        'cantidad': limite['cantidad_limite'],
                        'periodo': limite['periodo_dias'],
                        'unidad': limite['unidad_medida']
                    }
                
                return limites
                
            except Exception as e:
                print(f"Error al obtener límites desde DB: {str(e)}")
                # Fallback a límites por defecto en caso de error
                return {
                    'FTTH INSTALACIONES': {
                        'cinta_aislante': {'cantidad': 3, 'periodo': 15, 'unidad': 'días'},
                        'silicona': {'cantidad': 16, 'periodo': 7, 'unidad': 'días'},
                        'amarres_negros': {'cantidad': 50, 'periodo': 7, 'unidad': 'días'},
                        'amarres_blancos': {'cantidad': 50, 'periodo': 7, 'unidad': 'días'},
                        'grapas_blancas': {'cantidad': 100, 'periodo': 7, 'unidad': 'días'},
                        'grapas_negras': {'cantidad': 100, 'periodo': 7, 'unidad': 'días'}
                    }
                }
        
        # Obtener límites configurables desde la base de datos
        limites = obtener_limites_desde_db(cursor)
        
        for tecnico in tecnicos:
            # Determinar área de trabajo basada EXCLUSIVAMENTE en la carpeta asignada
            carpeta = tecnico.get('carpeta', '').upper() if tecnico.get('carpeta') else ''
            
            area_trabajo = None
            
            # Buscar coincidencia exacta en la carpeta del técnico
            if carpeta:
                for area in limites.keys():
                    if area in carpeta:
                        area_trabajo = area
                        print(f"INFO: Técnico {tecnico.get('nombre', 'N/A')} - Carpeta: '{carpeta}' -> Área asignada: '{area_trabajo}'")
                        break
            
            # Si no se encuentra carpeta válida, mostrar error y omitir técnico
            if area_trabajo is None:
                print(f"ERROR: Técnico {tecnico.get('nombre', 'N/A')} - Carpeta '{carpeta}' no está configurada para límites de ferretero")
                continue  # Omitir este técnico de la lista
            
            # Obtener asignaciones previas para este técnico
            cursor.execute("""
                SELECT 
                    fecha_asignacion,
                    silicona,
                    amarres_negros,
                    amarres_blancos,
                    cinta_aislante,
                    grapas_blancas,
                    grapas_negras
                FROM ferretero 
                WHERE id_codigo_consumidor = %s
                ORDER BY fecha_asignacion DESC
            """, (tecnico['id_codigo_consumidor'],))
            asignaciones_tecnico = cursor.fetchall()
            
            # Inicializar contadores para materiales en los períodos correspondientes
            contadores = {
                'cinta_aislante': 0,
                'silicona': 0,
                'amarres_negros': 0,
                'amarres_blancos': 0,
                'grapas_blancas': 0,
                'grapas_negras': 0
            }
            
            # Calcular consumo previo en los periodos correspondientes
            for asignacion in asignaciones_tecnico:
                fecha_asignacion = asignacion['fecha_asignacion']
                diferencia_dias = (fecha_actual - fecha_asignacion).days
                
                # Verificar límite de cintas
                if diferencia_dias <= limites[area_trabajo]['cinta_aislante']['periodo']:
                    contadores['cinta_aislante'] += int(asignacion.get('cinta_aislante', 0) or 0)
                    
                # Verificar límite de siliconas
                if diferencia_dias <= limites[area_trabajo]['silicona']['periodo']:
                    contadores['silicona'] += int(asignacion.get('silicona', 0) or 0)
                    
                # Verificar límite de amarres negros
                if diferencia_dias <= limites[area_trabajo]['amarres_negros']['periodo']:
                    contadores['amarres_negros'] += int(asignacion.get('amarres_negros', 0) or 0)
                    
                # Verificar límite de amarres blancos
                if diferencia_dias <= limites[area_trabajo]['amarres_blancos']['periodo']:
                    contadores['amarres_blancos'] += int(asignacion.get('amarres_blancos', 0) or 0)
                    
                # Verificar límite de grapas blancas
                if diferencia_dias <= limites[area_trabajo]['grapas_blancas']['periodo']:
                    contadores['grapas_blancas'] += int(asignacion.get('grapas_blancas', 0) or 0)
                    
                # Verificar límite de grapas negras
                if diferencia_dias <= limites[area_trabajo]['grapas_negras']['periodo']:
                    contadores['grapas_negras'] += int(asignacion.get('grapas_negras', 0) or 0)
            
            # Calcular límites disponibles
            limites_disponibles = {
                'area': area_trabajo,
                'cinta_aislante': max(0, limites[area_trabajo]['cinta_aislante']['cantidad'] - contadores['cinta_aislante']),
                'silicona': max(0, limites[area_trabajo]['silicona']['cantidad'] - contadores['silicona']),
                'amarres_negros': max(0, limites[area_trabajo]['amarres_negros']['cantidad'] - contadores['amarres_negros']),
                'amarres_blancos': max(0, limites[area_trabajo]['amarres_blancos']['cantidad'] - contadores['amarres_blancos']),
                'grapas_blancas': max(0, limites[area_trabajo]['grapas_blancas']['cantidad'] - contadores['grapas_blancas']),
                'grapas_negras': max(0, limites[area_trabajo]['grapas_negras']['cantidad'] - contadores['grapas_negras']),
                'periodos': {
                    'cinta_aislante': f"{limites[area_trabajo]['cinta_aislante']['periodo']} {limites[area_trabajo]['cinta_aislante']['unidad']}",
                    'silicona': f"{limites[area_trabajo]['silicona']['periodo']} {limites[area_trabajo]['silicona']['unidad']}",
                    'amarres_negros': f"{limites[area_trabajo]['amarres_negros']['periodo']} {limites[area_trabajo]['amarres_negros']['unidad']}",
                    'amarres_blancos': f"{limites[area_trabajo]['amarres_blancos']['periodo']} {limites[area_trabajo]['amarres_blancos']['unidad']}",
                    'grapas_blancas': f"{limites[area_trabajo]['grapas_blancas']['periodo']} {limites[area_trabajo]['grapas_blancas']['unidad']}",
                    'grapas_negras': f"{limites[area_trabajo]['grapas_negras']['periodo']} {limites[area_trabajo]['grapas_negras']['unidad']}"
                }
            }
            
            # Agregar información de límites al técnico
            tecnico_con_limites = {**tecnico, 'limites': limites_disponibles}
            tecnicos_con_limites.append(tecnico_con_limites)
        
        cursor.close()
        connection.close()

        return render_template('modulos/logistica/ferretero.html', 
                            asignaciones=asignaciones,
                            tecnicos=tecnicos_con_limites,
                            materiales_problematicos=materiales_problematicos,
                            fecha_seleccionada=fecha_seleccionada)

    except Exception as e:
        print(f"Error al obtener asignaciones ferretero: {str(e)}")
        return render_template('error.html', 
                           mensaje='Error al cargar las asignaciones de material ferretero',
                           error=str(e))

@app.route('/logistica/registrar_ferretero', methods=['POST'])
@login_required()
@role_required('logistica')
def registrar_ferretero():
    connection = None
    cursor = None
    try:
        # Obtener datos básicos
        id_codigo_consumidor = request.form.get('id_codigo_consumidor')
        fecha = request.form.get('fecha')
        
        # Obtener cantidades de materiales
        silicona = request.form.get('silicona', '0')
        amarres_negros = request.form.get('amarres_negros', '0')
        amarres_blancos = request.form.get('amarres_blancos', '0')
        cinta_aislante = request.form.get('cinta_aislante', '0')
        grapas_blancas = request.form.get('grapas_blancas', '0')
        grapas_negras = request.form.get('grapas_negras', '0')

        # Validar campos requeridos
        if not all([id_codigo_consumidor, fecha]):
            return jsonify({
                'status': 'error',
                'message': 'El ID del técnico y la fecha son requeridos'
            }), 400

        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'status': 'error',
                'message': 'Error de conexión a la base de datos'
            }), 500

        cursor = connection.cursor(dictionary=True)

        # Verificar que el técnico existe y obtener su área de trabajo (carpeta)
        cursor.execute('SELECT nombre, cargo, carpeta FROM recurso_operativo WHERE id_codigo_consumidor = %s', (id_codigo_consumidor,))
        tecnico = cursor.fetchone()
        
        if not tecnico:
            return jsonify({
                'status': 'error',
                'message': 'El técnico seleccionado no existe'
            }), 404
            
        # Obtener la fecha actual para cálculos
        fecha_actual = datetime.now()
        
        # Verificar asignaciones previas del técnico para controlar los límites
        cursor.execute("""
            SELECT 
                fecha_asignacion,
                silicona,
                amarres_negros,
                amarres_blancos,
                cinta_aislante,
                grapas_blancas,
                grapas_negras
            FROM ferretero 
            WHERE id_codigo_consumidor = %s
            ORDER BY fecha_asignacion DESC
        """, (id_codigo_consumidor,))
        asignaciones_previas = cursor.fetchall()
        
        # Determinar la carpeta del técnico (EXCLUSIVAMENTE)
        carpeta = tecnico.get('carpeta', '').upper() if tecnico.get('carpeta') else ''
        
        # Función para obtener límites desde la base de datos
        def obtener_limites_desde_db(cursor):
            """Obtiene los límites configurables desde la base de datos"""
            try:
                cursor.execute("""
                    SELECT 
                        area_trabajo,
                        material_tipo,
                        cantidad_limite,
                        periodo_dias,
                        unidad_medida
                    FROM limites_ferretero 
                    WHERE activo = TRUE
                    ORDER BY area_trabajo, material_tipo
                """)
                limites_db = cursor.fetchall()
                
                # Convertir a estructura de diccionario
                limites = {}
                for limite in limites_db:
                    area = limite['area_trabajo']
                    material = limite['material_tipo']
                    
                    if area not in limites:
                        limites[area] = {}
                    
                    limites[area][material] = {
                        'cantidad': limite['cantidad_limite'],
                        'periodo': limite['periodo_dias'],
                        'unidad': limite['unidad_medida']
                    }
                
                return limites
                
            except Exception as e:
                print(f"Error al obtener límites desde DB: {str(e)}")
                # Fallback a límites por defecto en caso de error
                return {
                    'FTTH INSTALACIONES': {
                        'cinta_aislante': {'cantidad': 3, 'periodo': 15, 'unidad': 'días'},
                        'silicona': {'cantidad': 16, 'periodo': 7, 'unidad': 'días'},
                        'amarres_negros': {'cantidad': 50, 'periodo': 7, 'unidad': 'días'},
                        'amarres_blancos': {'cantidad': 50, 'periodo': 7, 'unidad': 'días'},
                        'grapas_blancas': {'cantidad': 100, 'periodo': 7, 'unidad': 'días'},
                        'grapas_negras': {'cantidad': 100, 'periodo': 7, 'unidad': 'días'}
                    }
                }
        
        # Obtener límites configurables desde la base de datos
        limites = obtener_limites_desde_db(cursor)
        
        # Determinar área de trabajo basada EXCLUSIVAMENTE en la carpeta asignada
        area_trabajo = None
        
        # Buscar coincidencia exacta en la carpeta del técnico
        if carpeta:
            for area in limites.keys():
                if area in carpeta:
                    area_trabajo = area
                    print(f"INFO: Técnico {tecnico.get('nombre', 'N/A')} - Carpeta: '{carpeta}' -> Área asignada: '{area_trabajo}'")
                    break
        
        # Si no se encuentra carpeta válida, rechazar la asignación
        if area_trabajo is None:
            error_msg = f"La carpeta '{carpeta}' del técnico {tecnico.get('nombre', 'N/A')} no está configurada para límites de ferretero. Carpetas válidas: {', '.join(limites.keys())}"
            print(f"ERROR: {error_msg}")
            return jsonify({
                'status': 'error',
                'message': error_msg
            }), 400
        
        # Inicializar contadores para materiales en los períodos correspondientes
        contadores = {
            'cinta_aislante': 0,
            'silicona': 0,
            'amarres_negros': 0,
            'amarres_blancos': 0,
            'grapas_blancas': 0,
            'grapas_negras': 0
        }
        
        # Calcular consumo previo en los periodos correspondientes
        for asignacion in asignaciones_previas:
            fecha_asignacion = asignacion['fecha_asignacion']
            diferencia_dias = (fecha_actual - fecha_asignacion).days
            
            # Verificar límite de cintas
            if diferencia_dias <= limites[area_trabajo]['cinta_aislante']['periodo']:
                contadores['cinta_aislante'] += int(asignacion.get('cinta_aislante', 0) or 0)
                
            # Verificar límite de siliconas
            if diferencia_dias <= limites[area_trabajo]['silicona']['periodo']:
                contadores['silicona'] += int(asignacion.get('silicona', 0) or 0)
                
            # Verificar límite de amarres negros
            if diferencia_dias <= limites[area_trabajo]['amarres_negros']['periodo']:
                contadores['amarres_negros'] += int(asignacion.get('amarres_negros', 0) or 0)
                
            # Verificar límite de amarres blancos
            if diferencia_dias <= limites[area_trabajo]['amarres_blancos']['periodo']:
                contadores['amarres_blancos'] += int(asignacion.get('amarres_blancos', 0) or 0)
                
            # Verificar límite de grapas blancas
            if diferencia_dias <= limites[area_trabajo]['grapas_blancas']['periodo']:
                contadores['grapas_blancas'] += int(asignacion.get('grapas_blancas', 0) or 0)
                
            # Verificar límite de grapas negras
            if diferencia_dias <= limites[area_trabajo]['grapas_negras']['periodo']:
                contadores['grapas_negras'] += int(asignacion.get('grapas_negras', 0) or 0)
        
        # Calcular cantidades de la asignación actual
        cintas_solicitadas = int(cinta_aislante or 0)
        siliconas_solicitadas = int(silicona or 0)
        amarres_negros_solicitados = int(amarres_negros or 0)
        amarres_blancos_solicitados = int(amarres_blancos or 0)
        grapas_blancas_solicitadas = int(grapas_blancas or 0)
        grapas_negras_solicitadas = int(grapas_negras or 0)
        
        # Validaciones de límites con asignación parcial
        materiales_rechazados = []
        materiales_asignados = []
        
        # Validar cintas
        if contadores['cinta_aislante'] + cintas_solicitadas > limites[area_trabajo]['cinta_aislante']['cantidad']:
            limite = limites[area_trabajo]['cinta_aislante']
            materiales_rechazados.append(f"Cinta aislante: excede el límite de {limite['cantidad']} cada {limite['periodo']} {limite['unidad']} para {area_trabajo}. Ya se han asignado {contadores['cinta_aislante']}.")
            cintas_solicitadas = 0
            cinta_aislante = 0
        elif cintas_solicitadas > 0:
            materiales_asignados.append(f"Cinta aislante: {cintas_solicitadas} unidades")
        
        # Validar siliconas
        if contadores['silicona'] + siliconas_solicitadas > limites[area_trabajo]['silicona']['cantidad']:
            limite = limites[area_trabajo]['silicona']
            materiales_rechazados.append(f"Silicona: excede el límite de {limite['cantidad']} cada {limite['periodo']} {limite['unidad']} para {area_trabajo}. Ya se han asignado {contadores['silicona']}.")
            siliconas_solicitadas = 0
            silicona = 0
        elif siliconas_solicitadas > 0:
            materiales_asignados.append(f"Silicona: {siliconas_solicitadas} unidades")
        
        # Validar amarres negros
        if contadores['amarres_negros'] + amarres_negros_solicitados > limites[area_trabajo]['amarres_negros']['cantidad']:
            limite = limites[area_trabajo]['amarres_negros']
            materiales_rechazados.append(f"Amarres negros: excede el límite de {limite['cantidad']} cada {limite['periodo']} {limite['unidad']} para {area_trabajo}. Ya se han asignado {contadores['amarres_negros']}.")
            amarres_negros_solicitados = 0
            amarres_negros = 0
        elif amarres_negros_solicitados > 0:
            materiales_asignados.append(f"Amarres negros: {amarres_negros_solicitados} unidades")
        
        # Validar amarres blancos
        if contadores['amarres_blancos'] + amarres_blancos_solicitados > limites[area_trabajo]['amarres_blancos']['cantidad']:
            limite = limites[area_trabajo]['amarres_blancos']
            materiales_rechazados.append(f"Amarres blancos: excede el límite de {limite['cantidad']} cada {limite['periodo']} {limite['unidad']} para {area_trabajo}. Ya se han asignado {contadores['amarres_blancos']}.")
            amarres_blancos_solicitados = 0
            amarres_blancos = 0
        elif amarres_blancos_solicitados > 0:
            materiales_asignados.append(f"Amarres blancos: {amarres_blancos_solicitados} unidades")
        
        # Validar grapas blancas
        if contadores['grapas_blancas'] + grapas_blancas_solicitadas > limites[area_trabajo]['grapas_blancas']['cantidad']:
            limite = limites[area_trabajo]['grapas_blancas']
            materiales_rechazados.append(f"Grapas blancas: excede el límite de {limite['cantidad']} cada {limite['periodo']} {limite['unidad']} para {area_trabajo}. Ya se han asignado {contadores['grapas_blancas']}.")
            grapas_blancas_solicitadas = 0
            grapas_blancas = 0
        elif grapas_blancas_solicitadas > 0:
            materiales_asignados.append(f"Grapas blancas: {grapas_blancas_solicitadas} unidades")
        
        # Validar grapas negras
        if contadores['grapas_negras'] + grapas_negras_solicitadas > limites[area_trabajo]['grapas_negras']['cantidad']:
            limite = limites[area_trabajo]['grapas_negras']
            materiales_rechazados.append(f"Grapas negras: excede el límite de {limite['cantidad']} cada {limite['periodo']} {limite['unidad']} para {area_trabajo}. Ya se han asignado {contadores['grapas_negras']}.")
            grapas_negras_solicitadas = 0
            grapas_negras = 0
        elif grapas_negras_solicitadas > 0:
            materiales_asignados.append(f"Grapas negras: {grapas_negras_solicitadas} unidades")
        
        # Si todos los materiales fueron rechazados, rechazar la asignación completa
        if materiales_rechazados and not materiales_asignados:
            return jsonify({
                'status': 'error',
                'message': 'Todos los materiales solicitados exceden los límites permitidos',
                'materiales_rechazados': materiales_rechazados
            }), 400
        
        # Validar stock disponible antes de la asignación
        cursor.execute("""
            SELECT codigo_material, cantidad_disponible 
            FROM stock_general 
            WHERE codigo_material IN ('silicona', 'amarres_negros', 'amarres_blancos', 'cinta_aislante', 'grapas_blancas', 'grapas_negras')
        """)
        stock_results = cursor.fetchall()
        
        # Convertir a diccionario para fácil acceso
        stock_actual = {}
        for row in stock_results:
            stock_actual[row['codigo_material']] = row['cantidad_disponible']
        
        if not stock_actual:
            return jsonify({
                'status': 'error',
                'message': 'No se pudo obtener información del stock actual'
            }), 500
        
        # Verificar stock suficiente para cada material
        errores_stock = []
        
        if siliconas_solicitadas > 0 and stock_actual['silicona'] < siliconas_solicitadas:
            errores_stock.append(f"Stock insuficiente de silicona. Disponible: {stock_actual['silicona']}, Solicitado: {siliconas_solicitadas}")
        
        if amarres_negros_solicitados > 0 and stock_actual['amarres_negros'] < amarres_negros_solicitados:
            errores_stock.append(f"Stock insuficiente de amarres negros. Disponible: {stock_actual['amarres_negros']}, Solicitado: {amarres_negros_solicitados}")
        
        if amarres_blancos_solicitados > 0 and stock_actual['amarres_blancos'] < amarres_blancos_solicitados:
            errores_stock.append(f"Stock insuficiente de amarres blancos. Disponible: {stock_actual['amarres_blancos']}, Solicitado: {amarres_blancos_solicitados}")
        
        if cintas_solicitadas > 0 and stock_actual['cinta_aislante'] < cintas_solicitadas:
            errores_stock.append(f"Stock insuficiente de cinta aislante. Disponible: {stock_actual['cinta_aislante']}, Solicitado: {cintas_solicitadas}")
        
        if grapas_blancas_solicitadas > 0 and stock_actual['grapas_blancas'] < grapas_blancas_solicitadas:
            errores_stock.append(f"Stock insuficiente de grapas blancas. Disponible: {stock_actual['grapas_blancas']}, Solicitado: {grapas_blancas_solicitadas}")
        
        if grapas_negras_solicitadas > 0 and stock_actual['grapas_negras'] < grapas_negras_solicitadas:
            errores_stock.append(f"Stock insuficiente de grapas negras. Disponible: {stock_actual['grapas_negras']}, Solicitado: {grapas_negras_solicitadas}")
        
        # Si hay errores de stock, rechazar la asignación
        if errores_stock:
            return jsonify({
                'status': 'error',
                'message': 'Stock insuficiente para completar la asignación',
                'detalles': errores_stock
            }), 400
        
        # Insertar la asignación
        cursor.execute("""
            INSERT INTO ferretero (
                id_codigo_consumidor, 
                fecha_asignacion,
                silicona,
                amarres_negros,
                amarres_blancos,
                cinta_aislante,
                grapas_blancas,
                grapas_negras
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            id_codigo_consumidor,
            fecha,
            silicona,
            amarres_negros,
            amarres_blancos,
            cinta_aislante,
            grapas_blancas,
            grapas_negras
        ))
        
        # NOTA: El stock y movimientos se actualizan automáticamente mediante el trigger actualizar_stock_asignacion
        # No es necesario actualizar manualmente aquí para evitar descuento duplicado
        
        connection.commit()

        # Preparar mensaje de respuesta con información detallada
        if materiales_rechazados:
            # Asignación parcial
            mensaje = f'Asignación parcial completada para {area_trabajo}'
            return jsonify({
                'status': 'partial_success',
                'message': mensaje,
                'materiales_asignados': materiales_asignados,
                'materiales_rechazados': materiales_rechazados,
                'detalle': f'Se asignaron {len(materiales_asignados)} materiales y se rechazaron {len(materiales_rechazados)} por exceder límites'
            }), 201
        else:
            # Asignación completa
            return jsonify({
                'status': 'success',
                'message': f'Material ferretero asignado exitosamente para {area_trabajo}',
                'materiales_asignados': materiales_asignados
            }), 201

    except mysql.connector.Error as e:
        print(f"Error MySQL: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error en la base de datos: {str(e)}'
        }), 500
    except Exception as e:
        print(f"Error general: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error al registrar la asignación: {str(e)}'
        }), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/logistica/exportar_ferretero_csv')
@login_required()
@role_required('logistica')
def exportar_ferretero_csv():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener todas las asignaciones con información del técnico
        cursor.execute("""
            SELECT 
                f.*,
                r.nombre as nombre_tecnico,
                r.recurso_operativo_cedula as cedula_tecnico,
                r.cargo as cargo_tecnico
            FROM ferretero f 
            LEFT JOIN recurso_operativo r ON f.id_codigo_consumidor = r.id_codigo_consumidor 
            ORDER BY f.fecha_asignacion DESC
        """)
        asignaciones = cursor.fetchall()
        
        # Crear archivo CSV en memoria
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Escribir encabezados
        writer.writerow([
            'ID Asignación',
            'Fecha',
            'Técnico',
            'Cédula',
            'Cargo',
            'Silicona',
            'Amarres Negros',
            'Amarres Blancos',
            'Cinta Aislante',
            'Grapas Blancas',
            'Grapas Negras'
        ])
        
        # Escribir datos
        for asignacion in asignaciones:
            writer.writerow([
                asignacion['id_ferretero'],
                asignacion['fecha_asignacion'].strftime('%Y-%m-%d %H:%M:%S'),
                asignacion.get('nombre_tecnico', 'No asignado'),
                asignacion.get('cedula_tecnico', 'No disponible'),
                asignacion.get('cargo_tecnico', 'No especificado'),
                asignacion.get('silicona', '0'),
                asignacion.get('amarres_negros', '0'),
                asignacion.get('amarres_blancos', '0'),
                asignacion.get('cinta_aislante', '0'),
                asignacion.get('grapas_blancas', '0'),
                asignacion.get('grapas_negras', '0')
            ])
        
        # Preparar respuesta
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            mimetype='text/csv; charset=utf-8-sig',
            as_attachment=True,
            download_name=f'material_ferretero_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )
        
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/logistica/estadisticas_ferretero_page')
@login_required()
@role_required('logistica')
def estadisticas_ferretero_page():
    try:
        connection = get_db_connection()
        if connection is None:
            return render_template('error.html', 
                               mensaje='Error de conexión a la base de datos',
                               error='No se pudo establecer conexión con la base de datos')
        
        # Obtener lista de técnicos disponibles para filtros
        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT id_codigo_consumidor, nombre, recurso_operativo_cedula, cargo, carpeta 
            FROM recurso_operativo 
            WHERE cargo LIKE '%TECNICO%' OR cargo LIKE '%TÉCNICO%'
            ORDER BY nombre
        """)
        tecnicos = cursor.fetchall()
        
        cursor.close()
        connection.close()
        
        return render_template('modulos/logistica/estadisticas_ferretero.html', tecnicos=tecnicos)
    
    except Exception as e:
        print(f"Error al cargar página de estadísticas ferretero: {str(e)}")
        return render_template('error.html', 
                           mensaje='Error al cargar la página de estadísticas de material ferretero',
                           error=str(e))

@app.route('/logistica/estadisticas_ferretero')
@login_required()
@role_required('logistica')
def estadisticas_ferretero():
    connection = None
    cursor = None
    try:
        # Obtener parámetros de filtro
        mes = request.args.get('mes', '')
        material = request.args.get('material', '')
        area = request.args.get('area', '')
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'status': 'error',
                'message': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Construir la consulta base
        query = """
            SELECT 
                f.*,
                r.nombre,
                r.recurso_operativo_cedula,
                r.cargo,
                r.carpeta
            FROM ferretero f 
            LEFT JOIN recurso_operativo r ON f.id_codigo_consumidor = r.id_codigo_consumidor 
            WHERE 1=1
        """
        params = []
        
        # Aplicar filtro por mes (independiente del año)
        if mes and mes != 'todos':
            # Extraer el mes del formato 'YYYY-MM' o usar directamente si es solo el número
            try:
                if '-' in mes:
                    # Formato 'YYYY-MM'
                    mes_numero = int(mes.split('-')[1])
                else:
                    # Solo el número del mes
                    mes_numero = int(mes)
                query += " AND MONTH(f.fecha_asignacion) = %s"
                params.append(mes_numero)
            except (ValueError, IndexError) as e:
                print(f"Error al procesar el parámetro mes '{mes}': {e}")
                # Continuar sin filtro de mes si hay error
                pass
            
        # Asegurarse de que se muestren datos incluso si son de años futuros
        # No filtrar por año para permitir ver datos de cualquier año
        
        # Aplicar filtro por área
        if area and area != 'todos':
            query += " AND (r.carpeta LIKE %s OR r.cargo LIKE %s)"
            area_param = f'%{area}%'
            params.append(area_param)
            params.append(area_param)
        
        # Ejecutar consulta
        cursor.execute(query, params)
        asignaciones = cursor.fetchall()
        
        # Procesar los datos para estadísticas
        estadisticas_por_tecnico = {}
        areas_distribucion = {}
        
        for asignacion in asignaciones:
            id_tecnico = asignacion['id_codigo_consumidor']
            nombre = asignacion.get('nombre', 'Técnico sin nombre')
            
            # Determinar área de trabajo
            carpeta = asignacion.get('carpeta', '')
            carpeta = carpeta.upper() if carpeta else ''
            cargo = asignacion.get('cargo', '')
            cargo = cargo.upper() if cargo else ''
            
            area_trabajo = 'No especificada'
            areas_posibles = ['FTTH INSTALACIONES', 'INSTALACIONES DOBLES', 'POSTVENTA', 
                             'MANTENIMIENTO FTTH', 'ARREGLOS HFC', 'CONDUCTOR']
            
            for area_posible in areas_posibles:
                if area_posible in carpeta or area_posible in cargo:
                    area_trabajo = area_posible
                    break
            
            # Inicializar estadísticas para este técnico si no existe
            if id_tecnico not in estadisticas_por_tecnico:
                estadisticas_por_tecnico[id_tecnico] = {
                    'nombre': nombre,
                    'area': area_trabajo,
                    'silicona': 0,
                    'amarres_negros': 0,
                    'amarres_blancos': 0,
                    'cinta_aislante': 0,
                    'grapas_blancas': 0,
                    'grapas_negras': 0,
                    'total_asignaciones': 0
                }
            
            # Inicializar estadísticas para esta área si no existe
            if area_trabajo not in areas_distribucion:
                areas_distribucion[area_trabajo] = 0
            
            # Actualizar contadores
            estadisticas_por_tecnico[id_tecnico]['silicona'] += int(asignacion.get('silicona', 0) or 0)
            estadisticas_por_tecnico[id_tecnico]['amarres_negros'] += int(asignacion.get('amarres_negros', 0) or 0)
            estadisticas_por_tecnico[id_tecnico]['amarres_blancos'] += int(asignacion.get('amarres_blancos', 0) or 0)
            estadisticas_por_tecnico[id_tecnico]['cinta_aislante'] += int(asignacion.get('cinta_aislante', 0) or 0)
            estadisticas_por_tecnico[id_tecnico]['grapas_blancas'] += int(asignacion.get('grapas_blancas', 0) or 0)
            estadisticas_por_tecnico[id_tecnico]['grapas_negras'] += int(asignacion.get('grapas_negras', 0) or 0)
            estadisticas_por_tecnico[id_tecnico]['total_asignaciones'] += 1
            
            # Actualizar distribución por área
            areas_distribucion[area_trabajo] += 1
        
        # Aplicar filtro por material después de procesar todos los datos
        if material and material != 'todos':
            estadisticas_filtradas = {}
            for id_tecnico, stats in estadisticas_por_tecnico.items():
                incluir_tecnico = False
                if material == 'silicona' and stats['silicona'] > 0:
                    incluir_tecnico = True
                elif material == 'amarres' and (stats['amarres_negros'] + stats['amarres_blancos']) > 0:
                    incluir_tecnico = True
                elif material == 'cinta' and stats['cinta_aislante'] > 0:
                    incluir_tecnico = True
                elif material == 'grapas' and (stats['grapas_blancas'] + stats['grapas_negras']) > 0:
                    incluir_tecnico = True
                
                if incluir_tecnico:
                    estadisticas_filtradas[id_tecnico] = stats
            
            estadisticas_por_tecnico = estadisticas_filtradas
        
        # Calcular promedios y determinar técnicos por encima del promedio
        estadisticas_lista = list(estadisticas_por_tecnico.values())
        
        # Verificar si hay datos para mostrar
        if not estadisticas_lista:
            # No hay datos para los filtros seleccionados
            print(f"No hay datos para los filtros: mes={mes}, material={material}, area={area}")
            return jsonify({
                'status': 'success',
                'estadisticas': [],
                'top_tecnicos': [],
                'distribucion_area': [],
                'message': 'No hay datos disponibles para los filtros seleccionados'
            })
        
        # Calcular promedio de asignaciones por técnico
        total_asignaciones = sum(item['total_asignaciones'] for item in estadisticas_lista)
        promedio_asignaciones = total_asignaciones / len(estadisticas_lista) if estadisticas_lista else 0
        
        # Marcar técnicos por encima del promedio
        for item in estadisticas_lista:
            item['promedio_mensual'] = round(item['total_asignaciones'], 2)
            item['por_encima_promedio'] = item['total_asignaciones'] > promedio_asignaciones
            item['muy_por_encima_promedio'] = item['total_asignaciones'] > (promedio_asignaciones * 1.5)
        
        # Ordenar según el filtro de material aplicado
        if material and material != 'todos':
            # Ordenar por el material específico seleccionado
            if material == 'silicona':
                estadisticas_lista.sort(key=lambda x: x['silicona'], reverse=True)
            elif material == 'amarres':
                estadisticas_lista.sort(key=lambda x: (x['amarres_negros'] + x['amarres_blancos']), reverse=True)
            elif material == 'cinta':
                estadisticas_lista.sort(key=lambda x: x['cinta_aislante'], reverse=True)
            elif material == 'grapas':
                estadisticas_lista.sort(key=lambda x: (x['grapas_blancas'] + x['grapas_negras']), reverse=True)
        else:
            # Ordenar por total de asignaciones (descendente) cuando no hay filtro de material
            estadisticas_lista.sort(key=lambda x: x['total_asignaciones'], reverse=True)
        
        # Preparar top 5 técnicos
        top_tecnicos = estadisticas_lista[:5] if len(estadisticas_lista) >= 5 else estadisticas_lista
        
        # Calcular porcentajes para el top 5
        max_asignaciones = max([item['total_asignaciones'] for item in top_tecnicos]) if top_tecnicos else 1
        for item in top_tecnicos:
            item['porcentaje'] = round((item['total_asignaciones'] / max_asignaciones) * 100)
        
        # Preparar distribución por área
        distribucion_area = [{'area': area, 'total': total} for area, total in areas_distribucion.items()]
        
        return jsonify({
            'status': 'success',
            'estadisticas': estadisticas_lista,
            'top_tecnicos': top_tecnicos,
            'distribucion_area': distribucion_area
        })
        
    except Exception as e:
        print(f"Error al obtener estadísticas ferretero: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/logistica/exportar_estadisticas_ferretero')
@login_required()
@role_required('logistica')
def exportar_estadisticas_ferretero():
    connection = None
    cursor = None
    try:
        # Obtener parámetros de filtro (igual que en estadisticas_ferretero)
        mes = request.args.get('mes', '')
        material = request.args.get('material', '')
        area = request.args.get('area', '')
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Construir la consulta base (igual que en estadisticas_ferretero)
        query = """
            SELECT 
                f.*,
                r.nombre,
                r.recurso_operativo_cedula,
                r.cargo,
                r.carpeta
            FROM ferretero f 
            LEFT JOIN recurso_operativo r ON f.id_codigo_consumidor = r.id_codigo_consumidor 
            WHERE 1=1
        """
        params = []
        
        # Aplicar filtro por mes (independiente del año)
        if mes and mes != 'todos':
            # Extraer el mes del formato 'YYYY-MM' o usar directamente si es solo el número
            try:
                if '-' in mes:
                    # Formato 'YYYY-MM'
                    mes_numero = int(mes.split('-')[1])
                else:
                    # Solo el número del mes
                    mes_numero = int(mes)
                query += " AND MONTH(f.fecha_asignacion) = %s"
                params.append(mes_numero)
            except (ValueError, IndexError) as e:
                print(f"Error al procesar el parámetro mes '{mes}': {e}")
                # Continuar sin filtro de mes si hay error
                pass
            
        # Asegurarse de que se muestren datos incluso si son de años futuros
        # No filtrar por año para permitir ver datos de cualquier año
        
        # Aplicar filtro por área
        if area and area != 'todos':
            query += " AND (r.carpeta LIKE %s OR r.cargo LIKE %s)"
            area_param = f'%{area}%'
            params.append(area_param)
            params.append(area_param)
        
        # Ejecutar consulta
        cursor.execute(query, params)
        asignaciones = cursor.fetchall()
        
        # Procesar los datos para estadísticas (igual que en estadisticas_ferretero)
        estadisticas_por_tecnico = {}
        
        for asignacion in asignaciones:
            id_tecnico = asignacion['id_codigo_consumidor']
            nombre = asignacion.get('nombre', 'Técnico sin nombre')
            
            # Determinar área de trabajo
            carpeta = asignacion.get('carpeta', '').upper() if asignacion.get('carpeta') else ''
            cargo = asignacion.get('cargo', '').upper()
            
            area_trabajo = 'No especificada'
            areas_posibles = ['FTTH INSTALACIONES', 'INSTALACIONES DOBLES', 'POSTVENTA', 
                             'MANTENIMIENTO FTTH', 'ARREGLOS HFC', 'CONDUCTOR']
            
            for area_posible in areas_posibles:
                if area_posible in carpeta or area_posible in cargo:
                    area_trabajo = area_posible
                    break
            
            # Inicializar estadísticas para este técnico si no existe
            if id_tecnico not in estadisticas_por_tecnico:
                estadisticas_por_tecnico[id_tecnico] = {
                    'nombre': nombre,
                    'cedula': asignacion.get('recurso_operativo_cedula', 'No disponible'),
                    'area': area_trabajo,
                    'silicona': 0,
                    'amarres_negros': 0,
                    'amarres_blancos': 0,
                    'cinta_aislante': 0,
                    'grapas_blancas': 0,
                    'grapas_negras': 0,
                    'total_asignaciones': 0
                }
            
            # Actualizar contadores
            estadisticas_por_tecnico[id_tecnico]['silicona'] += int(asignacion.get('silicona', 0) or 0)
            estadisticas_por_tecnico[id_tecnico]['amarres_negros'] += int(asignacion.get('amarres_negros', 0) or 0)
            estadisticas_por_tecnico[id_tecnico]['amarres_blancos'] += int(asignacion.get('amarres_blancos', 0) or 0)
            estadisticas_por_tecnico[id_tecnico]['cinta_aislante'] += int(asignacion.get('cinta_aislante', 0) or 0)
            estadisticas_por_tecnico[id_tecnico]['grapas_blancas'] += int(asignacion.get('grapas_blancas', 0) or 0)
            estadisticas_por_tecnico[id_tecnico]['grapas_negras'] += int(asignacion.get('grapas_negras', 0) or 0)
            estadisticas_por_tecnico[id_tecnico]['total_asignaciones'] += 1
        
        # Aplicar filtro por material después de procesar todos los datos
        if material and material != 'todos':
            estadisticas_filtradas = {}
            for id_tecnico, stats in estadisticas_por_tecnico.items():
                incluir_tecnico = False
                if material == 'silicona' and stats['silicona'] > 0:
                    incluir_tecnico = True
                elif material == 'amarres' and (stats['amarres_negros'] + stats['amarres_blancos']) > 0:
                    incluir_tecnico = True
                elif material == 'cintas' and stats['cinta_aislante'] > 0:
                    incluir_tecnico = True
                elif material == 'grapas' and (stats['grapas_blancas'] + stats['grapas_negras']) > 0:
                    incluir_tecnico = True
                
                if incluir_tecnico:
                    estadisticas_filtradas[id_tecnico] = stats
            
            estadisticas_por_tecnico = estadisticas_filtradas
        
        # Calcular promedios y determinar técnicos por encima del promedio
        estadisticas_lista = list(estadisticas_por_tecnico.values())
        
        # Calcular promedio de asignaciones por técnico
        total_asignaciones = sum(item['total_asignaciones'] for item in estadisticas_lista)
        promedio_asignaciones = total_asignaciones / len(estadisticas_lista) if estadisticas_lista else 0
        
        # Marcar técnicos por encima del promedio
        for item in estadisticas_lista:
            item['promedio_mensual'] = round(item['total_asignaciones'], 2)
            item['por_encima_promedio'] = item['total_asignaciones'] > promedio_asignaciones
            item['muy_por_encima_promedio'] = item['total_asignaciones'] > (promedio_asignaciones * 1.5)
        
        # Ordenar por total de asignaciones (descendente)
        estadisticas_lista.sort(key=lambda x: x['total_asignaciones'], reverse=True)
        
        # Crear un DataFrame de pandas para generar el Excel
        df = pd.DataFrame(estadisticas_lista)
        
        # Renombrar columnas para el Excel
        columnas = {
            'nombre': 'Nombre',
            'cedula': 'Cédula',
            'area': 'Área',
            'silicona': 'Silicona',
            'amarres_negros': 'Amarres Negros',
            'amarres_blancos': 'Amarres Blancos',
            'cinta_aislante': 'Cinta Aislante',
            'grapas_blancas': 'Grapas Blancas',
            'grapas_negras': 'Grapas Negras',
            'total_asignaciones': 'Total Asignaciones',
            'promedio_mensual': 'Promedio Mensual'
        }
        df = df.rename(columns=columnas)
        
        # Seleccionar y ordenar columnas para el Excel
        columnas_excel = [
            'Nombre', 'Cédula', 'Área', 'Silicona', 'Amarres Negros', 'Amarres Blancos',
            'Cinta Aislante', 'Grapas Blancas', 'Grapas Negras', 'Total Asignaciones', 'Promedio Mensual'
        ]
        df = df[columnas_excel]
        
        # Crear un objeto BytesIO para guardar el Excel
        output = io.BytesIO()
        
        # Crear un objeto ExcelWriter
        with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
            df.to_excel(writer, sheet_name='Estadísticas', index=False)
            
            # Obtener el objeto workbook y worksheet
            workbook = writer.book
            worksheet = writer.sheets['Estadísticas']
            
            # Definir formatos
            header_format = workbook.add_format({
                'bold': True,
                'text_wrap': True,
                'valign': 'top',
                'fg_color': '#D7E4BC',
                'border': 1
            })
            
            # Aplicar formato a los encabezados
            for col_num, value in enumerate(df.columns.values):
                worksheet.write(0, col_num, value, header_format)
                
            # Ajustar ancho de columnas
            for i, col in enumerate(df.columns):
                column_width = max(df[col].astype(str).map(len).max(), len(col) + 2)
                worksheet.set_column(i, i, column_width)
        
        # Preparar respuesta
        output.seek(0)
        
        # Generar nombre de archivo con fecha actual
        fecha_actual = datetime.now().strftime("%d-%m-%Y")
        nombre_archivo = f'Estadisticas_Ferretero_{fecha_actual}.xlsx'
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre_archivo
        )
        
    except Exception as e:
        print(f"Error al exportar estadísticas ferretero a Excel: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

# ===== RUTAS PARA GESTIÓN DE STOCK DE MATERIAL FERRETERO =====

@app.route('/logistica/stock_ferretero')
@login_required()
@role_required('logistica')
def obtener_stock_ferretero():
    """Obtener el stock actual de material ferretero con cálculos de inventario"""
    connection = None
    cursor = None
    try:
        # Obtener parámetro de mes del filtro
        mes_filtro = request.args.get('mes', '')
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener stock actual
        cursor.execute("""
            SELECT 
                codigo_material as material_tipo,
                cantidad_disponible as cantidad_actual,
                cantidad_minima,
                fecha_actualizacion
            FROM stock_general 
            ORDER BY codigo_material
        """)
        stock = cursor.fetchall()
        
        # Calcular total asignado por mes (filtrado o actual)
        if mes_filtro and mes_filtro != 'todos':
            # Filtrar por mes específico
            cursor.execute("""
                SELECT 
                    SUM(silicona) as total_silicona,
                    SUM(amarres_negros) as total_amarres_negros,
                    SUM(amarres_blancos) as total_amarres_blancos,
                    SUM(cinta_aislante) as total_cinta_aislante,
                    SUM(grapas_blancas) as total_grapas_blancas,
                    SUM(grapas_negras) as total_grapas_negras
                FROM ferretero 
                WHERE MONTH(fecha_asignacion) = %s 
                AND YEAR(fecha_asignacion) = YEAR(CURDATE())
            """, (mes_filtro,))
        else:
            # Usar mes actual por defecto
            cursor.execute("""
                SELECT 
                    SUM(silicona) as total_silicona,
                    SUM(amarres_negros) as total_amarres_negros,
                    SUM(amarres_blancos) as total_amarres_blancos,
                    SUM(cinta_aislante) as total_cinta_aislante,
                    SUM(grapas_blancas) as total_grapas_blancas,
                    SUM(grapas_negras) as total_grapas_negras
                FROM ferretero 
                WHERE MONTH(fecha_asignacion) = MONTH(CURDATE()) 
                AND YEAR(fecha_asignacion) = YEAR(CURDATE())
            """)
        asignado_mes = cursor.fetchone()
        
        # Obtener total de entradas por material
        cursor.execute("""
            SELECT 
                material_tipo,
                SUM(cantidad_entrada) as total_entradas,
                COUNT(*) as numero_entradas,
                MAX(fecha_entrada) as ultima_entrada,
                SUM(precio_total) as valor_total_entradas
            FROM entradas_ferretero
            GROUP BY material_tipo
            ORDER BY material_tipo
        """)
        total_entradas = cursor.fetchall()
        
        # Obtener total asignado histórico por material
        cursor.execute("""
            SELECT 
                'silicona' as material_tipo,
                COALESCE(SUM(silicona), 0) as total_asignado
            FROM ferretero WHERE silicona > 0
            UNION ALL
            SELECT 
                'amarres_negros' as material_tipo,
                COALESCE(SUM(amarres_negros), 0) as total_asignado
            FROM ferretero WHERE amarres_negros > 0
            UNION ALL
            SELECT 
                'amarres_blancos' as material_tipo,
                COALESCE(SUM(amarres_blancos), 0) as total_asignado
            FROM ferretero WHERE amarres_blancos > 0
            UNION ALL
            SELECT 
                'cinta_aislante' as material_tipo,
                COALESCE(SUM(cinta_aislante), 0) as total_asignado
            FROM ferretero WHERE cinta_aislante > 0
            UNION ALL
            SELECT 
                'grapas_blancas' as material_tipo,
                COALESCE(SUM(grapas_blancas), 0) as total_asignado
            FROM ferretero WHERE grapas_blancas > 0
            UNION ALL
            SELECT 
                'grapas_negras' as material_tipo,
                COALESCE(SUM(grapas_negras), 0) as total_asignado
            FROM ferretero WHERE grapas_negras > 0
            ORDER BY material_tipo
        """)
        total_asignado_historico = cursor.fetchall()
        
        # Obtener datos para calcular promedio de consumo diario
        cursor.execute("""
            SELECT 
                'silicona' as material_tipo,
                COALESCE(SUM(silicona), 0) as total_consumido,
                COUNT(DISTINCT DATE(fecha_asignacion)) as dias_con_consumo
            FROM ferretero 
            WHERE silicona > 0 AND fecha_asignacion >= DATE_SUB(CURDATE(), INTERVAL 90 DAY)
            UNION ALL
            SELECT 
                'amarres_negros' as material_tipo,
                COALESCE(SUM(amarres_negros), 0) as total_consumido,
                COUNT(DISTINCT DATE(fecha_asignacion)) as dias_con_consumo
            FROM ferretero 
            WHERE amarres_negros > 0 AND fecha_asignacion >= DATE_SUB(CURDATE(), INTERVAL 90 DAY)
            UNION ALL
            SELECT 
                'amarres_blancos' as material_tipo,
                COALESCE(SUM(amarres_blancos), 0) as total_consumido,
                COUNT(DISTINCT DATE(fecha_asignacion)) as dias_con_consumo
            FROM ferretero 
            WHERE amarres_blancos > 0 AND fecha_asignacion >= DATE_SUB(CURDATE(), INTERVAL 90 DAY)
            UNION ALL
            SELECT 
                'cinta_aislante' as material_tipo,
                COALESCE(SUM(cinta_aislante), 0) as total_consumido,
                COUNT(DISTINCT DATE(fecha_asignacion)) as dias_con_consumo
            FROM ferretero 
            WHERE cinta_aislante > 0 AND fecha_asignacion >= DATE_SUB(CURDATE(), INTERVAL 90 DAY)
            UNION ALL
            SELECT 
                'grapas_blancas' as material_tipo,
                COALESCE(SUM(grapas_blancas), 0) as total_consumido,
                COUNT(DISTINCT DATE(fecha_asignacion)) as dias_con_consumo
            FROM ferretero 
            WHERE grapas_blancas > 0 AND fecha_asignacion >= DATE_SUB(CURDATE(), INTERVAL 90 DAY)
            UNION ALL
            SELECT 
                'grapas_negras' as material_tipo,
                COALESCE(SUM(grapas_negras), 0) as total_consumido,
                COUNT(DISTINCT DATE(fecha_asignacion)) as dias_con_consumo
            FROM ferretero 
            WHERE grapas_negras > 0 AND fecha_asignacion >= DATE_SUB(CURDATE(), INTERVAL 90 DAY)
            ORDER BY material_tipo
        """)
        consumo_historico = cursor.fetchall()
        
        # Calcular resumen de inventario
        resumen_inventario = []
        materiales = ['silicona', 'amarres_negros', 'amarres_blancos', 'cinta_aislante', 'grapas_blancas', 'grapas_negras']
        
        for material in materiales:
            # Buscar datos de entrada
            entrada = next((item for item in total_entradas if item['material_tipo'] == material), None)
            total_recibido = float(entrada['total_entradas']) if entrada and entrada['total_entradas'] else 0.0
            valor_entradas = float(entrada['valor_total_entradas']) if entrada and entrada['valor_total_entradas'] else 0.0
            ultima_entrada = entrada['ultima_entrada'] if entrada else None
            numero_entradas = int(entrada['numero_entradas']) if entrada and entrada['numero_entradas'] else 0
            
            # Buscar total asignado
            asignado = next((item for item in total_asignado_historico if item['material_tipo'] == material), None)
            total_asignado = float(asignado['total_asignado']) if asignado and asignado['total_asignado'] else 0.0
            
            # Buscar stock actual
            stock_item = next((item for item in stock if item['material_tipo'] == material), None)
            stock_actual = float(stock_item['cantidad_actual']) if stock_item and stock_item['cantidad_actual'] else 0.0
            stock_minimo = float(stock_item['cantidad_minima']) if stock_item and stock_item['cantidad_minima'] else 0.0
            
            # Buscar datos de consumo histórico
            consumo = next((item for item in consumo_historico if item['material_tipo'] == material), None)
            total_consumido = float(consumo['total_consumido']) if consumo and consumo['total_consumido'] else 0.0
            dias_con_consumo = int(consumo['dias_con_consumo']) if consumo and consumo['dias_con_consumo'] else 0
            
            # Calcular promedio de consumo diario y días de alcance
            if dias_con_consumo > 0:
                promedio_consumo_diario = total_consumido / dias_con_consumo
            else:
                promedio_consumo_diario = 0.0
            
            if promedio_consumo_diario > 0:
                dias_alcance = stock_actual / promedio_consumo_diario
            else:
                dias_alcance = float('inf') if stock_actual > 0 else 0
            
            # Formatear días de alcance para mostrar
            if dias_alcance == float('inf'):
                dias_alcance_display = "∞"
            elif dias_alcance > 999:
                dias_alcance_display = ">999"
            else:
                dias_alcance_display = round(dias_alcance, 1)
            
            # Calcular diferencia teórica vs real (mantener para compatibilidad)
            diferencia_teorica = total_recibido - total_asignado
            diferencia_real = diferencia_teorica - stock_actual
            
            resumen_inventario.append({
                'material_tipo': material,
                'total_recibido': int(total_recibido),
                'total_asignado': int(total_asignado),
                'stock_actual': int(stock_actual),
                'stock_minimo': int(stock_minimo),
                'diferencia_teorica': int(diferencia_teorica),
                'diferencia_real': int(diferencia_real),
                'dias_alcance': dias_alcance_display,
                'promedio_consumo_diario': round(promedio_consumo_diario, 2),
                'valor_entradas': float(valor_entradas),
                'ultima_entrada': ultima_entrada,
                'numero_entradas': numero_entradas,
                'estado_stock': 'Crítico' if stock_actual <= stock_minimo else 'Normal'
            })
        
        return jsonify({
            'status': 'success',
            'stock': stock,
            'asignado_mes_actual': asignado_mes,
            'resumen_inventario': resumen_inventario,
            'total_entradas': total_entradas,
            'total_asignado_historico': total_asignado_historico
        })
        
    except Exception as e:
        print(f"Error al obtener stock ferretero: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/stock/materiales')
def verificar_disponibilidad_materiales():
    """Endpoint simple para verificar disponibilidad de stock de materiales ferretero"""
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener stock actual de todos los materiales
        cursor.execute("""
            SELECT 
                codigo_material as material_tipo,
                cantidad_disponible,
                cantidad_minima
            FROM stock_general 
            WHERE activo = 1
            ORDER BY codigo_material
        """)
        stock_data = cursor.fetchall()
        
        # Crear diccionario con disponibilidad por material
        disponibilidad = {}
        for item in stock_data:
            material = item['material_tipo']
            cantidad = float(item['cantidad_disponible']) if item['cantidad_disponible'] else 0.0
            disponibilidad[material] = {
                'disponible': cantidad > 0,
                'cantidad': int(cantidad),
                'cantidad_minima': int(item['cantidad_minima']) if item['cantidad_minima'] else 0
            }
        
        # Asegurar que todos los materiales estén incluidos
        materiales_requeridos = ['silicona', 'cinta_aislante', 'amarres_negros', 'amarres_blancos', 'grapas_blancas', 'grapas_negras']
        for material in materiales_requeridos:
            if material not in disponibilidad:
                disponibilidad[material] = {
                    'disponible': False,
                    'cantidad': 0,
                    'cantidad_minima': 0
                }
        
        return jsonify({
            'status': 'success',
            'disponibilidad': disponibilidad
        })
        
    except Exception as e:
        print(f"Error al verificar disponibilidad de materiales: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/comparacion_mensual_materiales')
def obtener_comparacion_mensual_materiales():
    """Endpoint para obtener datos de comparación mensual de materiales"""
    connection = None
    cursor = None
    try:
        # Obtener parámetros
        material = request.args.get('material', 'silicona')
        anio = request.args.get('anio', '2025')
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Consulta para obtener asignaciones mensuales del material específico desde la tabla ferretero
        query = """
            SELECT 
                MONTH(fecha_asignacion) as mes,
                SUM(CASE 
                    WHEN %s = 'silicona' THEN COALESCE(silicona, 0)
                    WHEN %s = 'amarres_negros' THEN COALESCE(amarres_negros, 0)
                    WHEN %s = 'amarres_blancos' THEN COALESCE(amarres_blancos, 0)
                    WHEN %s = 'cinta_aislante' THEN COALESCE(cinta_aislante, 0)
                    WHEN %s = 'grapas_blancas' THEN COALESCE(grapas_blancas, 0)
                    WHEN %s = 'grapas_negras' THEN COALESCE(grapas_negras, 0)
                    ELSE 0
                END) as cantidad_asignada
            FROM ferretero 
            WHERE YEAR(fecha_asignacion) = %s
            AND (
                (%s = 'silicona' AND silicona > 0) OR
                (%s = 'amarres_negros' AND amarres_negros > 0) OR
                (%s = 'amarres_blancos' AND amarres_blancos > 0) OR
                (%s = 'cinta_aislante' AND cinta_aislante > 0) OR
                (%s = 'grapas_blancas' AND grapas_blancas > 0) OR
                (%s = 'grapas_negras' AND grapas_negras > 0)
            )
            GROUP BY MONTH(fecha_asignacion)
            ORDER BY mes
        """
        
        cursor.execute(query, (material, material, material, material, material, material, anio, material, material, material, material, material, material))
        datos_mensuales = cursor.fetchall()
        
        # Transformar datos para el formato esperado por el frontend
        datos_formateados = []
        for dato in datos_mensuales:
            datos_formateados.append({
                'mes': dato['mes'],
                'cantidad_asignada': int(dato['cantidad_asignada']) if dato['cantidad_asignada'] else 0
            })
        
        # Si no hay datos, crear estructura vacía
        if not datos_formateados:
            datos_formateados = []
            for mes in range(1, 13):
                datos_formateados.append({
                    'mes': mes,
                    'cantidad_asignada': 0
                })
        
        return jsonify({
            'status': 'success',
            'datos_mensuales': datos_formateados,
            'material': material,
            'anio': anio
        })
        
    except Exception as e:
        print(f"Error al obtener comparación mensual: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/logistica/entradas_ferretero', methods=['POST'])
@login_required()
@role_required('logistica')
def registrar_entrada_ferretero():
    """Registrar entrada de material ferretero"""
    connection = None
    cursor = None
    try:
        data = request.get_json()
        
        # Validar datos requeridos
        required_fields = ['material_tipo', 'cantidad', 'precio_unitario']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({
                    'status': 'error',
                    'message': f'El campo {field} es requerido'
                }), 400
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Insertar entrada
        cursor.execute("""
            INSERT INTO entradas_ferretero (
                material_tipo, cantidad_entrada, precio_unitario, precio_total,
                proveedor, numero_factura, observaciones, fecha_entrada, usuario_registro
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, NOW(), %s)
        """, (
            data['material_tipo'],
            data['cantidad'],
            data['precio_unitario'],
            float(data['cantidad']) * float(data['precio_unitario']),
            data.get('proveedor', ''),
            data.get('numero_factura', ''),
            data.get('observaciones', ''),
            session['user_id']
        ))
        
        # El trigger actualizar_stock_entrada se encarga automáticamente de:
        # 1. Actualizar el stock en la tabla stock_ferretero
        # 2. Registrar el movimiento en movimientos_stock_ferretero
        # Por lo tanto, no necesitamos hacer actualizaciones manuales aquí
        
        connection.commit()
        
        return jsonify({
            'status': 'success',
            'message': 'Entrada registrada correctamente'
        })
        
    except Exception as e:
        if connection:
            connection.rollback()
        print(f"Error al registrar entrada ferretero: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/logistica/movimientos_ferretero')
@login_required()
@role_required('logistica')
def obtener_movimientos_ferretero():
    """Obtener movimientos de stock de material ferretero"""
    connection = None
    cursor = None
    try:
        # Obtener parámetros de filtro
        material = request.args.get('material', '')
        fecha_inicio = request.args.get('fecha_inicio', '')
        fecha_fin = request.args.get('fecha_fin', '')
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Construir consulta con filtros
        query = """
            SELECT 
                material_tipo,
                tipo_movimiento,
                cantidad,
                fecha_movimiento,
                referencia_id,
                observaciones
            FROM movimientos_stock_ferretero 
            WHERE 1=1
        """
        params = []
        
        if material and material != 'todos':
            query += " AND material_tipo = %s"
            params.append(material)
            
        if fecha_inicio:
            query += " AND DATE(fecha_movimiento) >= %s"
            params.append(fecha_inicio)
            
        if fecha_fin:
            query += " AND DATE(fecha_movimiento) <= %s"
            params.append(fecha_fin)
            
        query += " ORDER BY fecha_movimiento DESC LIMIT 100"
        
        cursor.execute(query, params)
        movimientos = cursor.fetchall()
        
        return jsonify({
            'status': 'success',
            'movimientos': movimientos
        })
        
    except Exception as e:
        print(f"Error al obtener movimientos ferretero: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

def verificar_stock_bajo():
    """Verificar materiales con stock bajo o agotado"""
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return []
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener materiales con problemas de stock
        cursor.execute("""
            SELECT 
                material_tipo,
                cantidad_disponible,
                cantidad_minima,
                CASE 
                    WHEN cantidad_disponible = 0 THEN 'agotado'
                    WHEN cantidad_disponible <= cantidad_minima THEN 'bajo'
                    ELSE 'normal'
                END as estado_stock
            FROM stock_ferretero 
            WHERE cantidad_disponible <= cantidad_minima
            ORDER BY 
                CASE 
                    WHEN cantidad_disponible = 0 THEN 1
                    ELSE 2
                END,
                material_tipo
        """)
        
        materiales_problematicos = cursor.fetchall()
        
        # Formatear nombres de materiales para mostrar
        nombres_materiales = {
            'silicona': 'Silicona',
            'amarres_negros': 'Amarres Negros',
            'amarres_blancos': 'Amarres Blancos',
            'cinta_aislante': 'Cinta Aislante',
            'grapas_blancas': 'Grapas Blancas',
            'grapas_negras': 'Grapas Negras'
        }
        
        for material in materiales_problematicos:
            material['nombre_display'] = nombres_materiales.get(material['material_tipo'], material['material_tipo'])
        
        return materiales_problematicos
        
    except Exception as e:
        print(f"Error al verificar stock bajo: {str(e)}")
        return []
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/logistica/suministros_ferretero', methods=['GET'])
@login_required()
@role_required('logistica')
def obtener_suministros_ferretero():
    """Obtener suministros de la familia 'Material Ferretero'"""
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener suministros de material ferretero
        query = """
        SELECT 
            id_suministros,
            suministros_codigo,
            suministros_descripcion,
            suministros_cantidad,
            suministros_costo_unitario,
            fecha_registro
        FROM suministros 
        WHERE suministros_familia = 'Material Ferretero' 
        AND suministros_estado = 'Activo'
        ORDER BY suministros_descripcion
        """
        
        cursor.execute(query)
        suministros = cursor.fetchall()
        
        return jsonify({
            'status': 'success',
            'suministros': suministros
        })
        
    except Exception as e:
        print(f"Error al obtener suministros ferretero: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500
    
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/logistica/transferir_suministro_ferretero', methods=['POST'])
@login_required()
@role_required('logistica')
def transferir_suministro_ferretero():
    """Transferir suministro a stock ferretero"""
    connection = None
    cursor = None
    try:
        data = request.get_json()
        id_suministro = data.get('id_suministro')
        cantidad_transferir = data.get('cantidad')
        material_tipo = data.get('material_tipo')  # silicona, amarres_negros, etc.
        
        if not all([id_suministro, cantidad_transferir, material_tipo]):
            return jsonify({
                'status': 'error',
                'message': 'Faltan datos requeridos'
            }), 400
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'status': 'error',
                'message': 'Error de conexión a la base de datos'
            }), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # Verificar que el suministro existe y tiene cantidad suficiente
        cursor.execute("""
            SELECT suministros_cantidad, suministros_descripcion 
            FROM suministros 
            WHERE id_suministros = %s AND suministros_familia = 'Material Ferretero'
        """, (id_suministro,))
        
        suministro = cursor.fetchone()
        if not suministro:
            return jsonify({
                'status': 'error',
                'message': 'Suministro no encontrado'
            }), 404
        
        if suministro['suministros_cantidad'] < cantidad_transferir:
            return jsonify({
                'status': 'error',
                'message': 'Cantidad insuficiente en suministros'
            }), 400
        
        # Iniciar transacción
        cursor.execute("START TRANSACTION")
        
        # Actualizar cantidad en suministros
        nueva_cantidad_suministro = suministro['suministros_cantidad'] - cantidad_transferir
        cursor.execute("""
            UPDATE suministros 
            SET suministros_cantidad = %s 
            WHERE id_suministros = %s
        """, (nueva_cantidad_suministro, id_suministro))
        
        # Registrar entrada en entradas_ferretero
        cursor.execute("""
            INSERT INTO entradas_ferretero 
            (material_tipo, cantidad_entrada, precio_unitario, proveedor, numero_factura, observaciones, fecha_entrada, usuario_registro)
            VALUES (%s, %s, 0, 'Transferencia desde Suministros', %s, %s, NOW(), %s)
        """, (
            material_tipo, 
            cantidad_transferir, 
            f'SUM-{id_suministro}',
            f'Transferido desde: {suministro["suministros_descripcion"]}',
            session['user_id']
        ))
        
        # Confirmar transacción
        connection.commit()
        
        return jsonify({
            'status': 'success',
            'message': f'Transferencia exitosa: {cantidad_transferir} unidades de {material_tipo}'
        })
        
    except Exception as e:
        print(f"Error al transferir suministro: {str(e)}")
        if connection:
            connection.rollback()
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500
    
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/logistica/actualizar_stock_ferretero', methods=['POST'])
@login_required()
@role_required('logistica')
def actualizar_stock_ferretero():
    """Endpoint para actualizar el stock del ferretero después de asignaciones"""
    try:
        # Este endpoint simplemente confirma que el stock debe ser actualizado
        # La actualización real se hace automáticamente por los triggers de la base de datos
        return jsonify({
            'status': 'success',
            'message': 'Stock actualizado correctamente'
        })
        
    except Exception as e:
        print(f"Error al actualizar stock ferretero: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': str(e)
        }), 500

@app.route('/logistica/automotor')
@login_required()
@role_required('logistica')
def ver_parque_automotor():
    try:
        connection = get_db_connection()
        if connection is None:
            return render_template('error.html', 
                               mensaje='Error de conexión a la base de datos',
                               error='No se pudo establecer conexión con la base de datos')
                               
        cursor = connection.cursor(dictionary=True)
        
        # Obtener todos los vehículos con información del técnico asignado
        cursor.execute("""
            SELECT pa.*, r.nombre as tecnico_nombre, r.recurso_operativo_cedula, r.cargo
            FROM parque_automotor pa 
            LEFT JOIN recurso_operativo r ON pa.id_codigo_consumidor = r.id_codigo_consumidor 
            ORDER BY pa.fecha_asignacion DESC
        """)
        vehiculos = cursor.fetchall()

        # Convertir fechas de vencimiento a objetos datetime.date
        for vehiculo in vehiculos:
            if isinstance(vehiculo['soat_vencimiento'], datetime):
                vehiculo['soat_vencimiento'] = vehiculo['soat_vencimiento'].date()
            if isinstance(vehiculo['tecnomecanica_vencimiento'], datetime):
                vehiculo['tecnomecanica_vencimiento'] = vehiculo['tecnomecanica_vencimiento'].date()

        # Obtener lista de técnicos disponibles
        cursor.execute("""
            SELECT id_codigo_consumidor, nombre, recurso_operativo_cedula, cargo 
            FROM recurso_operativo 
            WHERE cargo LIKE '%TECNICO%' OR cargo LIKE '%TÉCNICO%'
            ORDER BY nombre
        """)
        tecnicos = cursor.fetchall()
        
        cursor.close()
        connection.close()

        return render_template('modulos/logistica/automotor.html', 
                            vehiculos=vehiculos,
                            tecnicos=tecnicos,
                            fecha_actual=datetime.now().date())

    except Exception as e:
        print(f"Error al obtener parque automotor: {str(e)}")
        return render_template('error.html', 
                           mensaje='Error al cargar el parque automotor',
                           error=str(e))

@app.route('/logistica/registrar_vehiculo', methods=['POST'])
@login_required()
@role_required('logistica')
def registrar_vehiculo():
    connection = None
    cursor = None
    try:
        # Obtener datos del formulario (IDs corregidos para coincidir con el frontend)
        placa = request.form.get('placa_vehiculo')
        tipo_vehiculo = request.form.get('tipo_vehiculo')
        marca = request.form.get('marca_vehiculo')
        modelo = request.form.get('modelo_vehiculo')
        color = request.form.get('color')
        supervisor = request.form.get('supervisor')
        id_codigo_consumidor = request.form.get('id_codigo_consumidor')
        fecha_asignacion = request.form.get('fecha_asignacion')
        soat_vencimiento = request.form.get('fecha_vencimiento_soat')
        tecnomecanica_vencimiento = request.form.get('fecha_vencimiento_tecnomecanica')
        observaciones = request.form.get('observaciones')

        # Validar campos requeridos
        if not all([placa, tipo_vehiculo, marca, modelo, color, fecha_asignacion]):
            return jsonify({
                'status': 'error',
                'message': 'Todos los campos marcados con * son requeridos'
            }), 400

        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'status': 'error',
                'message': 'Error de conexión a la base de datos'
            }), 500

        cursor = connection.cursor(dictionary=True)

        # Verificar si la placa ya existe
        cursor.execute('SELECT placa FROM parque_automotor WHERE placa = %s', (placa,))
        if cursor.fetchone():
            return jsonify({
                'status': 'error',
                'message': 'Ya existe un vehículo registrado con esta placa'
            }), 400

        # Insertar el nuevo vehículo
        cursor.execute("""
            INSERT INTO parque_automotor (
                placa, tipo_vehiculo, marca, modelo, color, supervisor,
                id_codigo_consumidor, fecha_asignacion,
                soat_vencimiento, tecnomecanica_vencimiento, observaciones
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            placa, tipo_vehiculo, marca, modelo, color, supervisor,
            id_codigo_consumidor, fecha_asignacion,
            soat_vencimiento, tecnomecanica_vencimiento, observaciones
        ))
        
        connection.commit()

        return jsonify({
            'status': 'success',
            'message': 'Vehículo registrado exitosamente'
        }), 201

    except mysql.connector.Error as e:
        print(f"Error MySQL: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error en la base de datos: {str(e)}'
        }), 500
    except Exception as e:
        print(f"Error general: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error al registrar el vehículo: {str(e)}'
        }), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/logistica/actualizar_vehiculo/<int:id_parque_automotor>', methods=['POST'])
@login_required()
@role_required('logistica')
def actualizar_vehiculo(id_parque_automotor):
    connection = None
    cursor = None
    try:
        # Obtener datos del formulario (IDs corregidos para coincidir con el frontend)
        placa = request.form.get('placa_vehiculo')
        tipo_vehiculo = request.form.get('tipo_vehiculo')
        marca = request.form.get('marca_vehiculo')
        modelo = request.form.get('modelo_vehiculo')
        color = request.form.get('color')
        supervisor = request.form.get('supervisor')
        id_codigo_consumidor = request.form.get('id_codigo_consumidor')
        fecha_asignacion = request.form.get('fecha_asignacion')
        estado = request.form.get('estado', 'Activo')
        soat_vencimiento = request.form.get('fecha_vencimiento_soat')
        tecnomecanica_vencimiento = request.form.get('fecha_vencimiento_tecnomecanica')
        observaciones = request.form.get('observaciones', '')
        
        # Campos adicionales válidos
        vin = request.form.get('vin')
        parque_automotorcol = request.form.get('parque_automotorcol')
        licencia = request.form.get('licencia')
        cedula_propietario = request.form.get('cedula_propietario')
        nombre_propietario = request.form.get('nombre_propietario')
        kilometraje_actual = request.form.get('kilometraje_actual')
        proximo_mantenimiento_km = request.form.get('proximo_mantenimiento_km')
        fecha_ultimo_mantenimiento = request.form.get('fecha_ultimo_mantenimiento')
        fecha_actualizacion = request.form.get('fecha_actualizacion')
        
        # Campos de inspección física
        estado_carroceria = request.form.get('estado_carroceria')
        estado_llantas = request.form.get('estado_llantas')
        estado_frenos = request.form.get('estado_frenos')
        estado_motor = request.form.get('estado_motor')
        estado_luces = request.form.get('estado_luces')
        estado_espejos = request.form.get('estado_espejos')
        estado_vidrios = request.form.get('estado_vidrios')
        estado_asientos = request.form.get('estado_asientos')
        
        # Campos de elementos de seguridad
        cinturon_seguridad = request.form.get('cinturon_seguridad')
        extintor = request.form.get('extintor')
        botiquin = request.form.get('botiquin')
        triangulos_seguridad = request.form.get('triangulos_seguridad')
        llanta_repuesto = request.form.get('llanta_repuesto')
        herramientas = request.form.get('herramientas')
        gato = request.form.get('gato')
        cruceta = request.form.get('cruceta')
        
        # Campos operativos válidos
        centro_de_trabajo = request.form.get('centro_de_trabajo')
        ciudad = request.form.get('ciudad')
        licencia_conduccion = request.form.get('licencia_conduccion')

        # Validar campos requeridos
        if not all([placa, tipo_vehiculo, marca, modelo, color, fecha_asignacion]):
            return jsonify({
                'status': 'error',
                'message': 'Todos los campos marcados con * son requeridos'
            }), 400

        # Convertir id_codigo_consumidor a None si está vacío
        id_codigo_consumidor = None if not id_codigo_consumidor else id_codigo_consumidor

        # Convertir fechas vacías a None
        soat_vencimiento = None if not soat_vencimiento else soat_vencimiento
        tecnomecanica_vencimiento = None if not tecnomecanica_vencimiento else tecnomecanica_vencimiento

        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'status': 'error',
                'message': 'Error de conexión a la base de datos'
            }), 500

        cursor = connection.cursor(dictionary=True)

        # Verificar si el vehículo existe
        cursor.execute('SELECT id_parque_automotor FROM parque_automotor WHERE id_parque_automotor = %s', (id_parque_automotor,))
        if not cursor.fetchone():
            return jsonify({
                'status': 'error',
                'message': 'El vehículo no existe'
            }), 404

        # Verificar si la placa ya existe para otro vehículo
        cursor.execute('SELECT id_parque_automotor FROM parque_automotor WHERE placa = %s AND id_parque_automotor != %s', (placa, id_parque_automotor))
        if cursor.fetchone():
            return jsonify({
                'status': 'error',
                'message': 'Ya existe otro vehículo registrado con esta placa'
            }), 400

        # Convertir fechas vacías a None para los campos válidos
        fecha_ultimo_mantenimiento = None if not fecha_ultimo_mantenimiento else fecha_ultimo_mantenimiento
        fecha_actualizacion = None if not fecha_actualizacion else fecha_actualizacion
        
        # Convertir valores numéricos vacíos a None
        kilometraje_actual = None if not kilometraje_actual else kilometraje_actual
        proximo_mantenimiento_km = None if not proximo_mantenimiento_km else proximo_mantenimiento_km

        # Actualizar el vehículo (solo campos que existen en la tabla)
        cursor.execute("""
            UPDATE parque_automotor SET
                placa = %s,
                tipo_vehiculo = %s,
                marca = %s,
                modelo = %s,
                color = %s,
                supervisor = %s,
                id_codigo_consumidor = %s,
                fecha_asignacion = %s,
                estado = %s,
                soat_vencimiento = %s,
                tecnomecanica_vencimiento = %s,
                observaciones = %s,
                vin = %s,
                parque_automotorcol = %s,
                licencia = %s,
                cedula_propietario = %s,
                nombre_propietario = %s,
                kilometraje_actual = %s,
                proximo_mantenimiento_km = %s,
                fecha_ultimo_mantenimiento = %s,
                fecha_actualizacion = %s,
                estado_carroceria = %s,
                estado_llantas = %s,
                estado_frenos = %s,
                estado_motor = %s,
                estado_luces = %s,
                estado_espejos = %s,
                estado_vidrios = %s,
                estado_asientos = %s,
                cinturon_seguridad = %s,
                extintor = %s,
                botiquin = %s,
                triangulos_seguridad = %s,
                llanta_repuesto = %s,
                herramientas = %s,
                gato = %s,
                cruceta = %s,
                centro_de_trabajo = %s,
                ciudad = %s,
                licencia_conduccion = %s
            WHERE id_parque_automotor = %s
        """, (
            placa, tipo_vehiculo, marca, modelo, color, supervisor,
            id_codigo_consumidor, fecha_asignacion, estado,
            soat_vencimiento, tecnomecanica_vencimiento, observaciones,
            vin, parque_automotorcol, licencia, cedula_propietario, nombre_propietario,
            kilometraje_actual, proximo_mantenimiento_km, fecha_ultimo_mantenimiento,
            fecha_actualizacion, estado_carroceria, estado_llantas, estado_frenos, estado_motor,
            estado_luces, estado_espejos, estado_vidrios, estado_asientos,
            cinturon_seguridad, extintor, botiquin, triangulos_seguridad,
            llanta_repuesto, herramientas, gato, cruceta,
            centro_de_trabajo, ciudad, licencia_conduccion, id_parque_automotor
        ))
        
        connection.commit()

        return jsonify({
            'status': 'success',
            'message': 'Vehículo actualizado exitosamente'
        })

    except mysql.connector.Error as e:
        print(f"Error MySQL en actualización: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error al actualizar en la base de datos: {str(e)}'
        }), 500
    except Exception as e:
        print(f"Error general en actualizar_vehiculo: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error al actualizar el vehículo: {str(e)}'
        }), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/logistica/exportar_automotor_csv')
@login_required()
@role_required('logistica')
def exportar_automotor_csv():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener todos los vehículos con información del técnico
        cursor.execute("""
            SELECT 
                pa.placa,
                pa.tipo_vehiculo,
                pa.marca,
                pa.modelo,
                pa.color,
                r.nombre as nombre_tecnico,
                r.recurso_operativo_cedula as cedula_tecnico,
                r.cargo as cargo_tecnico,
                pa.fecha_asignacion,
                pa.estado,
                pa.soat_vencimiento,
                pa.tecnomecanica_vencimiento,
                pa.observaciones
            FROM parque_automotor pa 
            LEFT JOIN recurso_operativo r ON pa.id_codigo_consumidor = r.id_codigo_consumidor 
            ORDER BY pa.fecha_asignacion DESC
        """)
        vehiculos = cursor.fetchall()
        
        # Crear archivo CSV en memoria
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Escribir encabezados
        writer.writerow([
            'Placa',
            'Tipo Vehículo',
            'Marca',
            'Modelo',
            'Color',
            'Técnico Asignado',
            'Cédula Técnico',
            'Cargo Técnico',
            'Fecha Asignación',
            'Estado',
            'Vencimiento SOAT',
            'Vencimiento Tecnomecánica',
            'Observaciones'
        ])
        
        # Escribir datos
        for vehiculo in vehiculos:
            writer.writerow([
                vehiculo['placa'],
                vehiculo['tipo_vehiculo'],
                vehiculo['marca'],
                vehiculo['modelo'],
                vehiculo['color'],
                vehiculo.get('nombre_tecnico', 'No asignado'),
                vehiculo.get('cedula_tecnico', 'No disponible'),
                vehiculo.get('cargo_tecnico', 'No especificado'),
                vehiculo['fecha_asignacion'].strftime('%Y-%m-%d') if vehiculo['fecha_asignacion'] else '',
                vehiculo.get('estado', 'Activo'),
                vehiculo['soat_vencimiento'].strftime('%Y-%m-%d') if vehiculo['soat_vencimiento'] else '',
                vehiculo['tecnomecanica_vencimiento'].strftime('%Y-%m-%d') if vehiculo['tecnomecanica_vencimiento'] else '',
                vehiculo.get('observaciones', '')
            ])
        
        # Preparar respuesta
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            mimetype='text/csv; charset=utf-8-sig',
            as_attachment=True,
            download_name=f'parque_automotor_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )
        
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

# Alias route for the template's expected endpoint name
@app.route('/logistica/automotor/exportar_csv')
@login_required()
@role_required('logistica')
def exportar_vehiculos_csv():
    """CSV export route with the name expected by the template"""
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener todos los vehículos con información del técnico
        cursor.execute("""
            SELECT 
                pa.placa,
                pa.tipo_vehiculo,
                pa.marca,
                pa.modelo,
                pa.color,
                r.nombre as nombre_tecnico,
                r.recurso_operativo_cedula as cedula_tecnico,
                r.cargo as cargo_tecnico,
                pa.fecha_asignacion,
                pa.estado,
                pa.soat_vencimiento,
                pa.tecnomecanica_vencimiento,
                pa.observaciones
            FROM parque_automotor pa 
            LEFT JOIN recurso_operativo r ON pa.id_codigo_consumidor = r.id_codigo_consumidor 
            ORDER BY pa.fecha_asignacion DESC
        """)
        vehiculos = cursor.fetchall()
        
        # Crear archivo CSV en memoria
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Escribir encabezados
        writer.writerow([
            'Placa',
            'Tipo Vehículo',
            'Marca',
            'Modelo',
            'Color',
            'Técnico Asignado',
            'Cédula Técnico',
            'Cargo Técnico',
            'Fecha Asignación',
            'Estado',
            'Vencimiento SOAT',
            'Vencimiento Tecnomecánica',
            'Observaciones'
        ])
        
        # Escribir datos
        for vehiculo in vehiculos:
            writer.writerow([
                vehiculo['placa'],
                vehiculo['tipo_vehiculo'],
                vehiculo['marca'],
                vehiculo['modelo'],
                vehiculo['color'],
                vehiculo.get('nombre_tecnico', 'No asignado'),
                vehiculo.get('cedula_tecnico', 'No disponible'),
                vehiculo.get('cargo_tecnico', 'No especificado'),
                vehiculo['fecha_asignacion'].strftime('%Y-%m-%d') if vehiculo['fecha_asignacion'] else '',
                vehiculo.get('estado', 'Activo'),
                vehiculo['soat_vencimiento'].strftime('%Y-%m-%d') if vehiculo['soat_vencimiento'] else '',
                vehiculo['tecnomecanica_vencimiento'].strftime('%Y-%m-%d') if vehiculo['tecnomecanica_vencimiento'] else '',
                vehiculo.get('observaciones', '')
            ])
        
        # Preparar respuesta
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),
            mimetype='text/csv; charset=utf-8-sig',
            as_attachment=True,
            download_name=f'vehiculos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        )
        
    except Error as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/logistica/ultima_asignacion')
@login_required()
@role_required('logistica')
def obtener_ultima_asignacion():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'status': 'error',
                'message': 'Error de conexión a la base de datos'
            }), 500

        cursor = connection.cursor(dictionary=True)
        
        # Obtener la última asignación creada (por ID, asumiendo que es autoincremental)
        cursor.execute("""
            SELECT id_asignacion 
            FROM asignacion 
            ORDER BY id_asignacion DESC 
            LIMIT 1
        """)
        
        resultado = cursor.fetchone()
        
        if resultado:
            return jsonify({
                'status': 'success',
                'id_asignacion': resultado['id_asignacion'],
                'message': 'Última asignación encontrada'
            })
        else:
            return jsonify({
                'status': 'error',
                'message': 'No se encontraron asignaciones'
            }), 404

    except mysql.connector.Error as e:
        print(f"Error MySQL: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error en la base de datos: {str(e)}'
        }), 500
    except Exception as e:
        print(f"Error general: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error al obtener la última asignación: {str(e)}'
        }), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/logistica/registrar_asignacion_con_firma', methods=['POST'])
@login_required()
@role_required('logistica')
def registrar_asignacion_con_firma():
    connection = None
    cursor = None
    try:
        # Obtener datos básicos
        id_codigo_consumidor = request.form.get('id_codigo_consumidor')
        fecha = request.form.get('fecha')
        cargo = request.form.get('cargo')
        firma = request.form.get('firma')
        id_asignador = request.form.get('id_asignador')

        # Validar campos requeridos
        if not all([id_codigo_consumidor, fecha, cargo, firma, id_asignador]):
            return jsonify({
                'status': 'error',
                'message': 'Faltan campos requeridos para la asignación con firma'
            }), 400

        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'status': 'error',
                'message': 'Error de conexión a la base de datos'
            }), 500

        cursor = connection.cursor(dictionary=True)

        # Verificar que el técnico existe
        cursor.execute('SELECT nombre FROM recurso_operativo WHERE id_codigo_consumidor = %s', (id_codigo_consumidor,))
        tecnico = cursor.fetchone()
        
        if not tecnico:
            return jsonify({
                'status': 'error',
                'message': 'El técnico seleccionado no existe'
            }), 404

        # Obtener la estructura de la tabla
        cursor.execute("""
            SELECT COLUMN_NAME 
            FROM INFORMATION_SCHEMA.COLUMNS 
            WHERE TABLE_SCHEMA = DATABASE()
            AND TABLE_NAME = 'asignacion'
        """)
        columnas_existentes = {row['COLUMN_NAME'].lower() for row in cursor.fetchall()}
        
        print("Columnas existentes:", columnas_existentes)  # Debug log

        # Campos base que siempre deben existir
        campos = ['id_codigo_consumidor', 'asignacion_fecha', 'asignacion_cargo']
        valores = [id_codigo_consumidor, fecha, cargo]

        # Lista de todas las herramientas posibles
        herramientas = [
            'adaptador_mandril', 'alicate', 'barra_45cm', 'bisturi_metalico',
            'broca_3_8', 'broca_386_ran', 'broca_126_ran',
            'broca_metalmadera_14', 'broca_metalmadera_38', 'broca_metalmadera_516',
            'caja_de_herramientas', 'cajon_rojo', 'cinta_de_senal', 'cono_retractil',
            'cortafrio', 'destor_de_estrella', 'destor_de_pala', 'destor_tester',
            'espatula', 'exten_de_corr_10_mts', 'llave_locking_male', 'llave_reliance',
            'llave_torque_rg_11', 'llave_torque_rg_6', 'llaves_mandril',
            'mandril_para_taladro', 'martillo_de_una', 'multimetro',
            'pelacable_rg_6y_rg_11', 'pinza_de_punta', 'pistola_de_silicona',
            'planillero', 'ponchadora_rg_6_y_rg_11', 'ponchadora_rj_45_y_rj11',
            'probador_de_tonos', 'probador_de_tonos_utp', 'puntas_para_multimetro',
            'sonda_metalica', 'tacos_de_madera', 'taladro_percutor',
            'telefono_de_pruebas', 'power_miter', 'bfl_laser', 'cortadora',
            'stripper_fibra', 'pelachaqueta'
        ]

        # Procesar cada herramienta
        for herramienta in herramientas:
            nombre_campo = f'asignacion_{herramienta}'
            if nombre_campo.lower() in columnas_existentes:
                campos.append(nombre_campo)
                valor = request.form.get(herramienta, '0')
                valores.append(valor if valor == '1' else '0')
            else:
                print(f"Campo ignorado (no existe en la tabla): {nombre_campo}")

        # Agregar estado por defecto si existe la columna
        if 'asignacion_estado' in columnas_existentes:
            campos.append('asignacion_estado')
            valores.append('1')
            
        # Agregar campos de firma y asignador
        if 'asignacion_firma' in columnas_existentes:
            campos.append('asignacion_firma')
            valores.append(firma)
            
        if 'id_asignador' in columnas_existentes:
            campos.append('id_asignador')
            valores.append(id_asignador)

        # Construir y ejecutar la consulta SQL
        sql = f"""
            INSERT INTO asignacion ({', '.join(campos)})
            VALUES ({', '.join(['%s'] * len(valores))})
        """
        
        print("SQL Query:", sql)  # Debug log
        print("Valores:", valores)  # Debug log

        cursor.execute(sql, tuple(valores))
        connection.commit()
        
        # Obtener el ID de la asignación recién creada
        cursor.execute("SELECT LAST_INSERT_ID() as id_asignacion")
        id_asignacion = cursor.fetchone()['id_asignacion']

        return jsonify({
            'status': 'success',
            'message': 'Asignación con firma registrada exitosamente',
            'id_asignacion': id_asignacion
        }), 201

    except mysql.connector.Error as e:
        print(f"Error MySQL: {str(e)}")  # Debug log
        return jsonify({
            'status': 'error',
            'message': f'Error en la base de datos: {str(e)}'
        }), 500
    except Exception as e:
        print(f"Error general: {str(e)}")  # Debug log
        return jsonify({
            'status': 'error',
            'message': f'Error al registrar la asignación con firma: {str(e)}'
        }), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/logistica/guardar_firma', methods=['POST'])
@login_required()
@role_required('logistica')
def guardar_firma():
    connection = None
    cursor = None
    try:
        # Obtener datos del JSON
        data = request.get_json()
        if not data:
            return jsonify({
                'status': 'error',
                'message': 'No se recibieron datos JSON'
            }), 400
            
        id_asignacion = data.get('id_asignacion')
        firma = data.get('firma')
        id_asignador = data.get('id_asignador')

        # Validar campos requeridos
        if not all([id_asignacion, firma, id_asignador]):
            return jsonify({
                'status': 'error',
                'message': 'Faltan campos requeridos (id_asignacion, firma, id_asignador)'
            }), 400

        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'status': 'error',
                'message': 'Error de conexión a la base de datos'
            }), 500

        cursor = connection.cursor(dictionary=True)

        # Verificar que la asignación existe
        cursor.execute('SELECT id_asignacion FROM asignacion WHERE id_asignacion = %s', (id_asignacion,))
        asignacion = cursor.fetchone()
        
        if not asignacion:
            return jsonify({
                'status': 'error',
                'message': 'La asignación no existe'
            }), 404

        # Actualizar la asignación con la firma
        cursor.execute("""
            UPDATE asignacion 
            SET asignacion_firma = %s, id_asignador = %s 
            WHERE id_asignacion = %s
        """, (firma, id_asignador, id_asignacion))
        
        connection.commit()

        return jsonify({
            'status': 'success',
            'message': 'Firma guardada exitosamente'
        })

    except mysql.connector.Error as e:
        print(f"Error MySQL: {str(e)}")  # Debug log
        return jsonify({
            'status': 'error',
            'message': f'Error en la base de datos: {str(e)}'
        }), 500
    except Exception as e:
        print(f"Error general: {str(e)}")  # Debug log
        return jsonify({
            'status': 'error',
            'message': f'Error al guardar la firma: {str(e)}'
        }), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/logistica/guardar_asignacion', methods=['POST'])
@login_required(role=['administrativo', 'logistica'])
def guardar_asignacion():
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'status': 'error', 'message': 'Error de conexión a la base de datos'})
            
        cursor = connection.cursor()
        
        # Obtener datos del formulario
        id_codigo_consumidor = request.form.get('id_codigo_consumidor')
        fecha = request.form.get('fecha')
        cargo = request.form.get('cargo')
        
        # Manejar la imagen si se proporcionó una
        imagen_path = None
        if 'imagen' in request.files:
            imagen = request.files['imagen']
            if imagen and imagen.filename:
                # Generar nombre único para la imagen
                extension = os.path.splitext(imagen.filename)[1]
                nuevo_nombre = f"asignacion_{datetime.now().strftime('%Y%m%d_%H%M%S')}{extension}"
                
                # Asegurar que el directorio existe
                upload_dir = os.path.join(app.root_path, 'static', 'uploads', 'asignacion')
                os.makedirs(upload_dir, exist_ok=True)
                
                # Guardar la imagen
                imagen_path = os.path.join('uploads', 'asignacion', nuevo_nombre)
                imagen.save(os.path.join(app.root_path, 'static', imagen_path))
        
        # Insertar asignación sin imagen (campo no existe en la tabla)
        cursor.execute("""
            INSERT INTO asignacion (
                id_codigo_consumidor, asignacion_fecha, asignacion_cargo, 
                asignacion_estado
            ) VALUES (%s, %s, %s, %s)
        """, (
            id_codigo_consumidor, fecha, cargo, '1'
        ))
        
        id_asignacion = cursor.lastrowid
        
        # Procesar herramientas usando la nueva tabla asignacion_herramientas
        for key, value in request.form.items():
            if key not in ['id_codigo_consumidor', 'fecha', 'cargo'] and value == '1' and key.strip():
                # Verificar que la clave no esté vacía
                if key.strip():
                    try:
                        # Mantener compatibilidad con el enfoque anterior
                        campo = f"asignacion_{key}"
                        if len(campo) > 11:  # 'asignacion_' tiene 11 caracteres
                            try:
                                # Actualizar la columna en la tabla asignacion si existe
                                query = f"""
                                    UPDATE asignacion 
                                    SET {campo} = '1'
                                    WHERE id_asignacion = %s
                                """
                                cursor.execute(query, (id_asignacion,))
                            except mysql.connector.Error as e:
                                # Si hay error, es probable que la columna no exista
                                # Continuamos con el nuevo enfoque
                                print(f"Columna {campo} no encontrada: {str(e)}")
                        
                        # Insertar en la nueva tabla asignacion_herramientas
                        cursor.execute("""
                            INSERT INTO asignacion_herramientas 
                            (id_asignacion, codigo, descripcion, estado) 
                            VALUES (%s, %s, %s, %s)
                        """, (
                            id_asignacion, 
                            key,  # Usar la clave como código
                            key.replace('_', ' ').title(),  # Convertir a formato legible
                            '1'  # Estado activo
                        ))
                    except mysql.connector.Error as e:
                        # Registrar el error pero continuar con otras herramientas
                        print(f"Error al guardar herramienta {key}: {str(e)}")
                        continue
        
        # Si hay imagen, guardarla en la tabla asignacion_herramientas como un registro especial
        if imagen_path:
            try:
                cursor.execute("""
                    INSERT INTO asignacion_herramientas 
                    (id_asignacion, codigo, descripcion, estado) 
                    VALUES (%s, %s, %s, %s)
                """, (
                    id_asignacion, 
                    'imagen',  
                    imagen_path,  # Guardar la ruta de la imagen
                    '1'  # Estado activo
                ))
            except mysql.connector.Error as e:
                print(f"Error al guardar imagen: {str(e)}")
        
        connection.commit()
        return jsonify({
            'status': 'success',
            'message': 'Asignación guardada correctamente',
            'id_asignacion': id_asignacion
        })
        
    except mysql.connector.Error as e:
        return jsonify({'status': 'error', 'message': f'Error al guardar la asignación: {str(e)}'})
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

# Modificar la ruta de obtener detalles para incluir la imagen
@app.route('/logistica/guardar_asignacion_simple', methods=['POST'])
@login_required(role=['administrativo', 'logistica'])
def guardar_asignacion_simple():
    connection = None
    cursor = None
    try:
        # Obtener datos básicos del formulario
        id_codigo_consumidor = request.form.get('id_codigo_consumidor')
        fecha = request.form.get('fecha')
        cargo = request.form.get('cargo')
        
        # Validar datos requeridos
        if not id_codigo_consumidor or not fecha or not cargo:
            return jsonify({
                'status': 'error',
                'message': 'Faltan campos obligatorios'
            }), 400
            
        # Obtener conexión a la base de datos
        connection = get_db_connection()
        if not connection:
            return jsonify({
                'status': 'error',
                'message': 'Error al conectar con la base de datos'
            }), 500
            
        # Obtener información del técnico
        cursor = connection.cursor(dictionary=True)
        
        query_tecnico = """
            SELECT nombre, recurso_operativo_cedula 
            FROM recurso_operativo 
            WHERE id_codigo_consumidor = %s
        """
        cursor.execute(query_tecnico, (id_codigo_consumidor,))
        tecnico = cursor.fetchone()
        
        if not tecnico:
            return jsonify({
                'status': 'error',
                'message': 'Técnico no encontrado'
            }), 404
            
        # Convertir fecha a formato datetime
        try:
            fecha_obj = datetime.strptime(fecha, '%Y-%m-%dT%H:%M')
            fecha_formateada = fecha_obj.strftime('%Y-%m-%d %H:%M:%S')
        except ValueError:
            return jsonify({
                'status': 'error',
                'message': 'Formato de fecha inválido'
            }), 400
        
        # Insertar asignación básica
        query = """
            INSERT INTO asignacion (
                id_codigo_consumidor, 
                asignacion_cedula, 
                asignacion_nombre, 
                asignacion_fecha, 
                asignacion_cargo, 
                asignacion_estado
            ) VALUES (%s, %s, %s, %s, %s, %s)
        """
        
        # Ejecutar consulta
        cursor.execute(query, [
            id_codigo_consumidor,
            tecnico['recurso_operativo_cedula'],  # cedula
            tecnico['nombre'],  # nombre
            fecha_formateada,
            cargo,
            '1'  # Estado activo
        ])
        
        # Obtener ID de la asignación
        id_asignacion = cursor.lastrowid
        
        # Procesar herramientas seleccionadas
        for key, value in request.form.items():
            if key not in ['id_codigo_consumidor', 'fecha', 'cargo'] and value == '1' and key.strip():
                try:
                    # Insertar en la tabla asignacion_herramientas
                    cursor.execute("""
                        INSERT INTO asignacion_herramientas 
                        (id_asignacion, codigo, descripcion, estado) 
                        VALUES (%s, %s, %s, %s)
                    """, (
                        id_asignacion, 
                        key,  # Usar la clave como código
                        key.replace('_', ' ').title(),  # Convertir a formato legible
                        '1'  # Estado activo
                    ))
                except mysql.connector.Error as e:
                    # Registrar el error pero continuar con otras herramientas
                    print(f"Error al guardar herramienta {key}: {str(e)}")
                    continue
        
        # Hacer commit y devolver respuesta exitosa
        connection.commit()
        return jsonify({
            'status': 'success',
            'message': 'Asignación básica guardada correctamente',
            'id_asignacion': id_asignacion
        }), 200
            
    except mysql.connector.Error as e:
        if connection:
            connection.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Error al guardar la asignación simple: {str(e)}'
        }), 500
    except Exception as e:
        if connection:
            connection.rollback()
        return jsonify({
            'status': 'error',
            'message': f'Error al procesar la solicitud: {str(e)}'
        }), 500
    finally:
        # Asegurarse de cerrar cursor y conexión
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/logistica/asignacion/<int:id>')
@login_required(role=['administrativo', 'logistica'])
def obtener_detalle_asignacion(id):
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'status': 'error', 'message': 'Error de conexión a la base de datos'})
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener información básica
        cursor.execute("""
            SELECT a.*, ro.nombre, ro.recurso_operativo_cedula
            FROM asignacion a
            JOIN recurso_operativo ro ON a.id_codigo_consumidor = ro.id_codigo_consumidor
            WHERE a.id_asignacion = %s
        """, (id,))
        
        asignacion = cursor.fetchone()
        if not asignacion:
            return jsonify({'status': 'error', 'message': 'Asignación no encontrada'})
            
        # Buscar imagen en la tabla asignacion_herramientas
        cursor.execute("""
            SELECT descripcion
            FROM asignacion_herramientas
            WHERE id_asignacion = %s AND codigo = 'imagen'
        """, (id,))
        
        imagen_result = cursor.fetchone()
        imagen_path = imagen_result['descripcion'] if imagen_result else None
        imagen_url = url_for('static', filename=imagen_path) if imagen_path else None
        
        info_basica = {
            'tecnico': asignacion['nombre'],
            'cedula': asignacion['recurso_operativo_cedula'],
            'cargo': asignacion['asignacion_cargo'],
            'fecha': asignacion['asignacion_fecha'].strftime('%Y-%m-%d %H:%M:%S'),
            'estado': 'Activo' if asignacion['asignacion_estado'] == '1' else 'Inactivo',
            'imagen': imagen_url
        }
        
        # Obtener detalles de herramientas
        cursor.execute("""
            SELECT codigo, descripcion, estado
            FROM asignacion_herramientas
            WHERE id_asignacion = %s
        """, (id,))
        
        herramientas = cursor.fetchall()
        
        # Clasificar herramientas
        herramientas_basicas = {}
        brocas = {}
        herramientas_red = {}
        
        for h in herramientas:
            nombre = h['descripcion']
            estado = h['estado']
            codigo = h['codigo']
            
            if 'broca' in nombre.lower():
                brocas[nombre] = {'estado': estado, 'codigo': codigo}
            elif any(red in nombre.lower() for red in ['cajon', 'cinta', 'cono', 'llave', 'ponchadora']):
                herramientas_red[nombre] = {'estado': estado, 'codigo': codigo}
            else:
                herramientas_basicas[nombre] = {'estado': estado, 'codigo': codigo}
        
        return jsonify({
            'status': 'success',
            'data': {
                'info_basica': info_basica,
                'herramientas_basicas': herramientas_basicas,
                'brocas': brocas,
                'herramientas_red': herramientas_red
            }
        })
        
    except mysql.connector.Error as e:
        return jsonify({'status': 'error', 'message': f'Error de base de datos: {str(e)}'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Error inesperado: {str(e)}'})
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'connection' in locals() and connection.is_connected():
            connection.close()

@app.route('/logistica/asignacion/<int:id_asignacion>/pdf', methods=['GET'])
@login_required(role=['administrativo', 'logistica'])
def generar_pdf_asignacion(id_asignacion):
    try:
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import Image as RLImage
        import base64
        import re
        from io import BytesIO
        
        # Obtener parámetros de la URL
        mostrar_firma = request.args.get('firmado', 'false').lower() == 'true'
        
        connection = get_db_connection()
        if not connection:
            return jsonify({'status': 'error', 'message': 'Error de conexión a la base de datos'})
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener información de la asignación incluyendo la firma
        cursor.execute("""
            SELECT a.*, ro.nombre, ro.recurso_operativo_cedula, fa.firma_imagen
            FROM asignacion a
            JOIN recurso_operativo ro ON a.id_codigo_consumidor = ro.id_codigo_consumidor
            LEFT JOIN firmas_asignaciones fa ON a.id_asignacion = fa.id_asignacion
            WHERE a.id_asignacion = %s
        """, (id_asignacion,))
        
        asignacion = cursor.fetchone()
        if not asignacion:
            return jsonify({'status': 'error', 'message': 'Asignación no encontrada'})
            
        # Obtener la firma si existe
        firma_imagen = asignacion.get('firma_imagen') if mostrar_firma else None
        
        # Obtener la imagen de la asignación si existe
        cursor.execute("""
            SELECT descripcion
            FROM asignacion_herramientas
            WHERE id_asignacion = %s AND codigo = 'imagen'
        """, (id_asignacion,))
        
        imagen_result = cursor.fetchone()
        imagen_path = imagen_result['descripcion'] if imagen_result else None
        
        # Crear PDF con ReportLab
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        contenido = []
        
        # Definir todos los estilos necesarios
        styles = getSampleStyleSheet()
        estilo_heading = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading1'],
            fontSize=14,
            spaceAfter=12,
            spaceBefore=12,
            textColor=colors.black,
            alignment=1  # Centrado
        )
        
        estilo_normal = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=10,
            spaceBefore=6,
            spaceAfter=6
        )
        
        estilo_titulo = ParagraphStyle(
            'CustomTitle',
            parent=styles['Title'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Centrado
        )
        
        # Título
        contenido.append(Paragraph("Formato de Asignación de Herramientas", estilo_titulo))
        
        # Información básica
        data = [
            ["Técnico:", asignacion['nombre']],
            ["Cédula:", asignacion['recurso_operativo_cedula']],
            ["Cargo:", asignacion['asignacion_cargo']],
            ["Fecha:", asignacion['asignacion_fecha'].strftime('%Y-%m-%d %H:%M:%S')],
            ["Estado:", "Activo" if asignacion['asignacion_estado'] == '1' else "Inactivo"]
        ]
        
        # Crear tabla de información básica
        table = Table(data, colWidths=[100, 400])
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('PADDING', (0, 0), (-1, -1), 4)
        ]))
        contenido.append(table)
        contenido.append(Spacer(1, 10))
        
        # Agregar imagen de la asignación si existe
        if imagen_path:
            try:
                contenido.append(Paragraph("Imagen de Herramientas", estilo_heading))
                contenido.append(Spacer(1, 10))
                
                # Ruta completa de la imagen
                ruta_completa = os.path.join(app.root_path, 'static', imagen_path)
                
                if os.path.exists(ruta_completa):
                    # Crear imagen para ReportLab con tamaño controlado
                    imagen_asignacion = RLImage(ruta_completa, width=400, height=200, kind='proportional')
                    contenido.append(imagen_asignacion)
                    app.logger.info(f"Imagen añadida al PDF correctamente: {ruta_completa}")
                else:
                    contenido.append(Paragraph("(Imagen no encontrada)", styles['Normal']))
                    app.logger.warning(f"Imagen no encontrada en la ruta: {ruta_completa}")
                
                contenido.append(Spacer(1, 10))
            except Exception as e:
                app.logger.error(f"Error al procesar imagen para PDF: {str(e)}")
                contenido.append(Paragraph("(Error al cargar imagen)", styles['Normal']))
                import traceback
                app.logger.error(traceback.format_exc())
        
        # Sección de herramientas
        contenido.append(Paragraph("Herramientas Asignadas", estilo_heading))
        contenido.append(Spacer(1, 10))
        
        # Lista de herramientas conocidas y sus descripciones
        herramientas_mapping = {
            'adaptador_mandril': 'Adaptador Mandril',
            'alicate': 'Alicate',
            'barra_45cm': 'Barra 45cm',
            'bisturi_metalico': 'Bisturí Metálico',
            'broca_3_8': 'Broca 3/8',
            'broca_386_ran': 'Broca 3/8 6 Ranuras',
            'broca_126_ran': 'Broca 1/2 6 Ranuras',
            'broca_metalmadera_14': 'Broca Metal/Madera 1/4',
            'broca_metalmadera_38': 'Broca Metal/Madera 3/8',
            'broca_metalmadera_516': 'Broca Metal/Madera 5/16',
            'caja_de_herramientas': 'Caja de Herramientas',
            'cajon_rojo': 'Cajón Rojo',
            'cinta_de_senal': 'Cinta de Señal',
            'cono_retractil': 'Cono Retráctil',
            'cortafrio': 'Cortafrío',
            'destor_de_estrella': 'Destornillador de Estrella',
            'destor_de_pala': 'Destornillador de Pala',
            'destor_tester': 'Destornillador Tester',
            'espatula': 'Espátula',
            'exten_de_corr_10_mts': 'Extensión de Corriente 10mts',
            'llave_locking_male': 'Llave Locking Male',
            'llave_reliance': 'Llave Reliance',
            'llave_torque_rg_11': 'Llave Torque RG-11',
            'llave_torque_rg_6': 'Llave Torque RG-6',
            'llaves_mandril': 'Llaves Mandril',
            'mandril_para_taladro': 'Mandril para Taladro',
            'martillo_de_una': 'Martillo de Una',
            'multimetro': 'Multímetro',
            'pelacable_rg_6y_rg_11': 'Pelacable RG-6 y RG-11',
            'pinza_de_punta': 'Pinza de Punta',
            'pistola_de_silicona': 'Pistola de Silicona',
            'planillero': 'Planillero',
            'ponchadora_rg_6_y_rg_11': 'Ponchadora RG-6 y RG-11',
            'ponchadora_rj_45_y_rj11': 'Ponchadora RJ-45 y RJ-11',
            'probador_de_tonos': 'Probador de Tonos',
            'probador_de_tonos_utp': 'Probador de Tonos UTP',
            'puntas_para_multimetro': 'Puntas para Multímetro',
            'sonda_metalica': 'Sonda Metálica',
            'tacos_de_madera': 'Tacos de Madera',
            'taladro_percutor': 'Taladro Percutor',
            'telefono_de_pruebas': 'Teléfono de Pruebas',
            'power_miter': 'Power Miter',
            'bfl_laser': 'BFL Laser',
            'cortadora': 'Cortadora',
            'stripper_fibra': 'Stripper Fibra',
            'pelachaqueta': 'Pelachaqueta'
        }
        
        # Obtener herramientas de la nueva tabla asignacion_herramientas
        cursor.execute("""
            SELECT codigo, descripcion, estado
            FROM asignacion_herramientas
            WHERE id_asignacion = %s AND codigo != 'imagen'
        """, (id_asignacion,))
        
        herramientas_db = cursor.fetchall()
        
        # Crear lista de herramientas asignadas
        herramientas_asignadas = []
        
        # Primero intentar usar las herramientas de la nueva tabla
        if herramientas_db:
            for herramienta in herramientas_db:
                codigo = herramienta['codigo']
                descripcion = herramienta['descripcion']
                estado = 'Asignada' if herramienta['estado'] == '1' else 'Inactiva'
                
                herramientas_asignadas.append([
                    codigo,
                    descripcion,
                    estado
                ])
        else:
            # Compatibilidad con el enfoque anterior
            for campo, valor in asignacion.items():
                if campo.startswith('asignacion_') and campo not in ['asignacion_fecha', 'asignacion_cargo', 'asignacion_estado', 'asignacion_firma']:
                    nombre_herramienta = campo.replace('asignacion_', '')
                    if valor == '1' and nombre_herramienta in herramientas_mapping:
                        herramientas_asignadas.append([
                            nombre_herramienta,
                            herramientas_mapping[nombre_herramienta],
                            'Asignada'
                        ])
        
        if herramientas_asignadas:
            # Crear tabla de herramientas
            headers = ['Código', 'Descripción', 'Estado']
            data_herramientas = [headers] + herramientas_asignadas
            
            tabla_herramientas = Table(data_herramientas, colWidths=[100, 300, 100])
            tabla_herramientas.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('PADDING', (0, 0), (-1, -1), 4),
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey)
            ]))
            contenido.append(tabla_herramientas)
        else:
            contenido.append(Paragraph("No hay herramientas asignadas", styles['Normal']))
        
        contenido.append(Spacer(1, 30))
        
        # Sección de firma si se solicitó
        if mostrar_firma and firma_imagen:
            try:
                contenido.append(Paragraph("Firma de Aceptación", estilo_heading))
                contenido.append(Spacer(1, 10))
                
                # Manejamos la imagen en memoria, evitando archivos temporales
                if firma_imagen.startswith('data:image'):
                    # Extraer la parte base64
                    firma_base64 = re.sub(r'^data:image/[^;]+;base64,', '', firma_imagen)
                    
                    # Decodificar a bytes
                    firma_bytes = base64.b64decode(firma_base64)
                    
                    # Crear buffer de memoria para la imagen
                    firma_buffer = BytesIO(firma_bytes)
                    
                    # Crear imagen para ReportLab
                    firma_img = RLImage(firma_buffer, width=200, height=100)
                    
                    # Añadir al contenido
                    contenido.append(firma_img)
                    app.logger.info("Firma añadida al PDF correctamente")
                else:
                    contenido.append(Paragraph("(Formato de firma no válido)", styles['Normal']))
                    app.logger.warning("Formato de firma no válido para incluir en PDF")
            except Exception as e:
                app.logger.error(f"Error al procesar firma para PDF: {str(e)}")
                contenido.append(Paragraph("(Error al cargar firma)", styles['Normal']))
                import traceback
                app.logger.error(traceback.format_exc())
        
        contenido.append(Spacer(1, 20))
        
        # Pie de página con fecha de generación
        fecha_generacion = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        pie = Paragraph(f"Documento generado el {fecha_generacion}", styles['Normal'])
        contenido.append(pie)
        
        # Construir el PDF
        doc.build(contenido)
        
        # Obtener los datos del buffer
        pdf_data = buffer.getvalue()
        buffer.close()
        
        # Crear respuesta
        response = make_response(pdf_data)
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'inline; filename=asignacion_{id_asignacion}.pdf'
        
        app.logger.info(f"PDF generado correctamente para asignación {id_asignacion}")
        return response
        
    except Exception as e:
        app.logger.error(f"Error al generar PDF: {str(e)}")
        import traceback
        app.logger.error(traceback.format_exc())
        return jsonify({
            'status': 'error',
            'message': f'Error al generar el PDF: {str(e)}'
        }), 500

@app.route('/logistica/asignacion/<int:id_asignacion>/firmar', methods=['POST'])
@login_required()
@role_required('logistica')
def firmar_asignacion(id_asignacion):
    try:
        # Verificar que el usuario tiene derechos para firmar
        usuario_id = session.get('user_id')
        app.logger.info(f"Intento de firma para asignación {id_asignacion} por usuario {usuario_id}")
        
        # Obtener la conexión a la base de datos
        connection = get_db_connection()
        if not connection:
            app.logger.error("Error al conectar con la base de datos al intentar firmar")
            return jsonify({
                'status': 'error',
                'message': 'Error al conectar con la base de datos'
            }), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener información de la asignación para verificar permisos
        query = """
            SELECT a.*, r.nombre, r.recurso_operativo_cedula
            FROM asignacion a
            JOIN recurso_operativo r ON a.id_codigo_consumidor = r.id_codigo_consumidor
            WHERE a.id_asignacion = %s
        """
        cursor.execute(query, (id_asignacion,))
        asignacion = cursor.fetchone()
        
        if not asignacion:
            cursor.close()
            connection.close()
            app.logger.warning(f"Intento de firma para asignación inexistente: {id_asignacion}")
            return jsonify({
                'status': 'error',
                'message': 'Asignación no encontrada'
            }), 404
        
        # Obtener los datos de la firma manuscrita
        data = request.json
        if not data or 'signature' not in data:
            app.logger.warning("Datos de firma incompletos o inválidos")
            return jsonify({
                'status': 'error',
                'message': 'No se recibieron datos de firma'
            }), 400
        
        # Validar formato de la firma
        firma_imagen = data['signature']
        if not firma_imagen.startswith('data:image/'):
            app.logger.warning("Formato de imagen de firma inválido")
            return jsonify({
                'status': 'error',
                'message': 'Formato de imagen de firma inválido'
            }), 400
        
        # Verificar el tamaño de la firma (para evitar datos demasiado grandes)
        if len(firma_imagen) > 500000:  # Limitar a ~500KB
            app.logger.warning("Imagen de firma demasiado grande")
            return jsonify({
                'status': 'error',
                'message': 'La imagen de firma es demasiado grande'
            }), 400
            
        fecha_firma = datetime.now()
        app.logger.info(f"Firma recibida correctamente para asignación {id_asignacion}")
        
        # Verificar que la firma se puede procesar correctamente
        try:
            # Extraer la parte base64 de la firma
            if 'data:image' in firma_imagen:
                imagen_base64 = re.sub(r'^data:image/[^;]+;base64,', '', firma_imagen)
            
            # Verificar que la decodificación funciona
            imagen_bytes = base64.b64decode(imagen_base64)
            
            # Verificar que la imagen es válida creando un objeto Image con ella
            with PILImage.open(BytesIO(imagen_bytes)) as img:
                formato = img.format
                width, height = img.size
                app.logger.info(f"Imagen de firma validada: formato={formato}, dimensiones={width}x{height}")
        except Exception as e:
            app.logger.error(f"Error al validar la imagen de firma: {str(e)}")
            return jsonify({
                'status': 'error',
                'message': f'Error al procesar la imagen de firma: {str(e)}'
            }), 400
        
        # Primero verificar si existe la tabla para firmas
        try:
            check_table_query = """
                CREATE TABLE IF NOT EXISTS firmas_asignaciones (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    id_asignacion INT NOT NULL,
                    id_usuario INT NOT NULL,
                    firma_imagen LONGTEXT,
                    fecha_firma DATETIME,
                    FOREIGN KEY (id_asignacion) REFERENCES asignacion(id_asignacion)
                )
            """
            cursor.execute(check_table_query)
            connection.commit()
            app.logger.info("Tabla firmas_asignaciones verificada/creada correctamente")
        except Exception as e:
            app.logger.error(f"Error al verificar tabla de firmas: {str(e)}")
            # Continuar aún si falla la verificación, ya que podría existir
        
        # Guardar la firma en la base de datos
        firma_guardada = False
        try:
            # Verificar si ya existe una firma para esta asignación
            cursor.execute("SELECT id FROM firmas_asignaciones WHERE id_asignacion = %s", (id_asignacion,))
            firma_existente = cursor.fetchone()
            
            if firma_existente:
                # Actualizar firma existente
                update_query = """
                    UPDATE firmas_asignaciones 
                    SET firma_imagen = %s, fecha_firma = %s, id_usuario = %s
                    WHERE id_asignacion = %s
                """
                cursor.execute(update_query, (firma_imagen, fecha_firma, usuario_id, id_asignacion))
                app.logger.info(f"Firma actualizada para asignación {id_asignacion}")
            else:
                # Insertar nueva firma
                insert_query = """
                    INSERT INTO firmas_asignaciones 
                    (id_asignacion, id_usuario, firma_imagen, fecha_firma) 
                    VALUES (%s, %s, %s, %s)
                """
                cursor.execute(insert_query, (id_asignacion, usuario_id, firma_imagen, fecha_firma))
                app.logger.info(f"Nueva firma guardada para asignación {id_asignacion}")
            
            connection.commit()
            firma_guardada = True
        except Exception as e:
            app.logger.error(f"Error al guardar firma en base de datos: {str(e)}")
            connection.rollback()
            # Si falla, seguimos para intentar generar el PDF igualmente
            import traceback
            app.logger.error(traceback.format_exc())
            
        # Generar un token JWT como respaldo o para verificación alternativa
        payload = {
            'asignacion_id': id_asignacion,
            'usuario_id': usuario_id,
            'fecha': fecha_firma.strftime('%Y-%m-%d %H:%M:%S'),
            'firma_guardada': firma_guardada,
            'iat': fecha_firma.timestamp(),
            'exp': (fecha_firma + timedelta(days=365)).timestamp()
        }
        
        # En un entorno real, este secreto estaría almacenado de manera segura
        clave_secreta = os.getenv('JWT_SECRET_KEY', 'clave_secreta_para_firmas')
        token = jwt.encode(payload, clave_secreta, algorithm='HS256')
        
        # Cerrar cursor y conexión
        cursor.close()
        connection.close()
        
        # Devolver URL al PDF firmado
        url_pdf_firmado = url_for('generar_pdf_asignacion', id_asignacion=id_asignacion, firmado='true', _external=True)
        
        app.logger.info(f"Proceso de firma completado exitosamente para asignación {id_asignacion}")
        return jsonify({
            'status': 'success',
            'message': 'Documento firmado correctamente',
            'token': token,
            'url_pdf': url_pdf_firmado,
            'firma_guardada': firma_guardada
        })
        
    except Exception as e:
        app.logger.error(f"Error inesperado al firmar documento: {str(e)}")
        
        # Log detallado para diagnóstico
        import traceback
        app.logger.error(traceback.format_exc())
        
        return jsonify({
            'status': 'error',
            'message': f'Error al firmar el documento: {str(e)}'
        }), 500

@app.route('/api/inventario/filtrar', methods=['POST'])
@login_required()
@role_required('logistica')
def filtrar_inventario():
    try:
        # Obtener parámetros del filtro
        filtros = request.json
        categoria = filtros.get('categoria', 'todos')
        ubicacion = filtros.get('ubicacion', 'todos')
        fecha_inicio = filtros.get('fecha_inicio')
        fecha_fin = filtros.get('fecha_fin')
        estado = filtros.get('estado', 'todos')
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # Construir consultas según filtros
        params = []
        where_clauses = []
        
        # Base de la consulta para movimientos
        movimientos_query = """
            SELECT 
                DATE(fecha_movimiento) as fecha,
                tipo_movimiento,
                item_afectado,
                cantidad,
                usuario,
                ubicacion
            FROM movimientos_inventario
            WHERE 1=1
        """
        
        # Aplicar filtros de fecha
        if fecha_inicio:
            where_clauses.append("fecha_movimiento >= %s")
            params.append(fecha_inicio)
        
        if fecha_fin:
            where_clauses.append("fecha_movimiento <= %s")
            params.append(fecha_fin)
        
        # Aplicar filtro de categoría
        if categoria != 'todos':
            # Mapeo de categorías a items
            categoria_items = {
                'herramientas_basicas': ['Adaptador Mandril', 'Alicate', 'Barra 45cm', 'Bisturí Metálico', 
                                         'Caja de Herramientas', 'Cortafrío', 'Destornillador Estrella', 
                                         'Destornillador Pala', 'Destornillador Tester', 'Martillo', 'Pinza de Punta'],
                'herramientas_especializadas': ['Multímetro', 'Taladro Percutor', 'Ponchadora RG6/RG11', 
                                               'Ponchadora RJ45/RJ11', 'Power Meter', 'BFL Laser', 
                                               'Cortadora', 'Stripper Fibra'],
               'material_ferretero': ['Silicona', 'Amarres Negros', 'Amarres Blancos', 
                                      'Cinta Aislante', 'Grapas Blancas', 'Grapas Negras']
            }
            
            if categoria in categoria_items:
                item_placeholders = ', '.join(['%s'] * len(categoria_items[categoria]))
                where_clauses.append(f"item_afectado IN ({item_placeholders})")
                params.extend(categoria_items[categoria])
        
        # Aplicar filtro de ubicación
        if ubicacion != 'todos':
            where_clauses.append("ubicacion = %s")
            params.append(ubicacion)
        
        # Construir consulta final
        if where_clauses:
            movimientos_query += " AND " + " AND ".join(where_clauses)
        
        movimientos_query += " ORDER BY fecha_movimiento DESC LIMIT 100"
        
        # Ejecutar consulta
        cursor.execute(movimientos_query, params)
        movimientos = cursor.fetchall()
        
        # Procesar datos para el gráfico de tendencia
        ultimos_30_dias = []
        hoy = datetime.now().date()
        for i in range(30, -1, -1):
            fecha = hoy - timedelta(days=i)
            ultimos_30_dias.append(fecha.strftime('%Y-%m-%d'))
        
        movimientos_por_dia = {}
        for fecha in ultimos_30_dias:
            movimientos_por_dia[fecha] = 0
            
        for movimiento in movimientos:
            fecha_str = movimiento['fecha'].strftime('%Y-%m-%d')
            if fecha_str in movimientos_por_dia:
                movimientos_por_dia[fecha_str] += 1
        
        # Preparar datos del gráfico de tendencia
        datos_tendencia = {
            'labels': ultimos_30_dias,
            'datos': [movimientos_por_dia.get(fecha, 0) for fecha in ultimos_30_dias]
        }
        
        # Simular estadísticas según los filtros
        # En una implementación real, estos datos se obtendrían de la base de datos
        estadisticas = {
            'total_items': len(movimientos),
            'entradas': sum(1 for m in movimientos if m['tipo_movimiento'] == 'entrada'),
            'salidas': sum(1 for m in movimientos if m['tipo_movimiento'] == 'salida'),
            'transferencias': sum(1 for m in movimientos if m['tipo_movimiento'] == 'transferencia'),
            'ajustes': sum(1 for m in movimientos if m['tipo_movimiento'] == 'ajuste')
        }
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'movimientos': movimientos,
            'tendencia': datos_tendencia,
            'estadisticas': estadisticas
        })
        
    except Exception as e:
        print(f"Error al filtrar inventario: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/inventario/exportar-csv', methods=['POST'])
@login_required()
@role_required('logistica')
def exportar_inventario_csv():
    try:
        # Obtener parámetros del filtro (similar a filtrar_inventario)
        filtros = request.json
        
        # Obtener datos según filtros...
        # (Código similar a filtrar_inventario para obtener los datos)
        
        # Para este ejemplo, usaremos datos simulados
        datos = []
        
        # Herramientas básicas
        for item in ['Adaptador Mandril', 'Alicate', 'Barra 45cm', 'Bisturí Metálico', 
                    'Caja de Herramientas', 'Cortafrío', 'Destornillador Estrella']:
            datos.append({
                'categoria': 'Herramientas Básicas',
                'item': item,
                'cantidad': 15,
                'ubicacion': 'Almacén Principal',
                'estado': 'Disponible'
            })
        
        # Herramientas especializadas
        for item in ['Multímetro', 'Taladro Percutor', 'Ponchadora RG6/RG11']:
            datos.append({
                'categoria': 'Herramientas Especializadas',
                'item': item,
                'cantidad': 8,
                'ubicacion': 'Bodega Norte',
                'estado': 'Disponible'
            })
        
        # Material ferretero
        for item in ['Silicona', 'Amarres Negros', 'Amarres Blancos']:
            datos.append({
                'categoria': 'Material Ferretero',
                'item': item,
                'cantidad': 50,
                'ubicacion': 'Bodega Sur',
                'estado': 'Disponible'
            })
        
        # Crear DataFrame y exportar a CSV
        df = pd.DataFrame(datos)
        csv_data = df.to_csv(index=False, sep=';')
        
        # Crear un objeto de archivo en memoria
        output = io.StringIO()
        output.write(csv_data)
        output.seek(0)
        
        # Devolver el archivo CSV
        return Response(
            output.getvalue(),
            mimetype="text/csv",
            headers={"Content-disposition": "attachment; filename=inventario.csv"}
        )
        
    except Exception as e:
        print(f"Error al exportar inventario a CSV: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/suministros', methods=['GET', 'POST'])
def suministros():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            flash('Error de conexión a la base de datos.', 'error')
            return redirect(url_for('dashboard'))
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener todos los suministros
        cursor.execute("SELECT * FROM suministros")
        suministros_data = cursor.fetchall()
        
        return render_template('modulos/administrativo/suministros.html', suministros=suministros_data)
    
    except mysql.connector.Error as e:
        flash(f'Error de base de datos: {str(e)}', 'error')
        return redirect(url_for('dashboard'))
    
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/guardar_suministro', methods=['POST'])
@login_required()
def guardar_suministro():
    connection = None
    cursor = None
    try:
        # Obtener datos del formulario
        suministros_codigo = request.form.get('suministros_codigo')
        suministros_descripcion = request.form.get('suministros_descripcion')
        suministros_unidad_medida = request.form.get('suministros_unidad_medida')
        suministros_familia = request.form.get('suministros_familia')
        suministros_cliente = request.form.get('suministros_cliente')
        suministros_tipo = request.form.get('suministros_tipo')
        suministros_estado = request.form.get('suministros_estado', 'Activo')
        suministros_requiere_serial = request.form.get('suministros_requiere_serial', 'no')
        suministros_serial = request.form.get('suministros_serial', '')
        suministros_costo_unitario = request.form.get('suministros_costo_unitario')
        suministros_cantidad = request.form.get('suministros_cantidad')
        fecha_registro = request.form.get('fecha_registro')
        id_codigo_consumidor = request.form.get('id_codigo_consumidor', session.get('user_id'))
        
        # Validar datos requeridos
        if not all([suministros_codigo, suministros_descripcion, suministros_unidad_medida, 
                    suministros_familia, suministros_costo_unitario, suministros_cantidad]):
            flash('Por favor complete todos los campos requeridos.', 'error')
            return redirect(url_for('suministros'))
        
        # Conectar a la base de datos
        connection = get_db_connection()
        if connection is None:
            flash('Error de conexión a la base de datos.', 'error')
            return redirect(url_for('suministros'))
        
        cursor = connection.cursor()
        
        # Insertar nuevo suministro
        query = """
        INSERT INTO suministros (
            suministros_codigo, suministros_descripcion, suministros_unidad_medida,
            suministros_familia, suministros_cliente, suministros_tipo, suministros_estado,
            suministros_requiere_serial, suministros_serial, suministros_costo_unitario,
            suministros_cantidad, fecha_registro, id_codigo_consumidor
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        values = (
            suministros_codigo, suministros_descripcion, suministros_unidad_medida,
            suministros_familia, suministros_cliente, suministros_tipo, suministros_estado,
            suministros_requiere_serial, suministros_serial, suministros_costo_unitario,
            suministros_cantidad, fecha_registro, id_codigo_consumidor
        )
        
        cursor.execute(query, values)
        connection.commit()
        
        flash('Suministro guardado exitosamente.', 'success')
        return redirect(url_for('suministros'))
        
    except mysql.connector.Error as e:
        flash(f'Error al guardar suministro: {str(e)}', 'error')
        return redirect(url_for('suministros'))
    
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/obtener_suministro/<int:id>', methods=['GET'])
@login_required()
def obtener_suministro(id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener suministro por ID
        cursor.execute("SELECT * FROM suministros WHERE id_suministros = %s", (id,))
        suministro = cursor.fetchone()
        
        if not suministro:
            return jsonify({'error': 'Suministro no encontrado'}), 404
        
        return jsonify(suministro)
        
    except mysql.connector.Error as e:
        return jsonify({'error': str(e)}), 500
    
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/actualizar_suministro', methods=['POST'])
@login_required()
def actualizar_suministro():
    connection = None
    cursor = None
    try:
        # Obtener datos del formulario
        id_suministros = request.form.get('id_suministros')
        suministros_codigo = request.form.get('suministros_codigo')
        suministros_descripcion = request.form.get('suministros_descripcion')
        suministros_unidad_medida = request.form.get('suministros_unidad_medida')
        suministros_familia = request.form.get('suministros_familia')
        suministros_cliente = request.form.get('suministros_cliente')
        suministros_tipo = request.form.get('suministros_tipo')
        suministros_estado = request.form.get('suministros_estado', 'Activo')
        suministros_requiere_serial = request.form.get('suministros_requiere_serial', 'no')
        suministros_serial = request.form.get('suministros_serial', '')
        suministros_costo_unitario = request.form.get('suministros_costo_unitario')
        suministros_cantidad = request.form.get('suministros_cantidad')
        
        # Validar datos requeridos
        if not all([id_suministros, suministros_codigo, suministros_descripcion, suministros_unidad_medida, 
                    suministros_familia, suministros_costo_unitario, suministros_cantidad]):
            flash('Por favor complete todos los campos requeridos.', 'error')
            return redirect(url_for('suministros'))
        
        # Conectar a la base de datos
        connection = get_db_connection()
        if connection is None:
            flash('Error de conexión a la base de datos.', 'error')
            return redirect(url_for('suministros'))
        
        cursor = connection.cursor()
        
        # Actualizar suministro
        query = """
        UPDATE suministros SET 
            suministros_codigo = %s, 
            suministros_descripcion = %s, 
            suministros_unidad_medida = %s,
            suministros_familia = %s, 
            suministros_cliente = %s, 
            suministros_tipo = %s, 
            suministros_estado = %s,
            suministros_requiere_serial = %s, 
            suministros_serial = %s, 
            suministros_costo_unitario = %s,
            suministros_cantidad = %s
        WHERE id_suministros = %s
        """
        values = (
            suministros_codigo, suministros_descripcion, suministros_unidad_medida,
            suministros_familia, suministros_cliente, suministros_tipo, suministros_estado,
            suministros_requiere_serial, suministros_serial, suministros_costo_unitario,
            suministros_cantidad, id_suministros
        )
        
        cursor.execute(query, values)
        connection.commit()
        
        flash('Suministro actualizado exitosamente.', 'success')
        return redirect(url_for('suministros'))
        
    except mysql.connector.Error as e:
        flash(f'Error al actualizar suministro: {str(e)}', 'error')
        return redirect(url_for('suministros'))
    
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/eliminar_suministro/<int:id>', methods=['POST'])
@login_required()
def eliminar_suministro(id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500
        
        cursor = connection.cursor()
        
        # Eliminar suministro
        cursor.execute("DELETE FROM suministros WHERE id_suministros = %s", (id,))
        connection.commit()
        
        return jsonify({'success': True, 'message': 'Suministro eliminado correctamente'})
        
    except mysql.connector.Error as e:
        return jsonify({'success': False, 'message': f'Error al eliminar suministro: {str(e)}'}), 500
    
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/admin/usuarios', methods=['GET'])
@login_required(role='administrativo')
def usuarios():
    connection = None
    cursor = None
    try:
        page = request.args.get('page', 1, type=int)
        items_per_page = 10  # Número de usuarios por página (solo para referencia en la plantilla)
        
        connection = get_db_connection()
        if connection is None:
            flash('Error de conexión a la base de datos.', 'error')
            return redirect(url_for('dashboard'))
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener total de usuarios para calcular el número de páginas
        cursor.execute("SELECT COUNT(*) as total FROM recurso_operativo")
        total_users = cursor.fetchone()['total']
        total_pages = (total_users + items_per_page - 1) // items_per_page  # Redondeo hacia arriba
        
        # Obtener TODOS los usuarios sin paginación para permitir filtrado completo
        query = """
            SELECT 
                id_codigo_consumidor,
                recurso_operativo_cedula,
                nombre,
                id_roles,
                estado,
                cargo
            FROM recurso_operativo
            ORDER BY id_codigo_consumidor
        """
        cursor.execute(query)
        usuarios = cursor.fetchall()
        
        # Convertir id_roles a nombres legibles
        for usuario in usuarios:
            usuario['role'] = ROLES.get(str(usuario['id_roles']), 'Desconocido')
        
        return render_template('modulos/administrativo/usuarios.html', 
                              usuarios=usuarios, 
                              current_page=page, 
                              total_pages=total_pages,
                              items_per_page=items_per_page,
                              ROLES=ROLES)  # Pasar el diccionario ROLES al contexto
        
    except mysql.connector.Error as e:
        flash(f'Error de base de datos: {str(e)}', 'error')
        return redirect(url_for('dashboard'))
    
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/obtener_usuario/<int:id>', methods=['GET'])
@login_required(role='administrativo')
def obtener_usuario(id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener usuario por ID incluyendo los nuevos campos
        cursor.execute("""
            SELECT 
                id_codigo_consumidor,
                recurso_operativo_cedula,
                nombre,
                id_roles,
                estado,
                cargo,
                carpeta,
                cliente,
                ciudad,
                super,
                analista,
                fecha_ingreso,
                fecha_retiro
            FROM recurso_operativo 
            WHERE id_codigo_consumidor = %s
        """, (id,))
        
        usuario = cursor.fetchone()
        
        if not usuario:
            return jsonify({'error': 'Usuario no encontrado'}), 404
        
        # Formatear fechas para input type="date" (YYYY-MM-DD)
        if usuario['fecha_ingreso']:
            if isinstance(usuario['fecha_ingreso'], datetime):
                usuario['fecha_ingreso'] = usuario['fecha_ingreso'].strftime('%Y-%m-%d')
        
        if usuario['fecha_retiro']:
            if isinstance(usuario['fecha_retiro'], datetime):
                usuario['fecha_retiro'] = usuario['fecha_retiro'].strftime('%Y-%m-%d')
        
        return jsonify(usuario)
        
    except mysql.connector.Error as e:
        return jsonify({'error': str(e)}), 500
    
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/obtener_opciones_usuario', methods=['GET'])
@login_required(role='administrativo')
def obtener_opciones_usuario():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener valores únicos para carpeta
        cursor.execute("SELECT DISTINCT carpeta FROM recurso_operativo WHERE carpeta IS NOT NULL AND carpeta != '' ORDER BY carpeta")
        carpetas = [row['carpeta'] for row in cursor.fetchall()]
        
        # Obtener valores únicos para cliente
        cursor.execute("SELECT DISTINCT cliente FROM recurso_operativo WHERE cliente IS NOT NULL AND cliente != '' ORDER BY cliente")
        clientes = [row['cliente'] for row in cursor.fetchall()]
        
        # Obtener valores únicos para ciudad
        cursor.execute("SELECT DISTINCT ciudad FROM recurso_operativo WHERE ciudad IS NOT NULL AND ciudad != '' ORDER BY ciudad")
        ciudades = [row['ciudad'] for row in cursor.fetchall()]
        
        # Obtener valores únicos para super
        cursor.execute("SELECT DISTINCT super FROM recurso_operativo WHERE super IS NOT NULL AND super != '' ORDER BY super")
        supers = [row['super'] for row in cursor.fetchall()]
        
        return jsonify({
            'carpetas': carpetas,
            'clientes': clientes,
            'ciudades': ciudades,
            'supers': supers
        })
        
    except mysql.connector.Error as e:
        return jsonify({'error': str(e)}), 500
    
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/actualizar_usuario', methods=['POST'])
@login_required(role='administrativo')
def actualizar_usuario():
    connection = None
    cursor = None
    try:
        # Obtener datos del formulario
        id_codigo_consumidor = request.form.get('id_codigo_consumidor')
        recurso_operativo_cedula = request.form.get('recurso_operativo_cedula')
        nombre = request.form.get('nombre')
        id_roles = request.form.get('id_roles')
        estado = request.form.get('estado', 'Activo')
        cargo = request.form.get('cargo', '')
        carpeta = request.form.get('carpeta', '')
        cliente = request.form.get('cliente', '')
        ciudad = request.form.get('ciudad', '')
        super_valor = request.form.get('super', '')
        password = request.form.get('password', '')
        analista = request.form.get('analista', '')
        fecha_ingreso = request.form.get('fecha_ingreso', '')
        fecha_retiro = request.form.get('fecha_retiro', '')
        
        # Validar datos requeridos
        if not all([id_codigo_consumidor, recurso_operativo_cedula, nombre, id_roles]):
            return jsonify({'success': False, 'message': 'Por favor complete todos los campos requeridos.'})
        
        # Conectar a la base de datos
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos.'})
        
        cursor = connection.cursor()
        
        # Preparar la consulta base
        query_fields = [
            'recurso_operativo_cedula = %s',
            'nombre = %s',
            'id_roles = %s',
            'estado = %s',
            'cargo = %s',
            'carpeta = %s',
            'cliente = %s',
            'ciudad = %s',
            'super = %s',
            'analista = %s',
            'fecha_ingreso = %s',
            'fecha_retiro = %s'
        ]
        
        values = [
            recurso_operativo_cedula, 
            nombre, 
            id_roles,
            estado,
            cargo,
            carpeta,
            cliente,
            ciudad,
            super_valor,
            analista,
            fecha_ingreso if fecha_ingreso else None,
            fecha_retiro if fecha_retiro else None
        ]
        
        # Si se proporciona una nueva contraseña, agregarla a la actualización
        if password and password.strip():
            # Encriptar la nueva contraseña con bcrypt
            hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
            query_fields.append('password = %s')
            values.append(hashed_password)
        
        # Agregar el ID al final para la cláusula WHERE
        values.append(id_codigo_consumidor)
        
        # Construir la consulta final
        query = f"""
        UPDATE recurso_operativo SET 
            {', '.join(query_fields)}
        WHERE id_codigo_consumidor = %s
        """
        
        cursor.execute(query, values)
        connection.commit()
        
        return jsonify({'success': True, 'message': 'Usuario actualizado exitosamente.'})
        
    except mysql.connector.Error as e:
        return jsonify({'success': False, 'message': f'Error al actualizar usuario: {str(e)}'})
    
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/eliminar_usuario/<int:id>', methods=['POST'])
@login_required(role='administrativo')
def eliminar_usuario(id):
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'})
            
        cursor = connection.cursor()
        
        # Verificar si el usuario existe
        cursor.execute("SELECT id_codigo_consumidor FROM recurso_operativo WHERE id_codigo_consumidor = %s", (id,))
        if not cursor.fetchone():
            return jsonify({'success': False, 'message': 'Usuario no encontrado'})
        
        # Eliminar el usuario
        cursor.execute("DELETE FROM recurso_operativo WHERE id_codigo_consumidor = %s", (id,))
        connection.commit()
        
        return jsonify({'success': True, 'message': 'Usuario eliminado correctamente'})
        
    except mysql.connector.Error as e:
        if connection:
            connection.rollback()
        return jsonify({'success': False, 'message': f'Error al eliminar usuario: {str(e)}'})
    
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/estadisticas_inventario')
@login_required(role='administrativo')
def estadisticas_inventario():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            flash('Error de conexión a la base de datos', 'danger')
            return redirect(url_for('dashboard'))
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener usuarios para el filtro
        cursor.execute("""
            SELECT id_codigo_consumidor, nombre
            FROM recurso_operativo
            WHERE estado = 'Activo'
            ORDER BY nombre
        """)
        usuarios = cursor.fetchall()
        
        # Obtener datos iniciales para los contadores usando las tablas existentes
        # Usaremos asignacion para herramientas y ferretero para materiales
        cursor.execute("""
            SELECT 
                COUNT(*) as total_elementos,
                SUM(CASE WHEN asignacion_estado = '1' THEN 1 ELSE 0 END) as elementos_asignados,
                SUM(CASE WHEN asignacion_estado = '0' THEN 1 ELSE 0 END) as elementos_disponibles,
                0 as elementos_mantenimiento
            FROM asignacion
        """)
        contadores_asignacion = cursor.fetchone()
        
        cursor.execute("""
            SELECT 
                COUNT(*) as total_elementos
            FROM ferretero
        """)
        contadores_ferretero = cursor.fetchone()
        
        # Combinar contadores
        contadores = {
            'total_elementos': (contadores_asignacion['total_elementos'] or 0) + (contadores_ferretero['total_elementos'] or 0),
            'elementos_asignados': contadores_asignacion['elementos_asignados'] or 0,
            'elementos_disponibles': contadores_asignacion['elementos_disponibles'] or 0,
            'elementos_mantenimiento': 0  # No hay datos de mantenimiento en las tablas actuales
        }
        
        # Obtener top 10 de herramientas (usando asignacion para herramientas básicas)
        cursor.execute("""
            SELECT 
                'Adaptador Mandril' as descripcion,
                'HD-001' as codigo,
                COUNT(*) as cantidad_total,
                SUM(CASE WHEN asignacion_estado = '1' THEN 1 ELSE 0 END) as asignados,
                SUM(CASE WHEN asignacion_estado = '0' THEN 1 ELSE 0 END) as disponibles,
                ROUND((SUM(asignacion_adaptador_mandril) / COUNT(*)) * 100) as frecuencia_uso
            FROM asignacion 
            WHERE asignacion_adaptador_mandril > 0
            UNION ALL
            SELECT 
                'Alicate' as descripcion,
                'HD-002' as codigo,
                COUNT(*) as cantidad_total,
                SUM(CASE WHEN asignacion_estado = '1' THEN 1 ELSE 0 END) as asignados,
                SUM(CASE WHEN asignacion_estado = '0' THEN 1 ELSE 0 END) as disponibles,
                ROUND((SUM(asignacion_alicate) / COUNT(*)) * 100) as frecuencia_uso
            FROM asignacion 
            WHERE asignacion_alicate > 0
            UNION ALL
            SELECT 
                'Barra 45cm' as descripcion,
                'HD-003' as codigo,
                COUNT(*) as cantidad_total,
                SUM(CASE WHEN asignacion_estado = '1' THEN 1 ELSE 0 END) as asignados,
                SUM(CASE WHEN asignacion_estado = '0' THEN 1 ELSE 0 END) as disponibles,
                ROUND((SUM(asignacion_barra_45cm) / COUNT(*)) * 100) as frecuencia_uso
            FROM asignacion 
            WHERE asignacion_barra_45cm > 0
            ORDER BY frecuencia_uso DESC
            LIMIT 10
        """)
        top_herramientas = cursor.fetchall()
        
        # Obtener top 10 de herramientas especializadas
        cursor.execute("""
            SELECT 
                'Multímetro' as descripcion,
                'HE-001' as codigo,
                COUNT(*) as cantidad_total,
                SUM(CASE WHEN asignacion_estado = '1' THEN 1 ELSE 0 END) as asignados,
                SUM(CASE WHEN asignacion_estado = '0' THEN 1 ELSE 0 END) as disponibles,
                ROUND((SUM(asignacion_multimetro) / COUNT(*)) * 100) as frecuencia_uso
            FROM asignacion 
            WHERE asignacion_multimetro > 0
            UNION ALL
            SELECT 
                'Taladro Percutor' as descripcion,
                'HE-002' as codigo,
                COUNT(*) as cantidad_total,
                SUM(CASE WHEN asignacion_estado = '1' THEN 1 ELSE 0 END) as asignados,
                SUM(CASE WHEN asignacion_estado = '0' THEN 1 ELSE 0 END) as disponibles,
                ROUND((SUM(asignacion_taladro_percutor) / COUNT(*)) * 100) as frecuencia_uso
            FROM asignacion 
            WHERE asignacion_taladro_percutor > 0
            ORDER BY frecuencia_uso DESC
            LIMIT 10
        """)
        top_dotaciones = cursor.fetchall()
        
        # Obtener top 10 de EPPs (Elementos de Protección Personal)
        cursor.execute("""
            SELECT 
                'Arnés' as descripcion,
                'EPP-001' as codigo,
                COUNT(*) as cantidad_total,
                SUM(CASE WHEN asignacion_estado = '1' THEN 1 ELSE 0 END) as asignados,
                SUM(CASE WHEN asignacion_estado = '0' THEN 1 ELSE 0 END) as disponibles,
                ROUND((SUM(asignacion_arnes) / COUNT(*)) * 100) as frecuencia_uso
            FROM asignacion 
            WHERE asignacion_arnes > 0
            UNION ALL
            SELECT 
                'Eslinga' as descripcion,
                'EPP-002' as codigo,
                COUNT(*) as cantidad_total,
                SUM(CASE WHEN asignacion_estado = '1' THEN 1 ELSE 0 END) as asignados,
                SUM(CASE WHEN asignacion_estado = '0' THEN 1 ELSE 0 END) as disponibles,
                ROUND((SUM(asignacion_eslinga) / COUNT(*)) * 100) as frecuencia_uso
            FROM asignacion 
            WHERE asignacion_eslinga > 0
            ORDER BY frecuencia_uso DESC
            LIMIT 10
        """)
        top_epps = cursor.fetchall()
        
        # Obtener top 10 de ferretero
        cursor.execute("""
            SELECT 
                'Silicona' as descripcion,
                'F-001' as codigo,
                COUNT(*) as cantidad_total,
                SUM(silicona) as asignados,
                0 as disponibles,
                ROUND((SUM(silicona) / COUNT(*)) * 100) as frecuencia_uso
            FROM ferretero 
            WHERE silicona > 0
            UNION ALL
            SELECT 
                'Amarres Negros' as descripcion,
                'F-002' as codigo,
                COUNT(*) as cantidad_total,
                SUM(amarres_negros) as asignados,
                0 as disponibles,
                ROUND((SUM(amarres_negros) / COUNT(*)) * 100) as frecuencia_uso
            FROM ferretero 
            WHERE amarres_negros > 0
            UNION ALL
            SELECT 
                'Amarres Blancos' as descripcion,
                'F-003' as codigo,
                COUNT(*) as cantidad_total,
                SUM(amarres_blancos) as asignados,
                0 as disponibles,
                ROUND((SUM(amarres_blancos) / COUNT(*)) * 100) as frecuencia_uso
            FROM ferretero 
            WHERE amarres_blancos > 0
            UNION ALL
            SELECT 
                'Cinta Aislante' as descripcion,
                'F-004' as codigo,
                COUNT(*) as cantidad_total,
                SUM(cinta_aislante) as asignados,
                0 as disponibles,
                ROUND((SUM(cinta_aislante) / COUNT(*)) * 100) as frecuencia_uso
            FROM ferretero 
            WHERE cinta_aislante > 0
            ORDER BY frecuencia_uso DESC
            LIMIT 10
        """)
        top_ferretero = cursor.fetchall()
        
        # Obtener datos para el gráfico de distribución por categorías
        cursor.execute("""
            SELECT 
                SUM(CASE 
                    WHEN asignacion_adaptador_mandril > 0 OR 
                         asignacion_alicate > 0 OR 
                         asignacion_barra_45cm > 0 OR 
                         asignacion_bisturi_metalico > 0 OR 
                         asignacion_caja_de_herramientas > 0 OR 
                         asignacion_cortafrio > 0 OR 
                         asignacion_destor_de_estrella > 0 OR 
                         asignacion_destor_de_pala > 0 OR 
                         asignacion_destor_tester > 0 OR 
                         asignacion_martillo_de_una > 0 OR 
                         asignacion_pinza_de_punta > 0 
                    THEN 1 ELSE 0 END) as herramientas,
                SUM(CASE 
                    WHEN asignacion_multimetro > 0 OR 
                         asignacion_taladro_percutor > 0 OR 
                         asignacion_ponchadora_rg_6_y_rg_11 > 0 OR 
                         asignacion_ponchadora_rj_45_y_rj11 > 0 OR 
                         asignacion_power_miter > 0 OR 
                         asignacion_bfl_laser > 0 OR 
                         asignacion_cortadora > 0 OR 
                         asignacion_stripper_fibra > 0 
                    THEN 1 ELSE 0 END) as dotaciones,
                SUM(CASE 
                    WHEN asignacion_arnes > 0 OR 
                         asignacion_eslinga > 0 OR 
                         asignacion_casco_tipo_ii > 0 OR 
                         asignacion_arana_casco > 0 OR 
                         asignacion_barbuquejo > 0 OR 
                         asignacion_guantes_vaqueta > 0 OR 
                         asignacion_gafas > 0 OR 
                         asignacion_linea_vida > 0 
                    THEN 1 ELSE 0 END) as epps
            FROM asignacion
        """)
        categorias_asignacion = cursor.fetchone()
        
        cursor.execute("""
            SELECT COUNT(*) as ferretero
            FROM ferretero
        """)
        categorias_ferretero = cursor.fetchone()
        
        # Formatear datos para gráficos
        categorias = {
            'herramientas': categorias_asignacion['herramientas'] or 0,
            'dotaciones': categorias_asignacion['dotaciones'] or 0,
            'epps': categorias_asignacion['epps'] or 0,
            'ferretero': categorias_ferretero['ferretero'] or 0
        }
        
        # Convertir valores a float para evitar problemas con Decimal
        datos_categorias = json.dumps([
            float(categorias['herramientas'] or 0), 
            float(categorias['dotaciones'] or 0), 
            float(categorias['epps'] or 0), 
            float(categorias['ferretero'] or 0)
        ])
        
        # Obtener datos para el gráfico de distribución por estado
        cursor.execute("""
            SELECT 
                SUM(CASE WHEN asignacion_estado = '1' THEN 1 ELSE 0 END) as asignados,
                SUM(CASE WHEN asignacion_estado = '0' THEN 1 ELSE 0 END) as disponibles,
                0 as en_mantenimiento,
                0 as de_baja
            FROM asignacion
        """)
        estados = cursor.fetchone()
        
        # Convertir valores a float para evitar problemas con Decimal
        datos_estados = json.dumps([
            float(estados['disponibles'] or 0), 
            float(estados['asignados'] or 0), 
            float(estados['en_mantenimiento'] or 0), 
            float(estados['de_baja'] or 0)
        ])
        
        # Obtener datos para el gráfico de elementos más asignados
        cursor.execute("""
            SELECT 
                'Adaptador Mandril' as descripcion,
                SUM(asignacion_adaptador_mandril) as total_asignaciones
            FROM asignacion 
            WHERE asignacion_adaptador_mandril > 0
            UNION ALL
            SELECT 
                'Alicate' as descripcion,
                SUM(asignacion_alicate) as total_asignaciones
            FROM asignacion 
            WHERE asignacion_alicate > 0
            UNION ALL
            SELECT 
                'Barra 45cm' as descripcion,
                SUM(asignacion_barra_45cm) as total_asignaciones
            FROM asignacion 
            WHERE asignacion_barra_45cm > 0
            UNION ALL
            SELECT 
                'Silicona' as descripcion,
                SUM(silicona) as total_asignaciones
            FROM ferretero 
            WHERE silicona > 0
            UNION ALL
            SELECT 
                'Amarres Negros' as descripcion,
                SUM(amarres_negros) as total_asignaciones
            FROM ferretero 
            WHERE amarres_negros > 0
            UNION ALL
            SELECT 
                'Cinta Aislante' as descripcion,
                SUM(cinta_aislante) as total_asignaciones
            FROM ferretero 
            WHERE cinta_aislante > 0
            ORDER BY total_asignaciones DESC
            LIMIT 6
        """)
        mas_asignados = cursor.fetchall()
        
        
        if mas_asignados:
            labels_mas_asignados = json.dumps([item['descripcion'] for item in mas_asignados])
            datos_mas_asignados = json.dumps([float(item['total_asignaciones'] or 0) for item in mas_asignados])
        else:
            labels_mas_asignados = json.dumps(["Taladro", "Guantes", "Martillo", "Pulidora", "Destornillador", "Llaves"])
            datos_mas_asignados = json.dumps([120, 98, 85, 74, 65, 60])
        
        # Obtener datos para el gráfico de tendencia de asignaciones
        cursor.execute("""
            SELECT 
                DATE_FORMAT(asignacion_fecha, '%b') as mes,
                MIN(asignacion_fecha) as fecha_referencia,
                COUNT(*) as total
            FROM asignacion
            WHERE asignacion_fecha >= DATE_SUB(NOW(), INTERVAL 6 MONTH)
            GROUP BY DATE_FORMAT(asignacion_fecha, '%Y-%m')
            ORDER BY MIN(asignacion_fecha)
        """)
        tendencia_asignacion = cursor.fetchall()
        
        cursor.execute("""
            SELECT 
                DATE_FORMAT(fecha_asignacion, '%b') as mes,
                MIN(fecha_asignacion) as fecha_referencia,
                COUNT(*) as total
            FROM ferretero
            WHERE fecha_asignacion >= DATE_SUB(NOW(), INTERVAL 6 MONTH)
            GROUP BY DATE_FORMAT(fecha_asignacion, '%Y-%m')
            ORDER BY MIN(fecha_asignacion)
        """)
        tendencia_ferretero = cursor.fetchall()
        
        # Combinar tendencias
        tendencia_meses = {}
        
        # Agregar datos de asignación
        for item in tendencia_asignacion:
            mes = item['mes']
            if mes not in tendencia_meses:
                tendencia_meses[mes] = 0
            tendencia_meses[mes] += item['total']
        
        # Agregar datos de ferretero
        for item in tendencia_ferretero:
            mes = item['mes']
            if mes not in tendencia_meses:
                tendencia_meses[mes] = 0
            tendencia_meses[mes] += item['total']
        
        # Convertir a lista ordenada por mes
        tendencia = [{'mes': k, 'total': v} for k, v in tendencia_meses.items()]
        tendencia.sort(key=lambda x: {'Ene': 1, 'Feb': 2, 'Mar': 3, 'Abr': 4, 'May': 5, 'Jun': 6, 
                                  'Jul': 7, 'Ago': 8, 'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dic': 12}.get(x['mes'], 0))
        
        if tendencia:
            labels_tendencia = json.dumps([item['mes'] for item in tendencia])
            datos_tendencia = json.dumps([float(item['total'] or 0) for item in tendencia])
        else:
            labels_tendencia = json.dumps(["Ene", "Feb", "Mar", "Abr", "May", "Jun"])
            datos_tendencia = json.dumps([65, 72, 78, 90, 85, 95])
        
        return render_template('modulos/administrativo/estadisticas_inventario.html',
                               usuarios=usuarios,
                               total_elementos=contadores['total_elementos'],
                               elementos_disponibles=contadores['elementos_disponibles'],
                               elementos_asignados=contadores['elementos_asignados'],
                               elementos_mantenimiento=contadores['elementos_mantenimiento'],
                               top_herramientas=top_herramientas,
                               top_dotaciones=top_dotaciones,
                               top_epps=top_epps,
                               top_ferretero=top_ferretero,
                               datos_categorias=datos_categorias,
                               datos_estados=datos_estados,
                               labels_mas_asignados=labels_mas_asignados,
                               datos_mas_asignados=datos_mas_asignados,
                               labels_tendencia=labels_tendencia,
                               datos_tendencia=datos_tendencia)
        
    except mysql.connector.Error as e:
        flash(f'Error al cargar datos: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))
    
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/estadisticas/inventario', methods=['GET'])
@login_required(role='administrativo')
def api_estadisticas_inventario():
    from decimal import Decimal
    
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'success': False,
                'message': 'Error de conexión a la base de datos'
            })
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener parámetros de filtro
        dias = request.args.get('dias', default='30', type=str)
        categoria = request.args.get('categoria', default='todos', type=str)
        estado = request.args.get('estado', default='todos', type=str)
        usuario = request.args.get('usuario', default='todos', type=str)
        
        # Construir condiciones de filtro
        condicion_fecha = "1=1"
        if dias != 'todos':
            condicion_fecha = f"asignacion_fecha >= DATE_SUB(NOW(), INTERVAL {dias} DAY)"
            
        condicion_categoria = "1=1"
        if categoria != 'todos':
            if categoria == 'herramientas':
                condicion_categoria = """(
                    asignacion_adaptador_mandril > 0 OR 
                    asignacion_alicate > 0 OR 
                    asignacion_barra_45cm > 0 OR 
                    asignacion_bisturi_metalico > 0 OR 
                    asignacion_caja_de_herramientas > 0 OR 
                    asignacion_cortafrio > 0 OR 
                    asignacion_destor_de_estrella > 0 OR 
                    asignacion_destor_de_pala > 0 OR 
                    asignacion_destor_tester > 0 OR 
                    asignacion_martillo_de_una > 0 OR 
                    asignacion_pinza_de_punta > 0
                )"""
            elif categoria == 'dotaciones':
                condicion_categoria = """(
                    asignacion_multimetro > 0 OR 
                    asignacion_taladro_percutor > 0 OR 
                    asignacion_ponchadora_rg_6_y_rg_11 > 0 OR 
                    asignacion_ponchadora_rj_45_y_rj11 > 0 OR 
                    asignacion_power_miter > 0 OR 
                    asignacion_bfl_laser > 0 OR 
                    asignacion_cortadora > 0 OR 
                    asignacion_stripper_fibra > 0
                )"""
            elif categoria == 'epps':
                condicion_categoria = """(
                    asignacion_arnes > 0 OR 
                    asignacion_eslinga > 0 OR 
                    asignacion_casco_tipo_ii > 0 OR 
                    asignacion_arana_casco > 0 OR 
                    asignacion_barbuquejo > 0 OR 
                    asignacion_guantes_vaqueta > 0 OR 
                    asignacion_gafas > 0 OR 
                    asignacion_linea_vida > 0
                )"""
            
        condicion_estado = "1=1"
        if estado != 'todos':
            condicion_estado = f"asignacion_estado = '{estado}'"
            
        condicion_usuario = "1=1"
        if usuario != 'todos':
            condicion_usuario = f"recurso_id = '{usuario}'"
        
        # Obtener datos para los contadores 
        cursor.execute(f"""
            SELECT 
                COUNT(*) as total_elementos,
                SUM(CASE WHEN asignacion_estado = '1' THEN 1 ELSE 0 END) as elementos_asignados,
                SUM(CASE WHEN asignacion_estado = '0' THEN 1 ELSE 0 END) as elementos_disponibles,
                0 as elementos_mantenimiento
            FROM asignacion
            WHERE {condicion_fecha} AND {condicion_categoria} AND {condicion_estado} AND {condicion_usuario}
        """)
        contadores_asignacion = cursor.fetchone()
        
        # Si la categoría no es ferretero, incluir contadores de ferretero
        contadores_ferretero = {'total_elementos': 0}
        if categoria == 'todos' or categoria == 'ferretero':
            condicion_fecha_ferretero = "1=1"
            if dias != 'todos':
                condicion_fecha_ferretero = f"fecha_asignacion >= DATE_SUB(NOW(), INTERVAL {dias} DAY)"
                
            condicion_usuario_ferretero = "1=1"
            if usuario != 'todos':
                condicion_usuario_ferretero = f"id_tecnico = '{usuario}'"
                
            cursor.execute(f"""
                SELECT 
                    COUNT(*) as total_elementos
                FROM ferretero
                WHERE {condicion_fecha_ferretero} AND {condicion_usuario_ferretero}
            """)
            contadores_ferretero = cursor.fetchone()
        
        # Combinar contadores
        contadores = {
            'total_elementos': (contadores_asignacion['total_elementos'] or 0) + (contadores_ferretero['total_elementos'] or 0),
            'elementos_asignados': contadores_asignacion['elementos_asignados'] or 0,
            'elementos_disponibles': contadores_asignacion['elementos_disponibles'] or 0,
            'elementos_mantenimiento': 0
        }
        
        # Obtener Top 10 según la categoría seleccionada
        resultados = {}
        
        if categoria == 'todos' or categoria == 'herramientas':
            cursor.execute(f"""
                SELECT 
                    'Adaptador Mandril' as descripcion,
                    'HD-001' as codigo,
                    COUNT(*) as cantidad_total,
                    SUM(CASE WHEN asignacion_estado = '1' THEN 1 ELSE 0 END) as asignados,
                    SUM(CASE WHEN asignacion_estado = '0' THEN 1 ELSE 0 END) as disponibles,
                    ROUND((SUM(asignacion_adaptador_mandril) / COUNT(*)) * 100) as frecuencia_uso
                FROM asignacion 
                WHERE asignacion_adaptador_mandril > 0 AND {condicion_fecha} AND {condicion_estado} AND {condicion_usuario}
                UNION ALL
                SELECT 
                    'Alicate' as descripcion,
                    'HD-002' as codigo,
                    COUNT(*) as cantidad_total,
                    SUM(CASE WHEN asignacion_estado = '1' THEN 1 ELSE 0 END) as asignados,
                    SUM(CASE WHEN asignacion_estado = '0' THEN 1 ELSE 0 END) as disponibles,
                    ROUND((SUM(asignacion_alicate) / COUNT(*)) * 100) as frecuencia_uso
                FROM asignacion 
                WHERE asignacion_alicate > 0 AND {condicion_fecha} AND {condicion_estado} AND {condicion_usuario}
                UNION ALL
                SELECT 
                    'Barra 45cm' as descripcion,
                    'HD-003' as codigo,
                    COUNT(*) as cantidad_total,
                    SUM(CASE WHEN asignacion_estado = '1' THEN 1 ELSE 0 END) as asignados,
                    SUM(CASE WHEN asignacion_estado = '0' THEN 1 ELSE 0 END) as disponibles,
                    ROUND((SUM(asignacion_barra_45cm) / COUNT(*)) * 100) as frecuencia_uso
                FROM asignacion 
                WHERE asignacion_barra_45cm > 0 AND {condicion_fecha} AND {condicion_estado} AND {condicion_usuario}
                ORDER BY frecuencia_uso DESC
                LIMIT 10
            """)
            top_herramientas = cursor.fetchall()
            # Formatear los valores para JSON
            for item in top_herramientas:
                for k, v in item.items():
                    if v is None:
                        item[k] = 0
                    elif isinstance(v, Decimal):
                        item[k] = float(v)
            resultados['top_herramientas'] = top_herramientas
        
        if categoria == 'todos' or categoria == 'dotaciones':
            cursor.execute(f"""
                SELECT 
                    'Multímetro' as descripcion,
                    'HE-001' as codigo,
                    COUNT(*) as cantidad_total,
                    SUM(CASE WHEN asignacion_estado = '1' THEN 1 ELSE 0 END) as asignados,
                    SUM(CASE WHEN asignacion_estado = '0' THEN 1 ELSE 0 END) as disponibles,
                    ROUND((SUM(asignacion_multimetro) / COUNT(*)) * 100) as frecuencia_uso
                FROM asignacion 
                WHERE asignacion_multimetro > 0 AND {condicion_fecha} AND {condicion_estado} AND {condicion_usuario}
                UNION ALL
                SELECT 
                    'Taladro Percutor' as descripcion,
                    'HE-002' as codigo,
                    COUNT(*) as cantidad_total,
                    SUM(CASE WHEN asignacion_estado = '1' THEN 1 ELSE 0 END) as asignados,
                    SUM(CASE WHEN asignacion_estado = '0' THEN 1 ELSE 0 END) as disponibles,
                    ROUND((SUM(asignacion_taladro_percutor) / COUNT(*)) * 100) as frecuencia_uso
                FROM asignacion 
                WHERE asignacion_taladro_percutor > 0 AND {condicion_fecha} AND {condicion_estado} AND {condicion_usuario}
                ORDER BY frecuencia_uso DESC
                LIMIT 10
            """)
            top_dotaciones = cursor.fetchall()
            # Formatear los valores para JSON
            for item in top_dotaciones:
                for k, v in item.items():
                    if v is None:
                        item[k] = 0
                    elif isinstance(v, Decimal):
                        item[k] = float(v)
            resultados['top_dotaciones'] = top_dotaciones
        
        if categoria == 'todos' or categoria == 'epps':
            cursor.execute(f"""
                SELECT 
                    'Arnés' as descripcion,
                    'EPP-001' as codigo,
                    COUNT(*) as cantidad_total,
                    SUM(CASE WHEN asignacion_estado = '1' THEN 1 ELSE 0 END) as asignados,
                    SUM(CASE WHEN asignacion_estado = '0' THEN 1 ELSE 0 END) as disponibles,
                    ROUND((SUM(asignacion_arnes) / COUNT(*)) * 100) as frecuencia_uso
                FROM asignacion 
                WHERE asignacion_arnes > 0 AND {condicion_fecha} AND {condicion_estado} AND {condicion_usuario}
                UNION ALL
                SELECT 
                    'Eslinga' as descripcion,
                    'EPP-002' as codigo,
                    COUNT(*) as cantidad_total,
                    SUM(CASE WHEN asignacion_estado = '1' THEN 1 ELSE 0 END) as asignados,
                    SUM(CASE WHEN asignacion_estado = '0' THEN 1 ELSE 0 END) as disponibles,
                    ROUND((SUM(asignacion_eslinga) / COUNT(*)) * 100) as frecuencia_uso
                FROM asignacion 
                WHERE asignacion_eslinga > 0 AND {condicion_fecha} AND {condicion_estado} AND {condicion_usuario}
                ORDER BY frecuencia_uso DESC
                LIMIT 10
            """)
            top_epps = cursor.fetchall()
            # Formatear los valores para JSON
            for item in top_epps:
                for k, v in item.items():
                    if v is None:
                        item[k] = 0
                    elif isinstance(v, Decimal):
                        item[k] = float(v)
            resultados['top_epps'] = top_epps
        
        if categoria == 'todos' or categoria == 'ferretero':
            cursor.execute(f"""
                SELECT 
                    'Silicona' as descripcion,
                    'F-001' as codigo,
                    COUNT(*) as cantidad_total,
                    SUM(silicona) as asignados,
                    0 as disponibles,
                    ROUND((SUM(silicona) / COUNT(*)) * 100) as frecuencia_uso
                FROM ferretero 
                WHERE silicona > 0 AND {condicion_fecha} AND {condicion_estado} AND {condicion_usuario}
                UNION ALL
                SELECT 
                    'Amarres Negros' as descripcion,
                    'F-002' as codigo,
                    COUNT(*) as cantidad_total,
                    SUM(amarres_negros) as asignados,
                    0 as disponibles,
                    ROUND((SUM(amarres_negros) / COUNT(*)) * 100) as frecuencia_uso
                FROM ferretero 
                WHERE amarres_negros > 0 AND {condicion_fecha} AND {condicion_estado} AND {condicion_usuario}
                UNION ALL
                SELECT 
                    'Amarres Blancos' as descripcion,
                    'F-003' as codigo,
                    COUNT(*) as cantidad_total,
                    SUM(amarres_blancos) as asignados,
                    0 as disponibles,
                    ROUND((SUM(amarres_blancos) / COUNT(*)) * 100) as frecuencia_uso
                FROM ferretero 
                WHERE amarres_blancos > 0 AND {condicion_fecha} AND {condicion_estado} AND {condicion_usuario}
                UNION ALL
                SELECT 
                    'Cinta Aislante' as descripcion,
                    'F-004' as codigo,
                    COUNT(*) as cantidad_total,
                    SUM(cinta_aislante) as asignados,
                    0 as disponibles,
                    ROUND((SUM(cinta_aislante) / COUNT(*)) * 100) as frecuencia_uso
                FROM ferretero 
                WHERE cinta_aislante > 0 AND {condicion_fecha} AND {condicion_estado} AND {condicion_usuario}
                ORDER BY frecuencia_uso DESC
                LIMIT 10
            """)
            top_ferretero = cursor.fetchall()
            # Formatear los valores para JSON
            for item in top_ferretero:
                for k, v in item.items():
                    if v is None:
                        item[k] = 0
                    elif isinstance(v, Decimal):
                        item[k] = float(v)
            resultados['top_ferretero'] = top_ferretero
        
        return jsonify({
            'total_elementos': contadores['total_elementos'],
            'elementos_asignados': contadores['elementos_asignados'],
            'elementos_disponibles': contadores['elementos_disponibles'],
            'elementos_mantenimiento': contadores['elementos_mantenimiento'],
            'top_herramientas': resultados.get('top_herramientas', []),
            'top_dotaciones': resultados.get('top_dotaciones', []),
            'top_epps': resultados.get('top_epps', []),
            'top_ferretero': resultados.get('top_ferretero', []),
            'herramientas': categorias['herramientas'],
            'dotaciones': categorias['dotaciones'],
            'epps': categorias['epps'],
            'ferretero': categorias['ferretero'],
            'estadisticas': estadisticas,
            'tendencia': datos_tendencia,
            'mas_asignados': {
                'labels': labels_mas_asignados,
                'datos': datos_mas_asignados
            },
            'tendencia_asignacion': tendencia_asignacion,
            'tendencia_ferretero': tendencia_ferretero
        })
        
    except mysql.connector.Error as e:
        flash(f'Error al cargar datos: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))
    
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/asistencia', methods=['GET'])
@login_required(role='administrativo')
def ver_asistencia():
    try:
        connection = get_db_connection()
        if connection is None:
            flash('Error de conexión a la base de datos', 'danger')
            return redirect(url_for('dashboard'))
            
        cursor = connection.cursor(dictionary=True)
        
        # Crear tablas si no existen
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS tipificacion_asistencia (
                id_tipificacion INT AUTO_INCREMENT PRIMARY KEY,
                codigo_tipificacion VARCHAR(50) NOT NULL,
                nombre_tipificacion VARCHAR(200),
                estado CHAR(1) DEFAULT '1',
                fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS asistencia (
                id_asistencia INT AUTO_INCREMENT PRIMARY KEY,
                cedula VARCHAR(20),
                tecnico VARCHAR(100),
                carpeta_dia VARCHAR(50),
                carpeta VARCHAR(50),
                super VARCHAR(100),
                fecha_asistencia TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                id_codigo_consumidor INT
            )
        """)
        connection.commit()
        
        # Insertar tipificaciones por defecto si no existen
        #tipificaciones_default = [
        #    ('VACACIONES', 'Vacaciones'),
        #    ('PERMISO', 'Permiso'),
        #    ('INCAPACIDAD', 'Incapacidad'),
        #    ('CAPACITACION', 'Capacitación'),
        #    ('AUSENCIA', 'Ausencia'),
        #    ('PRESENTE', 'Presente')
        #]
        
        #for codigo, descripcion in tipificaciones_default:
        #    cursor.execute("""
        #        INSERT IGNORE INTO tipificacion_asistencia 
        #        (codigo_tipificacion, nombre_tipificacion) 
        #        VALUES (%s, %s)
        #    """, (codigo, descripcion))
        #connection.commit()
        
        # Obtener lista de técnicos
        cursor.execute("""
           SELECT id_codigo_consumidor, recurso_operativo_cedula, nombre, carpeta
            FROM capired.recurso_operativo
            WHERE estado = 'Activo'
            ORDER BY nombre
        """)
        tecnicos = cursor.fetchall()
        
        # Obtener lista de tipificaciones para carpeta_dia
        cursor.execute("""
            SELECT codigo_tipificacion, nombre_tipificacion
            FROM tipificacion_asistencia
            WHERE estado = '1'
            ORDER BY codigo_tipificacion
        """)
        carpetas_dia = cursor.fetchall()
        
        # Obtener lista de supervisores
        cursor.execute("""
            SELECT DISTINCT super
            FROM recurso_operativo 
            WHERE super IS NOT NULL AND super != ''
            ORDER BY super
        """)
        supervisores = cursor.fetchall()
        
        return render_template('modulos/administrativo/asistencia.html',
                           tecnicos=tecnicos,
                           carpetas_dia=carpetas_dia,
                           supervisores=supervisores)
                           
    except mysql.connector.Error as e:
        flash(f'Error al cargar datos: {str(e)}', 'danger')
        return redirect(url_for('dashboard'))
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/asistencia/guardar', methods=['POST'])
@login_required(role='administrativo')
def guardar_asistencias():
    try:
        data = request.get_json()
        if not data or 'asistencias' not in data:
            return jsonify({'success': False, 'message': 'No se recibieron datos de asistencia'})
            
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'})
            
        cursor = connection.cursor()
        
        # Primero, crear la tabla si no existe con AUTO_INCREMENT
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS asistencia (
                id_asistencia INT AUTO_INCREMENT PRIMARY KEY,
                cedula VARCHAR(20),
                tecnico VARCHAR(100),
                carpeta_dia VARCHAR(50),
                carpeta VARCHAR(50),
                super VARCHAR(100),
                fecha_asistencia TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                id_codigo_consumidor INT
            )
        """)
        connection.commit()
        
        # Insertar cada asistencia
        for asistencia in data['asistencias']:
            cursor.execute("""
                INSERT INTO asistencia (
                    cedula, tecnico, carpeta_dia, carpeta, super, 
                    id_codigo_consumidor
                ) VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                asistencia['cedula'],
                asistencia['tecnico'],
                asistencia['carpeta_dia'],
                asistencia['carpeta'],
                asistencia['super'],
                asistencia['id_codigo_consumidor'],
                
            ))
        
        connection.commit()
        return jsonify({'success': True, 'message': 'Asistencias guardadas correctamente'})
        
    except mysql.connector.Error as e:
        return jsonify({'success': False, 'message': f'Error al guardar asistencias: {str(e)}'})
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/operativo/asistencia/guardar', methods=['POST'])
@login_required(role='operativo')
def guardar_asistencias_operativo():
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True, buffered=True)
        
        # Obtener el nombre del usuario actual (será usado como supervisor en los registros)
        nombre_usuario_actual = session.get('user_name', '')
        if not nombre_usuario_actual:
            return jsonify({'success': False, 'message': 'No se encontró información del usuario'}), 400
        
        # DEBUG: Información del usuario y supervisor
        usuario_actual = session.get('user_cedula', '')
        print(f"DEBUG - Usuario actual: {usuario_actual}")
        print(f"DEBUG - ID código consumidor: {session['id_codigo_consumidor']}")
        print(f"DEBUG - Nombre usuario actual (supervisor): {nombre_usuario_actual}")
        
        # Verificar si ya existe un registro para hoy
        fecha_hoy = datetime.now().strftime('%Y-%m-%d')
        cursor.execute("""
            SELECT COUNT(*) as registros_hoy
            FROM asistencia 
            WHERE super = %s AND DATE(fecha_asistencia) = %s
        """, (nombre_usuario_actual, fecha_hoy))
        
        registro_existente = cursor.fetchone()
        if registro_existente and registro_existente['registros_hoy'] > 0:
            return jsonify({'success': False, 'message': 'Ya existe un registro de asistencia para el día de hoy'}), 400
        
        data = request.get_json()
        asistencias = data.get('asistencias', [])
        
        # Validar que se haya seleccionado al menos un técnico
        if not asistencias:
            return jsonify({'success': False, 'message': 'Debe seleccionar al menos un técnico'}), 400
        
        # Insertar cada técnico
        print(f"DEBUG - Iniciando inserción de {len(asistencias)} técnicos")
        print(f"DEBUG - Datos recibidos: {asistencias}")
        
        registros_insertados = 0
        for i, asistencia in enumerate(asistencias):
            print(f"DEBUG - Insertando técnico {i+1}: {asistencia}")
            try:
                # VALIDACIÓN DUAL: Verificar AMBAS condiciones (carpeta Y cargo)
                # Solo si AMBAS validaciones pasan, aplicar valores de presupuesto
                cedula_tecnico = asistencia.get('cedula', '')
                valor_presupuesto = 0
                valor_eventos = 0
                presupuesto_encontrado = False
                carpeta_valida = False
                cargo_valido = False
                
                if cedula_tecnico:
                    try:
                        # 1. Obtener carpeta y cargo del técnico desde recurso_operativo
                        cursor.execute("""
                            SELECT carpeta, cargo 
                            FROM recurso_operativo 
                            WHERE recurso_operativo_cedula = %s 
                            LIMIT 1
                        """, (cedula_tecnico,))
                        tecnico_row = cursor.fetchone()
                        
                        if tecnico_row and tecnico_row['carpeta'] and tecnico_row['cargo']:
                            carpeta_tecnico = tecnico_row['carpeta']
                            cargo_tecnico = tecnico_row['cargo']
                            print(f"DEBUG - Técnico {cedula_tecnico}: carpeta='{carpeta_tecnico}', cargo='{cargo_tecnico}'")

                            # Normalización de carpetas para coincidencias de presupuesto
                            def normalizar_carpeta(nombre):
                                if not nombre:
                                    return ''
                                n = str(nombre).strip()
                                # Unificar mayúsculas/minúsculas sin perder formato de la BD
                                n_upper = n.upper()
                                # Remover sufijo BACK si existe
                                if n_upper.endswith(' BACK'):
                                    n_upper = n_upper.replace(' BACK', '')
                                # Sinónimos y correcciones comunes
                                mapa = {
                                    'BROWFIELD': 'BROWNFIELD',
                                }
                                # Carpetas válidas esperadas (canónicas)
                                validas = [
                                    'POSTVENTA', 'POSTVENTA FTTH',
                                    'FTTH INSTALACIONES', 'INSTALACIONES DOBLES',
                                    'MANTENIMIENTO FTTH', 'ARREGLOS HFC',
                                    'BROWNFIELD'
                                ]
                                # Aplicar corrección de sinónimos
                                n_upper = mapa.get(n_upper, n_upper)
                                # Si tras quitar BACK coincide con alguna válida, devolver esa forma
                                for v in validas:
                                    if n_upper == v:
                                        return v
                                # Si no coincide exactamente, intentar igualar por palabras claves
                                if 'POSTVENTA FTTH' in n_upper:
                                    return 'POSTVENTA FTTH'
                                if 'POSTVENTA' == n_upper or n_upper.startswith('POSTVENTA'):
                                    return 'POSTVENTA'
                                if 'FTTH INSTALACIONES' in n_upper:
                                    return 'FTTH INSTALACIONES'
                                if 'INSTALACIONES DOBLES' in n_upper:
                                    return 'INSTALACIONES DOBLES'
                                if 'MANTENIMIENTO FTTH' in n_upper:
                                    return 'MANTENIMIENTO FTTH'
                                if 'ARREGLOS HFC' in n_upper:
                                    return 'ARREGLOS HFC'
                                if 'BROWNFIELD' in n_upper or 'BROWFIELD' in n_upper:
                                    return 'BROWNFIELD'
                                # Por defecto devolver nombre original limpio (mayúsculas)
                                return n_upper

                            carpeta_normalizada = normalizar_carpeta(carpeta_tecnico)
                            if carpeta_normalizada != carpeta_tecnico:
                                print(f"DEBUG - Normalización de carpeta: '{carpeta_tecnico}' -> '{carpeta_normalizada}'")
                            
                            # VALIDACIÓN DUAL: Verificar carpeta Y cargo en presupuesto_carpeta
                            cursor.execute("""
                                SELECT presupuesto_diario, presupuesto_eventos 
                                FROM presupuesto_carpeta 
                                WHERE presupuesto_carpeta = %s AND presupuesto_cargo = %s
                                LIMIT 1
                            """, (carpeta_normalizada, cargo_tecnico))
                            presupuesto_row = cursor.fetchone()
                            
                            if presupuesto_row:
                                print(f"DEBUG - ✅ COMBINACIÓN VÁLIDA: carpeta='{carpeta_tecnico}' + cargo='{cargo_tecnico}' encontrada en presupuesto_carpeta")
                                
                                # Aplicar valores de presupuesto_carpeta
                                # presupuesto_diario -> valor
                                if 'presupuesto_diario' in presupuesto_row and presupuesto_row['presupuesto_diario'] is not None:
                                    try:
                                        valor_presupuesto = int(float(presupuesto_row['presupuesto_diario']))
                                        presupuesto_encontrado = True
                                    except Exception:
                                        valor_presupuesto = 0
                                
                                # presupuesto_eventos -> eventos
                                if 'presupuesto_eventos' in presupuesto_row and presupuesto_row['presupuesto_eventos'] is not None:
                                    try:
                                        valor_eventos = int(float(presupuesto_row['presupuesto_eventos']))
                                        presupuesto_encontrado = True
                                    except Exception:
                                        valor_eventos = 0
                                
                                print(f"DEBUG - Valores aplicados: valor={valor_presupuesto}, eventos={valor_eventos}")
                            else:
                                print(f"DEBUG - ❌ COMBINACIÓN INVÁLIDA: carpeta='{carpeta_tecnico}' + cargo='{cargo_tecnico}' NO encontrada en presupuesto_carpeta")
                                print(f"DEBUG - Aplicando valores en 0")
                                valor_presupuesto = 0
                                valor_eventos = 0
                                presupuesto_encontrado = False
                        else:
                            print(f"DEBUG - No se encontró carpeta o cargo para técnico con cédula '{cedula_tecnico}' en recurso_operativo")
                    except Exception as e_presupuesto:
                        print(f"DEBUG - Error obteniendo presupuesto para técnico '{cedula_tecnico}': {str(e_presupuesto)}")
                else:
                    print(f"DEBUG - No se proporcionó cédula del técnico")
                
                # Log final del resultado
                if not presupuesto_encontrado:
                    print(f"DEBUG - No se encontró presupuesto válido para técnico '{cedula_tecnico}'. Usando valores por defecto: valor=0, eventos=0")

                cursor.execute("""
                    INSERT INTO asistencia (cedula, tecnico, carpeta_dia, carpeta, super, fecha_asistencia, id_codigo_consumidor, valor, eventos)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    asistencia.get('cedula', ''),
                    asistencia.get('tecnico', ''),
                    asistencia.get('carpeta_dia', ''),
                    carpeta_normalizada,
                    asistencia.get('super', nombre_usuario_actual),
                    asistencia.get('fecha_asistencia', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                    asistencia.get('id_codigo_consumidor', 0),
                    valor_presupuesto,
                    valor_eventos
                ))
                registros_insertados += 1
                print(f"DEBUG - Técnico {i+1} insertado exitosamente")
            except Exception as e:
                print(f"DEBUG - Error insertando técnico {i+1}: {str(e)}")
        
        print(f"DEBUG - Total registros insertados: {registros_insertados}")
        connection.commit()
        print(f"DEBUG - Commit realizado exitosamente")
        return jsonify({'success': True, 'message': f'Se registraron {len(asistencias)} asistencias correctamente. El formulario se ha bloqueado para evitar registros duplicados.'})
        
    except mysql.connector.Error as e:
        return jsonify({'success': False, 'message': f'Error al guardar asistencias: {str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

# Ruta para mostrar la página de detalle de preoperacionales por técnicos
@app.route('/operativo/detalle_preoperacionales_tecnicos')
@login_required(role='operativo')
def detalle_preoperacionales_tecnicos():
    try:
        # Verificar conexión a la base de datos
        connection = get_db_connection()
        if connection is None:
            flash('Error de conexión a la base de datos', 'error')
            return redirect(url_for('dashboard_operativo'))
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener información del usuario actual
        nombre_usuario_actual = session.get('user_name', '')
        fecha_hoy = datetime.now().strftime('%Y-%m-%d')
        
        # Verificar si tiene asistencia registrada para hoy (solo para información)
        cursor.execute("""
            SELECT COUNT(*) as tiene_asistencia
            FROM asistencia 
            WHERE super = %s AND DATE(fecha_asistencia) = %s
        """, (nombre_usuario_actual, fecha_hoy))
        
        resultado = cursor.fetchone()
        tiene_asistencia = resultado['tiene_asistencia'] > 0 if resultado else False
        
        cursor.close()
        connection.close()
        
        # Comentado: Validación de asistencia eliminada para permitir acceso sin registro de asistencia
        # if not tiene_asistencia:
        #     flash('Debe registrar asistencia antes de acceder a esta funcionalidad', 'warning')
        #     return redirect(url_for('dashboard_operativo'))
        
        return render_template('modulos/operativo/detalle_preoperacionales_tecnicos.html',
                             user_name=session.get('user_name', ''),
                             user_role=session.get('user_role', ''),
                             tiene_asistencia=tiene_asistencia)
        
    except Exception as e:
        print(f"Error en detalle_preoperacionales_tecnicos: {str(e)}")
        flash('Error al cargar la página', 'error')
        return redirect(url_for('dashboard_operativo'))

# NUEVO: Reportes Indicadores (Operativo)
@app.route('/operativo/indicadores')
@login_required(role='operativo')
def indicadores_operativo():
    try:
        connection = get_db_connection()
        if connection is None:
            flash('Error de conexión a la base de datos', 'error')
            return redirect(url_for('dashboard_operativo'))
        
        cursor = connection.cursor(dictionary=True)
        nombre_usuario_actual = session.get('user_name', '')
        fecha_hoy = datetime.now().strftime('%Y-%m-%d')
        
        # Verificar asistencia por cédula/ID de usuario para evitar dependencia del nombre
        id_codigo = session.get('id_codigo_consumidor', 0)
        cedula_usuario = session.get('user_cedula', '')
        cursor.execute("""
            SELECT COUNT(*) as tiene_asistencia
            FROM asistencia 
            WHERE DATE(fecha_asistencia) = %s
              AND (id_codigo_consumidor = %s OR cedula = %s)
        """, (fecha_hoy, id_codigo, cedula_usuario))
        resultado = cursor.fetchone()
        tiene_asistencia = resultado['tiene_asistencia'] > 0 if resultado else False

        # Fallback: comparación normalizada sin acentos y sin sensibilidad a mayúsculas
        if not tiene_asistencia:
            cursor.execute("""
                SELECT super
                FROM asistencia
                WHERE DATE(fecha_asistencia) = %s
            """, (fecha_hoy,))
            filas = cursor.fetchall()
            import unicodedata
            def _norm(s):
                return ''.join(c for c in unicodedata.normalize('NFD', s or '') if unicodedata.category(c) != 'Mn').lower().strip()
            tiene_asistencia = any(_norm(f['super']) == _norm(nombre_usuario_actual) for f in filas)
        
        cursor.close()
        connection.close()
        
        return render_template('modulos/operativo/indicadores.html',
                               user_name=session.get('user_name', ''),
                               user_role=session.get('user_role', ''),
                               tiene_asistencia=tiene_asistencia)
    except Exception as e:
        print(f"Error en indicadores_operativo: {str(e)}")
        flash('Error al cargar la página de indicadores', 'error')
        return redirect(url_for('dashboard_operativo'))

# NUEVO: Submódulos de Indicadores (Próximamente)
@app.route('/operativo/indicadores/logistica')
@app.route('/operativo/indicadores/sst')
@app.route('/operativo/indicadores/calidad')
@login_required(role='operativo')
def indicadores_proximamente():
    try:
        connection = get_db_connection()
        if connection is None:
            flash('Error de conexión a la base de datos', 'error')
            return redirect(url_for('dashboard_operativo'))
        
        cursor = connection.cursor(dictionary=True)
        nombre_usuario_actual = session.get('user_name', '')
        fecha_hoy = datetime.now().strftime('%Y-%m-%d')
        
        cursor.execute("""
            SELECT COUNT(*) as tiene_asistencia
            FROM asistencia 
            WHERE super = %s AND DATE(fecha_asistencia) = %s
        """, (nombre_usuario_actual, fecha_hoy))
        resultado = cursor.fetchone()
        tiene_asistencia = resultado['tiene_asistencia'] > 0 if resultado else False
        
        # Determinar submódulo desde la ruta
        modulo = request.path.split('/')[-1]
        if modulo not in ['logistica', 'sst', 'calidad']:
            modulo = 'logistica'
        
        cursor.close()
        connection.close()
        
        return render_template('modulos/operativo/indicadores_proximamente.html',
                               modulo=modulo,
                               user_name=session.get('user_name', ''),
                               user_role=session.get('user_role', ''),
                               tiene_asistencia=tiene_asistencia)
    except Exception as e:
        print(f"Error en indicadores_proximamente: {str(e)}")
        flash('Error al cargar el submódulo de indicadores', 'error')
        return redirect(url_for('indicadores_operativo'))

# NUEVO: Indicadores de Operaciones (vista con opciones)
@app.route('/operativo/indicadores/operaciones')
@login_required(role='operativo')
def indicadores_operaciones():
    try:
        connection = get_db_connection()
        if connection is None:
            flash('Error de conexión a la base de datos', 'error')
            return redirect(url_for('dashboard_operativo'))
        cursor = connection.cursor(dictionary=True)
        nombre_usuario_actual = session.get('user_name', '')
        fecha_hoy = datetime.now().strftime('%Y-%m-%d')
        # Verificar asistencia por cédula/ID de usuario
        id_codigo = session.get('id_codigo_consumidor', 0)
        cedula_usuario = session.get('user_cedula', '')
        cursor.execute("""
            SELECT COUNT(*) as tiene_asistencia
            FROM asistencia 
            WHERE DATE(fecha_asistencia) = %s
              AND (id_codigo_consumidor = %s OR cedula = %s)
        """, (fecha_hoy, id_codigo, cedula_usuario))
        resultado = cursor.fetchone()
        tiene_asistencia = resultado['tiene_asistencia'] > 0 if resultado else False

        # Fallback: comparación normalizada sin acentos y sin sensibilidad a mayúsculas
        if not tiene_asistencia:
            cursor.execute("""
                SELECT super
                FROM asistencia
                WHERE DATE(fecha_asistencia) = %s
            """, (fecha_hoy,))
            filas = cursor.fetchall()
            import unicodedata
            def _norm(s):
                return ''.join(c for c in unicodedata.normalize('NFD', s or '') if unicodedata.category(c) != 'Mn').lower().strip()
            tiene_asistencia = any(_norm(f['super']) == _norm(nombre_usuario_actual) for f in filas)

        cursor.close()
        connection.close()
        return render_template('modulos/operativo/indicadores_operaciones.html',
                               user_name=session.get('user_name', ''),
                               user_role=session.get('user_role', ''),
                               tiene_asistencia=tiene_asistencia)
    except Exception as e:
        print(f"Error en indicadores_operaciones: {str(e)}")
        flash('Error al cargar Indicadores de Operaciones', 'error')
        return redirect(url_for('indicadores_operativo'))

# NUEVO: Indicador de Inicio de Operacion (activo)
@app.route('/operativo/indicadores/operaciones/inicio-operacion')
@login_required(role='operativo')
def indicadores_operaciones_inicio():
    try:
        connection = get_db_connection()
        if connection is None:
            flash('Error de conexión a la base de datos', 'error')
            return redirect(url_for('dashboard_operativo'))
        cursor = connection.cursor(dictionary=True)
        nombre_usuario_actual = session.get('user_name', '')
        fecha_hoy = datetime.now().strftime('%Y-%m-%d')
        # Verificar asistencia por cédula/ID de usuario
        id_codigo = session.get('id_codigo_consumidor', 0)
        cedula_usuario = session.get('user_cedula', '')
        cursor.execute("""
            SELECT COUNT(*) as tiene_asistencia
            FROM asistencia 
            WHERE DATE(fecha_asistencia) = %s
              AND (id_codigo_consumidor = %s OR cedula = %s)
        """, (fecha_hoy, id_codigo, cedula_usuario))
        resultado = cursor.fetchone()
        tiene_asistencia = resultado['tiene_asistencia'] > 0 if resultado else False

        # Fallback: comparación normalizada sin acentos y sin sensibilidad a mayúsculas
        if not tiene_asistencia:
            cursor.execute("""
                SELECT super
                FROM asistencia
                WHERE DATE(fecha_asistencia) = %s
            """, (fecha_hoy,))
            filas = cursor.fetchall()
            import unicodedata
            def _norm(s):
                return ''.join(c for c in unicodedata.normalize('NFD', s or '') if unicodedata.category(c) != 'Mn').lower().strip()
            tiene_asistencia = any(_norm(f['super']) == _norm(nombre_usuario_actual) for f in filas)

        cursor.close()
        connection.close()
        return render_template('modulos/operativo/indicadores_operaciones_inicio.html',
                               user_name=session.get('user_name', ''),
                               user_role=session.get('user_role', ''),
                               tiene_asistencia=tiene_asistencia)
    except Exception as e:
        print(f"Error en indicadores_operaciones_inicio: {str(e)}")
        flash('Error al cargar Indicador de Inicio de Operación', 'error')
        return redirect(url_for('indicadores_operaciones'))

# API para obtener información del usuario actual
@app.route('/api/current_user', methods=['GET'])
@login_required_api(role='operativo')
def api_current_user():
    try:
        return jsonify({
            'success': True,
            'username': session.get('user_name', ''),
            'nombre': session.get('user_name', ''),
            'role': session.get('user_role', ''),
            'id_codigo_consumidor': session.get('id_codigo_consumidor', 0),
            'cedula': session.get('user_cedula', '')
        })
    except Exception as e:
        print(f"Error en api_current_user: {str(e)}")
        return jsonify({'success': False, 'message': 'Error al obtener información del usuario'}), 500

# NUEVO: Submódulo Equipos Disponibles
@app.route('/operativo/equipos_disponibles')
@login_required(role='operativo')
def equipos_disponibles():
    try:
        connection = get_db_connection()
        if connection is None:
            flash('Error de conexión a la base de datos', 'error')
            return redirect(url_for('dashboard_operativo'))
        
        cursor = connection.cursor(dictionary=True)
        nombre_usuario_actual = session.get('user_name', '')
        fecha_hoy = datetime.now().strftime('%Y-%m-%d')
        
        # Verificar asistencia por cédula/ID de usuario
        id_codigo = session.get('id_codigo_consumidor', 0)
        cedula_usuario = session.get('user_cedula', '')
        cursor.execute("""
            SELECT COUNT(*) as tiene_asistencia
            FROM asistencia 
            WHERE DATE(fecha_asistencia) = %s
              AND (id_codigo_consumidor = %s OR cedula = %s)
        """, (fecha_hoy, id_codigo, cedula_usuario))
        resultado = cursor.fetchone()
        tiene_asistencia = resultado['tiene_asistencia'] > 0 if resultado else False

        # Fallback: comparación normalizada sin acentos y sin sensibilidad a mayúsculas
        if not tiene_asistencia:
            cursor.execute("""
                SELECT super
                FROM asistencia
                WHERE DATE(fecha_asistencia) = %s
            """, (fecha_hoy,))
            filas = cursor.fetchall()
            import unicodedata
            def _norm(s):
                return ''.join(c for c in unicodedata.normalize('NFD', s or '') if unicodedata.category(c) != 'Mn').lower().strip()
            tiene_asistencia = any(_norm(f['super']) == _norm(nombre_usuario_actual) for f in filas)

        cursor.close()
        connection.close()
        
        return render_template('modulos/operativo/equipos_disponibles.html',
                               user_name=session.get('user_name', ''),
                               user_role=session.get('user_role', ''),
                               tiene_asistencia=tiene_asistencia,
                               api_equipos_url='/api/equipos_disponibles')
    except Exception as e:
        print(f"Error en equipos_disponibles: {str(e)}")
        flash('Error al cargar la página de equipos disponibles', 'error')
        return redirect(url_for('dashboard_operativo'))

# API para obtener equipos disponibles filtrados por supervisor
@app.route('/api/equipos_disponibles', methods=['GET'])
@login_required_api(role='operativo')
def api_equipos_disponibles():
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500
        
        cursor = connection.cursor(dictionary=True)
        nombre_supervisor = session.get('user_name', '')
        
        # Intentar consulta con columnas completas; si faltan, hacer fallback
        try:
            cursor.execute("""
                SELECT 
                    cedula,
                    codigo_tercero,
                    elemento,
                    familia,
                    serial,
                    estado,
                    fecha_ultimo_movimiento,
                    dias_ultimo,
                    tecnico,
                    super,
                    cuenta,
                    ot,
                    observacion,
                    fecha_creacion
                FROM disponibilidad_equipos 
                WHERE super = %s 
                ORDER BY tecnico, familia, elemento
            """, (nombre_supervisor,))
        except mysql.connector.Error as err:
            # Fallback si las columnas no existen en el entorno actual
            if err.errno == 1054 or 'Unknown column' in str(err):
                cursor.execute("""
                    SELECT 
                        id_disponibilidad_equipos AS id,
                        cedula,
                        codigo_tercero,
                        elemento,
                        familia,
                        serial,
                        estado,
                        tecnico,
                        super,
                        cuenta,
                        ot,
                        observacion,
                        fecha_creacion,
                        DATEDIFF(NOW(), fecha_creacion) AS dias_ultimo
                    FROM disponibilidad_equipos 
                    WHERE super = %s 
                    ORDER BY tecnico, familia, elemento
                """, (nombre_supervisor,))
            else:
                raise
        
        equipos = cursor.fetchall()
        
        # Normalizar tipos para JSON (fechas y decimales)
        def normalize(value):
            if isinstance(value, datetime):
                return value.strftime('%Y-%m-%d %H:%M:%S')
            try:
                from decimal import Decimal as _Decimal
                if isinstance(value, _Decimal):
                    return float(value)
            except Exception:
                pass
            return value
        
        equipos_serializados = []
        for row in equipos:
            row_serializado = {}
            for k, v in row.items():
                row_serializado[k] = normalize(v)
            equipos_serializados.append(row_serializado)
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'equipos': equipos_serializados,
            'total': len(equipos_serializados)
        })
        
    except Exception as e:
        app.logger.exception("Error en api_equipos_disponibles")
        return jsonify({'success': False, 'message': str(e)}), 500

# API para actualizar cuenta, OT y observación de un equipo
@app.route('/api/actualizar_equipo', methods=['POST'])
@login_required(role='operativo')
def api_actualizar_equipo():
    try:
        data = request.get_json(force=True) or {}
        serial = (data.get('serial') or '').strip()
        cuenta_in = (data.get('cuenta') or '').strip()
        ot_in = (data.get('ot') or '').strip()
        observ_in = (data.get('observacion') or '').strip()

        if not serial:
            return jsonify({'success': False, 'message': 'Serial del equipo es requerido'}), 400

        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500

        cursor = connection.cursor(dictionary=True)
        nombre_supervisor = session.get('user_name', '')

        print(f"[actualizar_equipo] payload serial={serial} cuenta='{cuenta_in}' ot='{ot_in}' observacion='{observ_in}'")

        # Verificar pertenencia y obtener valores actuales
        cursor.execute(
            """
            SELECT cuenta, ot, observacion
            FROM disponibilidad_equipos
            WHERE serial = %s AND super = %s
            LIMIT 1
            """,
            (serial, nombre_supervisor)
        )
        current = cursor.fetchone()
        if not current:
            cursor.close()
            connection.close()
            return jsonify({'success': False, 'message': 'No tiene permisos para actualizar este equipo'}), 403

        # Determinar qué campos se pueden actualizar (cuenta/OT solo si están vacíos; observación siempre editable)
        def is_empty(v):
            return v is None or (isinstance(v, str) and v.strip() == '')

        update_fields = []
        values = []
        if cuenta_in and is_empty(current.get('cuenta')):
            update_fields.append('cuenta = %s')
            values.append(cuenta_in)
        if ot_in and is_empty(current.get('ot')):
            update_fields.append('ot = %s')
            values.append(ot_in)
        if observ_in:
            update_fields.append('observacion = %s')
            values.append(observ_in)

        if not update_fields:
            cursor.close()
            connection.close()
            return jsonify({'success': False, 'message': 'No hay campos disponibles para actualizar (ya bloqueados).'}), 409

        update_sql = f"UPDATE disponibilidad_equipos SET {', '.join(update_fields)}, fecha_ultimo_movimiento = NOW() WHERE serial = %s AND super = %s"
        values.extend([serial, nombre_supervisor])
        print(f"[actualizar_equipo] SQL: {update_sql} | values={values}")
        try:
            cursor.execute(update_sql, tuple(values))
            connection.commit()
        except mysql.connector.Error as err:
            # Fallback si columna fecha_ultimo_movimiento no existe
            if getattr(err, 'errno', None) == 1054 or 'Unknown column' in str(err):
                update_sql = f"UPDATE disponibilidad_equipos SET {', '.join(update_fields)} WHERE serial = %s AND super = %s"
                print(f"[actualizar_equipo] Fallback SQL: {update_sql} | values={values}")
                cursor.execute(update_sql, tuple(values))
                connection.commit()
            else:
                raise

        # Actualizar días desde último movimiento, si la columna existe
        try:
            cursor.execute(
                """
                UPDATE disponibilidad_equipos
                SET dias_ultimo = DATEDIFF(NOW(), fecha_ultimo_movimiento)
                WHERE serial = %s
                """,
                (serial,)
            )
            connection.commit()
        except mysql.connector.Error as err:
            # Ignorar si la columna no existe
            if not (getattr(err, 'errno', None) == 1054 or 'Unknown column' in str(err)):
                raise

        cursor.close()
        connection.close()

        return jsonify({'success': True, 'message': 'Datos actualizados exitosamente'})

    except Exception as e:
        print(f"Error en api_actualizar_equipo: {str(e)}")
        try:
            if 'connection' in locals() and connection:
                connection.rollback()
        except Exception:
            pass
        return jsonify({'success': False, 'message': 'Error al actualizar el equipo'}), 500

# API para obtener datos de preoperacionales por técnicos
@app.route('/api/operativo/preoperacionales_tecnicos', methods=['GET'])
@login_required(role='operativo')
def api_preoperacionales_tecnicos():
    try:
        # Obtener parámetro de fecha
        fecha = request.args.get('fecha')
        
        # Validar fecha
        if not fecha:
            return jsonify({'success': False, 'message': 'Debe proporcionar una fecha'}), 400
        
        try:
            fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'success': False, 'message': 'Formato de fecha inválido. Use YYYY-MM-DD'}), 400
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True, buffered=True)
        
        # Obtener información del usuario actual (supervisor)
        nombre_usuario_actual = session.get('user_name', '')
        id_codigo_consumidor = session.get('id_codigo_consumidor', 0)
        
        # Obtener lista de técnicos asignados al supervisor actual
        query_tecnicos = """
            SELECT id_codigo_consumidor, nombre, recurso_operativo_cedula as documento
            FROM recurso_operativo
            WHERE super = %s AND estado = 'Activo'
        """
        
        cursor.execute(query_tecnicos, (nombre_usuario_actual,))
        tecnicos = cursor.fetchall()
        
        if not tecnicos:
            return jsonify({
                'success': True,
                'data': [],
                'message': f'No se encontraron técnicos asignados al supervisor {nombre_usuario_actual}'
            })
        
        # Para cada técnico, verificar asistencia y preoperacional en la fecha especificada
        result_tecnicos = []
        
        for tecnico in tecnicos:
            id_tecnico = tecnico['id_codigo_consumidor']
            
            # Verificar asistencia para la fecha
            cursor.execute("""
                SELECT a.id_asistencia, a.fecha_asistencia 
                FROM asistencia a
                JOIN tipificacion_asistencia t ON a.carpeta_dia = t.codigo_tipificacion
                WHERE a.id_codigo_consumidor = %s 
                AND DATE(a.fecha_asistencia) = %s 
                AND t.valor = '1'
            """, (id_tecnico, fecha_obj))
            
            asistencia = cursor.fetchone()
            tiene_asistencia = asistencia is not None
            
            # Verificar preoperacional para la fecha
            cursor.execute("""
                SELECT id_preoperacional, fecha
                FROM preoperacional
                WHERE id_codigo_consumidor = %s AND DATE(fecha) = %s
            """, (id_tecnico, fecha_obj))
            
            preoperacional = cursor.fetchone()
            tiene_preoperacional = preoperacional is not None
            
            # Determinar estado del diligenciamiento
            estado_diligenciamiento = 'No Aplica'
            if tiene_asistencia:
                if tiene_preoperacional:
                    estado_diligenciamiento = 'Completo'
                else:
                    estado_diligenciamiento = 'Falta Preoperacional'
            
            # Agregar todos los técnicos (con o sin asistencia/preoperacional)
            result_tecnicos.append({
                'tecnico': tecnico['nombre'],
                'asistencia': 'Registrada' if tiene_asistencia else 'No Registrada',
                'preoperacional': 'Registrado' if tiene_preoperacional else 'No Registrado',
                'estado': estado_diligenciamiento
            })
        
        return jsonify({
            'success': True,
            'data': result_tecnicos,
            'total': len(result_tecnicos)
        })
        
    except Exception as e:
        logging.error(f"Error en obtener_datos_grafica_operacion_recurso: {str(e)}")
        logging.error(f"Parámetros recibidos: fecha_inicio={request.args.get('fecha_inicio')}, fecha_fin={request.args.get('fecha_fin')}, supervisor={request.args.get('supervisor')}, agrupacion={request.args.get('agrupacion')}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
    finally:
        try:
            if 'cursor' in locals() and cursor:
                cursor.close()
        except:
            pass
        try:
            if 'connection' in locals() and connection and connection.is_connected():
                connection.close()
        except:
            pass

# NUEVO: API Operativo - Asistencia por fecha para Inicio de Operación
@app.route('/api/operativo/inicio-operacion/asistencia', methods=['GET'])
@login_required(role='operativo')
def api_operativo_inicio_asistencia():
    """Lista registros de asistencia del día filtrados por el supervisor logueado.
    Campos: Cédula, Técnico, Carpeta, Super, Carpeta_Día, Tipificacion, Valor, Estado, Novedad.
    """
    try:
        # Obtener fecha; por defecto hoy (zona Bogotá)
        fecha_param = request.args.get('fecha')
        if fecha_param:
            try:
                fecha_consulta = datetime.strptime(fecha_param, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'success': False, 'message': 'Formato de fecha inválido. Use YYYY-MM-DD'}), 400
        else:
            fecha_consulta = get_bogota_datetime().date()

        # Supervisor actual desde sesión
        supervisor_actual = session.get('user_name', '')
        if not supervisor_actual:
            return jsonify({'success': False, 'message': 'No se encontró supervisor en sesión'}), 400

        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500
        cursor = connection.cursor(dictionary=True)

        # Consulta principal uniendo tipificación para obtener nombre del evento
        # CORREGIDA: Usando subconsulta para eliminar duplicados y tomar solo el registro más reciente por técnico
        consulta = """
            SELECT 
                a.cedula,
                a.tecnico,
                a.carpeta,
                a.super,
                a.carpeta_dia,
                COALESCE(t.nombre_tipificacion, a.carpeta_dia) AS carpeta_dia_nombre,
                a.eventos AS eventos,
                COALESCE(pc.presupuesto_diario, 0) AS presupuesto_diario,
                a.valor,
                a.estado,
                a.novedad
            FROM asistencia a
            LEFT JOIN tipificacion_asistencia t ON a.carpeta_dia = t.codigo_tipificacion
            LEFT JOIN presupuesto_carpeta pc ON t.nombre_tipificacion = pc.presupuesto_carpeta
            WHERE a.super = %s AND DATE(a.fecha_asistencia) = %s
            AND a.id_asistencia = (
                SELECT MAX(a2.id_asistencia)
                FROM asistencia a2
                WHERE a2.cedula = a.cedula 
                AND a2.super = %s 
                AND DATE(a2.fecha_asistencia) = %s
            )
            ORDER BY a.tecnico
        """
        cursor.execute(consulta, (supervisor_actual, fecha_consulta, supervisor_actual, fecha_consulta))
        registros = cursor.fetchall()

        # Estadísticas para tarjetas
        # 1) Total técnicos: calcular basándose en los técnicos únicos que realmente aparecen en asistencia
        # Esto es más preciso que usar recurso_operativo ya que puede haber transferencias temporales
        cursor.execute("""
            SELECT COUNT(DISTINCT cedula) AS total
            FROM asistencia
            WHERE super = %s AND DATE(fecha_asistencia) = %s
        """, (supervisor_actual, fecha_consulta))
        result = cursor.fetchone()
        total_tecnicos = result['total'] if result else 0

        # 2) Sin asistencia: contar solo las siguientes tipificaciones del día
        categorias_sin = [
            'AUSENCIA INJUSTIFICADA',
            'AUSENCIA INJUS AUXILIAR',
            'INCAPACIDAD ARL',
            'DOMINICAL O FESTIVO',
            'DOMINCAL O FESTIVO',
            'LICENCIA MATERNIDAD O PATERNIDAD',
            'LICENCIA LUTO',
            'SUSPENSION',
            'PERMISO',
            'VACACIONES',
            'DIA FAMILIA',
            'Restriccion medica',
            'RESTRICCION MEDICA',
            'RENUNCIA',
            'INCAPACIDAD MEDICA'
        ]
        placeholders = ','.join(['%s'] * len(categorias_sin))
        query_detalle = f"""
            SELECT t.nombre_tipificacion AS tipo, COUNT(DISTINCT a.cedula) AS cantidad
            FROM asistencia a
            LEFT JOIN tipificacion_asistencia t ON a.carpeta_dia = t.codigo_tipificacion
            WHERE a.super = %s AND DATE(a.fecha_asistencia) = %s
              AND t.nombre_tipificacion IN ({placeholders})
              AND a.id_asistencia = (
                  SELECT MAX(a2.id_asistencia)
                  FROM asistencia a2
                  WHERE a2.cedula = a.cedula 
                  AND a2.super = a.super 
                  AND DATE(a2.fecha_asistencia) = DATE(a.fecha_asistencia)
              )
            GROUP BY t.nombre_tipificacion
        """
        params_detalle = [supervisor_actual, fecha_consulta] + categorias_sin
        cursor.execute(query_detalle, tuple(params_detalle))
        detalle_rows = cursor.fetchall() or []
        detalle_map = {row['tipo']: row['cantidad'] for row in detalle_rows}
        detalle_list = [{'nombre': k, 'cantidad': v} for k, v in detalle_map.items()]
        sin_asistencia = sum(detalle_map.values())

        # 3) Con asistencia = total técnicos - sin asistencia (no negativo)
        con_asistencia = max(total_tecnicos - sin_asistencia, 0)

        # 4) Ok's Día y Presupuesto Día: calcular solo sobre registros únicos (sin duplicados)
        # Deduplicar registros por cédula para que coincida con lo que se muestra en la tabla
        registros_unicos = {}
        for r in registros:
            cedula = r.get('cedula')
            if cedula not in registros_unicos:
                registros_unicos[cedula] = r
        
        # Calcular totales solo sobre registros únicos
        oks_dia = 0
        presupuesto_dia = 0.0
        cumple_count = 0
        no_cumple_count = 0
        
        for r in registros_unicos.values():
            try:
                oks_dia += int(r.get('eventos') or r.get('tipificacion') or 0)
            except Exception:
                pass
            try:
                presupuesto_dia += float(r.get('valor') or 0)
            except Exception:
                pass
            
            # Contar estados de cumplimiento
            estado = (r.get('estado') or '').lower().strip()
            if estado in ['cumple', 'novedad']:
                cumple_count += 1
            elif estado in ['nocumple', 'no cumple', 'no aplica']:
                no_cumple_count += 1
        
        # Calcular presupuesto mensual basado en el primer día del mes con datos
        # Obtener el primer día del mes
        primer_dia_mes = fecha_consulta.replace(day=1)
        
        # Buscar el presupuesto del primer día del mes que tenga datos
        consulta_presupuesto_mes = """
            SELECT SUM(a.valor) as presupuesto_primer_dia
            FROM asistencia a
            WHERE a.super = %s 
            AND DATE(a.fecha_asistencia) = (
                SELECT MIN(DATE(a2.fecha_asistencia))
                FROM asistencia a2
                WHERE a2.super = %s 
                AND DATE(a2.fecha_asistencia) >= %s
                AND a2.valor > 0
            )
            AND a.valor > 0
        """
        cursor.execute(consulta_presupuesto_mes, (supervisor_actual, supervisor_actual, primer_dia_mes))
        resultado_presupuesto = cursor.fetchone()
        
        if resultado_presupuesto and resultado_presupuesto['presupuesto_primer_dia']:
            # Usar el presupuesto del primer día del mes con datos y multiplicar por 26
            presupuesto_mes = float(resultado_presupuesto['presupuesto_primer_dia']) * 26
        else:
            # Fallback: usar presupuesto del día actual multiplicado por 26
            presupuesto_mes = presupuesto_dia * 26

        # Calcular cumplimiento (porcentaje y fracción)
        total_cumplimiento = cumple_count + no_cumple_count
        cumplimiento_porcentaje = 0
        cumplimiento_fraccion = f"0/{total_cumplimiento}"
        
        if total_cumplimiento > 0:
            cumplimiento_porcentaje = round((cumple_count / total_cumplimiento) * 100, 1)
            cumplimiento_fraccion = f"{cumple_count}/{total_cumplimiento}"

        # Calcular cumplimiento mensual (acumulado del mes)
        # Obtener el primer y último día del mes
        primer_dia_mes = fecha_consulta.replace(day=1)
        if fecha_consulta.month == 12:
            ultimo_dia_mes = fecha_consulta.replace(year=fecha_consulta.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            ultimo_dia_mes = fecha_consulta.replace(month=fecha_consulta.month + 1, day=1) - timedelta(days=1)

        # Consulta para obtener todos los registros del mes
        consulta_mes = """
            SELECT 
                a.cedula,
                a.estado,
                DATE(a.fecha_asistencia) as fecha_dia
            FROM asistencia a
            WHERE a.super = %s 
            AND DATE(a.fecha_asistencia) BETWEEN %s AND %s
            AND a.id_asistencia = (
                SELECT MAX(a2.id_asistencia)
                FROM asistencia a2
                WHERE a2.cedula = a.cedula 
                AND a2.super = %s 
                AND DATE(a2.fecha_asistencia) = DATE(a.fecha_asistencia)
            )
            ORDER BY a.fecha_asistencia, a.tecnico
        """
        cursor.execute(consulta_mes, (supervisor_actual, primer_dia_mes, ultimo_dia_mes, supervisor_actual))
        registros_mes = cursor.fetchall()

        # Contar cumplimiento mensual
        cumple_mes_count = 0
        no_cumple_mes_count = 0
        
        for r in registros_mes:
            estado = (r.get('estado') or '').lower().strip()
            if estado in ['cumple', 'novedad']:
                cumple_mes_count += 1
            elif estado in ['nocumple', 'no cumple', 'no aplica']:
                no_cumple_mes_count += 1

        # Calcular cumplimiento mensual (porcentaje y fracción)
        total_cumplimiento_mes = cumple_mes_count + no_cumple_mes_count
        cumplimiento_mes_porcentaje = 0
        cumplimiento_mes_fraccion = f"0/{total_cumplimiento_mes}"
        
        if total_cumplimiento_mes > 0:
            cumplimiento_mes_porcentaje = round((cumple_mes_count / total_cumplimiento_mes) * 100, 1)
            cumplimiento_mes_fraccion = f"{cumple_mes_count}/{total_cumplimiento_mes}"

        stats = {
            'total_tecnicos': total_tecnicos,
            'sin_asistencia': sin_asistencia,
            'con_asistencia': con_asistencia,
            'detalle': detalle_map,
            'sin_asistencia_detalle': detalle_list,
            'oks_dia': oks_dia,
            'presupuesto_dia': presupuesto_dia,
            'presupuesto_mes': presupuesto_mes,
            'cumple': cumple_count,
            'no_cumple': no_cumple_count,
            'cumplimiento_porcentaje': cumplimiento_porcentaje,
            'cumplimiento_fraccion': cumplimiento_fraccion,
            'cumplimiento_total': total_cumplimiento,
            'cumple_mes': cumple_mes_count,
            'no_cumple_mes': no_cumple_mes_count,
            'cumplimiento_mes_porcentaje': cumplimiento_mes_porcentaje,
            'cumplimiento_mes_fraccion': cumplimiento_mes_fraccion,
            'cumplimiento_mes_total': total_cumplimiento_mes
        }

        # Devolver solo los registros únicos para que coincida con los cálculos
        registros_finales = list(registros_unicos.values())
        
        return jsonify({
            'success': True,
            'registros': registros_finales,
            'total': len(registros_finales),
            'fecha': fecha_consulta.strftime('%Y-%m-%d'),
            'supervisor': supervisor_actual,
            'stats': stats
        })
    except Exception as e:
        logging.error(f"Error en api_operativo_inicio_asistencia: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
    finally:
        try:
            if 'cursor' in locals() and cursor:
                cursor.close()
        except:
            pass
        try:
            if 'connection' in locals() and connection and connection.is_connected():
                connection.close()
        except:
            pass

# API para consultar registros de asistencia por supervisor y fecha (para edición)
@app.route('/api/asistencia/consultar', methods=['GET'])
@login_required(role='administrativo')
def consultar_asistencia():
    """Consultar registros de asistencia por supervisor y fecha, incluyendo técnicos sin registro"""
    try:
        supervisor = request.args.get('supervisor')
        fecha = request.args.get('fecha')
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        if not supervisor:
            return jsonify({'success': False, 'message': 'Supervisor es requerido'}), 400
            
        # Validar que se proporcione fecha única o rango de fechas
        if not fecha and (not fecha_inicio or not fecha_fin):
            return jsonify({'success': False, 'message': 'Debe proporcionar fecha única o rango de fechas (fecha_inicio y fecha_fin)'}), 400
            
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True, buffered=True)
        
        # Para consulta de fecha única, incluir todos los técnicos del supervisor
        if fecha:
            # Validar formato de fecha única
            try:
                fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'success': False, 'message': 'Formato de fecha inválido. Use YYYY-MM-DD'}), 400
            
            # Obtener solo los técnicos que tienen directamente al supervisor especificado
            cursor.execute("""
                SELECT 
                    id_codigo_consumidor,
                    recurso_operativo_cedula as cedula,
                    nombre as tecnico,
                    carpeta,
                    super
                FROM recurso_operativo 
                WHERE super = %s AND estado = 'Activo'
                ORDER BY nombre
            """, (supervisor,))
            
            tecnicos = cursor.fetchall()
            
            # Para cada técnico, buscar si tiene registro de asistencia para la fecha
            registros = []
            for tecnico in tecnicos:
                cursor.execute("""
                    SELECT 
                        id_asistencia,
                        cedula,
                        tecnico,
                        carpeta_dia,
                        carpeta,
                        super,
                        fecha_asistencia,
                        id_codigo_consumidor,
                        hora_inicio,
                        estado,
                        novedad
                    FROM asistencia 
                    WHERE id_codigo_consumidor = %s AND DATE(fecha_asistencia) = %s
                """, (tecnico['id_codigo_consumidor'], fecha_obj))
                
                registro_asistencia = cursor.fetchone()
                
                if registro_asistencia:
                    # Técnico tiene registro de asistencia
                    registros.append(registro_asistencia)
                else:
                    # Técnico no tiene registro, crear entrada vacía
                    registros.append({
                        'id_asistencia': None,
                        'cedula': tecnico['cedula'],
                        'tecnico': tecnico['tecnico'],
                        'carpeta_dia': None,
                        'carpeta': tecnico['carpeta'],
                        'super': supervisor,  # Usar el supervisor consultado, no el del técnico
                        'fecha_asistencia': fecha_obj.strftime('%Y-%m-%d'),
                        'id_codigo_consumidor': tecnico['id_codigo_consumidor'],
                        'hora_inicio': None,
                        'estado': None,
                        'novedad': None
                    })
        else:
            # Para rango de fechas, mantener comportamiento original (solo registros existentes)
            # Validar formato de rango de fechas
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'success': False, 'message': 'Formato de fechas inválido. Use YYYY-MM-DD'}), 400
                
            if fecha_inicio_obj > fecha_fin_obj:
                return jsonify({'success': False, 'message': 'La fecha de inicio no puede ser mayor que la fecha de fin'}), 400
                
            # Consultar solo registros de técnicos que tienen directamente al supervisor especificado
            cursor.execute("""
                SELECT 
                    id_asistencia,
                    cedula,
                    tecnico,
                    carpeta_dia,
                    carpeta,
                    super,
                    fecha_asistencia,
                    id_codigo_consumidor,
                    hora_inicio,
                    estado,
                    novedad
                FROM asistencia 
                WHERE super = %s AND DATE(fecha_asistencia) BETWEEN %s AND %s
                ORDER BY tecnico, fecha_asistencia
            """, (supervisor, fecha_inicio_obj, fecha_fin_obj))
            
            registros = cursor.fetchall()
        
        # Procesar los registros para formatear hora_inicio correctamente
        for registro in registros:
            if registro['hora_inicio']:
                # El campo time se devuelve como timedelta desde MySQL
                if hasattr(registro['hora_inicio'], 'total_seconds'):
                    # Es un timedelta
                    total_seconds = int(registro['hora_inicio'].total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    registro['hora_inicio'] = f"{hours:02d}:{minutes:02d}"
                else:
                    # Es un string o datetime.time
                    registro['hora_inicio'] = str(registro['hora_inicio'])[:5]  # HH:MM
        
        return jsonify({
            'success': True,
            'registros': registros,
            'total': len(registros)
        })
        
    except Exception as e:
        logging.error(f"Error en obtener_datos_grafica_carpeta_dia: {str(e)}")
        logging.error(f"Parámetros recibidos: fecha_inicio={request.args.get('fecha_inicio')}, fecha_fin={request.args.get('fecha_fin')}, supervisor={request.args.get('supervisor')}, agrupacion={request.args.get('agrupacion')}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
    finally:
        if 'cursor' in locals() and cursor:
            try:
                cursor.close()
            except:
                pass
        if 'connection' in locals() and connection and connection.is_connected():
            try:
                connection.close()
            except:
                pass

# API para exportar registros de asistencia a Excel
@app.route('/api/asistencia/exportar_excel', methods=['GET'])
@login_required(role='administrativo')
def exportar_asistencia_excel():
    """Exportar registros de asistencia a Excel con filtros de supervisor y fecha/rango"""
    try:
        supervisor = request.args.get('supervisor')
        fecha = request.args.get('fecha')
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        if not supervisor:
            return jsonify({'success': False, 'message': 'Supervisor es requerido'}), 400
            
        # Validar que se proporcione fecha única o rango de fechas
        if not fecha and (not fecha_inicio or not fecha_fin):
            return jsonify({'success': False, 'message': 'Debe proporcionar fecha única o rango de fechas'}), 400
            
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Construir consulta según el tipo de filtro de fecha
        if fecha:
            # Validar formato de fecha única
            try:
                fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
                fecha_inicio_obj = fecha_obj
                fecha_fin_obj = fecha_obj
            except ValueError:
                return jsonify({'success': False, 'message': 'Formato de fecha inválido. Use YYYY-MM-DD'}), 400
                
            # Consultar registros para fecha específica
            if supervisor == 'TODOS':
                cursor.execute("""
                    SELECT 
                        cedula,
                        tecnico,
                        carpeta_dia,
                        carpeta,
                        super,
                        DATE(fecha_asistencia) as fecha,
                        TIME(fecha_asistencia) as hora
                    FROM asistencia 
                    WHERE DATE(fecha_asistencia) = %s
                    ORDER BY super, tecnico, fecha_asistencia
                """, (fecha_obj,))
            else:
                cursor.execute("""
                    SELECT 
                        cedula,
                        tecnico,
                        carpeta_dia,
                        carpeta,
                        super,
                        DATE(fecha_asistencia) as fecha,
                        TIME(fecha_asistencia) as hora
                    FROM asistencia 
                    WHERE super = %s AND DATE(fecha_asistencia) = %s
                    ORDER BY tecnico, fecha_asistencia
                """, (supervisor, fecha_obj))
        else:
            # Validar formato de rango de fechas
            try:
                fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({'success': False, 'message': 'Formato de fechas inválido. Use YYYY-MM-DD'}), 400
                
            if fecha_inicio_obj > fecha_fin_obj:
                return jsonify({'success': False, 'message': 'La fecha de inicio no puede ser mayor que la fecha de fin'}), 400
                
            # Consultar registros para rango de fechas
            if supervisor == 'TODOS':
                cursor.execute("""
                    SELECT 
                        cedula,
                        tecnico,
                        carpeta_dia,
                        carpeta,
                        super,
                        DATE(fecha_asistencia) as fecha,
                        TIME(fecha_asistencia) as hora
                    FROM asistencia 
                    WHERE DATE(fecha_asistencia) BETWEEN %s AND %s
                    ORDER BY super, tecnico, fecha_asistencia
                """, (fecha_inicio_obj, fecha_fin_obj))
            else:
                cursor.execute("""
                    SELECT 
                        cedula,
                        tecnico,
                        carpeta_dia,
                        carpeta,
                        super,
                        DATE(fecha_asistencia) as fecha,
                        TIME(fecha_asistencia) as hora
                    FROM asistencia 
                    WHERE super = %s AND DATE(fecha_asistencia) BETWEEN %s AND %s
                    ORDER BY tecnico, fecha_asistencia
                """, (supervisor, fecha_inicio_obj, fecha_fin_obj))
        
        registros = cursor.fetchall()
        
        if not registros:
            return jsonify({'success': False, 'message': 'No se encontraron registros para exportar'}), 404
        
        # Crear DataFrame con los registros
        df_registros = pd.DataFrame(registros)
        
        # Obtener lista de técnicos únicos para la tabla de novedades
        tecnicos_unicos = df_registros['tecnico'].unique()
        
        # Obtener todas las tipificaciones disponibles
        cursor.execute("SELECT codigo_tipificacion as codigo, nombre_tipificacion as descripcion FROM tipificacion_asistencia ORDER BY codigo_tipificacion")
        tipificaciones = cursor.fetchall()
        
        # Crear tabla de técnicos con novedades del período
        tecnicos_novedades = []
        for tecnico in sorted(tecnicos_unicos):
            # Obtener registros del técnico en el período
            registros_tecnico = df_registros[df_registros['tecnico'] == tecnico]
            
            # Contar novedades por tipo
            novedades_count = registros_tecnico['carpeta_dia'].value_counts().to_dict()
            
            # Crear fila para el técnico
            fila_tecnico = {
                'Técnico': tecnico,
                'Total Registros': len(registros_tecnico)
            }
            
            # Agregar columnas para cada tipificación
            for tip in tipificaciones:
                codigo = tip['codigo']
                descripcion = tip['descripcion']
                count = novedades_count.get(codigo, 0)
                fila_tecnico[f'{codigo} - {descripcion}'] = count
            
            tecnicos_novedades.append(fila_tecnico)
        
        df_novedades = pd.DataFrame(tecnicos_novedades)
        
        # Crear archivo Excel en memoria
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Hoja 1: Registros de asistencia
            df_registros.to_excel(writer, sheet_name='Registros Asistencia', index=False)
            
            # Hoja 2: Tabla de técnicos con novedades
            df_novedades.to_excel(writer, sheet_name='Resumen Técnicos', index=False)
            
            # Obtener workbook para formatear
            workbook = writer.book
            
            # Formatear hoja de registros
            worksheet_registros = writer.sheets['Registros Asistencia']
            for column in worksheet_registros.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet_registros.column_dimensions[column_letter].width = adjusted_width
            
            # Formatear hoja de resumen
            worksheet_resumen = writer.sheets['Resumen Técnicos']
            for column in worksheet_resumen.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet_resumen.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Generar nombre del archivo
        fecha_str = fecha if fecha else f"{fecha_inicio}_{fecha_fin}"
        filename = f"asistencia_{supervisor}_{fecha_str}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al exportar: {str(e)}'}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

# API para actualizar un registro de asistencia
@app.route('/api/asistencia/actualizar', methods=['PUT'])
@login_required(role='administrativo')
def actualizar_asistencia():
    """Actualizar un registro de asistencia específico"""
    try:
        data = request.get_json()
        
        # Validar campos requeridos
        if not data.get('id_asistencia'):
            return jsonify({'success': False, 'message': 'ID de asistencia es requerido'}), 400
            
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor()
        
        # Obtener la fecha del registro de asistencia
        cursor.execute("""
            SELECT DATE(fecha_asistencia) as fecha_registro
            FROM asistencia 
            WHERE id_asistencia = %s
        """, (data.get('id_asistencia'),))
        
        resultado = cursor.fetchone()
        if not resultado:
            return jsonify({
                'success': False,
                'message': 'No se encontró el registro de asistencia'
            }), 404
            
        fecha_registro = resultado[0]
        
        # Obtener fecha actual en zona horaria de Bogotá
        bogota_tz = pytz.timezone('America/Bogota')
        fecha_actual = datetime.now(bogota_tz).date()
        
        # Calcular la diferencia en días
        diferencia_dias = (fecha_actual - fecha_registro).days
        
        # Validar que la fecha esté dentro del rango permitido (3 días hacia atrás)
        if diferencia_dias > 3:
            return jsonify({
                'success': False,
                'message': f'No se puede editar asistencia de más de 3 días atrás. La asistencia es del {fecha_registro.strftime("%d/%m/%Y")} y han pasado {diferencia_dias} días.'
            }), 400
        
        if diferencia_dias < 0:
            return jsonify({
                'success': False,
                'message': 'No se puede editar asistencia de fechas futuras.'
            }), 400
        
        # Actualizar el registro
        cursor.execute("""
            UPDATE asistencia 
            SET 
                cedula = %s,
                tecnico = %s,
                carpeta_dia = %s,
                carpeta = %s,
                super = %s,
                id_codigo_consumidor = %s
            WHERE id_asistencia = %s
        """, (
            data.get('cedula', ''),
            data.get('tecnico', ''),
            data.get('carpeta_dia', ''),
            data.get('carpeta', ''),
            data.get('super', ''),
            data.get('id_codigo_consumidor', 0),
            data.get('id_asistencia')
        ))
        
        if cursor.rowcount > 0:
            connection.commit()
            return jsonify({
                'success': True,
                'message': 'Registro actualizado correctamente'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'No se encontró el registro para actualizar'
            }), 404
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

# API para eliminar un registro de asistencia
@app.route('/api/asistencia/eliminar', methods=['DELETE'])
@login_required(role='administrativo')
def eliminar_asistencia():
    """Eliminar un registro de asistencia específico"""
    try:
        data = request.get_json()
        
        # Validar campos requeridos
        if not data.get('id_asistencia'):
            return jsonify({'success': False, 'message': 'ID de asistencia es requerido'}), 400
            
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor()
        
        # Eliminar el registro
        cursor.execute("""
            DELETE FROM asistencia 
            WHERE id_asistencia = %s
        """, (data.get('id_asistencia'),))
        
        if cursor.rowcount > 0:
            connection.commit()
            return jsonify({
                'success': True,
                'message': 'Registro eliminado correctamente'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'No se encontró el registro para eliminar'
            }), 404
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

# API para obtener valores únicos de tipificacion_asistencia
@app.route('/api/asistencia/tipificacion', methods=['GET'])
@login_required(role='administrativo')
def obtener_tipificacion_asistencia():
    """Obtener valores únicos de la tabla tipificacion_asistencia"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener valores únicos de codigo_tipificacion
        cursor.execute("""
            SELECT codigo_tipificacion, nombre_tipificacion as descripcion 
            FROM tipificacion_asistencia 
            WHERE codigo_tipificacion IS NOT NULL AND codigo_tipificacion != ''
            ORDER BY codigo_tipificacion
        """)
        
        tipificaciones = cursor.fetchall()
        
        return jsonify({
            'success': True,
            'tipificaciones': [{'codigo': t['codigo_tipificacion'], 'descripcion': t['descripcion']} for t in tipificaciones]
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

# API para obtener lista de supervisores
@app.route('/api/asistencia/supervisores', methods=['GET'])
@login_required(role='administrativo')
def obtener_supervisores_asistencia():
    """Obtener lista de supervisores únicos de la tabla recurso_operativo"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener supervisores únicos de recurso_operativo donde carpeta='SUPERVISORES'
        cursor.execute("""
            SELECT DISTINCT nombre as supervisor
            FROM recurso_operativo 
            WHERE carpeta = 'SUPERVISORES' AND nombre IS NOT NULL AND nombre != ''
            ORDER BY nombre
        """)
        
        supervisores = cursor.fetchall()
        
        return jsonify({
            'success': True,
            'supervisores': [s['supervisor'] for s in supervisores]
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

# API para obtener resumen agrupado de asistencia por tipificación
@app.route('/api/asistencia/resumen_agrupado', methods=['GET'])
@login_required(role='administrativo')
def obtener_resumen_agrupado_asistencia():
    """Obtener resumen de asistencia agrupado por grupos de tipificación y operación por recurso operativo"""
    try:
        # Obtener parámetros de filtro
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        supervisor_filtro = request.args.get('supervisor')
        
        # Si no se proporcionan fechas, usar la fecha actual
        if not fecha_inicio or not fecha_fin:
            fecha_actual = get_bogota_datetime().date()
            fecha_inicio = fecha_fin = fecha_actual
        else:
            try:
                fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                
                # Validaciones adicionales de rango de fechas
                if fecha_inicio > fecha_fin:
                    return jsonify({
                        'success': False,
                        'message': 'La fecha de inicio no puede ser mayor que la fecha de fin'
                    }), 400
                
                # Validar que las fechas no sean futuras
                fecha_actual = get_bogota_datetime().date()
                if fecha_inicio > fecha_actual or fecha_fin > fecha_actual:
                    return jsonify({
                        'success': False,
                        'message': 'No se pueden consultar fechas futuras'
                    }), 400
                
                # Validar rango máximo (1 año)
                diferencia_dias = (fecha_fin - fecha_inicio).days
                if diferencia_dias > 365:
                    return jsonify({
                        'success': False,
                        'message': 'El rango de fechas no puede ser mayor a 1 año'
                    }), 400
                    
            except ValueError:
                return jsonify({
                    'success': False,
                    'message': 'Formato de fecha inválido. Use YYYY-MM-DD'
                }), 400
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # ===== NUEVA FUNCIONALIDAD: OPERACIÓN X RECURSO OPERATIVO =====
        # Consulta para obtener datos agrupados por grupo y carpeta (excluyendo ausencias)
        # Se normaliza la carpeta para agrupar variantes como BACK y la falta de ortografía BROWFIELD
        query_operacion = """
            SELECT 
                CASE 
                    WHEN UPPER(a.carpeta) LIKE 'ARREGLOS HFC%'
                         OR UPPER(a.carpeta) LIKE 'MANTENIMIENTO FTTH%'
                         OR UPPER(a.carpeta) = 'DX' THEN 'ARREGLOS'
                    WHEN UPPER(a.carpeta) LIKE 'BROWNFIELD%'
                         OR UPPER(a.carpeta) LIKE 'BROWFIELD%'
                         OR UPPER(a.carpeta) LIKE 'FTTH INSTALACIONES%'
                         OR UPPER(a.carpeta) LIKE 'INSTALACIONES DOBLES%' THEN 'INSTALACIONES'
                    WHEN UPPER(a.carpeta) LIKE 'POSTVENTA%'
                         OR UPPER(a.carpeta) LIKE 'POSTVENTA FTTH%' THEN 'POSTVENTA'
                    ELSE 'OTROS'
                END as grupo,
                CASE 
                    WHEN UPPER(a.carpeta) LIKE 'ARREGLOS HFC%' THEN 'ARREGLOS HFC'
                    WHEN UPPER(a.carpeta) LIKE 'MANTENIMIENTO FTTH%' THEN 'MANTENIMIENTO FTTH'
                    WHEN UPPER(a.carpeta) = 'DX' THEN 'DX'
                    WHEN UPPER(a.carpeta) LIKE 'BROWNFIELD%' OR UPPER(a.carpeta) LIKE 'BROWFIELD%' THEN 'BROWNFIELD'
                    WHEN UPPER(a.carpeta) LIKE 'FTTH INSTALACIONES%' THEN 'FTTH INSTALACIONES'
                    WHEN UPPER(a.carpeta) LIKE 'INSTALACIONES DOBLES%' THEN 'INSTALACIONES DOBLES'
                    WHEN UPPER(a.carpeta) LIKE 'POSTVENTA FTTH%' THEN 'POSTVENTA FTTH'
                    WHEN UPPER(a.carpeta) LIKE 'POSTVENTA%' THEN 'POSTVENTA'
                    ELSE a.carpeta
                END as carpeta,
                COUNT(DISTINCT a.id_codigo_consumidor) as total_tecnicos
            FROM asistencia a
            INNER JOIN tipificacion_asistencia t ON a.carpeta_dia = t.codigo_tipificacion
            WHERE DATE(a.fecha_asistencia) BETWEEN %s AND %s
                AND t.grupo IS NOT NULL 
                AND t.grupo != ''
                AND t.grupo NOT IN ('AUSENCIA INJUSTIFICADA', 'AUSENCIA JUSTIFICADA')
                AND t.grupo IN ('ARREGLOS', 'INSTALACIONES', 'POSTVENTA')
                AND a.carpeta IS NOT NULL
                AND a.carpeta != ''
                AND (
                    UPPER(a.carpeta) LIKE 'ARREGLOS HFC%'
                    OR UPPER(a.carpeta) LIKE 'MANTENIMIENTO FTTH%'
                    OR UPPER(a.carpeta) = 'DX'
                    OR UPPER(a.carpeta) LIKE 'BROWNFIELD%'
                    OR UPPER(a.carpeta) LIKE 'BROWFIELD%'
                    OR UPPER(a.carpeta) LIKE 'FTTH INSTALACIONES%'
                    OR UPPER(a.carpeta) LIKE 'INSTALACIONES DOBLES%'
                    OR UPPER(a.carpeta) LIKE 'POSTVENTA%'
                    OR UPPER(a.carpeta) LIKE 'POSTVENTA FTTH%'
                )
        """
        
        params_operacion = [fecha_inicio, fecha_fin]
        
        # Agregar filtro por supervisor si se proporciona
        if supervisor_filtro:
            query_operacion += " AND a.super = %s"
            params_operacion.append(supervisor_filtro)
        
        query_operacion += """
            GROUP BY 
                CASE 
                    WHEN UPPER(a.carpeta) LIKE 'ARREGLOS HFC%'
                         OR UPPER(a.carpeta) LIKE 'MANTENIMIENTO FTTH%'
                         OR UPPER(a.carpeta) = 'DX' THEN 'ARREGLOS'
                    WHEN UPPER(a.carpeta) LIKE 'BROWNFIELD%'
                         OR UPPER(a.carpeta) LIKE 'BROWFIELD%'
                         OR UPPER(a.carpeta) LIKE 'FTTH INSTALACIONES%'
                         OR UPPER(a.carpeta) LIKE 'INSTALACIONES DOBLES%' THEN 'INSTALACIONES'
                    WHEN UPPER(a.carpeta) LIKE 'POSTVENTA%'
                         OR UPPER(a.carpeta) LIKE 'POSTVENTA FTTH%' THEN 'POSTVENTA'
                    ELSE 'OTROS'
                END,
                CASE 
                    WHEN UPPER(a.carpeta) LIKE 'ARREGLOS HFC%' THEN 'ARREGLOS HFC'
                    WHEN UPPER(a.carpeta) LIKE 'MANTENIMIENTO FTTH%' THEN 'MANTENIMIENTO FTTH'
                    WHEN UPPER(a.carpeta) = 'DX' THEN 'DX'
                    WHEN UPPER(a.carpeta) LIKE 'BROWNFIELD%' OR UPPER(a.carpeta) LIKE 'BROWFIELD%' THEN 'BROWNFIELD'
                    WHEN UPPER(a.carpeta) LIKE 'FTTH INSTALACIONES%' THEN 'FTTH INSTALACIONES'
                    WHEN UPPER(a.carpeta) LIKE 'INSTALACIONES DOBLES%' THEN 'INSTALACIONES DOBLES'
                    WHEN UPPER(a.carpeta) LIKE 'POSTVENTA FTTH%' THEN 'POSTVENTA FTTH'
                    WHEN UPPER(a.carpeta) LIKE 'POSTVENTA%' THEN 'POSTVENTA'
                    ELSE a.carpeta
                END
            ORDER BY grupo, carpeta
        """
        
        # Ejecutar consulta para operación por recurso operativo
        cursor.execute(query_operacion, tuple(params_operacion))
        resultados_operacion = cursor.fetchall()
        
        # Calcular total operativo (excluyendo ausencias)
        query_total_operativo = """
            SELECT COUNT(DISTINCT a.id_codigo_consumidor) as total_operativo
            FROM asistencia a
            INNER JOIN tipificacion_asistencia t ON a.carpeta_dia = t.codigo_tipificacion
            WHERE DATE(a.fecha_asistencia) BETWEEN %s AND %s
                AND t.grupo IS NOT NULL 
                AND t.grupo != ''
                AND t.grupo NOT IN ('AUSENCIA INJUSTIFICADA', 'AUSENCIA JUSTIFICADA')
                AND t.grupo IN ('ARREGLOS', 'INSTALACIONES', 'POSTVENTA')
                AND a.carpeta IS NOT NULL
                AND a.carpeta != ''
                AND (
                    UPPER(a.carpeta) LIKE 'ARREGLOS HFC%'
                    OR UPPER(a.carpeta) LIKE 'MANTENIMIENTO FTTH%'
                    OR UPPER(a.carpeta) = 'DX'
                    OR UPPER(a.carpeta) LIKE 'BROWNFIELD%'
                    OR UPPER(a.carpeta) LIKE 'BROWFIELD%'
                    OR UPPER(a.carpeta) LIKE 'FTTH INSTALACIONES%'
                    OR UPPER(a.carpeta) LIKE 'INSTALACIONES DOBLES%'
                    OR UPPER(a.carpeta) LIKE 'POSTVENTA%'
                    OR UPPER(a.carpeta) LIKE 'POSTVENTA FTTH%'
                )
        """
        
        params_total_operativo = [fecha_inicio, fecha_fin]
        if supervisor_filtro:
            query_total_operativo += " AND a.super = %s"
            params_total_operativo.append(supervisor_filtro)
        
        cursor.execute(query_total_operativo, tuple(params_total_operativo))
        total_operativo = cursor.fetchone()['total_operativo'] or 0
        
        # Procesar resultados de operación por recurso operativo
        operacion_recurso = []
        for resultado in resultados_operacion:
            grupo = resultado['grupo']
            carpeta = resultado['carpeta']
            total_tecnicos = resultado['total_tecnicos']
            
            # Calcular porcentaje respecto al total operativo (sin ausencias)
            porcentaje = round((total_tecnicos * 100) / total_operativo, 2) if total_operativo > 0 else 0
            
            operacion_recurso.append({
                'grupo': grupo,
                'carpeta': carpeta,
                'total_tecnicos': total_tecnicos,
                'porcentaje': porcentaje
            })
        
        # ===== FUNCIONALIDAD EXISTENTE: RESUMEN POR GRUPOS =====
        # Consulta base para obtener datos agrupados (solo registros con grupo válido)
        query_base = """
            SELECT 
                t.grupo,
                t.nombre_tipificacion as carpeta,
                COUNT(DISTINCT a.id_codigo_consumidor) as total_tecnicos
            FROM asistencia a
            INNER JOIN tipificacion_asistencia t ON a.carpeta_dia = t.codigo_tipificacion
            WHERE DATE(a.fecha_asistencia) BETWEEN %s AND %s
                AND t.grupo IS NOT NULL 
                AND t.grupo != ''
                AND t.grupo IN ('ARREGLOS', 'AUSENCIA INJUSTIFICADA', 'AUSENCIA JUSTIFICADA', 'INSTALACIONES', 'POSTVENTA')
        """
        
        params = [fecha_inicio, fecha_fin]
        
        # Agregar filtro por supervisor si se proporciona
        if supervisor_filtro:
            query_base += " AND a.super = %s"
            params.append(supervisor_filtro)
        
        query_base += """
            GROUP BY t.grupo, t.nombre_tipificacion
            ORDER BY t.grupo, t.nombre_tipificacion
        """
        
        # Ejecutar consulta principal
        cursor.execute(query_base, tuple(params))
        resultados = cursor.fetchall()
        
        # Calcular total de técnicos para porcentajes (solo registros con grupo válido)
        query_total = """
            SELECT COUNT(DISTINCT a.id_codigo_consumidor) as total_general
            FROM asistencia a
            INNER JOIN tipificacion_asistencia t ON a.carpeta_dia = t.codigo_tipificacion
            WHERE DATE(a.fecha_asistencia) BETWEEN %s AND %s
                AND t.grupo IS NOT NULL 
                AND t.grupo != ''
                AND t.grupo IN ('ARREGLOS', 'AUSENCIA INJUSTIFICADA', 'AUSENCIA JUSTIFICADA', 'INSTALACIONES', 'POSTVENTA')
        """
        
        params_total = [fecha_inicio, fecha_fin]
        if supervisor_filtro:
            query_total += " AND a.super = %s"
            params_total.append(supervisor_filtro)
        
        cursor.execute(query_total, tuple(params_total))
        total_general = cursor.fetchone()['total_general'] or 0
        
        # Calcular total operativo para la tabla tradicional (excluyendo ausencias)
        query_total_general_operativo = """
            SELECT COUNT(DISTINCT a.id_codigo_consumidor) as total_general_operativo
            FROM asistencia a
            INNER JOIN tipificacion_asistencia t ON a.carpeta_dia = t.codigo_tipificacion
            WHERE DATE(a.fecha_asistencia) BETWEEN %s AND %s
                AND t.grupo IS NOT NULL 
                AND t.grupo != ''
                AND t.grupo NOT IN ('AUSENCIA INJUSTIFICADA', 'AUSENCIA JUSTIFICADA')
                AND t.grupo IN ('ARREGLOS', 'INSTALACIONES', 'POSTVENTA')
        """
        
        params_total_general_operativo = [fecha_inicio, fecha_fin]
        if supervisor_filtro:
            query_total_general_operativo += " AND a.super = %s"
            params_total_general_operativo.append(supervisor_filtro)
        
        cursor.execute(query_total_general_operativo, tuple(params_total_general_operativo))
        total_general_operativo = cursor.fetchone()['total_general_operativo'] or 0
        
        # Procesar resultados y calcular porcentajes
        resumen_agrupado = []
        grupos_totales = {}
        
        # Primero, calcular totales por grupo
        for resultado in resultados:
            grupo = resultado['grupo']
            if grupo not in grupos_totales:
                grupos_totales[grupo] = 0
            grupos_totales[grupo] += resultado['total_tecnicos']
        
        # Luego, crear el resumen con porcentajes
        for resultado in resultados:
            grupo = resultado['grupo']
            carpeta = resultado['carpeta']
            total_tecnicos = resultado['total_tecnicos']
            
            # Calcular porcentaje respecto al total operativo (sin ausencias)
            porcentaje = round((total_tecnicos * 100) / total_general_operativo, 2) if total_general_operativo > 0 else 0
            
            resumen_agrupado.append({
                'grupo': grupo,
                'carpeta': carpeta,
                'total_tecnicos': total_tecnicos,
                'porcentaje': porcentaje
            })
        
        # Calcular resumen por grupos
        resumen_grupos = []
        for grupo, total_grupo in grupos_totales.items():
            porcentaje_grupo = round((total_grupo * 100) / total_general_operativo, 2) if total_general_operativo > 0 else 0
            resumen_grupos.append({
                'grupo': grupo,
                'total_tecnicos': total_grupo,
                'porcentaje': porcentaje_grupo
            })
        
        # ===== NUEVA FUNCIONALIDAD: OBTENER AUSENCIAS =====
        # Consulta para obtener ausencias para la tabla de operación
        query_ausencias_operacion = """
            SELECT 
                CASE 
                    WHEN t.grupo = 'AUSENCIA INJUSTIFICADA' THEN 'AUSENCIA INJUSTIFICADA'
                    WHEN t.grupo = 'AUSENCIA JUSTIFICADA' THEN 'AUSENCIA JUSTIFICADA'
                    ELSE t.grupo
                END as grupo,
                t.nombre_tipificacion as carpeta,
                COUNT(DISTINCT a.id_codigo_consumidor) as total_tecnicos
            FROM asistencia a
            INNER JOIN tipificacion_asistencia t ON a.carpeta_dia = t.codigo_tipificacion
            WHERE DATE(a.fecha_asistencia) BETWEEN %s AND %s
                AND t.grupo IS NOT NULL 
                AND t.grupo != ''
                AND t.grupo IN ('AUSENCIA INJUSTIFICADA', 'AUSENCIA JUSTIFICADA')
        """
        
        params_ausencias_operacion = [fecha_inicio, fecha_fin]
        
        # Agregar filtro por supervisor si se proporciona
        if supervisor_filtro:
            query_ausencias_operacion += " AND a.super = %s"
            params_ausencias_operacion.append(supervisor_filtro)
        
        query_ausencias_operacion += """
            GROUP BY CASE 
                WHEN t.grupo = 'AUSENCIA INJUSTIFICADA' THEN 'AUSENCIA INJUSTIFICADA'
                WHEN t.grupo = 'AUSENCIA JUSTIFICADA' THEN 'AUSENCIA JUSTIFICADA'
                ELSE t.grupo
            END, t.nombre_tipificacion
            ORDER BY grupo, t.nombre_tipificacion
        """
        
        # Ejecutar consulta para ausencias de operación
        cursor.execute(query_ausencias_operacion, tuple(params_ausencias_operacion))
        resultados_ausencias_operacion = cursor.fetchall()
        
        # Procesar ausencias para la tabla de operación
        ausencias_operacion = []
        total_ausencias = 0
        for resultado in resultados_ausencias_operacion:
            grupo = resultado['grupo']
            carpeta = resultado['carpeta']
            total_tecnicos = resultado['total_tecnicos']
            total_ausencias += total_tecnicos
            
            # Calcular porcentaje respecto al total operativo (sin ausencias)
            porcentaje = round((total_tecnicos * 100) / total_operativo, 2) if total_operativo > 0 else 0
            
            ausencias_operacion.append({
                'grupo': grupo,
                'carpeta': carpeta,
                'total_tecnicos': total_tecnicos,
                'porcentaje': porcentaje,
                'es_ausencia': True
            })
        
        # Los porcentajes de operación ya están calculados correctamente usando solo total_operativo
        
        # Ordenar resumen_agrupado para que las ausencias aparezcan al final
        resumen_agrupado_ordenado = []
        ausencias_detalladas = []
        
        for item in resumen_agrupado:
            if item['grupo'] in ['AUSENCIA INJUSTIFICADA', 'AUSENCIA JUSTIFICADA']:
                item['es_ausencia'] = True
                ausencias_detalladas.append(item)
            else:
                resumen_agrupado_ordenado.append(item)
        
        # Agregar ausencias al final
        resumen_agrupado_ordenado.extend(ausencias_detalladas)
        
        return jsonify({
            'success': True,
            'data': {
                # Nueva funcionalidad: Operación por Recurso Operativo
                'operacion_recurso': operacion_recurso,
                'ausencias_operacion': ausencias_operacion,
                'total_operativo': total_operativo,
                'total_ausencias': total_ausencias,
                # Funcionalidad existente
                'detallado': resumen_agrupado_ordenado,
                'resumen_grupos': resumen_grupos,
                'total_general': total_general,
                'total_general_operativo': total_general_operativo,
                'fecha_inicio': fecha_inicio.strftime('%Y-%m-%d'),
                'fecha_fin': fecha_fin.strftime('%Y-%m-%d'),
                'supervisor_filtro': supervisor_filtro
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

@app.route('/api/indicadores/cumplimiento')
@login_required(role='administrativo')
def obtener_indicadores_cumplimiento():
    try:
        # Obtener los parámetros de fecha
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        supervisor_filtro = request.args.get('supervisor')
        
        # Para compatibilidad con versiones anteriores
        fecha = request.args.get('fecha')
        
        # Mostrar información de debug sobre parámetros recibidos
        print(f"Parámetros recibidos en API:")
        print(f"- fecha_inicio: {fecha_inicio}")
        print(f"- fecha_fin: {fecha_fin}")
        print(f"- supervisor: {supervisor_filtro}")
        print(f"- fecha (compatibilidad): {fecha}")
        
        # Validar fechas
        if fecha:
            # Modo compatibilidad - una sola fecha
            try:
                fecha_inicio = fecha_fin = datetime.strptime(fecha, '%Y-%m-%d').date()
                print(f"Usando modo compatibilidad con fecha: {fecha_inicio}")
            except ValueError:
                print(f"Error: Formato de fecha inválido: {fecha}")
                return jsonify({
                    'success': False,
                    'error': 'Formato de fecha inválido. Use YYYY-MM-DD'
                }), 400
        elif fecha_inicio and fecha_fin:
            # Nuevo modo - rango de fechas
            try:
                fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                print(f"Usando rango de fechas: {fecha_inicio} a {fecha_fin}")
            except ValueError:
                print(f"Error: Formato de fecha inválido en rango: {fecha_inicio} - {fecha_fin}")
                return jsonify({
                    'success': False,
                    'error': 'Formato de fecha inválido. Use YYYY-MM-DD'
                }), 400
        else:
            # Si no se proporcionan fechas, usar la fecha actual
            fecha_actual = get_bogota_datetime().date()
            fecha_inicio = fecha_fin = fecha_actual
            print(f"Usando fecha actual: {fecha_actual}")
        
        connection = get_db_connection()
        if connection is None:
            print("Error: No se pudo establecer conexión a la base de datos")
            return jsonify({
                'success': False,
                'error': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Verificar datos para el rango de fechas
        cursor.execute("SELECT COUNT(*) as total FROM asistencia WHERE DATE(fecha_asistencia) BETWEEN %s AND %s", 
                      (fecha_inicio, fecha_fin))
        total_asistencia = cursor.fetchone()['total']
        
        cursor.execute("SELECT COUNT(*) as total FROM preoperacional WHERE DATE(fecha) BETWEEN %s AND %s", 
                      (fecha_inicio, fecha_fin))
        total_preop = cursor.fetchone()['total']
        
        print(f"Datos encontrados en el rango {fecha_inicio} a {fecha_fin}:")
        print(f"- Total asistencia: {total_asistencia}")
        print(f"- Total preoperacional: {total_preop}")
        
        if total_asistencia == 0 and total_preop == 0:
            print("No hay datos para el rango de fechas")
            return jsonify({
                'success': True,
                'indicadores': [],
                'mensaje': f'No hay datos para el período {fecha_inicio} - {fecha_fin}'
            })
        
        # Obtener asistencia válida por supervisor
        query_asistencia = """
            SELECT a.super as supervisor, COUNT(*) as total_asistencia 
            FROM asistencia a 
            JOIN tipificacion_asistencia t ON a.carpeta_dia = t.codigo_tipificacion
            WHERE DATE(a.fecha_asistencia) BETWEEN %s AND %s AND t.valor = '1'
            GROUP BY a.super
        """
        params_asistencia = [fecha_inicio, fecha_fin]
        
        # Obtener preoperacionales por supervisor - SOLO de técnicos con asistencia
        query_preoperacional = """
            SELECT p.supervisor, COUNT(*) as total_preoperacional
            FROM preoperacional p
            INNER JOIN asistencia a ON p.id_codigo_consumidor = a.id_codigo_consumidor 
                AND DATE(p.fecha) = DATE(a.fecha_asistencia)
            INNER JOIN tipificacion_asistencia t ON a.carpeta_dia = t.codigo_tipificacion
            WHERE DATE(p.fecha) BETWEEN %s AND %s 
                AND t.valor = '1'
            GROUP BY p.supervisor
        """
        params_preoperacional = [fecha_inicio, fecha_fin]
        
        # Aplicar filtro por supervisor si existe
        if supervisor_filtro:
            print(f"Aplicando filtro por supervisor: {supervisor_filtro}")
            query_asistencia += " HAVING a.super = %s"
            params_asistencia.append(supervisor_filtro)
            
            query_preoperacional += " HAVING supervisor = %s"
            params_preoperacional.append(supervisor_filtro)
        
        print(f"Ejecutando consulta de asistencia: {query_asistencia}")
        print(f"Parámetros: {params_asistencia}")
        
        # Ejecutar consultas
        cursor.execute(query_asistencia, tuple(params_asistencia))
        asistencia_raw = cursor.fetchall()
        
        print(f"Ejecutando consulta de preoperacional: {query_preoperacional}")
        print(f"Parámetros: {params_preoperacional}")
        
        cursor.execute(query_preoperacional, tuple(params_preoperacional))
        preop_raw = cursor.fetchall()
        
        # Debugging detallado de supervisores
        print(f"🔍 SUPERVISORES RAW DE ASISTENCIA:")
        for row in asistencia_raw:
            supervisor = row['supervisor']
            print(f"  - '{supervisor}' (len: {len(supervisor) if supervisor else 0}, repr: {repr(supervisor)})")
        
        print(f"🔍 SUPERVISORES RAW DE PREOPERACIONAL:")
        for row in preop_raw:
            supervisor = row['supervisor']
            print(f"  - '{supervisor}' (len: {len(supervisor) if supervisor else 0}, repr: {repr(supervisor)})")
        
        # Normalizar nombres de supervisores (trim espacios y convertir a mayúsculas)
        def normalizar_supervisor(nombre):
            if not nombre:
                return None
            return nombre.strip().upper()
        
        # Procesar asistencia con normalización
        asistencia_por_supervisor = {}
        for row in asistencia_raw:
            supervisor_original = row['supervisor']
            supervisor_normalizado = normalizar_supervisor(supervisor_original)
            if supervisor_normalizado:
                if supervisor_normalizado in asistencia_por_supervisor:
                    asistencia_por_supervisor[supervisor_normalizado] += row['total_asistencia']
                    print(f"⚠️  DUPLICADO EN ASISTENCIA: '{supervisor_original}' -> '{supervisor_normalizado}' (sumando {row['total_asistencia']})")
                else:
                    asistencia_por_supervisor[supervisor_normalizado] = row['total_asistencia']
        
        # Procesar preoperacional con normalización
        preop_por_supervisor = {}
        for row in preop_raw:
            supervisor_original = row['supervisor']
            supervisor_normalizado = normalizar_supervisor(supervisor_original)
            if supervisor_normalizado:
                if supervisor_normalizado in preop_por_supervisor:
                    preop_por_supervisor[supervisor_normalizado] += row['total_preoperacional']
                    print(f"⚠️  DUPLICADO EN PREOPERACIONAL: '{supervisor_original}' -> '{supervisor_normalizado}' (sumando {row['total_preoperacional']})")
                else:
                    preop_por_supervisor[supervisor_normalizado] = row['total_preoperacional']
        
        print(f"✅ ASISTENCIA NORMALIZADA: {asistencia_por_supervisor}")
        print(f"✅ PREOPERACIONAL NORMALIZADA: {preop_por_supervisor}")
        
        # Calcular indicadores con supervisores únicos
        indicadores = []
        supervisores = set(list(asistencia_por_supervisor.keys()) + list(preop_por_supervisor.keys()))
        
        for supervisor in supervisores:
            if supervisor:  # Ignorar supervisores nulos
                total_asistencia = asistencia_por_supervisor.get(supervisor, 0)
                total_preoperacional = preop_por_supervisor.get(supervisor, 0)
                porcentaje = (total_preoperacional * 100) / total_asistencia if total_asistencia > 0 else 0
                
                indicador = {
                    'supervisor': supervisor,
                    'total_asistencia': total_asistencia,
                    'total_preoperacional': total_preoperacional,
                    'porcentaje_cumplimiento': porcentaje
                }
                indicadores.append(indicador)
        
        # Ordenar por porcentaje de cumplimiento descendente
        indicadores.sort(key=lambda x: x['porcentaje_cumplimiento'], reverse=True)
        
        print(f"Se calcularon {len(indicadores)} indicadores")
        
        cursor.close()
        connection.close()
        
        # Incluir información del período consultado en la respuesta
        return jsonify({
            'success': True,
            'indicadores': indicadores,
            'periodo': {
                'fecha_inicio': fecha_inicio.strftime('%Y-%m-%d'),
                'fecha_fin': fecha_fin.strftime('%Y-%m-%d')
            }
        })
        
    except Exception as e:
        import traceback
        print(f"Error en indicadores de cumplimiento: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/indicadores/detalle_tecnicos')
@login_required(role='administrativo')
def obtener_detalle_tecnicos():
    """Obtener detalle de técnicos por supervisor con estado de asistencia y preoperacional"""
    print(f"🔍 [DETALLE_TECNICOS] ===== ENDPOINT LLAMADO =====")
    print(f"🔍 [DETALLE_TECNICOS] Usuario actual: {session.get('user_id', 'No autenticado')}")
    print(f"🔍 [DETALLE_TECNICOS] Rol usuario: {session.get('role', 'Sin rol')}")
    print(f"🔍 [DETALLE_TECNICOS] Request method: {request.method}")
    print(f"🔍 [DETALLE_TECNICOS] Request URL: {request.url}")
    print(f"🔍 [DETALLE_TECNICOS] Request args: {dict(request.args)}")
    print(f"🔍 [DETALLE_TECNICOS] Iniciando endpoint...")
    try:
        # Obtener parámetros
        fecha = request.args.get('fecha')
        supervisor = request.args.get('supervisor')
        
        print(f"🔍 [DETALLE_TECNICOS] Parámetros recibidos:")
        print(f"  - fecha: '{fecha}'")
        print(f"  - supervisor: '{supervisor}'")
        
        # Validar parámetros
        if not fecha or not supervisor:
            return jsonify({
                'success': False,
                'error': 'Se requieren los parámetros fecha y supervisor'
            }), 400
        
        # Validar formato de fecha
        try:
            fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({
                'success': False,
                'error': 'Formato de fecha inválido. Use YYYY-MM-DD'
            }), 400
        
        print(f"🔍 [DETALLE_TECNICOS] Intentando conectar a la base de datos...")
        connection = get_db_connection()
        if connection is None:
            print(f"❌ [DETALLE_TECNICOS] Error: No se pudo conectar a la base de datos")
            return jsonify({
                'success': False,
                'error': 'Error de conexión a la base de datos'
            }), 500
        
        print(f"✅ [DETALLE_TECNICOS] Conexión a BD exitosa")
        cursor = connection.cursor(dictionary=True)
        
        # Obtener técnicos del supervisor (usando la consulta correcta)
        query_tecnicos = """
            SELECT id_codigo_consumidor, nombre, recurso_operativo_cedula as documento
            FROM recurso_operativo 
            WHERE super = %s AND estado = 'Activo' AND id_codigo_consumidor IS NOT NULL
            ORDER BY nombre
        """
        
        print(f"🔍 [DETALLE_TECNICOS] Ejecutando consulta: {query_tecnicos}")
        print(f"🔍 [DETALLE_TECNICOS] Con parámetro supervisor: '{supervisor}'")
        
        cursor.execute(query_tecnicos, (supervisor,))
        tecnicos_supervisor = cursor.fetchall()
        
        print(f"✅ [DETALLE_TECNICOS] Técnicos encontrados en BD: {len(tecnicos_supervisor)}")
        for tecnico in tecnicos_supervisor:
            print(f"  - {tecnico['nombre']} (ID: {tecnico['id_codigo_consumidor']}, Doc: {tecnico['documento']})")
        
        if not tecnicos_supervisor:
            print(f"❌ No se encontraron técnicos activos para el supervisor: '{supervisor}'")
            cursor.close()
            connection.close()
            return jsonify({
                'success': True,
                'tecnicos': [],
                'mensaje': f'No se encontraron técnicos activos para el supervisor {supervisor}'
            })
        
        # Para cada técnico, verificar asistencia y preoperacional
        tecnicos_detalle = []
        
        for tecnico in tecnicos_supervisor:
            id_tecnico = tecnico['id_codigo_consumidor']
            nombre_tecnico = tecnico['nombre']
            
            # Verificar asistencia válida
            cursor.execute("""
                SELECT a.*, t.valor as asistencia_valida
                FROM asistencia a
                JOIN tipificacion_asistencia t ON a.carpeta_dia = t.codigo_tipificacion
                WHERE a.id_codigo_consumidor = %s 
                    AND DATE(a.fecha_asistencia) = %s
                    AND t.valor = '1'
            """, (id_tecnico, fecha_obj))
            
            asistencia = cursor.fetchone()
            tiene_asistencia = asistencia is not None
            hora_asistencia = None
            
            if tiene_asistencia:
                # Convertir a hora de Bogotá para mostrar
                fecha_asistencia_utc = asistencia['fecha_asistencia']
                if fecha_asistencia_utc:
                    fecha_bogota = convert_to_bogota_time(fecha_asistencia_utc)
                    hora_asistencia = fecha_bogota.strftime('%H:%M')
            
            # Verificar preoperacional
            cursor.execute("""
                SELECT * FROM preoperacional 
                WHERE id_codigo_consumidor = %s 
                    AND DATE(fecha) = %s
            """, (id_tecnico, fecha_obj))
            
            preoperacional = cursor.fetchone()
            tiene_preoperacional = preoperacional is not None
            hora_preoperacional = None
            
            if tiene_preoperacional:
                # Convertir a hora de Bogotá para mostrar
                fecha_preop_utc = preoperacional['fecha']
                if fecha_preop_utc:
                    fecha_bogota = convert_to_bogota_time(fecha_preop_utc)
                    hora_preoperacional = fecha_bogota.strftime('%H:%M')
            
            # Determinar estado general
            if tiene_asistencia and tiene_preoperacional:
                estado = "Completo"
            elif tiene_asistencia:
                estado = "Solo Asistencia"
            elif tiene_preoperacional:
                estado = "Solo Preoperacional"
            else:
                estado = "Sin Registros"
            
            tecnicos_detalle.append({
                'id_codigo_consumidor': id_tecnico,
                'nombre': nombre_tecnico,
                'asistencia': tiene_asistencia,
                'preoperacional': tiene_preoperacional,
                'estado': estado,
                'hora_asistencia': hora_asistencia or 'N/A',
                'hora_preoperacional': hora_preoperacional or 'N/A',
                'hora_registro': hora_asistencia or hora_preoperacional or 'N/A'
            })
        
        cursor.close()
        connection.close()
        
        print(f"Se encontraron {len(tecnicos_detalle)} técnicos para el supervisor {supervisor}")
        
        return jsonify({
            'success': True,
            'tecnicos': tecnicos_detalle,
            'supervisor': supervisor,
            'fecha': fecha
        })
        
    except Exception as e:
        import traceback
        error_msg = str(e)
        traceback_str = traceback.format_exc()
        
        print(f"❌ [DETALLE_TECNICOS] ERROR CRÍTICO:")
        print(f"  - Tipo de error: {type(e).__name__}")
        print(f"  - Mensaje: {error_msg}")
        print(f"  - Traceback completo:")
        print(traceback_str)
        print(f"❌ [DETALLE_TECNICOS] FIN DEL ERROR")
        
        # Cerrar conexiones si existen
        try:
            if 'cursor' in locals():
                cursor.close()
            if 'connection' in locals():
                connection.close()
        except:
            pass
        
        return jsonify({
            'success': False,
            'error': f"Error interno del servidor: {error_msg}",
            'error_type': type(e).__name__
        }), 500

@app.route('/indicadores/api')
@login_required(role='administrativo')
def indicadores_api():
    # Obtener la lista de supervisores para el selector del formulario
    try:
        connection = get_db_connection()
        cursor = connection.cursor(dictionary=True)
        
        # Consulta para obtener supervisores únicos de recurso_operativo
        cursor.execute("SELECT DISTINCT super as supervisor FROM recurso_operativo WHERE super IS NOT NULL AND super != ''")
        supervisores_db = cursor.fetchall()
        
        supervisores = [sup['supervisor'] for sup in supervisores_db]
        
        cursor.close()
        connection.close()
    except Exception as e:
        logging.error(f"Error al obtener supervisores: {str(e)}")
        supervisores = []
    
    return render_template('modulos/administrativo/api_indicadores_cumplimiento.html', 
                          supervisores=supervisores)




@app.route('/api/cargos', methods=['GET'])
@login_required(role='administrativo')
def get_cargos():
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener cargos distintos de la tabla recurso_operativo
        cursor.execute("SELECT DISTINCT cargo FROM recurso_operativo WHERE cargo IS NOT NULL AND cargo != '' ORDER BY cargo")
        cargos = [row['cargo'] for row in cursor.fetchall()]
        
        return jsonify({'success': True, 'cargos': cargos})
        
    except Error as e:
        return jsonify({'success': False, 'message': f'Error al obtener cargos: {str(e)}'}), 500
        
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/analistas', methods=['GET'])
@login_required(role='administrativo')
def get_analistas():
    """Obtener lista de usuarios con cargo ANALISTA para el dropdown de analistas"""
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener usuarios con cargo ANALISTA que estén activos
        cursor.execute("""
            SELECT id_codigo_consumidor, nombre, recurso_operativo_cedula 
            FROM recurso_operativo 
            WHERE cargo = 'ANALISTA' AND estado = 'Activo' 
            ORDER BY nombre
        """)
        analistas = cursor.fetchall()
        
        return jsonify({'success': True, 'analistas': analistas})
        
    except Error as e:
        return jsonify({'success': False, 'message': f'Error al obtener analistas: {str(e)}'}), 500
        
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

# ============================================================================
# REST API ENDPOINTS FOR VEHICLE MANAGEMENT
# ============================================================================

@app.route('/api/vehiculos', methods=['GET'])
@login_required()
@role_required('logistica')
def api_obtener_vehiculos():
    """Obtener lista de vehículos con filtros opcionales"""
    try:
        print("Intentando obtener conexión a la base de datos...")
        connection = get_db_connection()
        if connection is None:
            print("La conexión es None")
            return jsonify({
                'success': False,
                'message': 'MySQL Connection not available'
            }), 500
        print("Conexión obtenida exitosamente")
            
        cursor = connection.cursor(dictionary=True)
        
        # Parámetros de filtro
        estado = request.args.get('estado')
        tipo_vehiculo = request.args.get('tipo_vehiculo')
        tecnico_id = request.args.get('tecnico_id')
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))
        
        # Construir consulta base
        query = """
            SELECT pa.*, r.nombre as tecnico_nombre, r.recurso_operativo_cedula, r.cargo,
                   CASE 
                       WHEN pa.soat_vencimiento <= CURDATE() THEN 'vencido'
                       WHEN pa.soat_vencimiento <= DATE_ADD(CURDATE(), INTERVAL 30 DAY) THEN 'por_vencer'
                       ELSE 'vigente'
                   END as estado_soat,
                   CASE 
                       WHEN pa.tecnomecanica_vencimiento <= CURDATE() THEN 'vencido'
                       WHEN pa.tecnomecanica_vencimiento <= DATE_ADD(CURDATE(), INTERVAL 30 DAY) THEN 'por_vencer'
                       ELSE 'vigente'
                   END as estado_tecnomecanica
            FROM parque_automotor pa 
            LEFT JOIN recurso_operativo r ON pa.id_codigo_consumidor = r.id_codigo_consumidor
        """
        
        conditions = []
        params = []
        
        if estado:
            conditions.append("pa.estado = %s")
            params.append(estado)
            
        if tipo_vehiculo:
            conditions.append("pa.tipo_vehiculo = %s")
            params.append(tipo_vehiculo)
            
        if tecnico_id:
            conditions.append("pa.id_codigo_consumidor = %s")
            params.append(tecnico_id)
            
        if conditions:
            query += " WHERE " + " AND ".join(conditions)
            
        query += " ORDER BY pa.fecha_asignacion DESC"
        
        # Contar total de registros
        count_query = query.replace(
            "SELECT pa.*, r.nombre, r.recurso_operativo_cedula, r.cargo, CASE WHEN pa.soat_vencimiento <= CURDATE() THEN 'vencido' WHEN pa.soat_vencimiento <= DATE_ADD(CURDATE(), INTERVAL 30 DAY) THEN 'por_vencer' ELSE 'vigente' END as estado_soat, CASE WHEN pa.tecnomecanica_vencimiento <= CURDATE() THEN 'vencido' WHEN pa.tecnomecanica_vencimiento <= DATE_ADD(CURDATE(), INTERVAL 30 DAY) THEN 'por_vencer' ELSE 'vigente' END as estado_tecnomecanica",
            "SELECT COUNT(*)"
        )
        
        cursor.execute(count_query, params)
        total = cursor.fetchone()['COUNT(*)']
        
        # Aplicar paginación
        offset = (page - 1) * per_page
        query += f" LIMIT {per_page} OFFSET {offset}"
        
        cursor.execute(query, params)
        vehiculos = cursor.fetchall()
        
        # Convertir fechas para serialización JSON
        for vehiculo in vehiculos:
            if vehiculo['fecha_asignacion']:
                vehiculo['fecha_asignacion'] = vehiculo['fecha_asignacion'].isoformat()
            if vehiculo['soat_vencimiento']:
                vehiculo['soat_vencimiento'] = vehiculo['soat_vencimiento'].isoformat()
            if vehiculo['tecnomecanica_vencimiento']:
                vehiculo['tecnomecanica_vencimiento'] = vehiculo['tecnomecanica_vencimiento'].isoformat()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'data': vehiculos,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total': total,
                'pages': (total + per_page - 1) // per_page
            }
        })
        
    except Exception as e:
        print(f"Error en API obtener vehículos: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error al obtener vehículos: {str(e)}'
        }), 500

@app.route('/api/vehiculos/<int:vehiculo_id>', methods=['GET'])
@login_required()
@role_required('logistica')
def api_obtener_vehiculo(vehiculo_id):
    """Obtener un vehículo específico por ID"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'success': False,
                'message': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT pa.*, r.nombre as tecnico_nombre, r.recurso_operativo_cedula, r.cargo
            FROM parque_automotor pa 
            LEFT JOIN recurso_operativo r ON pa.id_codigo_consumidor = r.id_codigo_consumidor
            WHERE pa.id_parque_automotor = %s
        """, (vehiculo_id,))
        
        vehiculo = cursor.fetchone()
        
        if not vehiculo:
            return jsonify({
                'success': False,
                'message': 'Vehículo no encontrado'
            }), 404
            
        # Convertir fechas para serialización JSON
        if vehiculo['fecha_asignacion']:
            vehiculo['fecha_asignacion'] = vehiculo['fecha_asignacion'].isoformat()
        if vehiculo['soat_vencimiento']:
            vehiculo['soat_vencimiento'] = vehiculo['soat_vencimiento'].isoformat()
        if vehiculo['tecnomecanica_vencimiento']:
            vehiculo['tecnomecanica_vencimiento'] = vehiculo['tecnomecanica_vencimiento'].isoformat()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'data': vehiculo
        })
        
    except Exception as e:
        print(f"Error en API obtener vehículo: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error al obtener vehículo: {str(e)}'
        }), 500

@app.route('/api/vehiculos', methods=['POST'])
@login_required()
@role_required('logistica')
def api_crear_vehiculo():
    """Crear un nuevo vehículo"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                'success': False,
                'message': 'No se recibieron datos JSON'
            }), 400
            
        # Validar campos requeridos
        required_fields = ['placa', 'tipo_vehiculo', 'marca', 'modelo', 'color', 'fecha_asignacion']
        for field in required_fields:
            if not data.get(field):
                return jsonify({
                    'success': False,
                    'message': f'El campo {field} es requerido'
                }), 400
                
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'success': False,
                'message': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Verificar si la placa ya existe
        cursor.execute('SELECT placa FROM parque_automotor WHERE placa = %s', (data['placa'],))
        if cursor.fetchone():
            return jsonify({
                'success': False,
                'message': 'Ya existe un vehículo registrado con esta placa'
            }), 400
            
        # Insertar el nuevo vehículo
        cursor.execute("""
            INSERT INTO parque_automotor (
                placa, tipo_vehiculo, marca, modelo, color, 
                id_codigo_consumidor, fecha_asignacion, estado,
                soat_vencimiento, tecnomecanica_vencimiento, observaciones,
                kilometraje_actual, proximo_mantenimiento_km
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            data['placa'], data['tipo_vehiculo'], data['marca'], data['modelo'], data['color'],
            data.get('id_codigo_consumidor'), data['fecha_asignacion'], data.get('estado', 'Activo'),
            data.get('soat_vencimiento'), data.get('tecnomecanica_vencimiento'), data.get('observaciones'),
            data.get('kilometraje_actual', 0), data.get('proximo_mantenimiento_km')
        ))
        
        vehiculo_id = cursor.lastrowid
        connection.commit()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': 'Vehículo creado exitosamente',
            'data': {'id': vehiculo_id}
        }), 201
        
    except Exception as e:
        print(f"Error en API crear vehículo: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error al crear vehículo: {str(e)}'
        }), 500

@app.route('/api/vehiculos/<int:vehiculo_id>', methods=['PUT'])
@login_required()
@role_required('logistica')
def api_actualizar_vehiculo(vehiculo_id):
    """Actualizar un vehículo existente"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({
                'success': False,
                'message': 'No se recibieron datos JSON'
            }), 400
            
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'success': False,
                'message': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Verificar si el vehículo existe
        cursor.execute('SELECT id_parque_automotor FROM parque_automotor WHERE id_parque_automotor = %s', (vehiculo_id,))
        if not cursor.fetchone():
            return jsonify({
                'success': False,
                'message': 'El vehículo no existe'
            }), 404
            
        # Verificar si la placa ya existe para otro vehículo
        if 'placa' in data:
            cursor.execute('SELECT id_parque_automotor FROM parque_automotor WHERE placa = %s AND id_parque_automotor != %s', 
                         (data['placa'], vehiculo_id))
            if cursor.fetchone():
                return jsonify({
                    'success': False,
                    'message': 'Ya existe otro vehículo registrado con esta placa'
                }), 400
                
        # Construir consulta de actualización dinámicamente
        update_fields = []
        params = []
        
        allowed_fields = [
            'placa', 'tipo_vehiculo', 'marca', 'modelo', 'color',
            'id_codigo_consumidor', 'fecha_asignacion', 'estado',
            'soat_vencimiento', 'tecnomecanica_vencimiento', 'observaciones',
            'kilometraje_actual', 'proximo_mantenimiento_km',
            # Campos adicionales del vehículo
            'vin', 'parque_automotorcol', 'licencia',
            # Información del propietario
            'cedula_propietario', 'nombre_propietario',
            # Campos de mantenimiento
            'proximo_mantenimiento', 'fecha_ultimo_mantenimiento', 'ultimo_mantenimiento',
            'fecha_actualizacion', 'numero_vin', 'propietario_cedula', 'propietario_nombre',
            # Inspección física
            'estado_carroceria', 'estado_llantas', 'estado_frenos', 'estado_motor',
            'estado_luces', 'estado_espejos', 'estado_vidrios', 'estado_asientos',
            # Elementos de seguridad
            'cinturon_seguridad', 'extintor', 'botiquin', 'triangulos_seguridad',
            'llanta_repuesto', 'herramientas', 'gato', 'cruceta',
            # Información operativa
            'centro_de_trabajo', 'ciudad', 'supervisor', 'fecha', 'kilometraje'
        ]
        
        for field in allowed_fields:
            if field in data:
                update_fields.append(f"{field} = %s")
                params.append(data[field])
                
        if not update_fields:
            return jsonify({
                'success': False,
                'message': 'No se proporcionaron campos para actualizar'
            }), 400
            
        params.append(vehiculo_id)
        
        query = f"UPDATE parque_automotor SET {', '.join(update_fields)} WHERE id_parque_automotor = %s"
        cursor.execute(query, params)
        
        connection.commit()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': 'Vehículo actualizado exitosamente'
        })
        
    except Exception as e:
        print(f"Error en API actualizar vehículo: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error al actualizar vehículo: {str(e)}'
        }), 500

@app.route('/api/vehiculos/<int:vehiculo_id>', methods=['DELETE'])
@login_required()
@role_required('logistica')
def api_eliminar_vehiculo(vehiculo_id):
    """Eliminar un vehículo (cambiar estado a Inactivo)"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'success': False,
                'message': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Verificar si el vehículo existe
        cursor.execute('SELECT id_parque_automotor FROM parque_automotor WHERE id_parque_automotor = %s', (vehiculo_id,))
        if not cursor.fetchone():
            return jsonify({
                'success': False,
                'message': 'El vehículo no existe'
            }), 404
            
        # Cambiar estado a Inactivo en lugar de eliminar físicamente
        cursor.execute('UPDATE parque_automotor SET estado = %s WHERE id_parque_automotor = %s', 
                      ('Inactivo', vehiculo_id))
        
        connection.commit()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': 'Vehículo eliminado exitosamente'
        })
        
    except Exception as e:
        print(f"Error en API eliminar vehículo: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error al eliminar vehículo: {str(e)}'
        }), 500

@app.route('/api/vehiculos/dashboard', methods=['GET'])
@login_required()
@role_required('logistica')
def api_dashboard_vehiculos():
    """Obtener estadísticas del dashboard de vehículos"""
    try:
        print("Dashboard: Intentando obtener conexión a la base de datos...")
        connection = get_db_connection()
        if connection is None:
            print("Dashboard: La conexión es None")
            return jsonify({
                'success': False,
                'message': 'MySQL Connection not available'
            }), 500
        print("Dashboard: Conexión obtenida exitosamente")
            
        cursor = connection.cursor(dictionary=True)
        
        # Estadísticas generales
        cursor.execute("""
            SELECT 
                COUNT(*) as total_vehiculos,
                SUM(CASE WHEN estado = 'Activo' THEN 1 ELSE 0 END) as vehiculos_activos,
                SUM(CASE WHEN estado = 'Mantenimiento' THEN 1 ELSE 0 END) as vehiculos_mantenimiento,
                SUM(CASE WHEN estado = 'Inactivo' THEN 1 ELSE 0 END) as vehiculos_inactivos
            FROM parque_automotor
        """)
        estadisticas = cursor.fetchone()
        
        # Alertas de vencimiento
        cursor.execute("""
            SELECT 
                SUM(CASE WHEN soat_vencimiento <= CURDATE() THEN 1 ELSE 0 END) as soat_vencidos,
                SUM(CASE WHEN soat_vencimiento <= DATE_ADD(CURDATE(), INTERVAL 30 DAY) AND soat_vencimiento > CURDATE() THEN 1 ELSE 0 END) as soat_por_vencer,
                SUM(CASE WHEN tecnomecanica_vencimiento <= CURDATE() THEN 1 ELSE 0 END) as tecnomecanica_vencidos,
                SUM(CASE WHEN tecnomecanica_vencimiento <= DATE_ADD(CURDATE(), INTERVAL 30 DAY) AND tecnomecanica_vencimiento > CURDATE() THEN 1 ELSE 0 END) as tecnomecanica_por_vencer
            FROM parque_automotor
            WHERE estado = 'Activo'
        """)
        alertas = cursor.fetchone()
        
        # Distribución por tipo de vehículo
        cursor.execute("""
            SELECT tipo_vehiculo, COUNT(*) as cantidad
            FROM parque_automotor
            WHERE estado = 'Activo'
            GROUP BY tipo_vehiculo
            ORDER BY cantidad DESC
        """)
        tipos_vehiculo = cursor.fetchall()
        
        # Obtener alertas recientes para mostrar en el dashboard
        cursor.execute("""
            SELECT 
                placa,
                'SOAT' as tipo_documento,
                soat_vencimiento as fecha_vencimiento,
                CASE 
                    WHEN soat_vencimiento <= CURDATE() THEN 'vencido'
                    ELSE 'por_vencer'
                END as tipo_alerta
            FROM parque_automotor 
            WHERE estado = 'Activo' 
                AND soat_vencimiento <= DATE_ADD(CURDATE(), INTERVAL 30 DAY)
            UNION ALL
            SELECT 
                placa,
                'Tecnomecánica' as tipo_documento,
                tecnomecanica_vencimiento as fecha_vencimiento,
                CASE 
                    WHEN tecnomecanica_vencimiento <= CURDATE() THEN 'vencido'
                    ELSE 'por_vencer'
                END as tipo_alerta
            FROM parque_automotor 
            WHERE estado = 'Activo' 
                AND tecnomecanica_vencimiento <= DATE_ADD(CURDATE(), INTERVAL 30 DAY)
            ORDER BY fecha_vencimiento ASC
            LIMIT 10
        """)
        alertas_recientes = cursor.fetchall()
        
        cursor.close()
        connection.close()
        
        # Calcular totales de vencimientos
        documentos_vencidos = (alertas['soat_vencidos'] or 0) + (alertas['tecnomecanica_vencidos'] or 0)
        vencimientos_proximos = (alertas['soat_por_vencer'] or 0) + (alertas['tecnomecanica_por_vencer'] or 0)
        
        return jsonify({
            'total_vehiculos': estadisticas['total_vehiculos'] or 0,
            'vehiculos_activos': estadisticas['vehiculos_activos'] or 0,
            'vencimientos_proximos': vencimientos_proximos,
            'documentos_vencidos': documentos_vencidos,
            'alertas_recientes': alertas_recientes,
            'tipos_vehiculo': tipos_vehiculo
        })
        
    except Exception as e:
        print(f"Error en API dashboard vehículos: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Error al obtener estadísticas: {str(e)}'
        }), 500

@app.route('/api/vehiculos/graficos', methods=['GET'])
def api_vehiculos_graficos():
    # Verificar autenticación
    if 'user_id' not in session:
        return jsonify({
            'success': False,
            'message': 'Autenticación requerida'
        }), 401
    
    # Verificar rol
    user_role = session.get('user_role')
    if user_role not in ['logistica', 'administrativo']:
        return jsonify({
            'success': False,
            'message': 'Permisos insuficientes'
        }), 403
    """Obtener datos para gráficos del módulo de vehículos"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'success': False,
                'message': 'MySQL Connection not available'
            }), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Datos para gráfico de tipos de vehículo
        cursor.execute("""
            SELECT tipo_vehiculo, COUNT(*) as cantidad
            FROM parque_automotor
            WHERE estado = 'Activo'
            GROUP BY tipo_vehiculo
            ORDER BY cantidad DESC
        """)
        tipo_vehiculo = cursor.fetchall()
        
        # Datos para gráfico de estado de documentos
        cursor.execute("""
            SELECT 
                'SOAT Vigente' as estado,
                COUNT(*) as cantidad
            FROM parque_automotor 
            WHERE estado = 'Activo' AND soat_vencimiento > CURDATE()
            UNION ALL
            SELECT 
                'SOAT Vencido' as estado,
                COUNT(*) as cantidad
            FROM parque_automotor 
            WHERE estado = 'Activo' AND soat_vencimiento <= CURDATE()
            UNION ALL
            SELECT 
                'Tecnomecánica Vigente' as estado,
                COUNT(*) as cantidad
            FROM parque_automotor 
            WHERE estado = 'Activo' AND tecnomecanica_vencimiento > CURDATE()
            UNION ALL
            SELECT 
                'Tecnomecánica Vencida' as estado,
                COUNT(*) as cantidad
            FROM parque_automotor 
            WHERE estado = 'Activo' AND tecnomecanica_vencimiento <= CURDATE()
        """)
        estado_documentos = cursor.fetchall()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'tipo_vehiculo': tipo_vehiculo,
            'estado_documentos': estado_documentos
        })
        
    except Exception as e:
        print(f"Error en API gráficos vehículos: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error al obtener datos de gráficos: {str(e)}'
        }), 500

@app.route('/api/vehiculos/plantilla', methods=['GET'])
@login_required()
@role_required('logistica')
def api_descargar_plantilla_vehiculos():
    """Generar y descargar plantilla Excel para importar vehículos"""
    try:
        import pandas as pd
        import io
        from flask import send_file
        
        # Definir las columnas de la plantilla
        columns = [
            'placa',
            'tipo_vehiculo', 
            'marca',
            'modelo',
            'color',
            'soat_vencimiento',
            'tecnomecanica_vencimiento',
            'observaciones',
            'kilometraje_actual',
            'proximo_mantenimiento_km'
        ]
        
        # Crear DataFrame con ejemplos
        data = {
            'placa': ['ABC123', 'XYZ789'],
            'tipo_vehiculo': ['Moto', 'Carro'],
            'marca': ['Honda', 'Toyota'],
            'modelo': ['CB150', 'Corolla'],
            'color': ['Rojo', 'Blanco'],
            'soat_vencimiento': ['2024-12-31', '2024-11-30'],
            'tecnomecanica_vencimiento': ['2024-10-15', '2024-09-20'],
            'observaciones': ['Ejemplo de observación', ''],
            'kilometraje_actual': [15000, 25000],
            'proximo_mantenimiento_km': [20000, 30000]
        }
        
        df = pd.DataFrame(data)
        
        # Crear archivo Excel en memoria
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Escribir datos de ejemplo
            df.to_excel(writer, sheet_name='Vehículos', index=False)
            
            # Crear hoja de instrucciones
            instrucciones = pd.DataFrame({
                'INSTRUCCIONES PARA IMPORTAR VEHÍCULOS': [
                    '1. Complete todas las columnas obligatorias marcadas con *',
                    '2. Las fechas deben estar en formato YYYY-MM-DD (ej: 2024-12-31)',
                    '3. Los tipos de vehículo válidos son: Moto, Carro, Camioneta, Camión',
                    '4. La placa debe ser única en el sistema',
                    '5. Elimine las filas de ejemplo antes de importar',
                    '6. Guarde el archivo en formato Excel (.xlsx)',
                    '',
                    'COLUMNAS OBLIGATORIAS:',
                    '- placa *',
                    '- tipo_vehiculo *', 
                    '- marca *',
                    '- modelo *',
                    '- color *',
                    '',
                    'COLUMNAS OPCIONALES:',
                    '- soat_vencimiento (formato: YYYY-MM-DD)',
                    '- tecnomecanica_vencimiento (formato: YYYY-MM-DD)',
                    '- observaciones',
                    '- kilometraje_actual (número)',
                    '- proximo_mantenimiento_km (número)'
                ]
            })
            instrucciones.to_excel(writer, sheet_name='Instrucciones', index=False)
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='plantilla_vehiculos.xlsx'
        )
        
    except ImportError:
        return jsonify({
            'success': False,
            'message': 'La librería pandas no está instalada. Contacte al administrador.'
        }), 500
        
    except Exception as e:
        print(f"Error al generar plantilla: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error al generar plantilla: {str(e)}'
        }), 500

@app.route('/api/vehiculos/importar', methods=['POST'])
@login_required()
@role_required('logistica')
def api_importar_vehiculos():
    """Importar vehículos desde archivo Excel"""
    try:
        if 'archivo' not in request.files:
            return jsonify({
                'success': False,
                'message': 'No se encontró el archivo'
            }), 400
            
        archivo = request.files['archivo']
        
        if archivo.filename == '':
            return jsonify({
                'success': False,
                'message': 'No se seleccionó ningún archivo'
            }), 400
            
        if not archivo.filename.endswith(('.xlsx', '.xls')):
            return jsonify({
                'success': False,
                'message': 'El archivo debe ser de formato Excel (.xlsx o .xls)'
            }), 400
            
        try:
            import pandas as pd
            import io
            
            # Leer el archivo Excel
            df = pd.read_excel(io.BytesIO(archivo.read()))
            
            # Validar columnas requeridas
            required_columns = ['placa', 'tipo_vehiculo', 'marca', 'modelo', 'color']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                return jsonify({
                    'success': False,
                    'message': f'Faltan las siguientes columnas: {", ".join(missing_columns)}'
                }), 400
                
            connection = get_db_connection()
            if connection is None:
                return jsonify({
                    'success': False,
                    'message': 'Error de conexión a la base de datos'
                }), 500
                
            cursor = connection.cursor(dictionary=True)
            
            vehiculos_insertados = 0
            vehiculos_error = []
            
            for index, row in df.iterrows():
                try:
                    # Verificar si la placa ya existe
                    cursor.execute('SELECT placa FROM parque_automotor WHERE placa = %s', (row['placa'],))
                    if cursor.fetchone():
                        vehiculos_error.append({
                            'fila': index + 2,  # +2 porque Excel empieza en 1 y hay header
                            'placa': row['placa'],
                            'error': 'La placa ya existe'
                        })
                        continue
                        
                    # Insertar vehículo
                    cursor.execute("""
                        INSERT INTO parque_automotor (
                            placa, tipo_vehiculo, marca, modelo, color, 
                            id_codigo_consumidor, fecha_asignacion, estado,
                            soat_vencimiento, tecnomecanica_vencimiento, observaciones,
                            kilometraje_actual, proximo_mantenimiento_km
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        row['placa'],
                        row['tipo_vehiculo'],
                        row['marca'],
                        row['modelo'],
                        row['color'],
                        row.get('id_codigo_consumidor'),
                        row.get('fecha_asignacion'),
                        row.get('estado', 'Activo'),
                        row.get('soat_vencimiento'),
                        row.get('tecnomecanica_vencimiento'),
                        row.get('observaciones'),
                        row.get('kilometraje_actual', 0),
                        row.get('proximo_mantenimiento_km')
                    ))
                    
                    vehiculos_insertados += 1
                    
                except Exception as e:
                    vehiculos_error.append({
                        'fila': index + 2,
                        'placa': row.get('placa', 'N/A'),
                        'error': str(e)
                    })
                    
            connection.commit()
            cursor.close()
            connection.close()
            
            return jsonify({
                'success': True,
                'message': f'Importación completada. {vehiculos_insertados} vehículos insertados.',
                'data': {
                    'vehiculos_insertados': vehiculos_insertados,
                    'vehiculos_error': vehiculos_error
                }
            })
            
        except ImportError:
            return jsonify({
                'success': False,
                'message': 'La librería pandas no está instalada. Contacte al administrador.'
            }), 500
            
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'Error al procesar el archivo: {str(e)}'
            }), 500
            
    except Exception as e:
        print(f"Error en API importar vehículos: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error al importar vehículos: {str(e)}'
        }), 500

@app.route('/api/vehiculos/alertas', methods=['GET'])
@login_required()
@role_required('logistica')
def api_alertas_vencimiento():
    """Obtener alertas de vencimiento de documentos"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'success': False,
                'message': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Parámetros de filtro
        tipo_alerta = request.args.get('tipo')  # 'soat', 'tecnomecanica', 'all'
        estado_alerta = request.args.get('estado')  # 'vencido', 'por_vencer', 'all'
        dias_anticipacion = int(request.args.get('dias', 30))
        
        # Construir consulta base
        query = """
            SELECT 
                pa.id_parque_automotor,
                pa.placa,
                pa.tipo_vehiculo,
                pa.marca,
                pa.modelo,
                r.nombre as tecnico_nombre,
                pa.soat_vencimiento,
                pa.tecnomecanica_vencimiento,
                CASE 
                    WHEN pa.soat_vencimiento <= CURDATE() THEN 'vencido'
                    WHEN pa.soat_vencimiento <= DATE_ADD(CURDATE(), INTERVAL %s DAY) THEN 'por_vencer'
                    ELSE 'vigente'
                END as estado_soat,
                CASE 
                    WHEN pa.tecnomecanica_vencimiento <= CURDATE() THEN 'vencido'
                    WHEN pa.tecnomecanica_vencimiento <= DATE_ADD(CURDATE(), INTERVAL %s DAY) THEN 'por_vencer'
                    ELSE 'vigente'
                END as estado_tecnomecanica,
                DATEDIFF(pa.soat_vencimiento, CURDATE()) as dias_soat,
                DATEDIFF(pa.tecnomecanica_vencimiento, CURDATE()) as dias_tecnomecanica
            FROM parque_automotor pa
            LEFT JOIN recurso_operativo r ON pa.id_codigo_consumidor = r.id_codigo_consumidor
            WHERE pa.estado = 'Activo'
        """
        
        params = [dias_anticipacion, dias_anticipacion]
        
        # Filtros adicionales
        conditions = []
        
        if tipo_alerta == 'soat':
            conditions.append("(pa.soat_vencimiento <= DATE_ADD(CURDATE(), INTERVAL %s DAY))")
            params.append(dias_anticipacion)
        elif tipo_alerta == 'tecnomecanica':
            conditions.append("(pa.tecnomecanica_vencimiento <= DATE_ADD(CURDATE(), INTERVAL %s DAY))")
            params.append(dias_anticipacion)
        else:  # 'all' o no especificado
            conditions.append("(pa.soat_vencimiento <= DATE_ADD(CURDATE(), INTERVAL %s DAY) OR pa.tecnomecanica_vencimiento <= DATE_ADD(CURDATE(), INTERVAL %s DAY))")
            params.extend([dias_anticipacion, dias_anticipacion])
            
        if estado_alerta == 'vencido':
            conditions.append("(pa.soat_vencimiento <= CURDATE() OR pa.tecnomecanica_vencimiento <= CURDATE())")
        elif estado_alerta == 'por_vencer':
            conditions.append("(pa.soat_vencimiento > CURDATE() AND pa.soat_vencimiento <= DATE_ADD(CURDATE(), INTERVAL %s DAY)) OR (pa.tecnomecanica_vencimiento > CURDATE() AND pa.tecnomecanica_vencimiento <= DATE_ADD(CURDATE(), INTERVAL %s DAY))")
            params.extend([dias_anticipacion, dias_anticipacion])
            
        if conditions:
            query += " AND (" + " AND ".join(conditions) + ")"
            
        query += " ORDER BY pa.soat_vencimiento ASC, pa.tecnomecanica_vencimiento ASC"
        
        cursor.execute(query, params)
        alertas = cursor.fetchall()
        
        # Convertir fechas para serialización JSON y estructurar alertas
        alertas_estructuradas = []
        
        for vehiculo in alertas:
            # Convertir fechas
            if vehiculo['soat_vencimiento']:
                vehiculo['soat_vencimiento'] = vehiculo['soat_vencimiento'].isoformat()
            if vehiculo['tecnomecanica_vencimiento']:
                vehiculo['tecnomecanica_vencimiento'] = vehiculo['tecnomecanica_vencimiento'].isoformat()
                
            # Crear alertas individuales para SOAT y Tecnomecánica
            if (tipo_alerta in ['soat', 'all'] and 
                vehiculo['estado_soat'] in ['vencido', 'por_vencer']):
                alertas_estructuradas.append({
                    'id_vehiculo': vehiculo['id_parque_automotor'],
                    'placa': vehiculo['placa'],
                    'tipo_vehiculo': vehiculo['tipo_vehiculo'],
                    'marca': vehiculo['marca'],
                    'modelo': vehiculo['modelo'],
                    'tecnico_nombre': vehiculo['tecnico_nombre'],
                    'tipo_documento': 'SOAT',
                    'fecha_vencimiento': vehiculo['soat_vencimiento'],
                    'estado': vehiculo['estado_soat'],
                    'dias_restantes': vehiculo['dias_soat']
                })
                
            if (tipo_alerta in ['tecnomecanica', 'all'] and 
                vehiculo['estado_tecnomecanica'] in ['vencido', 'por_vencer']):
                alertas_estructuradas.append({
                    'id_vehiculo': vehiculo['id_parque_automotor'],
                    'placa': vehiculo['placa'],
                    'tipo_vehiculo': vehiculo['tipo_vehiculo'],
                    'marca': vehiculo['marca'],
                    'modelo': vehiculo['modelo'],
                    'tecnico_nombre': vehiculo['tecnico_nombre'],
                    'tipo_documento': 'Tecnomecánica',
                    'fecha_vencimiento': vehiculo['tecnomecanica_vencimiento'],
                    'estado': vehiculo['estado_tecnomecanica'],
                    'dias_restantes': vehiculo['dias_tecnomecanica']
                })
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'data': alertas_estructuradas,
            'total': len(alertas_estructuradas)
        })
        
    except Exception as e:
        print(f"Error en API alertas vencimiento: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error al obtener alertas: {str(e)}'
        }), 500

@app.route('/api/vehiculos/vencimientos', methods=['GET'])
@login_required()
@role_required('logistica')
def api_vencimientos_vehiculos():
    """Obtener datos de vencimientos para la tabla de vencimientos"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'success': False,
                'message': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Parámetros de filtro
        dias_anticipacion = int(request.args.get('dias', 30))
        tipo_documento = request.args.get('tipo')  # 'soat', 'tecnomecanica', 'all'
        
        # Consulta para obtener vencimientos
        query = """
            SELECT 
                pa.id_parque_automotor,
                pa.placa,
                pa.tipo_vehiculo,
                pa.marca,
                pa.modelo,
                r.nombre as tecnico_nombre,
                pa.soat_vencimiento,
                pa.tecnomecanica_vencimiento,
                DATEDIFF(pa.soat_vencimiento, CURDATE()) as dias_soat,
                DATEDIFF(pa.tecnomecanica_vencimiento, CURDATE()) as dias_tecnomecanica
            FROM parque_automotor pa
            LEFT JOIN recurso_operativo r ON pa.id_codigo_consumidor = r.id_codigo_consumidor
            WHERE pa.estado = 'Activo'
            AND (
                pa.soat_vencimiento <= DATE_ADD(CURDATE(), INTERVAL %s DAY)
                OR pa.tecnomecanica_vencimiento <= DATE_ADD(CURDATE(), INTERVAL %s DAY)
            )
            ORDER BY 
                LEAST(
                    COALESCE(pa.soat_vencimiento, '9999-12-31'),
                    COALESCE(pa.tecnomecanica_vencimiento, '9999-12-31')
                ) ASC
        """
        
        cursor.execute(query, [dias_anticipacion, dias_anticipacion])
        vehiculos = cursor.fetchall()
        
        # Estructurar datos para la tabla de vencimientos
        vencimientos_data = []
        
        for vehiculo in vehiculos:
            # Procesar SOAT
            if (vehiculo['soat_vencimiento'] and 
                vehiculo['dias_soat'] <= dias_anticipacion):
                
                # Determinar estado del documento
                if vehiculo['dias_soat'] < 0:
                    estado = 'vencido'
                elif vehiculo['dias_soat'] <= 7:
                    estado = 'critico'
                elif vehiculo['dias_soat'] <= 30:
                    estado = 'proximo'
                else:
                    estado = 'vigente'
                
                vencimientos_data.append({
                    'placa': vehiculo['placa'],
                    'tipo_vehiculo': vehiculo['tipo_vehiculo'],
                    'tipo_documento': 'SOAT',
                    'fecha_vencimiento': vehiculo['soat_vencimiento'].isoformat() if vehiculo['soat_vencimiento'] else None,
                    'fecha_vencimiento_nueva': vehiculo['soat_vencimiento'].strftime('%Y-%m-%d') if vehiculo['soat_vencimiento'] else None,
                    'dias_restantes': vehiculo['dias_soat'],
                    'estado_documento': estado,
                    'tecnico_nombre': vehiculo['tecnico_nombre'],
                    'fecha_renovacion': None  # Se puede obtener del historial si existe
                })
            
            # Procesar Tecnomecánica
            if (vehiculo['tecnomecanica_vencimiento'] and 
                vehiculo['dias_tecnomecanica'] <= dias_anticipacion):
                
                # Determinar estado del documento
                if vehiculo['dias_tecnomecanica'] < 0:
                    estado = 'vencido'
                elif vehiculo['dias_tecnomecanica'] <= 7:
                    estado = 'critico'
                elif vehiculo['dias_tecnomecanica'] <= 30:
                    estado = 'proximo'
                else:
                    estado = 'vigente'
                
                vencimientos_data.append({
                    'placa': vehiculo['placa'],
                    'tipo_vehiculo': vehiculo['tipo_vehiculo'],
                    'tipo_documento': 'Tecnomecánica',
                    'fecha_vencimiento': vehiculo['tecnomecanica_vencimiento'].isoformat() if vehiculo['tecnomecanica_vencimiento'] else None,
                    'fecha_vencimiento_nueva': vehiculo['tecnomecanica_vencimiento'].strftime('%Y-%m-%d') if vehiculo['tecnomecanica_vencimiento'] else None,
                    'dias_restantes': vehiculo['dias_tecnomecanica'],
                    'estado_documento': estado,
                    'tecnico_nombre': vehiculo['tecnico_nombre'],
                    'fecha_renovacion': None  # Se puede obtener del historial si existe
                })
        
        # Filtrar por tipo de documento si se especifica
        if tipo_documento and tipo_documento != 'all':
            if tipo_documento == 'soat':
                vencimientos_data = [v for v in vencimientos_data if v['tipo_documento'] == 'SOAT']
            elif tipo_documento == 'tecnomecanica':
                vencimientos_data = [v for v in vencimientos_data if v['tipo_documento'] == 'Tecnomecánica']
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'data': vencimientos_data,
            'total': len(vencimientos_data)
        })
        
    except Exception as e:
        print(f"Error en API vencimientos: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error al obtener vencimientos: {str(e)}'
        }), 500

@app.route('/api/vehiculos/reportes', methods=['GET'])
@login_required()
@role_required('logistica')
def api_reportes_vehiculos():
    """Generar reportes de vehículos con filtros"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'success': False,
                'message': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Parámetros de filtro
        tipo_reporte = request.args.get('tipo', 'general')  # 'general', 'vencimientos', 'mantenimiento'
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        estado = request.args.get('estado')
        tipo_vehiculo = request.args.get('tipo_vehiculo')
        
        if tipo_reporte == 'general':
            # Reporte general de vehículos
            query = """
                SELECT 
                    pa.*,
                    r.nombre as tecnico_nombre,
                    r.recurso_operativo_cedula,
                    CASE 
                        WHEN pa.soat_vencimiento <= CURDATE() THEN 'Vencido'
                        WHEN pa.soat_vencimiento <= DATE_ADD(CURDATE(), INTERVAL 30 DAY) THEN 'Por vencer'
                        ELSE 'Vigente'
                    END as estado_soat,
                    CASE 
                        WHEN pa.tecnomecanica_vencimiento <= CURDATE() THEN 'Vencido'
                        WHEN pa.tecnomecanica_vencimiento <= DATE_ADD(CURDATE(), INTERVAL 30 DAY) THEN 'Por vencer'
                        ELSE 'Vigente'
                    END as estado_tecnomecanica
                FROM parque_automotor pa
                LEFT JOIN recurso_operativo r ON pa.id_codigo_consumidor = r.id_codigo_consumidor
                WHERE 1=1
            """
            
        elif tipo_reporte == 'vencimientos':
            # Reporte de vencimientos
            query = """
                SELECT 
                    pa.placa,
                    pa.tipo_vehiculo,
                    pa.marca,
                    pa.modelo,
                    r.nombre as tecnico_nombre,
                    pa.soat_vencimiento,
                    pa.tecnomecanica_vencimiento,
                    DATEDIFF(pa.soat_vencimiento, CURDATE()) as dias_soat,
                    DATEDIFF(pa.tecnomecanica_vencimiento, CURDATE()) as dias_tecnomecanica,
                    CASE 
                        WHEN pa.soat_vencimiento <= CURDATE() THEN 'Vencido'
                        WHEN pa.soat_vencimiento <= DATE_ADD(CURDATE(), INTERVAL 30 DAY) THEN 'Por vencer'
                        ELSE 'Vigente'
                    END as estado_soat,
                    CASE 
                        WHEN pa.tecnomecanica_vencimiento <= CURDATE() THEN 'Vencido'
                        WHEN pa.tecnomecanica_vencimiento <= DATE_ADD(CURDATE(), INTERVAL 30 DAY) THEN 'Por vencer'
                        ELSE 'Vigente'
                    END as estado_tecnomecanica
                FROM parque_automotor pa
                LEFT JOIN recurso_operativo r ON pa.id_codigo_consumidor = r.id_codigo_consumidor
                WHERE pa.estado = 'Activo'
                AND (pa.soat_vencimiento <= DATE_ADD(CURDATE(), INTERVAL 60 DAY) 
                     OR pa.tecnomecanica_vencimiento <= DATE_ADD(CURDATE(), INTERVAL 60 DAY))
            """
            
        else:  # mantenimiento
            # Reporte de mantenimiento
            query = """
                SELECT 
                    pa.placa,
                    pa.tipo_vehiculo,
                    pa.marca,
                    pa.modelo,
                    pa.kilometraje_actual,
                    pa.proximo_mantenimiento_km,
                    r.nombre as tecnico_nombre,
                    CASE 
                        WHEN pa.proximo_mantenimiento_km IS NOT NULL AND pa.kilometraje_actual >= pa.proximo_mantenimiento_km THEN 'Mantenimiento requerido'
                        WHEN pa.proximo_mantenimiento_km IS NOT NULL AND (pa.proximo_mantenimiento_km - pa.kilometraje_actual) <= 1000 THEN 'Próximo mantenimiento'
                        ELSE 'Al día'
                    END as estado_mantenimiento
                FROM parque_automotor pa
                LEFT JOIN recurso_operativo r ON pa.id_codigo_consumidor = r.id_codigo_consumidor
                WHERE pa.estado = 'Activo'
            """
            
        # Aplicar filtros adicionales
        conditions = []
        params = []
        
        if fecha_inicio and fecha_fin:
            conditions.append("pa.fecha_asignacion BETWEEN %s AND %s")
            params.extend([fecha_inicio, fecha_fin])
            
        if estado:
            conditions.append("pa.estado = %s")
            params.append(estado)
            
        if tipo_vehiculo:
            conditions.append("pa.tipo_vehiculo = %s")
            params.append(tipo_vehiculo)
            
        if conditions:
            query += " AND " + " AND ".join(conditions)
            
        query += " ORDER BY pa.placa"
        
        cursor.execute(query, params)
        datos = cursor.fetchall()
        
        # Convertir fechas para serialización JSON
        for item in datos:
            for key, value in item.items():
                if isinstance(value, date):
                    item[key] = value.isoformat()
                    
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'data': datos,
            'tipo_reporte': tipo_reporte,
            'total_registros': len(datos)
        })
        
    except Exception as e:
        print(f"Error en API reportes vehículos: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error al generar reporte: {str(e)}'
        }), 500

@app.route('/api/indicadores/estado_vehiculos')
@login_required(role='administrativo')
def obtener_estado_vehiculos():
    """Obtener estado de vehículos por supervisor basado en datos reales de la tabla preoperacional"""
    try:
        connection = get_db_connection()
        if connection is None:
            print("Error: No se pudo establecer conexión a la base de datos")
            return jsonify({
                'success': False,
                'error': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener parámetro de mes desde la URL (formato YYYY-MM)
        mes_param = request.args.get('mes')
        
        if mes_param:
            try:
                # Validar formato YYYY-MM
                año, mes = mes_param.split('-')
                año = int(año)
                mes = int(mes)
                
                if mes < 1 or mes > 12:
                    raise ValueError("Mes inválido")
                    
                primer_dia_mes = datetime(año, mes, 1).date()
                
                # Calcular último día del mes
                if mes == 12:
                    ultimo_dia_mes = datetime(año + 1, 1, 1).date() - timedelta(days=1)
                else:
                    ultimo_dia_mes = datetime(año, mes + 1, 1).date() - timedelta(days=1)
                    
            except (ValueError, IndexError):
                return jsonify({
                    'success': False,
                    'error': 'Formato de mes inválido. Use YYYY-MM'
                }), 400
        else:
            # Si no se especifica mes, usar el mes actual
            fecha_actual = get_bogota_datetime().date()
            primer_dia_mes = fecha_actual.replace(day=1)
            # Calcular último día del mes actual
            if fecha_actual.month == 12:
                ultimo_dia_mes = datetime(fecha_actual.year + 1, 1, 1).date() - timedelta(days=1)
            else:
                ultimo_dia_mes = datetime(fecha_actual.year, fecha_actual.month + 1, 1).date() - timedelta(days=1)
        
        print(f"Consultando estado de vehículos desde: {primer_dia_mes} hasta: {ultimo_dia_mes}")
        
        # Consultar preoperacionales del mes seleccionado con supervisores
        query = """
            SELECT 
                supervisor,
                estado_espejos,
                bocina_pito,
                frenos,
                encendido,
                estado_bateria,
                estado_amortiguadores,
                estado_llantas,
                luces_altas_bajas,
                direccionales_delanteras_traseras,
                estado_fisico_vehiculo_espejos,
                estado_fisico_vehiculo_bocina_pito,
                estado_fisico_vehiculo_frenos,
                estado_fisico_vehiculo_encendido,
                estado_fisico_vehiculo_bateria,
                estado_fisico_vehiculo_amortiguadores,
                estado_fisico_vehiculo_llantas,
                estado_fisico_vehiculo_luces_altas,
                estado_fisico_vehiculo_luces_bajas,
                estado_fisico_vehiculo_direccionales_delanteras,
                estado_fisico_vehiculo_direccionales_traseras
            FROM preoperacional 
            WHERE DATE(fecha) >= %s AND DATE(fecha) <= %s
                AND supervisor IS NOT NULL 
                AND supervisor != ''
        """
        
        cursor.execute(query, (primer_dia_mes, ultimo_dia_mes))
        preoperacionales = cursor.fetchall()
        
        print(f"Se encontraron {len(preoperacionales)} registros preoperacionales")
        
        # Agrupar por supervisor y clasificar vehículos
        supervisores_stats = {}
        
        for preop in preoperacionales:
            supervisor = preop['supervisor']
            
            if supervisor not in supervisores_stats:
                supervisores_stats[supervisor] = {
                    'bueno': 0,
                    'regular': 0,
                    'malo': 0,
                    'total': 0
                }
            
            # Campos a evaluar para determinar el estado del vehículo
            campos_estado = [
                'estado_espejos', 'bocina_pito', 'frenos', 'encendido', 'estado_bateria',
                'estado_amortiguadores', 'estado_llantas', 'luces_altas_bajas', 'direccionales_delanteras_traseras',
                'estado_fisico_vehiculo_espejos', 'estado_fisico_vehiculo_bocina_pito', 'estado_fisico_vehiculo_frenos',
                'estado_fisico_vehiculo_encendido', 'estado_fisico_vehiculo_bateria', 'estado_fisico_vehiculo_amortiguadores',
                'estado_fisico_vehiculo_llantas', 'estado_fisico_vehiculo_luces_altas', 'estado_fisico_vehiculo_luces_bajas',
                'estado_fisico_vehiculo_direccionales_delanteras', 'estado_fisico_vehiculo_direccionales_traseras'
            ]
            
            # Contar estados buenos, regulares y malos
            buenos = 0
            regulares = 0
            malos = 0
            total_campos = 0
            
            for campo in campos_estado:
                valor = preop.get(campo)
                if valor is not None and valor != '':
                    total_campos += 1
                    valor_lower = str(valor).lower().strip()
                    
                    if valor_lower in ['bueno', 'good', 'si', 'sí', 'yes', '1', 'ok']:
                        buenos += 1
                    elif valor_lower in ['regular', 'medio', 'parcial', 'partial']:
                        regulares += 1
                    elif valor_lower in ['malo', 'bad', 'no', '0', 'dañado', 'defectuoso']:
                        malos += 1
            
            # Clasificar el vehículo basado en el porcentaje de componentes en buen estado
            if total_campos > 0:
                porcentaje_bueno = (buenos / total_campos) * 100
                porcentaje_malo = (malos / total_campos) * 100
                
                if porcentaje_bueno >= 80:  # 80% o más componentes buenos
                    supervisores_stats[supervisor]['bueno'] += 1
                elif porcentaje_malo >= 30:  # 30% o más componentes malos
                    supervisores_stats[supervisor]['malo'] += 1
                else:  # Estado intermedio
                    supervisores_stats[supervisor]['regular'] += 1
            else:
                # Si no hay datos, clasificar como regular
                supervisores_stats[supervisor]['regular'] += 1
            
            supervisores_stats[supervisor]['total'] += 1
        
        # Convertir a formato de respuesta
        estadisticas = []
        for supervisor, stats in supervisores_stats.items():
            estadisticas.append({
                'supervisor': supervisor,
                'bueno': stats['bueno'],
                'regular': stats['regular'],
                'malo': stats['malo'],
                'total': stats['total']
            })
        
        # Ordenar por total descendente
        estadisticas.sort(key=lambda x: x['total'], reverse=True)
        
        print(f"Se calcularon estadísticas para {len(estadisticas)} supervisores")
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'estadisticas': estadisticas,
            'periodo': {
                'mes': primer_dia_mes.strftime('%Y-%m'),
                'desde': primer_dia_mes.strftime('%Y-%m-%d'),
                'hasta': ultimo_dia_mes.strftime('%Y-%m-%d')
            }
        })
        
    except Exception as e:
        import traceback
        print(f"Error en estado de vehículos: {str(e)}")
        print(f"Traceback: {traceback.format_exc()}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# APIs para Presupuesto
@app.route('/api/lider/presupuesto', methods=['GET'])
@login_required_lider_api()
def api_lider_presupuesto():
    """
    API para obtener datos de presupuesto mensual por supervisor.
    Devuelve estructura compatible con el frontend:
    {
        success: true,
        presupuestos: [{supervisor, mes, ano, presupuesto_mensual, registros}],
        resumen: {total, promedio, supervisores, mayor},
        total_registros: N
    }
    """
    connection = None
    cursor = None
    try:
        # Obtener parámetros de filtro opcionales
        mes = request.args.get('mes')
        # Aceptar 'ano' y también 'año' por compatibilidad
        ano = request.args.get('ano') or request.args.get('año')
        supervisor = request.args.get('supervisor')

        # Convertir tipos cuando sea posible
        try:
            mes = int(mes) if mes is not None and mes != '' else None
        except Exception:
            mes = None
        try:
            ano = int(ano) if ano is not None and ano != '' else None
        except Exception:
            ano = None

        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500

        cursor = connection.cursor(dictionary=True)

        # Construir consulta SQL base
        sql = """
            SELECT 
                super AS supervisor,
                MONTH(fecha_asistencia) AS mes,
                YEAR(fecha_asistencia) AS ano,
                SUM(valor) AS presupuesto_mensual,
                COUNT(*) AS registros
            FROM asistencia 
            WHERE super IS NOT NULL AND super != ''
        """

        params = []

        # Agregar filtros opcionales
        if mes is not None:
            sql += " AND MONTH(fecha_asistencia) = %s"
            params.append(mes)

        if ano is not None:
            sql += " AND YEAR(fecha_asistencia) = %s"
            params.append(ano)

        if supervisor:
            sql += " AND super = %s"
            params.append(supervisor)

        # Agrupar por supervisor, mes y año
        sql += """
            GROUP BY super, MONTH(fecha_asistencia), YEAR(fecha_asistencia)
            ORDER BY super, ano DESC, mes DESC
        """

        cursor.execute(sql, params)
        resultados = cursor.fetchall()

        # Formatear resultados y calcular resumen
        presupuestos = []
        total = 0.0
        mayor = 0.0
        supervisores_set = set()

        for row in resultados:
            presupuesto_mensual = float(row['presupuesto_mensual']) if row.get('presupuesto_mensual') is not None else 0.0
            presupuestos.append({
                'supervisor': row.get('supervisor') or '',
                'mes': row.get('mes') or 0,
                'ano': row.get('ano') or 0,
                'presupuesto_mensual': presupuesto_mensual,
                'registros': int(row.get('registros') or 0)
            })
            total += presupuesto_mensual
            mayor = max(mayor, presupuesto_mensual)
            if row.get('supervisor'):
                supervisores_set.add(row['supervisor'])

        total_supervisores = len(supervisores_set)
        promedio = total / total_supervisores if total_supervisores > 0 else 0.0

        return jsonify({
            'success': True,
            'presupuestos': presupuestos,
            'resumen': {
                'total': total,
                'promedio': promedio,
                'supervisores': total_supervisores,
                'mayor': mayor
            },
            'total_registros': len(presupuestos)
        })

    except mysql.connector.Error as e:
        print(f"ERROR DB - API Presupuesto: {str(e)}")
        return jsonify({'success': False, 'message': f'Error de base de datos: {str(e)}'}), 500
    except Exception as e:
        print(f"ERROR - API Presupuesto: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error interno: {str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/lider/presupuesto/supervisores', methods=['GET'])
@login_required_lider_api()
def api_lider_presupuesto_supervisores():
    """
    API para obtener lista de supervisores únicos de la tabla asistencia
    """
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener supervisores únicos
        cursor.execute("""
            SELECT DISTINCT super as supervisor
            FROM asistencia 
            WHERE super IS NOT NULL AND super != ''
            ORDER BY super
        """)
        
        resultados = cursor.fetchall()
        
        # Extraer solo los nombres de supervisores
        supervisores = [row['supervisor'] for row in resultados]
        
        return jsonify({
            'success': True,
            'supervisores': supervisores,
            'total': len(supervisores)
        })
        
    except mysql.connector.Error as e:
        print(f"ERROR DB - API Supervisores Presupuesto: {str(e)}")
        return jsonify({'error': f'Error de base de datos: {str(e)}'}), 500
    except Exception as e:
        print(f"ERROR - API Supervisores Presupuesto: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error interno: {str(e)}'}), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

# Rutas para el módulo de historial de seriales
@app.route('/logistica/historial_seriales')
@login_required()
@role_required('logistica')
def historial_seriales():
    """Página principal del módulo de historial de seriales"""
    return render_template('modulos/logistica/historial_seriales.html')

@app.route('/logistica/buscar_serial', methods=['POST'])
@login_required()
@role_required('logistica')
def buscar_serial():
    """Buscar historial de un serial específico"""
    try:
        # Manejar tanto JSON como form data
        if request.is_json:
            data = request.get_json()
            serial = data.get('serial', '').strip()
        else:
            serial = request.form.get('serial', '').strip()
        
        if not serial:
            return jsonify({
                'success': False,
                'message': 'El número de serial es requerido'
            })
        
        # Conectar a la base de datos capired
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Buscar el serial en la tabla qry
        cursor.execute("""
            SELECT id, serial, fecha, estado, ot, cuenta, codigo, descripcion
            FROM qry 
            WHERE serial = %s
            ORDER BY fecha DESC
        """, (serial,))
        
        resultados = cursor.fetchall()
        
        # Calcular estadísticas
        estadisticas = {
            'total_registros': len(resultados),
            'seriales_unicos': len(set(r['serial'] for r in resultados if r['serial'])),
            'cuentas_unicas': len(set(r['cuenta'] for r in resultados if r['cuenta'])),
            'ot_unicas': len(set(r['ot'] for r in resultados if r['ot']))
        }
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'data': resultados,
            'estadisticas': estadisticas
        })
        
    except Exception as e:
        print(f"Error al buscar serial: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error al buscar el serial: {str(e)}'
        })

@app.route('/logistica/buscar_cuenta', methods=['POST'])
@login_required()
@role_required('logistica')
def buscar_cuenta():
    """Buscar historial por cuenta y/o OT (búsqueda flexible)"""
    try:
        # Manejar tanto JSON como form data
        if request.is_json:
            data = request.get_json()
            cuenta = data.get('cuenta', '').strip()
            ot = data.get('ot', '').strip()
        else:
            cuenta = request.form.get('cuenta', '').strip()
            ot = request.form.get('ot', '').strip()
        
        # Validar que al menos uno de los campos esté presente
        if not cuenta and not ot:
            return jsonify({
                'success': False,
                'message': 'Debe proporcionar al menos una Cuenta o una OT para realizar la búsqueda'
            })
        
        # Conectar a la base de datos capired
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Construir consulta dinámica según los parámetros proporcionados
        if cuenta and ot:
            # Buscar por cuenta Y OT
            cursor.execute("""
                SELECT id, serial, fecha, estado, ot, cuenta, codigo, descripcion
                FROM qry 
                WHERE cuenta = %s AND ot = %s
                ORDER BY fecha DESC
            """, (cuenta, ot))
        elif cuenta:
            # Buscar solo por cuenta
            cursor.execute("""
                SELECT id, serial, fecha, estado, ot, cuenta, codigo, descripcion
                FROM qry 
                WHERE cuenta = %s
                ORDER BY fecha DESC
            """, (cuenta,))
        else:
            # Buscar solo por OT
            cursor.execute("""
                SELECT id, serial, fecha, estado, ot, cuenta, codigo, descripcion
                FROM qry 
                WHERE ot = %s
                ORDER BY fecha DESC
            """, (ot,))
        
        resultados = cursor.fetchall()
        
        # Calcular estadísticas
        estadisticas = {
            'total_registros': len(resultados),
            'seriales_unicos': len(set(r['serial'] for r in resultados if r['serial'])),
            'cuentas_unicas': len(set(r['cuenta'] for r in resultados if r['cuenta'])),
            'ot_unicas': len(set(r['ot'] for r in resultados if r['ot']))
        }
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'data': resultados,
            'estadisticas': estadisticas
        })
        
    except Exception as e:
        print(f"Error al buscar cuenta/OT: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error al buscar la cuenta/OT: {str(e)}'
        })

@app.route('/logistica/buscar_ot', methods=['POST'])
@login_required()
@role_required('logistica')
def buscar_ot():
    """Buscar historial por OT y/o cuenta (búsqueda flexible)"""
    try:
        # Manejar tanto JSON como form data
        if request.is_json:
            data = request.get_json()
            ot = data.get('ot', '').strip()
            cuenta = data.get('cuenta', '').strip()
        else:
            ot = request.form.get('ot', '').strip()
            cuenta = request.form.get('cuenta', '').strip()
        
        # Validar que al menos uno de los campos esté presente
        if not ot and not cuenta:
            return jsonify({
                'success': False,
                'message': 'Debe proporcionar al menos una OT o una Cuenta para realizar la búsqueda'
            })
        
        # Conectar a la base de datos capired
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Construir consulta dinámica según los parámetros proporcionados
        if ot and cuenta:
            # Buscar por OT Y cuenta
            cursor.execute("""
                SELECT id, serial, fecha, estado, ot, cuenta, codigo, descripcion
                FROM qry 
                WHERE ot = %s AND cuenta = %s
                ORDER BY fecha DESC
            """, (ot, cuenta))
        elif ot:
            # Buscar solo por OT
            cursor.execute("""
                SELECT id, serial, fecha, estado, ot, cuenta, codigo, descripcion
                FROM qry 
                WHERE ot = %s
                ORDER BY fecha DESC
            """, (ot,))
        else:
            # Buscar solo por cuenta
            cursor.execute("""
                SELECT id, serial, fecha, estado, ot, cuenta, codigo, descripcion
                FROM qry 
                WHERE cuenta = %s
                ORDER BY fecha DESC
            """, (cuenta,))
        
        resultados = cursor.fetchall()
        
        # Calcular estadísticas
        estadisticas = {
            'total_registros': len(resultados),
            'seriales_unicos': len(set(r['serial'] for r in resultados if r['serial'])),
            'cuentas_unicas': len(set(r['cuenta'] for r in resultados if r['cuenta'])),
            'ot_unicas': len(set(r['ot'] for r in resultados if r['ot']))
        }
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'data': resultados,
            'estadisticas': estadisticas
        })
        
    except Exception as e:
        print(f"Error al buscar OT/cuenta: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error al buscar la OT/cuenta: {str(e)}'
        })

@app.route('/logistica/buscar_masivo', methods=['POST'])
@login_required()
@role_required('logistica')
def buscar_masivo():
    """Buscar historial masivo desde archivo TXT"""
    try:
        if 'archivo_masivo' not in request.files:
            return jsonify({
                'success': False,
                'message': 'No se ha seleccionado ningún archivo'
            })
        
        archivo = request.files['archivo_masivo']
        tipo_busqueda = request.form.get('tipo_consulta_masiva', 'seriales')
        
        if archivo.filename == '':
            return jsonify({
                'success': False,
                'message': 'No se ha seleccionado ningún archivo'
            })
        
        # Leer el contenido del archivo
        contenido = archivo.read().decode('utf-8')
        lineas = [linea.strip() for linea in contenido.split('\n') if linea.strip()]
        
        if not lineas:
            return jsonify({
                'success': False,
                'message': 'El archivo está vacío o no contiene datos válidos'
            })
        
        # Conectar a la base de datos capired
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        resultados = []
        
        # Buscar cada elemento según el tipo
        for item in lineas:
            if tipo_busqueda == 'seriales':
                cursor.execute("""
                    SELECT id, serial, fecha, estado, ot, cuenta, codigo, descripcion
                    FROM qry 
                    WHERE serial = %s
                    ORDER BY fecha DESC
                """, (item,))
            elif tipo_busqueda == 'cuentas':
                cursor.execute("""
                    SELECT id, serial, fecha, estado, ot, cuenta, codigo, descripcion
                    FROM qry 
                    WHERE cuenta = %s
                    ORDER BY fecha DESC
                """, (item,))
            elif tipo_busqueda == 'ot':
                cursor.execute("""
                    SELECT id, serial, fecha, estado, ot, cuenta, codigo, descripcion
                    FROM qry 
                    WHERE ot = %s
                    ORDER BY fecha DESC
                """, (item,))
            
            resultados.extend(cursor.fetchall())
        
        cursor.close()
        connection.close()
        
        # Calcular estadísticas
        estadisticas = {
            'total_registros': len(resultados),
            'seriales_unicos': len(set(r['serial'] for r in resultados if r['serial'])),
            'cuentas_unicas': len(set(r['cuenta'] for r in resultados if r['cuenta'])),
            'ot_unicas': len(set(r['ot'] for r in resultados if r['ot']))
        }
        
        return jsonify({
            'success': True,
            'data': resultados,
            'estadisticas': estadisticas,
            'total': len(resultados),
            'items_procesados': len(lineas)
        })
        
    except Exception as e:
        print(f"Error en búsqueda masiva: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error en la búsqueda masiva: {str(e)}'
        })

@app.route('/logistica/cargar_qry', methods=['POST'])
@login_required()
@role_required('logistica')
def cargar_qry():
    """Cargar datos masivamente a la tabla qry desde archivo CSV/Excel"""
    connection = None
    cursor = None
    
    try:
        if 'archivo_qry' not in request.files:
            return jsonify({
                'success': False,
                'message': 'No se ha seleccionado ningún archivo'
            })
        
        archivo = request.files['archivo_qry']
        validar_datos = request.form.get('validar_datos', 'false').lower() == 'true'
        actualizar_existentes = request.form.get('actualizar_existentes', 'false').lower() == 'true'
        
        if archivo.filename == '':
            return jsonify({
                'success': False,
                'message': 'No se ha seleccionado ningún archivo'
            })
        
        # Validar extensión del archivo
        extension = archivo.filename.lower().split('.')[-1]
        if extension not in ['csv', 'xlsx', 'xls']:
            return jsonify({
                'success': False,
                'message': 'Formato de archivo no soportado. Use CSV o Excel (.xlsx, .xls)'
            })
        
        # Procesar archivo según su tipo
        datos = []
        errores = []
        
        try:
            if extension == 'csv':
                import csv
                import io
                
                # Intentar múltiples codificaciones para archivos CSV
                contenido_bytes = archivo.read()
                contenido = None
                
                # Lista de codificaciones a probar
                encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
                
                for encoding in encodings:
                    try:
                        contenido = contenido_bytes.decode(encoding)
                        break
                    except UnicodeDecodeError:
                        continue
                
                if contenido is None:
                    return jsonify({
                        'success': False,
                        'message': 'No se pudo decodificar el archivo CSV. Verifique la codificación del archivo.'
                    })
                
                reader = csv.DictReader(io.StringIO(contenido))
                datos = list(reader)
            else:
                import pandas as pd
                df = pd.read_excel(archivo)
                datos = df.to_dict('records')
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'Error al leer el archivo: {str(e)}'
            })
        
        if not datos:
            return jsonify({
                'success': False,
                'message': 'El archivo está vacío o no contiene datos válidos'
            })
        
        # Validar columnas requeridas
        columnas_requeridas = ['serial', 'fecha', 'estado', 'ot', 'cuenta', 'codigo', 'descripcion']
        primera_fila = datos[0]
        columnas_archivo = list(primera_fila.keys())
        
        # Verificar que todas las columnas requeridas estén presentes
        columnas_faltantes = [col for col in columnas_requeridas if col not in columnas_archivo]
        if columnas_faltantes:
            return jsonify({
                'success': False,
                'message': f'Faltan las siguientes columnas en el archivo: {", ".join(columnas_faltantes)}'
            })
        
        # Conectar a la base de datos usando la función centralizada
        from datetime import datetime
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'success': False,
                'message': 'Error de conexión a la base de datos. Por favor, inténtelo más tarde.'
            })
        
        cursor = connection.cursor()
        
        registros_procesados = 0
        registros_insertados = 0
        registros_actualizados = 0
        
        for i, fila in enumerate(datos, 1):
            try:
                # Extraer datos de la fila
                serial = str(fila.get('serial', '')).strip()
                fecha = fila.get('fecha', '')
                estado = str(fila.get('estado', '')).strip()
                ot = str(fila.get('ot', '')).strip()
                cuenta = str(fila.get('cuenta', '')).strip()
                codigo = str(fila.get('codigo', '')).strip()
                descripcion = str(fila.get('descripcion', '')).strip()
                
                # Validaciones básicas si está habilitado
                if validar_datos:
                    if not serial:
                        errores.append(f'Fila {i}: Serial es requerido')
                        continue
                    
                    # Validar formato de fecha si es necesario
                    if fecha:
                        try:
                            if isinstance(fecha, str):
                                # Intentar parsear diferentes formatos de fecha
                                from dateutil import parser
                                fecha = parser.parse(fecha).strftime('%Y-%m-%d')
                            elif hasattr(fecha, 'strftime'):
                                fecha = fecha.strftime('%Y-%m-%d')
                        except:
                            errores.append(f'Fila {i}: Formato de fecha inválido')
                            continue
                
                # Verificar si el registro ya existe (por serial)
                if actualizar_existentes:
                    cursor.execute("SELECT id FROM qry WHERE serial = %s", (serial,))
                    existe = cursor.fetchone()
                    
                    if existe:
                        # Actualizar registro existente
                        cursor.execute("""
                            UPDATE qry 
                            SET fecha = %s, estado = %s, ot = %s, cuenta = %s, codigo = %s, descripcion = %s
                            WHERE serial = %s
                        """, (fecha, estado, ot, cuenta, codigo, descripcion, serial))
                        registros_actualizados += 1
                    else:
                        # Insertar nuevo registro
                        cursor.execute("""
                            INSERT INTO qry (serial, fecha, estado, ot, cuenta, codigo, descripcion)
                            VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """, (serial, fecha, estado, ot, cuenta, codigo, descripcion))
                        registros_insertados += 1
                else:
                    # Solo insertar (puede generar duplicados)
                    cursor.execute("""
                        INSERT INTO qry (serial, fecha, estado, ot, cuenta, codigo, descripcion)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """, (serial, fecha, estado, ot, cuenta, codigo, descripcion))
                    registros_insertados += 1
                
                registros_procesados += 1
                
            except Exception as e:
                errores.append(f'Fila {i}: Error al procesar - {str(e)}')
                continue
        
        # Confirmar transacción
        connection.commit()
        
        # Preparar respuesta
        mensaje = f'Carga completada. Procesados: {registros_procesados}'
        if registros_insertados > 0:
            mensaje += f', Insertados: {registros_insertados}'
        if registros_actualizados > 0:
            mensaje += f', Actualizados: {registros_actualizados}'
        if errores:
            mensaje += f', Errores: {len(errores)}'
        
        # Asegurar que los mensajes de error no contengan caracteres problemáticos
        errores_limpios = []
        for error in errores[:10]:
            try:
                # Intentar codificar y decodificar para limpiar caracteres problemáticos
                error_limpio = str(error).encode('utf-8', errors='replace').decode('utf-8')
                errores_limpios.append(error_limpio)
            except:
                errores_limpios.append('Error de codificación en mensaje')
        
        return jsonify({
            'success': True,
            'message': mensaje,
            'registros_procesados': registros_procesados,
            'registros_insertados': registros_insertados,
            'registros_actualizados': registros_actualizados,
            'errores': errores_limpios
        })
        
    except mysql.connector.Error as db_error:
        if connection:
            connection.rollback()
        # Limpiar mensaje de error de base de datos
        try:
            error_msg = str(db_error).encode('utf-8', errors='replace').decode('utf-8')
        except:
            error_msg = 'Error de base de datos (problema de codificación)'
        
        return jsonify({
            'success': False,
            'message': f'Error de base de datos: {error_msg}'
        })
    
    except Exception as e:
        print(f"Error en carga de QRY: {str(e)}")
        # Limpiar mensaje de error general
        try:
            error_msg = str(e).encode('utf-8', errors='replace').decode('utf-8')
        except:
            error_msg = 'Error interno del servidor (problema de codificación)'
        
        return jsonify({
            'success': False,
            'message': f'Error interno del servidor: {error_msg}'
        })
    
    finally:
        if cursor:
            cursor.close()
        if connection:
            connection.close()

@app.route('/logistica/limites_tecnico/<int:id_codigo_consumidor>')
@login_required()
@role_required('logistica')
def obtener_limites_tecnico(id_codigo_consumidor):
    """Endpoint para obtener los límites actualizados de un técnico específico"""
    try:
        print(f"\n=== DEPURACIÓN LÍMITES TÉCNICO ===")
        print(f"ID recibido: {id_codigo_consumidor} (tipo: {type(id_codigo_consumidor)})")
        
        connection = get_db_connection()
        if connection is None:
            print("ERROR: No se pudo conectar a la base de datos")
            return jsonify({
                'success': False,
                'mensaje': 'Error de conexión a la base de datos'
            })
                               
        cursor = connection.cursor(dictionary=True)
        
        # Obtener información del técnico
        print(f"Ejecutando consulta para técnico ID: {id_codigo_consumidor}")
        cursor.execute("""
            SELECT id_codigo_consumidor, nombre, recurso_operativo_cedula, cargo, carpeta 
            FROM recurso_operativo 
            WHERE id_codigo_consumidor = %s
        """, (id_codigo_consumidor,))
        
        tecnico = cursor.fetchone()
        print(f"Resultado consulta técnico: {tecnico}")
        
        if not tecnico:
            print(f"ERROR: Técnico con ID {id_codigo_consumidor} no encontrado")
            return jsonify({
                'success': False,
                'mensaje': 'Técnico no encontrado'
            })
        
        # Definir límites según área de trabajo
        limites = {
            'FTTH INSTALACIONES': {
                'cinta_aislante': {'cantidad': 3, 'periodo': 15, 'unidad': 'días'},
                'silicona': {'cantidad': 16, 'periodo': 7, 'unidad': 'días'},
                'amarres_negros': {'cantidad': 50, 'periodo': 7, 'unidad': 'días'},
                'amarres_blancos': {'cantidad': 50, 'periodo': 7, 'unidad': 'días'},
                'grapas_blancas': {'cantidad': 100, 'periodo': 7, 'unidad': 'días'},
                'grapas_negras': {'cantidad': 100, 'periodo': 7, 'unidad': 'días'}
            },
            'INSTALACIONES DOBLES': {
                'cinta_aislante': {'cantidad': 3, 'periodo': 15, 'unidad': 'días'},
                'silicona': {'cantidad': 16, 'periodo': 7, 'unidad': 'días'},
                'amarres_negros': {'cantidad': 50, 'periodo': 7, 'unidad': 'días'},
                'amarres_blancos': {'cantidad': 50, 'periodo': 7, 'unidad': 'días'},
                'grapas_blancas': {'cantidad': 150, 'periodo': 7, 'unidad': 'días'},
                'grapas_negras': {'cantidad': 100, 'periodo': 7, 'unidad': 'días'}
            },
            'POSTVENTA': {
                'cinta_aislante': {'cantidad': 3, 'periodo': 15, 'unidad': 'días'},
                'silicona': {'cantidad': 16, 'periodo': 7, 'unidad': 'días'},
                'amarres_negros': {'cantidad': 50, 'periodo': 7, 'unidad': 'días'},
                'amarres_blancos': {'cantidad': 50, 'periodo': 7, 'unidad': 'días'},
                'grapas_blancas': {'cantidad': 100, 'periodo': 7, 'unidad': 'días'},
                'grapas_negras': {'cantidad': 100, 'periodo': 7, 'unidad': 'días'}
            },
            'MANTENIMIENTO FTTH': {
                'cinta_aislante': {'cantidad': 1, 'periodo': 15, 'unidad': 'días'},
                'silicona': {'cantidad': 8, 'periodo': 7, 'unidad': 'días'},
                'amarres_negros': {'cantidad': 50, 'periodo': 15, 'unidad': 'días'},
                'amarres_blancos': {'cantidad': 50, 'periodo': 15, 'unidad': 'días'},
                'grapas_blancas': {'cantidad': 100, 'periodo': 7, 'unidad': 'días'},
                'grapas_negras': {'cantidad': 100, 'periodo': 7, 'unidad': 'días'}
            },
            'ARREGLOS HFC': {
                'cinta_aislante': {'cantidad': 1, 'periodo': 15, 'unidad': 'días'},
                'silicona': {'cantidad': 8, 'periodo': 7, 'unidad': 'días'},
                'amarres_negros': {'cantidad': 50, 'periodo': 15, 'unidad': 'días'},
                'amarres_blancos': {'cantidad': 50, 'periodo': 15, 'unidad': 'días'},
                'grapas_blancas': {'cantidad': 100, 'periodo': 7, 'unidad': 'días'},
                'grapas_negras': {'cantidad': 100, 'periodo': 7, 'unidad': 'días'}
            },
            'CONDUCTOR': {
                'cinta_aislante': {'cantidad': 99, 'periodo': 15, 'unidad': 'días'},
                'silicona': {'cantidad': 99, 'periodo': 7, 'unidad': 'días'},
                'amarres_negros': {'cantidad': 99, 'periodo': 15, 'unidad': 'días'},
                'amarres_blancos': {'cantidad': 99, 'periodo': 15, 'unidad': 'días'},
                'grapas_blancas': {'cantidad': 99, 'periodo': 7, 'unidad': 'días'},
                'grapas_negras': {'cantidad': 99, 'periodo': 7, 'unidad': 'días'}
            },
            'SUPERVISORES': {
                'cinta_aislante': {'cantidad': 99, 'periodo': 15, 'unidad': 'días'},
                'silicona': {'cantidad': 99, 'periodo': 7, 'unidad': 'días'},
                'amarres_negros': {'cantidad': 99, 'periodo': 15, 'unidad': 'días'},
                'amarres_blancos': {'cantidad': 99, 'periodo': 15, 'unidad': 'días'},
                'grapas_blancas': {'cantidad': 99, 'periodo': 7, 'unidad': 'días'},
                'grapas_negras': {'cantidad': 99, 'periodo': 7, 'unidad': 'días'}
            },
            'BROWNFIELD': {
                'cinta_aislante': {'cantidad': 5, 'periodo': 15, 'unidad': 'días'},
                'silicona': {'cantidad': 16, 'periodo': 7, 'unidad': 'días'},
                'amarres_negros': {'cantidad': 50, 'periodo': 15, 'unidad': 'días'},
                'amarres_blancos': {'cantidad': 50, 'periodo': 15, 'unidad': 'días'},
                'grapas_blancas': {'cantidad': 200, 'periodo': 15, 'unidad': 'días'},
                'grapas_negras': {'cantidad': 200, 'periodo': 15, 'unidad': 'días'}
            },
        }
        
        # Determinar área de trabajo
        carpeta = tecnico.get('carpeta', '').upper() if tecnico.get('carpeta') else ''
        cargo = tecnico.get('cargo', '').upper()
        
        print(f"Datos del técnico:")
        print(f"  - Nombre: {tecnico.get('nombre')}")
        print(f"  - Carpeta: '{carpeta}' (original: '{tecnico.get('carpeta')}')")
        print(f"  - Cargo: '{cargo}' (original: '{tecnico.get('cargo')}')")
        
        area_trabajo = None
        
        # Primero intentar determinar por carpeta
        if carpeta:
            print(f"Buscando área por carpeta: '{carpeta}'")
            for area in limites.keys():
                print(f"  - Comparando con área: '{area}' -> {area in carpeta}")
                if area in carpeta:
                    area_trabajo = area
                    print(f"  - ¡ENCONTRADA! Área asignada: {area_trabajo}")
                    break
        
        # Si no se encontró por carpeta, intentar por cargo
        if area_trabajo is None:
            print(f"Buscando área por cargo: '{cargo}'")
            for area in limites.keys():
                print(f"  - Comparando con área: '{area}' -> {area in cargo}")
                if area in cargo:
                    area_trabajo = area
                    print(f"  - ¡ENCONTRADA! Área asignada: {area_trabajo}")
                    break
        
        # Si no se encuentra un área específica, usar límites por defecto
        if area_trabajo is None:
            print("No se encontró área específica, usando POSTVENTA por defecto")
            area_trabajo = 'POSTVENTA'
        
        print(f"Área de trabajo final: {area_trabajo}")
        
        # Obtener asignaciones previas para este técnico
        fecha_actual = datetime.now()
        print(f"\nConsultando asignaciones previas para técnico ID: {id_codigo_consumidor}")
        cursor.execute("""
            SELECT 
                fecha_asignacion,
                silicona,
                amarres_negros,
                amarres_blancos,
                cinta_aislante,
                grapas_blancas,
                grapas_negras
            FROM ferretero 
            WHERE id_codigo_consumidor = %s
            ORDER BY fecha_asignacion DESC
        """, (id_codigo_consumidor,))
        asignaciones_tecnico = cursor.fetchall()
        print(f"Asignaciones encontradas: {len(asignaciones_tecnico)}")
        
        if asignaciones_tecnico:
            print("Primeras 3 asignaciones:")
            for i, asig in enumerate(asignaciones_tecnico[:3]):
                print(f"  {i+1}. Fecha: {asig['fecha_asignacion']}, Silicona: {asig['silicona']}, Cintas: {asig['cinta_aislante']}")
        
        # Inicializar contadores para materiales en los períodos correspondientes
        contadores = {
            'cinta_aislante': 0,
            'silicona': 0,
            'amarres_negros': 0,
            'amarres_blancos': 0,
            'grapas_blancas': 0,
            'grapas_negras': 0
        }
        
        # Calcular consumo previo en los periodos correspondientes
        for asignacion in asignaciones_tecnico:
            fecha_asignacion = asignacion['fecha_asignacion']
            diferencia_dias = (fecha_actual - fecha_asignacion).days
            
            # Verificar límite de cintas
            if diferencia_dias <= limites[area_trabajo]['cinta_aislante']['periodo']:
                contadores['cinta_aislante'] += int(asignacion.get('cinta_aislante', 0) or 0)
                
            # Verificar límite de siliconas
            if diferencia_dias <= limites[area_trabajo]['silicona']['periodo']:
                contadores['silicona'] += int(asignacion.get('silicona', 0) or 0)
                
            # Verificar límite de amarres negros
            if diferencia_dias <= limites[area_trabajo]['amarres_negros']['periodo']:
                contadores['amarres_negros'] += int(asignacion.get('amarres_negros', 0) or 0)
                
            # Verificar límite de amarres blancos
            if diferencia_dias <= limites[area_trabajo]['amarres_blancos']['periodo']:
                contadores['amarres_blancos'] += int(asignacion.get('amarres_blancos', 0) or 0)
                
            # Verificar límite de grapas blancas
            if diferencia_dias <= limites[area_trabajo]['grapas_blancas']['periodo']:
                contadores['grapas_blancas'] += int(asignacion.get('grapas_blancas', 0) or 0)
                
            # Verificar límite de grapas negras
            if diferencia_dias <= limites[area_trabajo]['grapas_negras']['periodo']:
                contadores['grapas_negras'] += int(asignacion.get('grapas_negras', 0) or 0)
        
        # Calcular límites disponibles
        print(f"\nContadores finales:")
        print(f"  - Cintas consumidas: {contadores['cinta_aislante']} / {limites[area_trabajo]['cinta_aislante']['cantidad']}")
        print(f"  - Siliconas consumidas: {contadores['silicona']} / {limites[area_trabajo]['silicona']['cantidad']}")
        print(f"  - Amarres negros consumidos: {contadores['amarres_negros']} / {limites[area_trabajo]['amarres_negros']['cantidad']}")
        print(f"  - Amarres blancos consumidos: {contadores['amarres_blancos']} / {limites[area_trabajo]['amarres_blancos']['cantidad']}")
        print(f"  - Grapas blancas consumidas: {contadores['grapas_blancas']} / {limites[area_trabajo]['grapas_blancas']['cantidad']}")
        print(f"  - Grapas negras consumidas: {contadores['grapas_negras']} / {limites[area_trabajo]['grapas_negras']['cantidad']}")
        
        limites_disponibles = {
            'area': area_trabajo,
            'cinta_aislante': max(0, limites[area_trabajo]['cinta_aislante']['cantidad'] - contadores['cinta_aislante']),
            'silicona': max(0, limites[area_trabajo]['silicona']['cantidad'] - contadores['silicona']),
            'amarres_negros': max(0, limites[area_trabajo]['amarres_negros']['cantidad'] - contadores['amarres_negros']),
            'amarres_blancos': max(0, limites[area_trabajo]['amarres_blancos']['cantidad'] - contadores['amarres_blancos']),
            'grapas_blancas': max(0, limites[area_trabajo]['grapas_blancas']['cantidad'] - contadores['grapas_blancas']),
            'grapas_negras': max(0, limites[area_trabajo]['grapas_negras']['cantidad'] - contadores['grapas_negras']),
            'periodos': {
                'cinta_aislante': f"{limites[area_trabajo]['cinta_aislante']['periodo']} {limites[area_trabajo]['cinta_aislante']['unidad']}",
                'silicona': f"{limites[area_trabajo]['silicona']['periodo']} {limites[area_trabajo]['silicona']['unidad']}",
                'amarres_negros': f"{limites[area_trabajo]['amarres_negros']['periodo']} {limites[area_trabajo]['amarres_negros']['unidad']}",
                'amarres_blancos': f"{limites[area_trabajo]['amarres_blancos']['periodo']} {limites[area_trabajo]['amarres_blancos']['unidad']}",
                'grapas_blancas': f"{limites[area_trabajo]['grapas_blancas']['periodo']} {limites[area_trabajo]['grapas_blancas']['unidad']}",
                'grapas_negras': f"{limites[area_trabajo]['grapas_negras']['periodo']} {limites[area_trabajo]['grapas_negras']['unidad']}"
            }
        }
        
        print(f"\nLímites disponibles calculados:")
        print(f"  - Cintas disponibles: {limites_disponibles['cinta_aislante']}")
        print(f"  - Siliconas disponibles: {limites_disponibles['silicona']}")
        print(f"  - Amarres negros disponibles: {limites_disponibles['amarres_negros']}")
        print(f"  - Amarres blancos disponibles: {limites_disponibles['amarres_blancos']}")
        print(f"  - Grapas blancas disponibles: {limites_disponibles['grapas_blancas']}")
        print(f"  - Grapas negras disponibles: {limites_disponibles['grapas_negras']}")
        print(f"=== FIN DEPURACIÓN ===\n")
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'tecnico': tecnico,
            'limites': limites_disponibles
        })
        
    except Exception as e:
        print(f"\n=== ERROR EN LÍMITES TÉCNICO ===")
        print(f"ID técnico: {id_codigo_consumidor}")
        print(f"Error: {str(e)}")
        import traceback
        print(f"Traceback: {traceback.format_exc()}")
        print(f"=== FIN ERROR ===\n")
        return jsonify({
            'success': False,
            'mensaje': f'Error al obtener límites: {str(e)}'
        })

# Las rutas de asistencia operativa han sido eliminadas - ahora se redirige al módulo administrativo

# Importar las nuevas APIs de reportes
from api_reportes import (
    api_reportes_filtros_avanzados,
    api_reportes_generar_consolidado,
    api_reportes_validar_consistencia,
    api_reportes_configuracion,
    api_reportes_consolidacion_jerarquica,
    api_reportes_analisis_tendencias,
    api_reportes_exportar,
    api_reportes_programados,
    login_required_api
)

# Nuevas rutas API para el módulo de reportes optimizado
@app.route('/api/reportes/filtros/avanzados', methods=['GET', 'POST'])
@login_required_api()
def route_api_reportes_filtros_avanzados():
    return api_reportes_filtros_avanzados()

@app.route('/api/reportes/generar/consolidado', methods=['POST'])
@login_required_api()
def route_api_reportes_generar_consolidado():
    return api_reportes_generar_consolidado()

@app.route('/api/reportes/validar/consistencia', methods=['GET'])
@login_required_api()
def route_api_reportes_validar_consistencia():
    return api_reportes_validar_consistencia()

@app.route('/api/reportes/configuracion', methods=['GET', 'POST', 'PUT', 'DELETE'])
@login_required_api()
def route_api_reportes_configuracion():
    return api_reportes_configuracion()

@app.route('/api/reportes/consolidacion/jerarquica', methods=['POST'])
@login_required_api()
def route_api_reportes_consolidacion_jerarquica():
    return api_reportes_consolidacion_jerarquica()

@app.route('/api/reportes/analisis/tendencias', methods=['POST'])
@login_required_api()
def route_api_reportes_analisis_tendencias():
    return api_reportes_analisis_tendencias()

@app.route('/api/reportes/exportar', methods=['POST'])
@login_required_api()
def route_api_reportes_exportar():
    return api_reportes_exportar()

@app.route('/api/reportes/programados', methods=['GET', 'POST', 'PUT', 'DELETE'])
@login_required_api()
def route_api_reportes_programados():
    return api_reportes_programados()

@app.route('/logistica/estadisticas_generales', methods=['GET'])
@login_required()
@role_required('logistica')
def estadisticas_generales():
    """Obtener estadísticas generales de la tabla QRY"""
    try:
        # Conectar a la base de datos capired
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener total de registros
        cursor.execute("SELECT COUNT(*) as total FROM qry")
        total_registros = cursor.fetchone()['total']
        
        # Obtener seriales únicos
        cursor.execute("SELECT COUNT(DISTINCT serial) as unicos FROM qry WHERE serial IS NOT NULL AND serial != ''")
        seriales_unicos = cursor.fetchone()['unicos']
        
        # Obtener cuentas únicas
        cursor.execute("SELECT COUNT(DISTINCT cuenta) as unicos FROM qry WHERE cuenta IS NOT NULL AND cuenta != ''")
        cuentas_unicas = cursor.fetchone()['unicos']
        
        # Obtener OT únicas
        cursor.execute("SELECT COUNT(DISTINCT ot) as unicos FROM qry WHERE ot IS NOT NULL AND ot != ''")
        ot_unicas = cursor.fetchone()['unicos']
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'total_registros': total_registros,
            'seriales_unicos': seriales_unicos,
            'cuentas_unicas': cuentas_unicas,
            'ot_unicas': ot_unicas
        })
        
    except Exception as e:
        print(f"Error al obtener estadísticas generales: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Error al obtener estadísticas: {str(e)}'
        })

# Endpoint para obtener lista de supervisores para el módulo de logística
@app.route('/api/logistica/supervisores', methods=['GET'])
@login_required()
@role_required('logistica')
def obtener_supervisores_logistica():
    """Obtiene la lista de supervisores desde la tabla recurso_operativo"""
    connection = None
    cursor = None
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener supervisores filtrados por carpeta 'supervisores'
        cursor.execute("""
            SELECT 
                id_codigo_consumidor,
                nombre,
                recurso_operativo_cedula,
                cargo
            FROM recurso_operativo 
            WHERE carpeta = 'supervisores' AND estado = 'Activo'
            ORDER BY nombre ASC
        """)
        
        supervisores = cursor.fetchall()
        
        return jsonify({
            'status': 'success',
            'supervisores': supervisores
        })
        
    except mysql.connector.Error as e:
        print(f"Error MySQL al obtener supervisores: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error al obtener supervisores: {str(e)}'
        }), 500
    except Exception as e:
        print(f"Error general al obtener supervisores: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error al obtener supervisores: {str(e)}'
        }), 500
    finally:
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

# Definir la clase User para Flask-Login
class User(UserMixin):
    def __init__(self, id, nombre, role):
        self.id = id
        self.nombre = nombre
        self.role = role
    
    def has_role(self, role_id):
        return str(self.role) == str(role_id) or self.role == 'administrativo'

@login_manager.user_loader
def load_user(user_id):
    connection = get_db_connection()
    if connection is None:
        return None
    
    cursor = connection.cursor(dictionary=True)
    cursor.execute("SELECT id_codigo_consumidor, nombre, id_roles FROM recurso_operativo WHERE id_codigo_consumidor = %s", (user_id,))
    user_data = cursor.fetchone()
    cursor.close()
    connection.close()
    
    if user_data:
        return User(user_data['id_codigo_consumidor'], user_data['nombre'], ROLES.get(str(user_data['id_roles'])))
    return None

# ===== MÓDULO DE DEVOLUCIONES Y CAMBIOS DE DOTACIÓN =====

@app.route('/logistica/cambios_dotacion')
@login_required()
@role_required('logistica')
def cambios_dotacion():
    """Mostrar formulario de cambios de dotación"""
    try:
        connection = get_db_connection()
        if connection is None:
            flash('Error de conexión a la base de datos', 'danger')
            return redirect(url_for('logistica_dashboard'))
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener lista de técnicos activos
        cursor.execute("""
            SELECT id_codigo_consumidor, nombre, recurso_operativo_cedula
            FROM recurso_operativo 
            WHERE estado = 'Activo' AND carpeta IN ('FTTH INSTALACIONES', 'POSTVENTA', 'BROWFIELD', 'MANTENIMIENTO FTTH', 'ARREGLOS HFC', 'INSTALACIONES DOBLES')
            ORDER BY nombre ASC
        """)
        tecnicos = cursor.fetchall()
        
        cursor.close()
        connection.close()
        
        return render_template('modulos/logistica/cambios_dotacion.html', tecnicos=tecnicos)
        
    except Exception as e:
        print(f"Error en cambios_dotacion: {str(e)}")
        flash('Error al cargar el formulario de cambios', 'danger')
        return redirect(url_for('logistica_dashboard'))

@app.route('/logistica/devoluciones_dotacion')
@login_required()
@role_required('logistica')
def devoluciones_dotacion():
    """Mostrar formulario de devoluciones de dotación"""
    try:
        connection = get_db_connection()
        if connection is None:
            flash('Error de conexión a la base de datos', 'danger')
            return redirect(url_for('logistica_dashboard'))
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener lista de técnicos activos
        cursor.execute("""
            SELECT id_codigo_consumidor, nombre, recurso_operativo_cedula
            FROM recurso_operativo 
            WHERE estado = 'Activo' AND carpeta IN ('FTTH INSTALACIONES', 'POSTVENTA', 'BROWFIELD', 'MANTENIMIENTO FTTH', 'ARREGLOS HFC', 'INSTALACIONES DOBLES')
            ORDER BY nombre ASC
        """)
        tecnicos = cursor.fetchall()
        
        # Obtener lista de clientes activos con IDs únicos
        cursor.execute("""
            SELECT 
                ROW_NUMBER() OVER (ORDER BY cliente ASC) as id,
                cliente as nombre
            FROM (
                SELECT DISTINCT cliente
                FROM recurso_operativo 
                WHERE cliente IS NOT NULL AND cliente != '' AND estado = 'Activo'
            ) as clientes_unicos
            ORDER BY cliente ASC
        """)
        clientes = cursor.fetchall()
        
        cursor.close()
        connection.close()
        
        return render_template('modulos/logistica/devoluciones_dotacion.html', 
                             tecnicos=tecnicos, clientes=clientes)
        
    except Exception as e:
        print(f"Error en devoluciones_dotacion: {str(e)}")
        flash('Error al cargar el formulario de devoluciones', 'danger')
        return redirect(url_for('logistica_dashboard'))

@app.route('/logistica/registrar_cambio_dotacion', methods=['POST'])
@login_required()
@role_required('logistica')
def registrar_cambio_dotacion():
    """Procesar registro de cambio de dotación con gestión diferenciada de stock"""
    try:
        # Conectar a la base de datos capired
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener datos del formulario
        id_codigo_consumidor = request.form.get('id_codigo_consumidor')
        fecha_cambio = request.form.get('fecha_cambio')
        observaciones = request.form.get('observaciones', '')
        
        # Función auxiliar para convertir cantidad a entero o None
        def convertir_cantidad(valor):
            if valor is None or valor == '' or valor == '0':
                return None
            try:
                cantidad = int(valor)
                return cantidad if cantidad > 0 else None
            except (ValueError, TypeError):
                return None
        
        # Función auxiliar para convertir talla a valor válido o None si está vacío
        def convertir_talla(valor):
            if valor is None or valor == '' or valor.strip() == '':
                return None
            return valor.strip()
        
        # Obtener cantidades, tallas y estados de valoración de dotación
        elementos_dotacion = {
            'pantalon': {
                'cantidad': convertir_cantidad(request.form.get('pantalon')),
                'talla': convertir_talla(request.form.get('pantalon_talla')),
                'valorado': request.form.get('pantalon_valorado') == 'on'
            },
            'camisetagris': {
                'cantidad': convertir_cantidad(request.form.get('camisetagris')),
                'talla': convertir_talla(request.form.get('camiseta_gris_talla')),
                'valorado': request.form.get('camisetagris_valorado') == 'on'
            },
            'guerrera': {
                'cantidad': convertir_cantidad(request.form.get('guerrera')),
                'talla': convertir_talla(request.form.get('guerrera_talla')),
                'valorado': request.form.get('guerrera_valorado') == 'on'
            },
            'camisetapolo': {
                'cantidad': convertir_cantidad(request.form.get('camisetapolo')),
                'talla': convertir_talla(request.form.get('camiseta_polo_talla')),
                'valorado': request.form.get('camisetapolo_valorado') == 'on'
            },
            'guantes_nitrilo': {
                'cantidad': convertir_cantidad(request.form.get('guantes_nitrilo')),
                'talla': None,
                'valorado': request.form.get('guantesnitrilo_valorado') == 'on'
            },
            'guantes_carnaza': {
                'cantidad': convertir_cantidad(request.form.get('guantes_carnaza')),
                'talla': None,
                'valorado': request.form.get('guantescarnaza_valorado') == 'on'
            },
            'gafas': {
                'cantidad': convertir_cantidad(request.form.get('gafas')),
                'talla': None,
                'valorado': request.form.get('gafas_valorado') == 'on'
            },
            'gorra': {
                'cantidad': convertir_cantidad(request.form.get('gorra')),
                'talla': None,
                'valorado': request.form.get('gorra_valorado') == 'on'
            },
            'casco': {
                'cantidad': convertir_cantidad(request.form.get('casco')),
                'talla': None,
                'valorado': request.form.get('casco_valorado') == 'on'
            },
            'botas': {
                'cantidad': convertir_cantidad(request.form.get('botas')),
                'talla': convertir_talla(request.form.get('botas_talla')),
                'valorado': request.form.get('botas_valorado') == 'on'
            },
            'chaqueta': {
                'cantidad': convertir_cantidad(request.form.get('chaqueta')),
                'talla': convertir_talla(request.form.get('chaqueta_talla')),
                'valorado': request.form.get('chaqueta_valorado') == 'on'
            }
        }
        
        # Validaciones básicas
        if not id_codigo_consumidor or not fecha_cambio:
            flash('Técnico y fecha son campos obligatorios', 'danger')
            return redirect(url_for('cambios_dotacion'))
        
        # Obtener configuración de valoración de elementos
        cursor.execute("SELECT elemento, es_valorado FROM dotacion_elementos_config")
        config_valoracion = {row['elemento']: row['es_valorado'] for row in cursor.fetchall()}
        
        # Validar stock disponible antes de procesar usando el stock unificado por estado
        errores_stock = []
        for elemento, datos in elementos_dotacion.items():
            cantidad = datos['cantidad']
            if cantidad is not None and cantidad > 0:
                
                # Verificar stock disponible según estado valorado del formulario
                es_valorado = datos['valorado']
                tipo_stock = "VALORADO" if es_valorado else "NO VALORADO"
                
                # Obtener stock actual del elemento filtrado por estado valorado
                cursor.execute("""
                    SELECT COALESCE(SUM(cantidad), 0) as stock_disponible
                    FROM ingresos_dotaciones 
                    WHERE tipo_elemento = %s AND estado = %s
                """, (elemento, tipo_stock))
                
                ingresos_result = cursor.fetchone()
                stock_ingresos = ingresos_result['stock_disponible'] if ingresos_result else 0
                
                # Obtener salidas del elemento con el mismo estado
                cursor.execute("""
                    SELECT COALESCE(SUM(
                        CASE 
                            WHEN %s = 'pantalon' AND estado_pantalon = %s THEN pantalon
                            WHEN %s = 'camisetagris' AND estado_camiseta_gris = %s THEN camisetagris
                            WHEN %s = 'guerrera' AND estado_guerrera = %s THEN guerrera
                            WHEN %s = 'camisetapolo' AND estado_camiseta_polo = %s THEN camisetapolo
                            WHEN %s = 'guantes_nitrilo' AND estado_guantes_nitrilo = %s THEN guantes_nitrilo
                            WHEN %s = 'guantes_carnaza' AND estado_guantes_carnaza = %s THEN guantes_carnaza
                            WHEN %s = 'gafas' AND estado_gafas = %s THEN gafas
                            WHEN %s = 'gorra' AND estado_gorra = %s THEN gorra
                            WHEN %s = 'casco' AND estado_casco = %s THEN casco
                            WHEN %s = 'botas' AND estado_botas = %s THEN botas
                            WHEN %s = 'chaqueta' AND estado_chaqueta = %s THEN chaqueta
                            ELSE 0
                        END
                    ), 0) as total_salidas
                    FROM cambios_dotacion
                """, (elemento, tipo_stock, elemento, tipo_stock, elemento, tipo_stock, 
                      elemento, tipo_stock, elemento, tipo_stock, elemento, tipo_stock,
                      elemento, tipo_stock, elemento, tipo_stock, elemento, tipo_stock,
                      elemento, tipo_stock, elemento, tipo_stock))
                
                salidas_result = cursor.fetchone()
                stock_salidas = salidas_result['total_salidas'] if salidas_result else 0
                
                # Calcular stock disponible real por estado
                stock_disponible = stock_ingresos - stock_salidas
                
                if stock_disponible < cantidad:
                    nombre_elemento = elemento.replace('_', ' ').title()
                    errores_stock.append(
                        f"Stock insuficiente para {nombre_elemento} ({tipo_stock}): "
                        f"Disponible: {stock_disponible}, Solicitado: {cantidad}"
                    )
        
        # Si hay errores de stock, mostrarlos y no procesar
        if errores_stock:
            for error in errores_stock:
                flash(error, 'warning')
            return redirect(url_for('cambios_dotacion'))
        
        # Procesar el cambio de dotación
        # 1. Insertar en cambios_dotacion con estados de valoración
        query_cambio = """
            INSERT INTO cambios_dotacion (
                id_codigo_consumidor, fecha_cambio, pantalon, pantalon_talla, estado_pantalon,
                camisetagris, camiseta_gris_talla, estado_camiseta_gris, guerrera, guerrera_talla, estado_guerrera,
                camisetapolo, camiseta_polo_talla, estado_camiseta_polo, guantes_nitrilo, estado_guantes_nitrilo,
                guantes_carnaza, estado_guantes_carnaza, gafas, estado_gafas, gorra, estado_gorra,
                casco, estado_casco, botas, botas_talla, estado_botas, chaqueta, chaqueta_talla, estado_chaqueta, observaciones
            ) VALUES (
                %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
            )
        """
        
        cursor.execute(query_cambio, (
            id_codigo_consumidor, fecha_cambio, 
            elementos_dotacion['pantalon']['cantidad'], elementos_dotacion['pantalon']['talla'], 
            'VALORADO' if elementos_dotacion['pantalon']['valorado'] else 'NO VALORADO',
            elementos_dotacion['camisetagris']['cantidad'], elementos_dotacion['camisetagris']['talla'],
            'VALORADO' if elementos_dotacion['camisetagris']['valorado'] else 'NO VALORADO',
            elementos_dotacion['guerrera']['cantidad'], elementos_dotacion['guerrera']['talla'],
            'VALORADO' if elementos_dotacion['guerrera']['valorado'] else 'NO VALORADO',
            elementos_dotacion['camisetapolo']['cantidad'], elementos_dotacion['camisetapolo']['talla'],
            'VALORADO' if elementos_dotacion['camisetapolo']['valorado'] else 'NO VALORADO',
            elementos_dotacion['guantes_nitrilo']['cantidad'], 
            'VALORADO' if elementos_dotacion['guantes_nitrilo']['valorado'] else 'NO VALORADO',
            elementos_dotacion['guantes_carnaza']['cantidad'],
            'VALORADO' if elementos_dotacion['guantes_carnaza']['valorado'] else 'NO VALORADO',
            elementos_dotacion['gafas']['cantidad'],
            'VALORADO' if elementos_dotacion['gafas']['valorado'] else 'NO VALORADO',
            elementos_dotacion['gorra']['cantidad'],
            'VALORADO' if elementos_dotacion['gorra']['valorado'] else 'NO VALORADO',
            elementos_dotacion['casco']['cantidad'],
            'VALORADO' if elementos_dotacion['casco']['valorado'] else 'NO VALORADO',
            elementos_dotacion['botas']['cantidad'], elementos_dotacion['botas']['talla'],
            'VALORADO' if elementos_dotacion['botas']['valorado'] else 'NO VALORADO',
            elementos_dotacion['chaqueta']['cantidad'], elementos_dotacion['chaqueta']['talla'],
            'VALORADO' if elementos_dotacion['chaqueta']['valorado'] else 'NO VALORADO',
            observaciones
        ))
        
        # NOTA: Los cambios de dotación solo se registran en cambios_dotacion
        # NO se debe insertar en la tabla dotaciones para evitar contaminación cruzada
        
        connection.commit()
        
        # Preparar mensaje de éxito con detalles de estado valorado
        elementos_procesados = []
        for elemento, datos in elementos_dotacion.items():
            cantidad = datos['cantidad']
            if cantidad is not None and cantidad > 0:
                estado_valorado = "VALORADO" if datos['valorado'] else "NO VALORADO"
                nombre_elemento = elemento.replace('_', ' ').title()
                elementos_procesados.append(f"{nombre_elemento}: {cantidad} ({estado_valorado})")
        
        mensaje_exito = f"Cambio de dotación registrado exitosamente. Elementos procesados: {', '.join(elementos_procesados)}"
        flash(mensaje_exito, 'success')
        
        cursor.close()
        connection.close()
        
        return redirect(url_for('cambios_dotacion'))
        
    except Exception as e:
        print(f"Error al registrar cambio de dotación: {str(e)}")
        flash(f'Error al registrar cambio: {str(e)}', 'danger')
        return redirect(url_for('cambios_dotacion'))

@app.route('/api/cambios_dotacion/historial', methods=['GET'])
@login_required()
@role_required('logistica')
def api_cambios_dotacion_historial():
    """API para obtener el historial de cambios de dotación"""
    try:
        # Conectar a la base de datos capired
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Consulta mejorada con JOIN para obtener información completa del técnico
        # Incluye los campos de estado (valorado/no valorado) directamente de cambios_dotacion
        query = """
            SELECT 
                cd.id_cambio as id,
                cd.id_codigo_consumidor,
                ro.nombre as tecnico_nombre,
                ro.recurso_operativo_cedula as tecnico_cedula,
                cd.fecha_cambio,
                cd.pantalon,
                cd.pantalon_talla,
                cd.estado_pantalon,
                cd.camisetagris,
                cd.camiseta_gris_talla,
                cd.estado_camiseta_gris,
                cd.guerrera,
                cd.guerrera_talla,
                cd.estado_guerrera,
                cd.camisetapolo,
                cd.camiseta_polo_talla,
                cd.estado_camiseta_polo,
                cd.chaqueta,
                cd.chaqueta_talla,
                cd.estado_chaqueta,
                cd.guantes_nitrilo,
                cd.estado_guantes_nitrilo,
                cd.guantes_carnaza,
                cd.estado_guantes_carnaza,
                cd.gafas,
                cd.estado_gafas,
                cd.gorra,
                cd.estado_gorra,
                cd.casco,
                cd.estado_casco,
                cd.botas,
                cd.botas_talla,
                cd.estado_botas,
                cd.observaciones,
                cd.fecha_registro as created_at
            FROM cambios_dotacion cd
            LEFT JOIN recurso_operativo ro ON cd.id_codigo_consumidor = ro.id_codigo_consumidor
            ORDER BY cd.fecha_cambio DESC, cd.fecha_registro DESC
        """
        
        cursor.execute(query)
        cambios = cursor.fetchall()
        
        # Formatear los datos para el frontend con información mejorada
        historial = []
        for cambio in cambios:
            # Crear lista de elementos modificados con estado de valoración
            elementos_modificados = []
            
            # Mapeo de elementos con sus configuraciones incluyendo estados de valoración
            elementos_data = [
                ('pantalon', cambio['pantalon'], cambio['pantalon_talla'], cambio['estado_pantalon']),
                ('camisetagris', cambio['camisetagris'], cambio['camiseta_gris_talla'], cambio['estado_camiseta_gris']),
                ('guerrera', cambio['guerrera'], cambio['guerrera_talla'], cambio['estado_guerrera']),
                ('camisetapolo', cambio['camisetapolo'], cambio['camiseta_polo_talla'], cambio['estado_camiseta_polo']),
                ('chaqueta', cambio['chaqueta'], cambio['chaqueta_talla'], cambio['estado_chaqueta']),
                ('guantes_nitrilo', cambio['guantes_nitrilo'], None, cambio['estado_guantes_nitrilo']),
                ('guantes_carnaza', cambio['guantes_carnaza'], None, cambio['estado_guantes_carnaza']),
                ('gafas', cambio['gafas'], None, cambio['estado_gafas']),
                ('gorra', cambio['gorra'], None, cambio['estado_gorra']),
                ('casco', cambio['casco'], None, cambio['estado_casco']),
                ('botas', cambio['botas'], cambio['botas_talla'], cambio['estado_botas'])
            ]
            
            # Nombres descriptivos para elementos
            nombres_elementos = {
                'pantalon': 'Pantalón',
                'camisetagris': 'Camiseta Gris',
                'guerrera': 'Guerrera',
                'camisetapolo': 'Camiseta Polo',
                'chaqueta': 'Chaqueta',
                'guantes_nitrilo': 'Guantes Nitrilo',
                'guantes_carnaza': 'Guantes Carnaza',
                'gafas': 'Gafas',
                'gorra': 'Gorra',
                'casco': 'Casco',
                'botas': 'Botas'
            }
            
            for elemento, cantidad, talla, estado_valoracion in elementos_data:
                if cantidad and cantidad > 0:
                    # Usar el estado de valoración directamente de cambios_dotacion
                    es_valorado = estado_valoracion == 'VALORADO' if estado_valoracion else False
                    nombre_elemento = nombres_elementos.get(elemento, elemento.title())
                    
                    # Crear objeto con la estructura que espera el frontend
                    elemento_obj = {
                        'descripcion': nombre_elemento,
                        'elemento': nombre_elemento,
                        'cantidad': cantidad,
                        'es_valorado': es_valorado,
                        'estado_valoracion': estado_valoracion or 'NO VALORADO',  # Campo adicional para mostrar el estado exacto
                        'precio_unitario': None  # Se puede agregar lógica de precios aquí si es necesario
                    }
                    
                    if talla:
                        elemento_obj['talla'] = talla
                    
                    elementos_modificados.append(elemento_obj)
            
            # Formatear información completa del técnico
            tecnico_info = "Técnico no encontrado"
            if cambio['tecnico_nombre'] and cambio['tecnico_cedula']:
                tecnico_info = f"{cambio['tecnico_nombre']} (ID: {cambio['tecnico_cedula']})"
            elif cambio['tecnico_nombre']:
                tecnico_info = f"{cambio['tecnico_nombre']} (ID: {cambio['id_codigo_consumidor']})"
            elif cambio['tecnico_cedula']:
                tecnico_info = f"Cédula: {cambio['tecnico_cedula']}"
            else:
                tecnico_info = f"ID Sistema: {cambio['id_codigo_consumidor']}"
            
            # Enviar fechas en formato ISO para que JavaScript pueda parsearlas correctamente
            fecha_cambio_iso = None
            fecha_registro_iso = None
            
            if cambio['fecha_cambio']:
                fecha_cambio_iso = cambio['fecha_cambio'].isoformat()
            
            if cambio['created_at']:
                fecha_registro_iso = cambio['created_at'].isoformat()
            
            # Crear string de elementos para compatibilidad con versiones anteriores
            elementos_texto = ', '.join([f"{elem['descripcion']}: {elem['cantidad']}" + (f" (Talla: {elem['talla']})" if elem.get('talla') else "") for elem in elementos_modificados]) if elementos_modificados else 'Sin elementos especificados'
            
            historial.append({
                'id': cambio['id'],
                'tecnico': tecnico_info,
                'tecnico_nombre': cambio['tecnico_nombre'] or 'No disponible',
                'tecnico_cedula': cambio['tecnico_cedula'] or 'No disponible',
                'id_codigo_consumidor': cambio['id_codigo_consumidor'],
                'elementos_modificados': elementos_texto,  # String para compatibilidad
                'elementos_modificados_detalle': elementos_modificados,  # Array de objetos para el frontend
                'fecha_cambio': fecha_cambio_iso,
                'fecha_registro': fecha_registro_iso,
                'observaciones': cambio['observaciones'] or 'Sin observaciones'
            })
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'data': historial
        })
        
    except Exception as e:
        print(f"Error al obtener historial de cambios de dotación: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error al cargar el historial: {str(e)}'
        }), 500

@app.route('/api/cambios_dotacion/exportar', methods=['GET'])
@login_required()
@role_required('logistica')
def api_cambios_dotacion_exportar():
    """API para exportar cambios de dotación a Excel con dos hojas"""
    try:
        # Conectar a la base de datos capired
        import mysql.connector
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Consulta para obtener todos los cambios de dotación
        query = """
            SELECT 
                cd.id_cambio as id,
                cd.id_codigo_consumidor,
                ro.nombre as tecnico_nombre,
                ro.recurso_operativo_cedula as tecnico_cedula,
                cd.fecha_cambio,
                cd.pantalon,
                cd.pantalon_talla,
                cd.estado_pantalon,
                cd.camisetagris,
                cd.camiseta_gris_talla,
                cd.estado_camiseta_gris,
                cd.guerrera,
                cd.guerrera_talla,
                cd.estado_guerrera,
                cd.camisetapolo,
                cd.camiseta_polo_talla,
                cd.estado_camiseta_polo,
                cd.chaqueta,
                cd.chaqueta_talla,
                cd.estado_chaqueta,
                cd.guantes_nitrilo,
                cd.estado_guantes_nitrilo,
                cd.guantes_carnaza,
                cd.estado_guantes_carnaza,
                cd.gafas,
                cd.estado_gafas,
                cd.gorra,
                cd.estado_gorra,
                cd.casco,
                cd.estado_casco,
                cd.botas,
                cd.botas_talla,
                cd.estado_botas,
                cd.observaciones,
                cd.fecha_registro as created_at
            FROM cambios_dotacion cd
            LEFT JOIN recurso_operativo ro ON cd.id_codigo_consumidor = ro.id_codigo_consumidor
            ORDER BY cd.fecha_cambio DESC, cd.fecha_registro DESC
        """
        
        cursor.execute(query)
        cambios = cursor.fetchall()
        
        # Crear el libro de Excel
        wb = Workbook()
        
        # Eliminar la hoja por defecto
        wb.remove(wb.active)
        
        # Estilos para el Excel
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # HOJA 1: Resumen por Elementos
        ws1 = wb.create_sheet("Resumen por Elementos")
        
        # Configurar encabezados de la Hoja 1
        headers_resumen = ["Elemento", "Cantidad Total", "Cantidad Valorada", "Cantidad No Valorada"]
        for col, header in enumerate(headers_resumen, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Mapeo de elementos para el resumen
        elementos_resumen = {
            'Pantalón': {'total': 0, 'valorado': 0, 'no_valorado': 0},
            'Camiseta Gris': {'total': 0, 'valorado': 0, 'no_valorado': 0},
            'Guerrera': {'total': 0, 'valorado': 0, 'no_valorado': 0},
            'Camiseta Polo': {'total': 0, 'valorado': 0, 'no_valorado': 0},
            'Chaqueta': {'total': 0, 'valorado': 0, 'no_valorado': 0},
            'Guantes Nitrilo': {'total': 0, 'valorado': 0, 'no_valorado': 0},
            'Guantes Carnaza': {'total': 0, 'valorado': 0, 'no_valorado': 0},
            'Gafas': {'total': 0, 'valorado': 0, 'no_valorado': 0},
            'Gorra': {'total': 0, 'valorado': 0, 'no_valorado': 0},
            'Casco': {'total': 0, 'valorado': 0, 'no_valorado': 0},
            'Botas': {'total': 0, 'valorado': 0, 'no_valorado': 0}
        }
        
        # Procesar datos para el resumen
        for cambio in cambios:
            elementos_data = [
                ('Pantalón', cambio['pantalon'], cambio['estado_pantalon']),
                ('Camiseta Gris', cambio['camisetagris'], cambio['estado_camiseta_gris']),
                ('Guerrera', cambio['guerrera'], cambio['estado_guerrera']),
                ('Camiseta Polo', cambio['camisetapolo'], cambio['estado_camiseta_polo']),
                ('Chaqueta', cambio['chaqueta'], cambio['estado_chaqueta']),
                ('Guantes Nitrilo', cambio['guantes_nitrilo'], cambio['estado_guantes_nitrilo']),
                ('Guantes Carnaza', cambio['guantes_carnaza'], cambio['estado_guantes_carnaza']),
                ('Gafas', cambio['gafas'], cambio['estado_gafas']),
                ('Gorra', cambio['gorra'], cambio['estado_gorra']),
                ('Casco', cambio['casco'], cambio['estado_casco']),
                ('Botas', cambio['botas'], cambio['estado_botas'])
            ]
            
            for nombre_elemento, cantidad, estado in elementos_data:
                if cantidad and cantidad > 0:
                    elementos_resumen[nombre_elemento]['total'] += cantidad
                    if estado == 'VALORADO':
                        elementos_resumen[nombre_elemento]['valorado'] += cantidad
                    else:
                        elementos_resumen[nombre_elemento]['no_valorado'] += cantidad
        
        # Escribir datos del resumen
        row = 2
        for elemento, datos in elementos_resumen.items():
            ws1.cell(row=row, column=1, value=elemento).border = border
            ws1.cell(row=row, column=2, value=datos['total']).border = border
            ws1.cell(row=row, column=3, value=datos['valorado']).border = border
            ws1.cell(row=row, column=4, value=datos['no_valorado']).border = border
            row += 1
        
        # Ajustar ancho de columnas de la Hoja 1
        for col in range(1, 5):
            ws1.column_dimensions[get_column_letter(col)].width = 20
        
        # HOJA 2: Historial Completo
        ws2 = wb.create_sheet("Historial Completo")
        
        # Configurar encabezados de la Hoja 2
        headers_historial = [
            "ID", "Técnico", "Cédula", "Fecha Cambio", "Fecha Registro",
            "Pantalón", "Talla Pantalón", "Estado Pantalón",
            "Camiseta Gris", "Talla Camiseta Gris", "Estado Camiseta Gris",
            "Guerrera", "Talla Guerrera", "Estado Guerrera",
            "Camiseta Polo", "Talla Camiseta Polo", "Estado Camiseta Polo",
            "Chaqueta", "Talla Chaqueta", "Estado Chaqueta",
            "Guantes Nitrilo", "Estado Guantes Nitrilo",
            "Guantes Carnaza", "Estado Guantes Carnaza",
            "Gafas", "Estado Gafas",
            "Gorra", "Estado Gorra",
            "Casco", "Estado Casco",
            "Botas", "Talla Botas", "Estado Botas",
            "Observaciones"
        ]
        
        for col, header in enumerate(headers_historial, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Escribir datos del historial
        row = 2
        for cambio in cambios:
            ws2.cell(row=row, column=1, value=cambio['id']).border = border
            ws2.cell(row=row, column=2, value=cambio['tecnico_nombre'] or 'No disponible').border = border
            ws2.cell(row=row, column=3, value=cambio['tecnico_cedula'] or 'No disponible').border = border
            ws2.cell(row=row, column=4, value=cambio['fecha_cambio'].strftime('%Y-%m-%d') if cambio['fecha_cambio'] else '').border = border
            ws2.cell(row=row, column=5, value=cambio['created_at'].strftime('%Y-%m-%d %H:%M:%S') if cambio['created_at'] else '').border = border
            
            # Elementos de dotación
            col = 6
            elementos_data = [
                (cambio['pantalon'], cambio['pantalon_talla'], cambio['estado_pantalon']),
                (cambio['camisetagris'], cambio['camiseta_gris_talla'], cambio['estado_camiseta_gris']),
                (cambio['guerrera'], cambio['guerrera_talla'], cambio['estado_guerrera']),
                (cambio['camisetapolo'], cambio['camiseta_polo_talla'], cambio['estado_camiseta_polo']),
                (cambio['chaqueta'], cambio['chaqueta_talla'], cambio['estado_chaqueta']),
                (cambio['guantes_nitrilo'], None, cambio['estado_guantes_nitrilo']),
                (cambio['guantes_carnaza'], None, cambio['estado_guantes_carnaza']),
                (cambio['gafas'], None, cambio['estado_gafas']),
                (cambio['gorra'], None, cambio['estado_gorra']),
                (cambio['casco'], None, cambio['estado_casco']),
                (cambio['botas'], cambio['botas_talla'], cambio['estado_botas'])
            ]
            
            for cantidad, talla, estado in elementos_data:
                ws2.cell(row=row, column=col, value=cantidad or 0).border = border
                col += 1
                if talla is not None:  # Solo para elementos con talla
                    ws2.cell(row=row, column=col, value=talla or '').border = border
                    col += 1
                ws2.cell(row=row, column=col, value=estado or 'NO VALORADO').border = border
                col += 1
            
            ws2.cell(row=row, column=col, value=cambio['observaciones'] or '').border = border
            row += 1
        
        # Ajustar ancho de columnas de la Hoja 2
        for col in range(1, len(headers_historial) + 1):
            ws2.column_dimensions[get_column_letter(col)].width = 15
        
        cursor.close()
        connection.close()
        
        # Crear el archivo en memoria
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generar nombre del archivo con timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'cambios_dotacion_{timestamp}.xlsx'
        
        # Crear la respuesta
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        print(f"Error al exportar cambios de dotación: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error al exportar: {str(e)}'
        }), 500

@app.route('/api/cambios_dotacion/exportar_csv', methods=['GET'])
@login_required()
@role_required('logistica')
def api_cambios_dotacion_exportar_csv():
    """API para exportar cambios de dotación a CSV con una fila por elemento modificado"""
    try:
        # Conectar a la base de datos capired
        import mysql.connector
        import csv
        import io
        
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Consulta para obtener todos los cambios de dotación
        query = """
            SELECT 
                cd.id_cambio as id,
                cd.id_codigo_consumidor,
                ro.nombre as tecnico_nombre,
                ro.recurso_operativo_cedula as tecnico_cedula,
                cd.fecha_cambio,
                cd.pantalon,
                cd.pantalon_talla,
                cd.estado_pantalon,
                cd.camisetagris,
                cd.camiseta_gris_talla,
                cd.estado_camiseta_gris,
                cd.guerrera,
                cd.guerrera_talla,
                cd.estado_guerrera,
                cd.camisetapolo,
                cd.camiseta_polo_talla,
                cd.estado_camiseta_polo,
                cd.chaqueta,
                cd.chaqueta_talla,
                cd.estado_chaqueta,
                cd.guantes_nitrilo,
                cd.estado_guantes_nitrilo,
                cd.guantes_carnaza,
                cd.estado_guantes_carnaza,
                cd.gafas,
                cd.estado_gafas,
                cd.gorra,
                cd.estado_gorra,
                cd.casco,
                cd.estado_casco,
                cd.botas,
                cd.botas_talla,
                cd.estado_botas,
                cd.observaciones,
                cd.fecha_registro as created_at
            FROM cambios_dotacion cd
            LEFT JOIN recurso_operativo ro ON cd.id_codigo_consumidor = ro.id_codigo_consumidor
            ORDER BY cd.fecha_cambio DESC, cd.fecha_registro DESC
        """
        
        cursor.execute(query)
        cambios = cursor.fetchall()
        
        # Crear el archivo CSV en memoria
        output = io.StringIO()
        
        # Definir las nuevas columnas del CSV - una fila por elemento
        fieldnames = [
            'Tecnico',
            'Cedula',
            'Fecha_Cambio',
            'Elemento_Modificado',
            'Cantidad',
            'Talla',
            'Estado',
            'Observaciones',
            'Fecha_Registro'
        ]
        
        # Crear el writer CSV con delimitador de coma y quoting mínimo
        writer = csv.DictWriter(output, fieldnames=fieldnames, delimiter=',', quoting=csv.QUOTE_MINIMAL)
        
        # Escribir encabezados
        writer.writeheader()
        
        # Definir los elementos de dotación con sus campos correspondientes
        elementos_dotacion = [
            ('pantalon', 'pantalon', 'pantalon_talla', 'estado_pantalon'),
            ('camiseta_gris', 'camisetagris', 'camiseta_gris_talla', 'estado_camiseta_gris'),
            ('guerrera', 'guerrera', 'guerrera_talla', 'estado_guerrera'),
            ('camiseta_polo', 'camisetapolo', 'camiseta_polo_talla', 'estado_camiseta_polo'),
            ('chaqueta', 'chaqueta', 'chaqueta_talla', 'estado_chaqueta'),
            ('guantes_nitrilo', 'guantes_nitrilo', None, 'estado_guantes_nitrilo'),
            ('guantes_carnaza', 'guantes_carnaza', None, 'estado_guantes_carnaza'),
            ('gafas', 'gafas', None, 'estado_gafas'),
            ('gorra', 'gorra', None, 'estado_gorra'),
            ('casco', 'casco', None, 'estado_casco'),
            ('botas', 'botas', 'botas_talla', 'estado_botas')
        ]
        
        # Escribir datos - una fila por elemento modificado
        for cambio in cambios:
            # Datos comunes del técnico y cambio
            tecnico = str(cambio['tecnico_nombre'] or 'No disponible').strip()
            cedula = str(cambio['tecnico_cedula'] or 'No disponible').strip()
            fecha_cambio = cambio['fecha_cambio'].strftime('%d/%m/%Y') if cambio['fecha_cambio'] else ''
            observaciones = str(cambio['observaciones'] or '').strip()
            fecha_registro = cambio['created_at'].strftime('%d/%m/%Y %H:%M') if cambio['created_at'] else ''
            
            # Iterar por cada elemento de dotación
            for elemento_nombre, campo_cantidad, campo_talla, campo_estado in elementos_dotacion:
                cantidad = cambio.get(campo_cantidad, 0) or 0
                
                # Solo incluir elementos con cantidad > 0
                if cantidad > 0:
                    talla = ''
                    if campo_talla and cambio.get(campo_talla):
                        talla = str(cambio[campo_talla]).strip()
                    
                    estado = str(cambio.get(campo_estado, 'NO VALORADO') or 'NO VALORADO').strip()
                    
                    row = {
                        'Tecnico': tecnico,
                        'Cedula': cedula,
                        'Fecha_Cambio': fecha_cambio,
                        'Elemento_Modificado': elemento_nombre,
                        'Cantidad': str(cantidad).strip(),
                        'Talla': talla,
                        'Estado': estado,
                        'Observaciones': observaciones,
                        'Fecha_Registro': fecha_registro
                    }
                    writer.writerow(row)
        
        cursor.close()
        connection.close()
        
        # Obtener el contenido CSV sin BOM para evitar problemas de formato
        csv_content = output.getvalue()
        csv_bytes = csv_content.encode('utf-8')
        
        # Generar nombre del archivo con timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'cambios_dotacion_detallado_{timestamp}.csv'
        
        # Crear la respuesta
        response = make_response(csv_bytes)
        response.headers['Content-Type'] = 'text/csv; charset=utf-8'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        print(f"Error al exportar cambios de dotación CSV: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error al exportar CSV: {str(e)}'
        }), 500

@app.route('/api/validar_pin', methods=['POST'])
@login_required()
def api_validar_pin():
    """API para validar el PIN del usuario logueado - Optimizado para respuesta rápida"""
    connection = None
    cursor = None
    
    try:
        # Validación rápida de datos de entrada
        data = request.get_json()
        if not data or 'pin' not in data:
            return jsonify({
                'success': False,
                'error': 'PIN no proporcionado'
            }), 400
        
        pin_ingresado = data['pin'].strip()
        usuario_id = session.get('user_id')
        
        if not usuario_id:
            return jsonify({
                'success': False,
                'error': 'Usuario no autenticado'
            }), 401
        
        # Validación básica del PIN ingresado
        if not pin_ingresado or len(pin_ingresado) < 4:
            return jsonify({
                'success': False,
                'error': 'PIN debe tener al menos 4 dígitos'
            }), 400
        
        # Conectar a la base de datos con timeout optimizado
        connection = mysql.connector.connect(
            **db_config,
            connection_timeout=5,  # Timeout de conexión de 5 segundos
            autocommit=True
        )
        cursor = connection.cursor()
        
        # Consulta optimizada - usar la columna correcta id_codigo_consumidor
        query = """
            SELECT pin 
            FROM recurso_operativo 
            WHERE id_codigo_consumidor = %s
            LIMIT 1
        """
        
        # Ejecutar consulta con timeout
        cursor.execute(query, (usuario_id,))
        result = cursor.fetchone()
        
        # Verificar si el usuario existe
        if not result:
            return jsonify({
                'success': False,
                'error': 'Usuario no encontrado en recursos operativos'
            }), 404
        
        pin_usuario = result[0]
        
        # Verificación rápida si el usuario tiene PIN asignado
        if pin_usuario is None or pin_usuario == 0:
            return jsonify({
                'success': False,
                'error': 'No tiene PIN asignado. Contacte al administrador para configurar su PIN de seguridad.'
            }), 403
        
        # Validar el PIN (conversión a string para comparación)
        if str(pin_ingresado) == str(pin_usuario):
            return jsonify({
                'success': True,
                'message': 'PIN validado correctamente'
            })
        else:
            return jsonify({
                'success': False,
                'error': 'PIN incorrecto'
            }), 401
            
    except mysql.connector.Error as err:
        print(f"Error de base de datos al validar PIN: {err}")
        return jsonify({
            'success': False,
            'error': 'Error de conexión con la base de datos'
        }), 500
    except Exception as e:
        print(f"Error al validar PIN: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Error interno del servidor'
        }), 500
    finally:
        # Cerrar conexiones de forma segura
        if cursor:
            cursor.close()
        if connection and connection.is_connected():
            connection.close()

@app.route('/api/cambios_dotacion/<int:cambio_id>', methods=['PUT'])
@login_required()
@role_required('logistica')
def api_actualizar_cambio_dotacion(cambio_id):
    """API para actualizar un cambio de dotación existente"""
    try:
        # Obtener datos del formulario
        data = request.get_json()
        
        if not data:
            return jsonify({
                'success': False,
                'error': 'No se recibieron datos para actualizar'
            }), 400
        
        # Conectar a la base de datos capired
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor()
        
        # Verificar que el registro existe
        cursor.execute("SELECT id_cambio FROM cambios_dotacion WHERE id_cambio = %s", (cambio_id,))
        if not cursor.fetchone():
            cursor.close()
            connection.close()
            return jsonify({
                'success': False,
                'error': 'El registro no existe'
            }), 404
        
        # Preparar los datos para actualizar
        fecha_cambio = data.get('fecha_cambio')
        observaciones = data.get('observaciones', '')
        
        # Elementos de dotación con sus cantidades, tallas y estados
        elementos = {
            'pantalon': {
                'cantidad': int(data.get('pantalon', 0)) if data.get('pantalon') else 0,
                'talla': data.get('pantalon_talla', ''),
                'estado': 'VALORADO' if data.get('estado_pantalon') else 'NO VALORADO'
            },
            'camisetagris': {
                'cantidad': int(data.get('camisetagris', 0)) if data.get('camisetagris') else 0,
                'talla': data.get('camiseta_gris_talla', ''),
                'estado': 'VALORADO' if data.get('estado_camiseta_gris') else 'NO VALORADO'
            },
            'guerrera': {
                'cantidad': int(data.get('guerrera', 0)) if data.get('guerrera') else 0,
                'talla': data.get('guerrera_talla', ''),
                'estado': 'VALORADO' if data.get('estado_guerrera') else 'NO VALORADO'
            },
            'camisetapolo': {
                'cantidad': int(data.get('camisetapolo', 0)) if data.get('camisetapolo') else 0,
                'talla': data.get('camiseta_polo_talla', ''),
                'estado': 'VALORADO' if data.get('estado_camiseta_polo') else 'NO VALORADO'
            },
            'botas': {
                'cantidad': int(data.get('botas', 0)) if data.get('botas') else 0,
                'talla': data.get('botas_talla', ''),
                'estado': 'VALORADO' if data.get('estado_botas') else 'NO VALORADO'
            },
            'guantes_nitrilo': {
                'cantidad': int(data.get('guantes_nitrilo', 0)) if data.get('guantes_nitrilo') else 0,
                'talla': '',
                'estado': 'VALORADO' if data.get('estado_guantes_nitrilo') else 'NO VALORADO'
            },
            'guantes_carnaza': {
                'cantidad': int(data.get('guantes_carnaza', 0)) if data.get('guantes_carnaza') else 0,
                'talla': '',
                'estado': 'VALORADO' if data.get('estado_guantes_carnaza') else 'NO VALORADO'
            },
            'gafas': {
                'cantidad': int(data.get('gafas', 0)) if data.get('gafas') else 0,
                'talla': '',
                'estado': 'VALORADO' if data.get('estado_gafas') else 'NO VALORADO'
            },
            'gorra': {
                'cantidad': int(data.get('gorra', 0)) if data.get('gorra') else 0,
                'talla': '',
                'estado': 'VALORADO' if data.get('estado_gorra') else 'NO VALORADO'
            },
            'casco': {
                'cantidad': int(data.get('casco', 0)) if data.get('casco') else 0,
                'talla': '',
                'estado': 'VALORADO' if data.get('estado_casco') else 'NO VALORADO'
            },
            'chaqueta': {
                'cantidad': int(data.get('chaqueta', 0)) if data.get('chaqueta') else 0,
                'talla': data.get('chaqueta_talla', ''),
                'estado': 'VALORADO' if data.get('estado_chaqueta') else 'NO VALORADO'
            }
        }
        
        # Construir la consulta de actualización
        update_query = """
            UPDATE cambios_dotacion SET
                fecha_cambio = %s,
                pantalon = %s,
                pantalon_talla = %s,
                estado_pantalon = %s,
                camisetagris = %s,
                camiseta_gris_talla = %s,
                estado_camiseta_gris = %s,
                guerrera = %s,
                guerrera_talla = %s,
                estado_guerrera = %s,
                camisetapolo = %s,
                camiseta_polo_talla = %s,
                estado_camiseta_polo = %s,
                botas = %s,
                botas_talla = %s,
                estado_botas = %s,
                guantes_nitrilo = %s,
                estado_guantes_nitrilo = %s,
                guantes_carnaza = %s,
                estado_guantes_carnaza = %s,
                gafas = %s,
                estado_gafas = %s,
                gorra = %s,
                estado_gorra = %s,
                casco = %s,
                estado_casco = %s,
                chaqueta = %s,
                chaqueta_talla = %s,
                estado_chaqueta = %s,
                observaciones = %s
            WHERE id_cambio = %s
        """
        
        # Preparar los valores para la consulta
        valores = [
            fecha_cambio,
            elementos['pantalon']['cantidad'],
            elementos['pantalon']['talla'] if elementos['pantalon']['talla'] else None,
            elementos['pantalon']['estado'],
            elementos['camisetagris']['cantidad'],
            elementos['camisetagris']['talla'] if elementos['camisetagris']['talla'] else None,
            elementos['camisetagris']['estado'],
            elementos['guerrera']['cantidad'],
            elementos['guerrera']['talla'] if elementos['guerrera']['talla'] else None,
            elementos['guerrera']['estado'],
            elementos['camisetapolo']['cantidad'],
            elementos['camisetapolo']['talla'] if elementos['camisetapolo']['talla'] else None,
            elementos['camisetapolo']['estado'],
            elementos['botas']['cantidad'],
            elementos['botas']['talla'] if elementos['botas']['talla'] else None,
            elementos['botas']['estado'],
            elementos['guantes_nitrilo']['cantidad'],
            elementos['guantes_nitrilo']['estado'],
            elementos['guantes_carnaza']['cantidad'],
            elementos['guantes_carnaza']['estado'],
            elementos['gafas']['cantidad'],
            elementos['gafas']['estado'],
            elementos['gorra']['cantidad'],
            elementos['gorra']['estado'],
            elementos['casco']['cantidad'],
            elementos['casco']['estado'],
            elementos['chaqueta']['cantidad'],
            elementos['chaqueta']['talla'] if elementos['chaqueta']['talla'] else None,
            elementos['chaqueta']['estado'],
            observaciones,
            cambio_id
        ]
        
        # Ejecutar la actualización
        print(f"Ejecutando actualización para cambio_id: {cambio_id}")
        print(f"Consulta SQL: {update_query}")
        print(f"Valores: {valores}")
        
        cursor.execute(update_query, valores)
        connection.commit()
        
        # Verificar que se actualizó al menos un registro
        if cursor.rowcount == 0:
            cursor.close()
            connection.close()
            return jsonify({
                'success': False,
                'error': 'No se pudo actualizar el registro. Verifique que el ID sea válido.'
            }), 400
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': 'Cambio de dotación actualizado exitosamente'
        })
        
    except mysql.connector.Error as db_error:
        print(f"Error de base de datos al actualizar cambio de dotación: {str(db_error)}")
        if 'connection' in locals():
            connection.close()
        return jsonify({
            'success': False,
            'error': f'Error de base de datos: {str(db_error)}'
        }), 500
    except Exception as e:
        print(f"Error general al actualizar cambio de dotación: {str(e)}")
        if 'connection' in locals():
            connection.close()
        return jsonify({
            'success': False,
            'error': f'Error al actualizar el registro: {str(e)}'
        }), 500

@app.route('/api/indicadores_cambios_dotacion', methods=['GET'])
@login_required()
@role_required('logistica')
def api_indicadores_cambios_dotacion():
    """API para obtener datos de indicadores de cambios de dotación para gráficas"""
    try:
        # Conectar a la base de datos capired
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener parámetros de filtros
        estado_valorado = request.args.get('estado_valorado', 'todos')  # 'valorado', 'no_valorado', 'todos'
        mes = request.args.get('mes', '')  # 1-12 o vacío para todos
        elemento = request.args.get('elemento', 'todos')  # nombre del elemento o 'todos'
        anio = request.args.get('anio', str(datetime.now().year))  # año actual por defecto
        
        # Construir la consulta base
        base_query = """
            SELECT 
                MONTH(cd.fecha_cambio) as mes,
                YEAR(cd.fecha_cambio) as anio,
                'pantalon' as elemento_tipo,
                'Pantalón' as elemento_nombre,
                SUM(cd.pantalon) as cantidad_total,
                SUM(CASE WHEN cd.estado_pantalon = 'VALORADO' THEN cd.pantalon ELSE 0 END) as cantidad_valorado,
                SUM(CASE WHEN cd.estado_pantalon != 'VALORADO' OR cd.estado_pantalon IS NULL THEN cd.pantalon ELSE 0 END) as cantidad_no_valorado
            FROM cambios_dotacion cd
            WHERE cd.pantalon > 0 AND YEAR(cd.fecha_cambio) = %s
            GROUP BY MONTH(cd.fecha_cambio), YEAR(cd.fecha_cambio)
            
            UNION ALL
            
            SELECT 
                MONTH(cd.fecha_cambio) as mes,
                YEAR(cd.fecha_cambio) as anio,
                'camisetagris' as elemento_tipo,
                'Camiseta Gris' as elemento_nombre,
                SUM(cd.camisetagris) as cantidad_total,
                SUM(CASE WHEN cd.estado_camiseta_gris = 'VALORADO' THEN cd.camisetagris ELSE 0 END) as cantidad_valorado,
                SUM(CASE WHEN cd.estado_camiseta_gris != 'VALORADO' OR cd.estado_camiseta_gris IS NULL THEN cd.camisetagris ELSE 0 END) as cantidad_no_valorado
            FROM cambios_dotacion cd
            WHERE cd.camisetagris > 0 AND YEAR(cd.fecha_cambio) = %s
            GROUP BY MONTH(cd.fecha_cambio), YEAR(cd.fecha_cambio)
            
            UNION ALL
            
            SELECT 
                MONTH(cd.fecha_cambio) as mes,
                YEAR(cd.fecha_cambio) as anio,
                'guerrera' as elemento_tipo,
                'Guerrera' as elemento_nombre,
                SUM(cd.guerrera) as cantidad_total,
                SUM(CASE WHEN cd.estado_guerrera = 'VALORADO' THEN cd.guerrera ELSE 0 END) as cantidad_valorado,
                SUM(CASE WHEN cd.estado_guerrera != 'VALORADO' OR cd.estado_guerrera IS NULL THEN cd.guerrera ELSE 0 END) as cantidad_no_valorado
            FROM cambios_dotacion cd
            WHERE cd.guerrera > 0 AND YEAR(cd.fecha_cambio) = %s
            GROUP BY MONTH(cd.fecha_cambio), YEAR(cd.fecha_cambio)
            
            UNION ALL
            
            SELECT 
                MONTH(cd.fecha_cambio) as mes,
                YEAR(cd.fecha_cambio) as anio,
                'camisetapolo' as elemento_tipo,
                'Camiseta Polo' as elemento_nombre,
                SUM(cd.camisetapolo) as cantidad_total,
                SUM(CASE WHEN cd.estado_camiseta_polo = 'VALORADO' THEN cd.camisetapolo ELSE 0 END) as cantidad_valorado,
                SUM(CASE WHEN cd.estado_camiseta_polo != 'VALORADO' OR cd.estado_camiseta_polo IS NULL THEN cd.camisetapolo ELSE 0 END) as cantidad_no_valorado
            FROM cambios_dotacion cd
            WHERE cd.camisetapolo > 0 AND YEAR(cd.fecha_cambio) = %s
            GROUP BY MONTH(cd.fecha_cambio), YEAR(cd.fecha_cambio)
            
            UNION ALL
            
            SELECT 
                MONTH(cd.fecha_cambio) as mes,
                YEAR(cd.fecha_cambio) as anio,
                'guantes_nitrilo' as elemento_tipo,
                'Guantes Nitrilo' as elemento_nombre,
                SUM(cd.guantes_nitrilo) as cantidad_total,
                SUM(CASE WHEN cd.estado_guantes_nitrilo = 'VALORADO' THEN cd.guantes_nitrilo ELSE 0 END) as cantidad_valorado,
                SUM(CASE WHEN cd.estado_guantes_nitrilo != 'VALORADO' OR cd.estado_guantes_nitrilo IS NULL THEN cd.guantes_nitrilo ELSE 0 END) as cantidad_no_valorado
            FROM cambios_dotacion cd
            WHERE cd.guantes_nitrilo > 0 AND YEAR(cd.fecha_cambio) = %s
            GROUP BY MONTH(cd.fecha_cambio), YEAR(cd.fecha_cambio)
            
            UNION ALL
            
            SELECT 
                MONTH(cd.fecha_cambio) as mes,
                YEAR(cd.fecha_cambio) as anio,
                'guantes_carnaza' as elemento_tipo,
                'Guantes Carnaza' as elemento_nombre,
                SUM(cd.guantes_carnaza) as cantidad_total,
                SUM(CASE WHEN cd.estado_guantes_carnaza = 'VALORADO' THEN cd.guantes_carnaza ELSE 0 END) as cantidad_valorado,
                SUM(CASE WHEN cd.estado_guantes_carnaza != 'VALORADO' OR cd.estado_guantes_carnaza IS NULL THEN cd.guantes_carnaza ELSE 0 END) as cantidad_no_valorado
            FROM cambios_dotacion cd
            WHERE cd.guantes_carnaza > 0 AND YEAR(cd.fecha_cambio) = %s
            GROUP BY MONTH(cd.fecha_cambio), YEAR(cd.fecha_cambio)
            
            UNION ALL
            
            SELECT 
                MONTH(cd.fecha_cambio) as mes,
                YEAR(cd.fecha_cambio) as anio,
                'gafas' as elemento_tipo,
                'Gafas' as elemento_nombre,
                SUM(cd.gafas) as cantidad_total,
                SUM(CASE WHEN cd.estado_gafas = 'VALORADO' THEN cd.gafas ELSE 0 END) as cantidad_valorado,
                SUM(CASE WHEN cd.estado_gafas != 'VALORADO' OR cd.estado_gafas IS NULL THEN cd.gafas ELSE 0 END) as cantidad_no_valorado
            FROM cambios_dotacion cd
            WHERE cd.gafas > 0 AND YEAR(cd.fecha_cambio) = %s
            GROUP BY MONTH(cd.fecha_cambio), YEAR(cd.fecha_cambio)
            
            UNION ALL
            
            SELECT 
                MONTH(cd.fecha_cambio) as mes,
                YEAR(cd.fecha_cambio) as anio,
                'gorra' as elemento_tipo,
                'Gorra' as elemento_nombre,
                SUM(cd.gorra) as cantidad_total,
                SUM(CASE WHEN cd.estado_gorra = 'VALORADO' THEN cd.gorra ELSE 0 END) as cantidad_valorado,
                SUM(CASE WHEN cd.estado_gorra != 'VALORADO' OR cd.estado_gorra IS NULL THEN cd.gorra ELSE 0 END) as cantidad_no_valorado
            FROM cambios_dotacion cd
            WHERE cd.gorra > 0 AND YEAR(cd.fecha_cambio) = %s
            GROUP BY MONTH(cd.fecha_cambio), YEAR(cd.fecha_cambio)
            
            UNION ALL
            
            SELECT 
                MONTH(cd.fecha_cambio) as mes,
                YEAR(cd.fecha_cambio) as anio,
                'casco' as elemento_tipo,
                'Casco' as elemento_nombre,
                SUM(cd.casco) as cantidad_total,
                SUM(CASE WHEN cd.estado_casco = 'VALORADO' THEN cd.casco ELSE 0 END) as cantidad_valorado,
                SUM(CASE WHEN cd.estado_casco != 'VALORADO' OR cd.estado_casco IS NULL THEN cd.casco ELSE 0 END) as cantidad_no_valorado
            FROM cambios_dotacion cd
            WHERE cd.casco > 0 AND YEAR(cd.fecha_cambio) = %s
            GROUP BY MONTH(cd.fecha_cambio), YEAR(cd.fecha_cambio)
            
            UNION ALL
            
            SELECT 
                MONTH(cd.fecha_cambio) as mes,
                YEAR(cd.fecha_cambio) as anio,
                'botas' as elemento_tipo,
                'Botas' as elemento_nombre,
                SUM(cd.botas) as cantidad_total,
                SUM(CASE WHEN cd.estado_botas = 'VALORADO' THEN cd.botas ELSE 0 END) as cantidad_valorado,
                SUM(CASE WHEN cd.estado_botas != 'VALORADO' OR cd.estado_botas IS NULL THEN cd.botas ELSE 0 END) as cantidad_no_valorado
            FROM cambios_dotacion cd
            WHERE cd.botas > 0 AND YEAR(cd.fecha_cambio) = %s
            GROUP BY MONTH(cd.fecha_cambio), YEAR(cd.fecha_cambio)
            
            ORDER BY mes, elemento_nombre
        """
        
        # Ejecutar consulta con parámetros (año repetido para cada UNION)
        params = [anio] * 10  # 10 elementos diferentes
        cursor.execute(base_query, params)
        resultados = cursor.fetchall()
        
        # Procesar y filtrar resultados según los parámetros
        if mes and mes != '':
            # Filtro por mes específico - comportamiento actual
            datos_filtrados = []
            for row in resultados:
                # Aplicar filtro de mes si se especifica
                if str(row['mes']) != str(mes):
                    continue
                    
                # Aplicar filtro de elemento si se especifica
                if elemento != 'todos' and row['elemento_tipo'] != elemento:
                    continue
                    
                # Aplicar filtro de estado valorado
                if estado_valorado == 'valorado':
                    cantidad_mostrar = row['cantidad_valorado']
                elif estado_valorado == 'no_valorado':
                    cantidad_mostrar = row['cantidad_no_valorado']
                else:  # 'todos'
                    cantidad_mostrar = row['cantidad_total']
                
                # Solo incluir si hay cantidad para mostrar
                if cantidad_mostrar > 0:
                    datos_filtrados.append({
                        'mes': row['mes'],
                        'anio': row['anio'],
                        'elemento_tipo': row['elemento_tipo'],
                        'elemento_nombre': row['elemento_nombre'],
                        'cantidad': cantidad_mostrar,
                        'cantidad_total': row['cantidad_total'],
                        'cantidad_valorado': row['cantidad_valorado'],
                        'cantidad_no_valorado': row['cantidad_no_valorado']
                    })
        else:
            # Todos los meses - agregar datos por elemento
            elementos_agregados = {}
            
            for row in resultados:
                # Aplicar filtro de elemento si se especifica
                if elemento != 'todos' and row['elemento_tipo'] != elemento:
                    continue
                
                elemento_key = row['elemento_tipo']
                
                # Aplicar filtro de estado valorado
                if estado_valorado == 'valorado':
                    cantidad_mostrar = row['cantidad_valorado']
                elif estado_valorado == 'no_valorado':
                    cantidad_mostrar = row['cantidad_no_valorado']
                else:  # 'todos'
                    cantidad_mostrar = row['cantidad_total']
                
                # Agregar al elemento existente o crear nuevo
                if elemento_key in elementos_agregados:
                    elementos_agregados[elemento_key]['cantidad'] += cantidad_mostrar
                    elementos_agregados[elemento_key]['cantidad_total'] += row['cantidad_total']
                    elementos_agregados[elemento_key]['cantidad_valorado'] += row['cantidad_valorado']
                    elementos_agregados[elemento_key]['cantidad_no_valorado'] += row['cantidad_no_valorado']
                else:
                    elementos_agregados[elemento_key] = {
                        'mes': None,  # No hay mes específico para datos agregados
                        'anio': row['anio'],
                        'elemento_tipo': row['elemento_tipo'],
                        'elemento_nombre': row['elemento_nombre'],
                        'cantidad': cantidad_mostrar,
                        'cantidad_total': row['cantidad_total'],
                        'cantidad_valorado': row['cantidad_valorado'],
                        'cantidad_no_valorado': row['cantidad_no_valorado']
                    }
            
            # Convertir a lista y filtrar elementos con cantidad > 0
            datos_filtrados = []
            for elemento_data in elementos_agregados.values():
                if elemento_data['cantidad'] > 0:
                    datos_filtrados.append(elemento_data)
        
        # Obtener años disponibles para el filtro
        cursor.execute("""
            SELECT DISTINCT YEAR(fecha_cambio) as anio 
            FROM cambios_dotacion 
            WHERE fecha_cambio IS NOT NULL 
            ORDER BY anio DESC
        """)
        anios_disponibles = [row['anio'] for row in cursor.fetchall()]
        
        # Obtener elementos disponibles para el filtro
        elementos_disponibles = [
            {'tipo': 'pantalon', 'nombre': 'Pantalón'},
            {'tipo': 'camisetagris', 'nombre': 'Camiseta Gris'},
            {'tipo': 'guerrera', 'nombre': 'Guerrera'},
            {'tipo': 'camisetapolo', 'nombre': 'Camiseta Polo'},
            {'tipo': 'guantes_nitrilo', 'nombre': 'Guantes Nitrilo'},
            {'tipo': 'guantes_carnaza', 'nombre': 'Guantes Carnaza'},
            {'tipo': 'gafas', 'nombre': 'Gafas'},
            {'tipo': 'gorra', 'nombre': 'Gorra'},
            {'tipo': 'casco', 'nombre': 'Casco'},
            {'tipo': 'botas', 'nombre': 'Botas'}
        ]
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'data': datos_filtrados,
            'filtros': {
                'anios_disponibles': anios_disponibles,
                'elementos_disponibles': elementos_disponibles,
                'meses': [
                    {'numero': 1, 'nombre': 'Enero'},
                    {'numero': 2, 'nombre': 'Febrero'},
                    {'numero': 3, 'nombre': 'Marzo'},
                    {'numero': 4, 'nombre': 'Abril'},
                    {'numero': 5, 'nombre': 'Mayo'},
                    {'numero': 6, 'nombre': 'Junio'},
                    {'numero': 7, 'nombre': 'Julio'},
                    {'numero': 8, 'nombre': 'Agosto'},
                    {'numero': 9, 'nombre': 'Septiembre'},
                    {'numero': 10, 'nombre': 'Octubre'},
                    {'numero': 11, 'nombre': 'Noviembre'},
                    {'numero': 12, 'nombre': 'Diciembre'}
                ]
            },
            'filtros_aplicados': {
                'estado_valorado': estado_valorado,
                'mes': mes,
                'elemento': elemento,
                'anio': anio
            }
        })
        
    except Exception as e:
        print(f"Error al obtener indicadores de cambios de dotación: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error al cargar los indicadores: {str(e)}'
        }), 500

@app.route('/logistica/registrar_devolucion_dotacion', methods=['POST'])
@login_required()
@role_required('logistica')
def registrar_devolucion_dotacion():
    """Procesar registro de devolución de dotación"""
    try:
        # Conectar a la base de datos capired
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor()
        
        # Obtener datos del formulario
        tecnico_id = request.form.get('tecnico_id')
        cliente_id = request.form.get('cliente_id')
        fecha_devolucion = request.form.get('fecha_devolucion')
        motivo = request.form.get('motivo')
        observaciones = request.form.get('observaciones', '')
        estado = request.form.get('estado', 'REGISTRADA')
        created_by = session.get('user_id')
        
        # Validaciones básicas
        if not tecnico_id or not cliente_id or not motivo:
            flash('Técnico, cliente y motivo son campos obligatorios', 'danger')
            return redirect(url_for('devoluciones_dotacion'))
        
        # Obtener el nombre del cliente basado en el ID
        cursor_temp = connection.cursor(dictionary=True)
        cursor_temp.execute("""
            SELECT cliente as nombre
            FROM (
                SELECT 
                    ROW_NUMBER() OVER (ORDER BY cliente ASC) as id,
                    cliente
                FROM (
                    SELECT DISTINCT cliente
                    FROM recurso_operativo 
                    WHERE cliente IS NOT NULL AND cliente != '' AND estado = 'Activo'
                ) as clientes_unicos
            ) as clientes_numerados
            WHERE id = %s
        """, (cliente_id,))
        cliente_data = cursor_temp.fetchone()
        cursor_temp.close()
        
        if not cliente_data:
            flash('Cliente no válido', 'danger')
            return redirect(url_for('devoluciones_dotacion'))
        
        cliente_nombre = cliente_data['nombre']
        
        # Insertar en la base de datos
        query = """
            INSERT INTO devoluciones_dotacion (
                tecnico_id, cliente_id, fecha_devolucion, motivo, 
                observaciones, estado, created_by
            ) VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        
        cursor.execute(query, (
            tecnico_id, cliente_id, fecha_devolucion, motivo,
            observaciones, estado, created_by
        ))
        
        connection.commit()
        cursor.close()
        connection.close()
        
        flash('Devolución de dotación registrada exitosamente', 'success')
        return redirect(url_for('devoluciones_dotacion'))
        
    except Exception as e:
        print(f"Error al registrar devolución de dotación: {str(e)}")
        flash(f'Error al registrar devolución: {str(e)}', 'danger')
        return redirect(url_for('devoluciones_dotacion'))

@app.route('/api/devoluciones/listar', methods=['GET'])
@login_required()
@role_required('logistica')
def listar_devoluciones():
    """API para listar todas las devoluciones registradas"""
    try:
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener devoluciones con información del técnico
        query = """
            SELECT 
                d.id,
                d.tecnico_id,
                ro.nombre as tecnico_nombre,
                d.cliente_id,
                d.fecha_devolucion,
                d.motivo,
                d.observaciones,
                d.estado,
                d.created_at,
                d.updated_at,
                COUNT(dd.id) as total_elementos
            FROM devoluciones_dotacion d
            LEFT JOIN recurso_operativo ro ON d.tecnico_id = ro.id_codigo_consumidor
            LEFT JOIN devolucion_detalles dd ON d.id = dd.devolucion_id
            GROUP BY d.id
            ORDER BY d.created_at DESC
        """
        
        cursor.execute(query)
        devoluciones = cursor.fetchall()
        
        # Obtener nombres de clientes para cada devolución
        for devolucion in devoluciones:
            cursor.execute("""
                SELECT cliente as nombre
                FROM (
                    SELECT 
                        ROW_NUMBER() OVER (ORDER BY cliente ASC) as id,
                        cliente
                    FROM (
                        SELECT DISTINCT cliente
                        FROM recurso_operativo 
                        WHERE cliente IS NOT NULL AND cliente != '' AND estado = 'Activo'
                    ) as clientes_unicos
                ) as clientes_numerados
                WHERE id = %s
            """, (devolucion['cliente_id'],))
            cliente_data = cursor.fetchone()
            devolucion['cliente_nombre'] = cliente_data['nombre'] if cliente_data else 'Cliente no encontrado'
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'devoluciones': devoluciones
        })
        
    except Exception as e:
        print(f"Error al listar devoluciones: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/devoluciones/<int:devolucion_id>/detalles', methods=['GET'])
@app.route('/obtener_detalles_devolucion/<int:devolucion_id>', methods=['GET'])
@login_required()
@role_required('logistica')
def obtener_detalles_devolucion(devolucion_id):
    """API para obtener detalles de una devolución específica"""
    try:
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener información de la devolución
        cursor.execute("""
            SELECT 
                d.*,
                ro.nombre as tecnico_nombre
            FROM devoluciones_dotacion d
            LEFT JOIN recurso_operativo ro ON d.tecnico_id = ro.id_codigo_consumidor
            WHERE d.id = %s
        """, (devolucion_id,))
        
        devolucion = cursor.fetchone()
        
        if not devolucion:
            return jsonify({
                'success': False,
                'error': 'Devolución no encontrada'
            }), 404
        
        # Obtener nombre del cliente
        cursor.execute("""
            SELECT cliente as nombre
            FROM (
                SELECT 
                    ROW_NUMBER() OVER (ORDER BY cliente ASC) as id,
                    cliente
                FROM (
                    SELECT DISTINCT cliente
                    FROM recurso_operativo 
                    WHERE cliente IS NOT NULL AND cliente != '' AND estado = 'Activo'
                ) as clientes_unicos
            ) as clientes_numerados
            WHERE id = %s
        """, (devolucion['cliente_id'],))
        cliente_data = cursor.fetchone()
        devolucion['cliente_nombre'] = cliente_data['nombre'] if cliente_data else 'Cliente no encontrado'
        
        # Obtener detalles de elementos devueltos
        cursor.execute("""
            SELECT *
            FROM devolucion_detalles
            WHERE devolucion_id = %s
            ORDER BY created_at ASC
        """, (devolucion_id,))
        
        detalles = cursor.fetchall()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'devolucion': devolucion,
            'detalles': detalles
        })
        
    except Exception as e:
        print(f"Error al obtener detalles de devolución: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/devoluciones/<int:devolucion_id>/detalles', methods=['POST'])
@login_required()
@role_required('logistica')
def agregar_detalle_devolucion(devolucion_id):
    """API para agregar un elemento específico a una devolución"""
    try:
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor()
        
        # Obtener datos del formulario
        data = request.get_json() if request.is_json else request.form
        
        elemento = data.get('elemento')
        talla = data.get('talla')
        cantidad = data.get('cantidad', 1)
        estado_elemento = data.get('estado_elemento', 'USADO_BUENO')
        observaciones = data.get('observaciones', '')
        
        # Validaciones
        if not elemento:
            return jsonify({
                'success': False,
                'error': 'El elemento es obligatorio'
            }), 400
        
        # Verificar que la devolución existe
        cursor.execute("SELECT id FROM devoluciones_dotacion WHERE id = %s", (devolucion_id,))
        if not cursor.fetchone():
            return jsonify({
                'success': False,
                'error': 'Devolución no encontrada'
            }), 404
        
        # Insertar detalle
        query = """
            INSERT INTO devolucion_detalles 
            (devolucion_id, elemento, talla, cantidad, estado_elemento, observaciones)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        
        cursor.execute(query, (
            devolucion_id, elemento, talla, cantidad, estado_elemento, observaciones
        ))
        
        detalle_id = cursor.lastrowid
        connection.commit()
        
        # Obtener el elemento recién creado
        cursor.execute("""
            SELECT id, devolucion_id, elemento, talla, cantidad, estado_elemento, observaciones, created_at
            FROM devolucion_detalles WHERE id = %s
        """, (detalle_id,))
        
        elemento_creado = cursor.fetchone()
        cursor.close()
        connection.close()
        
        # Formatear el elemento para la respuesta
        elemento_data = {
            'id': elemento_creado[0],
            'devolucion_id': elemento_creado[1],
            'elemento': elemento_creado[2],
            'talla': elemento_creado[3],
            'cantidad': elemento_creado[4],
            'estado_elemento': elemento_creado[5],
            'observaciones': elemento_creado[6],
            'created_at': elemento_creado[7].isoformat() if elemento_creado[7] else None
        }
        
        return jsonify({
            'success': True,
            'detalle_id': detalle_id,
            'elemento': elemento_data,
            'message': 'Elemento agregado exitosamente'
        })
        
    except Exception as e:
        print(f"Error al agregar detalle de devolución: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/devoluciones/detalles/<int:detalle_id>', methods=['PUT'])
@login_required()
@role_required('logistica')
def actualizar_detalle_devolucion(detalle_id):
    """API para actualizar un detalle específico de devolución"""
    try:
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor()
        
        # Obtener datos del formulario
        data = request.get_json() if request.is_json else request.form
        
        elemento = data.get('elemento')
        talla = data.get('talla')
        cantidad = data.get('cantidad')
        estado_elemento = data.get('estado_elemento')
        observaciones = data.get('observaciones')
        
        # Construir query dinámicamente
        updates = []
        values = []
        
        if elemento:
            updates.append('elemento = %s')
            values.append(elemento)
        if talla is not None:
            updates.append('talla = %s')
            values.append(talla)
        if cantidad:
            updates.append('cantidad = %s')
            values.append(cantidad)
        if estado_elemento:
            updates.append('estado_elemento = %s')
            values.append(estado_elemento)
        if observaciones is not None:
            updates.append('observaciones = %s')
            values.append(observaciones)
        
        if not updates:
            return jsonify({
                'success': False,
                'error': 'No hay campos para actualizar'
            }), 400
        
        updates.append('updated_at = CURRENT_TIMESTAMP')
        values.append(detalle_id)
        
        query = f"UPDATE devolucion_detalles SET {', '.join(updates)} WHERE id = %s"
        
        cursor.execute(query, values)
        connection.commit()
        
        if cursor.rowcount == 0:
            return jsonify({
                'success': False,
                'error': 'Detalle no encontrado'
            }), 404
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': 'Detalle actualizado exitosamente'
        })
        
    except Exception as e:
        print(f"Error al actualizar detalle de devolución: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/devoluciones/detalles/<int:detalle_id>', methods=['DELETE'])
@login_required()
@role_required('logistica')
def eliminar_detalle_devolucion(detalle_id):
    """API para eliminar un detalle específico de devolución"""
    try:
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor()
        
        cursor.execute("DELETE FROM devolucion_detalles WHERE id = %s", (detalle_id,))
        connection.commit()
        
        if cursor.rowcount == 0:
            return jsonify({
                'success': False,
                'error': 'Detalle no encontrado'
            }), 404
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': 'Detalle eliminado exitosamente'
        })
        
    except Exception as e:
        print(f"Error al eliminar detalle de devolución: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

# ============================================================================
# SISTEMA DE GESTIÓN DE ESTADOS PARA DEVOLUCIONES
# ============================================================================

def validar_transicion_estado(estado_actual, estado_nuevo, usuario_id):
    """
    Valida si una transición de estado es permitida para un usuario específico
    
    Args:
        estado_actual (str): Estado actual de la devolución
        estado_nuevo (str): Estado al que se quiere transicionar
        usuario_id (int): ID del usuario que intenta hacer la transición
    
    Returns:
        dict: {'valida': bool, 'mensaje': str, 'rol_usuario': str}
    """
    try:
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener rol del usuario
        cursor.execute("""
            SELECT r.id_roles as rol_id, r.nombre_rol as rol_nombre
            FROM recurso_operativo ro
            JOIN roles r ON ro.id_roles = r.id_roles
            WHERE ro.id_codigo_consumidor = %s
        """, (usuario_id,))
        
        usuario_rol = cursor.fetchone()
        
        if not usuario_rol:
            return {
                'valida': False,
                'mensaje': 'Usuario no encontrado o sin rol asignado',
                'rol_usuario': None
            }
        
        # Verificar si la transición está permitida para el rol
        cursor.execute("""
            SELECT permitido
            FROM permisos_transicion
            WHERE rol_id = %s AND estado_origen = %s AND estado_destino = %s
        """, (usuario_rol['rol_id'], estado_actual, estado_nuevo))
        
        permiso = cursor.fetchone()
        
        cursor.close()
        connection.close()
        
        if not permiso:
            return {
                'valida': False,
                'mensaje': f'Transición de {estado_actual} a {estado_nuevo} no está configurada para el rol {usuario_rol["rol_nombre"]}',
                'rol_usuario': usuario_rol['rol_nombre']
            }
        
        if not permiso['permitido']:
            return {
                'valida': False,
                'mensaje': f'No tiene permisos para cambiar de {estado_actual} a {estado_nuevo}',
                'rol_usuario': usuario_rol['rol_nombre']
            }
        
        return {
            'valida': True,
            'mensaje': 'Transición permitida',
            'rol_usuario': usuario_rol['rol_nombre']
        }
        
    except Exception as e:
        print(f"Error al validar transición: {str(e)}")
        return {
            'valida': False,
            'mensaje': f'Error interno: {str(e)}',
            'rol_usuario': None
        }

def registrar_auditoria_estado(devolucion_id, estado_anterior, estado_nuevo, usuario_id, motivo_cambio):
    """
    Registra un cambio de estado en la tabla de auditoría
    
    Args:
        devolucion_id (int): ID de la devolución
        estado_anterior (str): Estado previo
        estado_nuevo (str): Nuevo estado
        usuario_id (int): ID del usuario que realizó el cambio
        motivo_cambio (str): Motivo del cambio
    
    Returns:
        bool: True si se registró exitosamente, False en caso contrario
    """
    try:
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor()
        
        query = """
            INSERT INTO auditoria_estados 
            (devolucion_id, estado_anterior, estado_nuevo, usuario_id, motivo_cambio)
            VALUES (%s, %s, %s, %s, %s)
        """
        
        cursor.execute(query, (
            devolucion_id, estado_anterior, estado_nuevo, usuario_id, motivo_cambio
        ))
        
        connection.commit()
        cursor.close()
        connection.close()
        
        return True
        
    except Exception as e:
        print(f"Error al registrar auditoría: {str(e)}")
        return False

def obtener_transiciones_validas(estado_actual, usuario_id):
    """
    Obtiene las transiciones válidas para un estado y usuario específico
    
    Args:
        estado_actual (str): Estado actual de la devolución
        usuario_id (int): ID del usuario
    
    Returns:
        list: Lista de estados a los que se puede transicionar
    """
    try:
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener rol del usuario
        cursor.execute("""
            SELECT r.id_roles as rol_id
            FROM recurso_operativo ro
            JOIN roles r ON ro.id_roles = r.id_roles
            WHERE ro.id_codigo_consumidor = %s
        """, (usuario_id,))
        
        usuario_rol = cursor.fetchone()
        
        if not usuario_rol:
            return []
        
        # Obtener transiciones permitidas
        cursor.execute("""
            SELECT estado_destino
            FROM permisos_transicion
            WHERE rol_id = %s AND estado_origen = %s AND permitido = TRUE
        """, (usuario_rol['rol_id'], estado_actual))
        
        transiciones = cursor.fetchall()
        
        cursor.close()
        connection.close()
        
        return [t['estado_destino'] for t in transiciones]
        
    except Exception as e:
        print(f"Error al obtener transiciones válidas: {str(e)}")
        return []

def enviar_notificacion_cambio_estado(devolucion_id, estado_anterior, estado_nuevo):
    """
    Envía notificaciones automáticas cuando cambia el estado
    
    Args:
        devolucion_id (int): ID de la devolución
        estado_anterior (str): Estado anterior
        estado_nuevo (str): Nuevo estado
    
    Returns:
        bool: True si se enviaron las notificaciones exitosamente
    """
    try:
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener configuración de notificaciones para este cambio de estado
        cursor.execute("""
            SELECT cn.*, pe.nombre as plantilla_email_nombre, ps.nombre as plantilla_sms_nombre
            FROM configuracion_notificaciones cn
            LEFT JOIN plantillas_email pe ON cn.plantilla_email_id = pe.id
            LEFT JOIN plantillas_sms ps ON cn.plantilla_sms_id = ps.id
            WHERE cn.evento_trigger = 'CAMBIO_ESTADO' 
            AND (cn.estado_origen = %s OR cn.estado_origen IS NULL)
            AND cn.estado_destino = %s
            AND cn.activo = TRUE
        """, (estado_anterior, estado_nuevo))
        
        configuraciones = cursor.fetchall()
        
        if not configuraciones:
            print(f"No hay configuraciones de notificación para transición {estado_anterior} -> {estado_nuevo}")
            cursor.close()
            connection.close()
            return True
        
        # Obtener datos de la devolución
        cursor.execute("""
            SELECT d.*, c.nombre as cliente_nombre
            FROM devoluciones_dotacion d
            LEFT JOIN clientes c ON d.cliente_id = c.id
            WHERE d.id = %s
        """, (devolucion_id,))
        
        devolucion_data = cursor.fetchone()
        
        if not devolucion_data:
            print(f"Devolución {devolucion_id} no encontrada")
            cursor.close()
            connection.close()
            return False
        
        # Procesar cada configuración de notificación
        for config in configuraciones:
            # Obtener usuarios destinatarios según roles
            roles_destinatarios = config['destinatarios_roles']
            if isinstance(roles_destinatarios, str):
                import json
                roles_destinatarios = json.loads(roles_destinatarios)
            
            # Convertir nombres de roles a IDs
            placeholders = ','.join(['%s'] * len(roles_destinatarios))
            cursor.execute(f"""
                SELECT ro.id_codigo_consumidor as id, ro.nombre, r.nombre_rol as rol_nombre
                FROM recurso_operativo ro
                JOIN roles r ON ro.id_roles = r.id_roles
                WHERE r.nombre_rol IN ({placeholders}) AND ro.estado = 'Activo'
            """, roles_destinatarios)
            
            usuarios_destinatarios = cursor.fetchall()
            
            # Preparar variables para plantillas
            variables = {
                'devolucion_id': devolucion_id,
                'cliente_nombre': devolucion_data['cliente_nombre'] or 'N/A',
                'fecha_registro': devolucion_data['created_at'].strftime('%d/%m/%Y %H:%M'),
                'observaciones': devolucion_data['observaciones'] or 'Sin observaciones',
                'estado_anterior': estado_anterior,
                'estado_nuevo': estado_nuevo
            }
            
            # Enviar notificaciones según tipo
            for usuario in usuarios_destinatarios:
                if config['tipo_notificacion'] in ['EMAIL', 'AMBOS'] and usuario['email']:
                    programar_notificacion_email(devolucion_id, usuario, config, variables)
                
                if config['tipo_notificacion'] in ['SMS', 'AMBOS'] and usuario['telefono']:
                    programar_notificacion_sms(devolucion_id, usuario, config, variables)
        
        connection.commit()
        cursor.close()
        connection.close()
        
        return True
        
    except Exception as e:
        print(f"Error al enviar notificaciones: {str(e)}")
        return False

def programar_notificacion_email(devolucion_id, usuario, config, variables):
    """
    Programa una notificación por email
    """
    try:
        import mysql.connector
        from datetime import datetime, timedelta
        
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener plantilla de email
        cursor.execute("""
            SELECT asunto, cuerpo_html, cuerpo_texto
            FROM plantillas_email
            WHERE id = %s AND activo = TRUE
        """, (config['plantilla_email_id'],))
        
        plantilla = cursor.fetchone()
        
        if not plantilla:
            print(f"Plantilla de email {config['plantilla_email_id']} no encontrada")
            return False
        
        # Reemplazar variables en la plantilla
        asunto = plantilla['asunto']
        cuerpo_html = plantilla['cuerpo_html']
        cuerpo_texto = plantilla['cuerpo_texto']
        
        for var, valor in variables.items():
            asunto = asunto.replace(f'{{{var}}}', str(valor))
            cuerpo_html = cuerpo_html.replace(f'{{{var}}}', str(valor))
            if cuerpo_texto:
                cuerpo_texto = cuerpo_texto.replace(f'{{{var}}}', str(valor))
        
        # Calcular fecha de envío (con delay si está configurado)
        fecha_programada = datetime.now()
        if config['delay_minutos'] > 0:
            fecha_programada += timedelta(minutes=config['delay_minutos'])
        
        # Insertar en historial de notificaciones
        cursor.execute("""
            INSERT INTO historial_notificaciones 
            (devolucion_id, tipo_notificacion, destinatario, asunto, mensaje, 
             fecha_programada, configuracion_id)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            devolucion_id, 'EMAIL', usuario['email'], asunto, 
            cuerpo_html, fecha_programada, config['id']
        ))
        
        connection.commit()
        cursor.close()
        connection.close()
        
        # Si no hay delay, enviar inmediatamente
        if config['delay_minutos'] == 0:
            enviar_email_inmediato(usuario['email'], asunto, cuerpo_html, cuerpo_texto)
        
        return True
        
    except Exception as e:
        print(f"Error al programar notificación email: {str(e)}")
        return False

def programar_notificacion_sms(devolucion_id, usuario, config, variables):
    """
    Programa una notificación por SMS
    """
    try:
        import mysql.connector
        from datetime import datetime, timedelta
        
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener plantilla de SMS
        cursor.execute("""
            SELECT mensaje
            FROM plantillas_sms
            WHERE id = %s AND activo = TRUE
        """, (config['plantilla_sms_id'],))
        
        plantilla = cursor.fetchone()
        
        if not plantilla:
            print(f"Plantilla de SMS {config['plantilla_sms_id']} no encontrada")
            return False
        
        # Reemplazar variables en la plantilla
        mensaje = plantilla['mensaje']
        for var, valor in variables.items():
            mensaje = mensaje.replace(f'{{{var}}}', str(valor))
        
        # Calcular fecha de envío
        fecha_programada = datetime.now()
        if config['delay_minutos'] > 0:
            fecha_programada += timedelta(minutes=config['delay_minutos'])
        
        # Insertar en historial de notificaciones
        cursor.execute("""
            INSERT INTO historial_notificaciones 
            (devolucion_id, tipo_notificacion, destinatario, mensaje, 
             fecha_programada, configuracion_id)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (
            devolucion_id, 'SMS', usuario['telefono'], mensaje, 
            fecha_programada, config['id']
        ))
        
        connection.commit()
        cursor.close()
        connection.close()
        
        # Si no hay delay, enviar inmediatamente
        if config['delay_minutos'] == 0:
            enviar_sms_inmediato(usuario['telefono'], mensaje)
        
        return True
        
    except Exception as e:
        print(f"Error al programar notificación SMS: {str(e)}")
        return False

def enviar_email_inmediato(destinatario, asunto, cuerpo_html, cuerpo_texto=None):
    """
    Envía un email inmediatamente usando la configuración SMTP
    """
    try:
        import smtplib
        import mysql.connector
        import json
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        
        # Obtener configuración SMTP
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT configuracion
            FROM configuracion_servicios
            WHERE servicio = 'SMTP' AND activo = TRUE
            LIMIT 1
        """)
        
        config_smtp = cursor.fetchone()
        cursor.close()
        connection.close()
        
        if not config_smtp:
            print("No hay configuración SMTP activa")
            return False
        
        smtp_config = json.loads(config_smtp['configuracion'])
        
        # Crear mensaje
        msg = MIMEMultipart('alternative')
        msg['Subject'] = asunto
        msg['From'] = smtp_config['from']
        msg['To'] = destinatario
        
        # Agregar contenido
        if cuerpo_texto:
            part1 = MIMEText(cuerpo_texto, 'plain', 'utf-8')
            msg.attach(part1)
        
        part2 = MIMEText(cuerpo_html, 'html', 'utf-8')
        msg.attach(part2)
        
        # Enviar email
        server = smtplib.SMTP(smtp_config['host'], smtp_config['port'])
        if not smtp_config.get('secure', False):
            server.starttls()
        
        server.login(smtp_config['auth']['user'], smtp_config['auth']['pass'])
        server.send_message(msg)
        server.quit()
        
        print(f"Email enviado exitosamente a {destinatario}")
        return True
        
    except Exception as e:
        print(f"Error al enviar email: {str(e)}")
        return False

def enviar_sms_inmediato(telefono, mensaje):
    """
    Envía un SMS inmediatamente (implementación básica)
    """
    try:
        # Aquí se implementaría la integración con un proveedor de SMS como Twilio
        # Por ahora, solo registramos el intento
        print(f"SMS programado para {telefono}: {mensaje}")
        
        # TODO: Implementar integración real con proveedor SMS
        # Ejemplo con Twilio:
        # from twilio.rest import Client
        # client = Client(account_sid, auth_token)
        # message = client.messages.create(
        #     body=mensaje,
        #     from_='+1234567890',
        #     to=telefono
        # )
        
        return True
        
    except Exception as e:
        print(f"Error al enviar SMS: {str(e)}")
        return False

# ============================================================================
# APIs DE CONFIGURACIÓN DE NOTIFICACIONES Y ROLES

@app.route('/api/configuracion/notificaciones', methods=['GET'])
def obtener_configuraciones_notificaciones():
    """
    Obtiene todas las configuraciones de notificaciones
    """
    try:
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT cn.*, 
                   pe.nombre as plantilla_email_nombre,
                   ps.nombre as plantilla_sms_nombre
            FROM configuracion_notificaciones cn
            LEFT JOIN plantillas_email pe ON cn.plantilla_email_id = pe.id
            LEFT JOIN plantillas_sms ps ON cn.plantilla_sms_id = ps.id
            ORDER BY cn.evento_trigger, cn.estado_destino
        """)
        
        configuraciones = cursor.fetchall()
        
        # Convertir JSON strings a objetos
        for config in configuraciones:
            if config['destinatarios_roles']:
                import json
                config['destinatarios_roles'] = json.loads(config['destinatarios_roles'])
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'configuraciones': configuraciones
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error al obtener configuraciones: {str(e)}'
        }), 500

@app.route('/api/configuracion/notificaciones', methods=['POST'])
def crear_configuracion_notificacion():
    """
    Crea una nueva configuración de notificación
    """
    try:
        data = request.get_json()
        
        # Validaciones
        required_fields = ['evento_trigger', 'estado_destino', 'tipo_notificacion', 'destinatarios_roles']
        for field in required_fields:
            if field not in data:
                return jsonify({
                    'success': False,
                    'error': f'Campo requerido: {field}'
                }), 400
        
        import mysql.connector
        import json
        
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor()
        
        # Convertir roles a JSON
        destinatarios_roles_json = json.dumps(data['destinatarios_roles'])
        
        cursor.execute("""
            INSERT INTO configuracion_notificaciones 
            (evento_trigger, estado_origen, estado_destino, tipo_notificacion,
             destinatarios_roles, plantilla_email_id, plantilla_sms_id, 
             delay_minutos, activo)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            data['evento_trigger'],
            data.get('estado_origen'),
            data['estado_destino'],
            data['tipo_notificacion'],
            destinatarios_roles_json,
            data.get('plantilla_email_id'),
            data.get('plantilla_sms_id'),
            data.get('delay_minutos', 0),
            data.get('activo', True)
        ))
        
        connection.commit()
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': 'Configuración creada exitosamente'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error al crear configuración: {str(e)}'
        }), 500

@app.route('/api/configuracion/notificaciones/<int:config_id>', methods=['PUT'])
def actualizar_configuracion_notificacion(config_id):
    """
    Actualiza una configuración de notificación existente
    """
    try:
        data = request.get_json()
        
        import mysql.connector
        import json
        
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor()
        
        # Construir query dinámicamente
        update_fields = []
        values = []
        
        if 'evento_trigger' in data:
            update_fields.append('evento_trigger = %s')
            values.append(data['evento_trigger'])
        
        if 'estado_origen' in data:
            update_fields.append('estado_origen = %s')
            values.append(data['estado_origen'])
        
        if 'estado_destino' in data:
            update_fields.append('estado_destino = %s')
            values.append(data['estado_destino'])
        
        if 'tipo_notificacion' in data:
            update_fields.append('tipo_notificacion = %s')
            values.append(data['tipo_notificacion'])
        
        if 'destinatarios_roles' in data:
            update_fields.append('destinatarios_roles = %s')
            values.append(json.dumps(data['destinatarios_roles']))
        
        if 'plantilla_email_id' in data:
            update_fields.append('plantilla_email_id = %s')
            values.append(data['plantilla_email_id'])
        
        if 'plantilla_sms_id' in data:
            update_fields.append('plantilla_sms_id = %s')
            values.append(data['plantilla_sms_id'])
        
        if 'delay_minutos' in data:
            update_fields.append('delay_minutos = %s')
            values.append(data['delay_minutos'])
        
        if 'activo' in data:
            update_fields.append('activo = %s')
            values.append(data['activo'])
        
        if not update_fields:
            return jsonify({
                'success': False,
                'error': 'No hay campos para actualizar'
            }), 400
        
        values.append(config_id)
        
        query = f"UPDATE configuracion_notificaciones SET {', '.join(update_fields)} WHERE id = %s"
        cursor.execute(query, values)
        
        connection.commit()
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': 'Configuración actualizada exitosamente'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error al actualizar configuración: {str(e)}'
        }), 500

@app.route('/api/configuracion/plantillas/email', methods=['GET'])
def obtener_plantillas_email():
    """
    Obtiene todas las plantillas de email
    """
    try:
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, nombre, asunto, descripcion, activo, created_at
            FROM plantillas_email
            ORDER BY nombre
        """)
        
        plantillas = cursor.fetchall()
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'plantillas': plantillas
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error al obtener plantillas: {str(e)}'
        }), 500

@app.route('/api/configuracion/plantillas/sms', methods=['GET'])
def obtener_plantillas_sms():
    """
    Obtiene todas las plantillas de SMS
    """
    try:
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, nombre, descripcion, activo, created_at
            FROM plantillas_sms
            ORDER BY nombre
        """)
        
        plantillas = cursor.fetchall()
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'plantillas': plantillas
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error al obtener plantillas: {str(e)}'
        }), 500

@app.route('/api/configuracion/roles', methods=['GET'])
def obtener_roles_sistema():
    """
    Obtiene todos los roles del sistema con sus permisos
    """
    try:
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener roles
        cursor.execute("""
            SELECT r.*, COUNT(u.id) as usuarios_count
            FROM roles r
            LEFT JOIN usuarios u ON r.id = u.rol_id AND u.activo = 1
            GROUP BY r.id
            ORDER BY r.nombre
        """)
        
        roles = cursor.fetchall()
        
        # Obtener permisos para cada rol
        for rol in roles:
            cursor.execute("""
                SELECT p.nombre, p.descripcion, rp.puede_leer, rp.puede_escribir, 
                       rp.puede_actualizar, rp.puede_eliminar
                FROM permisos_roles rp
                JOIN permisos p ON rp.permiso_id = p.id
                WHERE rp.rol_id = %s
                ORDER BY p.nombre
            """, (rol['id'],))
            
            rol['permisos'] = cursor.fetchall()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'roles': roles
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error al obtener roles: {str(e)}'
        }), 500

@app.route('/api/configuracion/permisos', methods=['GET'])
def obtener_permisos_disponibles():
    """
    Obtiene todos los permisos disponibles en el sistema
    """
    try:
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        cursor.execute("""
            SELECT id, nombre, descripcion, modulo
            FROM permisos
            ORDER BY modulo, nombre
        """)
        
        permisos = cursor.fetchall()
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'permisos': permisos
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error al obtener permisos: {str(e)}'
        }), 500

@app.route('/api/configuracion/roles/<int:rol_id>/permisos', methods=['PUT'])
def actualizar_permisos_rol(rol_id):
    """
    Actualiza los permisos de un rol específico
    """
    try:
        data = request.get_json()
        permisos = data.get('permisos', [])
        
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor()
        
        # Eliminar permisos existentes
        cursor.execute("DELETE FROM permisos_roles WHERE rol_id = %s", (rol_id,))
        
        # Insertar nuevos permisos
        for permiso in permisos:
            cursor.execute("""
                INSERT INTO permisos_roles 
                (rol_id, permiso_id, puede_leer, puede_escribir, puede_actualizar, puede_eliminar)
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                rol_id,
                permiso['permiso_id'],
                permiso.get('puede_leer', False),
                permiso.get('puede_escribir', False),
                permiso.get('puede_actualizar', False),
                permiso.get('puede_eliminar', False)
            ))
        
        connection.commit()
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'message': 'Permisos actualizados exitosamente'
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error al actualizar permisos: {str(e)}'
        }), 500

# ============================================================================
# APIs DE CONFIGURACIÓN DE HORA PREOPERACIONAL
# ============================================================================

@app.route('/api/configuracion/hora-preoperacional', methods=['GET'])
def obtener_hora_preoperacional():
    """
    API para obtener la hora límite actual del preoperacional
    """
    try:
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener la hora límite actual
        cursor.execute("""
            SELECT id_hora_preoperacional, hora_limite_preoperacional
            FROM hora_preoperacional 
            ORDER BY id_hora_preoperacional DESC 
            LIMIT 1
        """)
        
        resultado = cursor.fetchone()
        
        if resultado:
            # Convertir time a string para JSON
            hora_limite = str(resultado['hora_limite_preoperacional'])
            
            return jsonify({
                'success': True,
                'data': {
                    'id': resultado['id_hora_preoperacional'],
                    'hora_limite': hora_limite,
                    'hora_limite_formatted': hora_limite[:5]  # HH:MM format
                }
            })
        else:
            return jsonify({
                'success': False,
                'error': 'No se encontró configuración de hora límite'
            }), 404
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error al obtener hora límite: {str(e)}'
        }), 500
    finally:
        if 'connection' in locals() and connection.is_connected():
            cursor.close()
            connection.close()

@app.route('/api/configuracion/hora-preoperacional', methods=['PUT'])
@login_required()
def actualizar_hora_preoperacional():
    """
    API para actualizar la hora límite del preoperacional (solo administradores)
    """
    try:
        # Verificar que el usuario sea administrador
        if session.get('user_role') != 'administrativo':
            return jsonify({
                'success': False,
                'error': 'No tienes permisos para realizar esta acción'
            }), 403
        
        data = request.get_json()
        nueva_hora = data.get('hora_limite')
        
        if not nueva_hora:
            return jsonify({
                'success': False,
                'error': 'La hora límite es obligatoria'
            }), 400
        
        # Validar formato de hora (HH:MM)
        import re
        if not re.match(r'^([01]?[0-9]|2[0-3]):[0-5][0-9]$', nueva_hora):
            return jsonify({
                'success': False,
                'error': 'Formato de hora inválido. Use HH:MM'
            }), 400
        
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor()
        
        # Actualizar la hora límite
        cursor.execute("""
            UPDATE hora_preoperacional 
            SET hora_limite_preoperacional = %s
            WHERE id_hora_preoperacional = 1
        """, (nueva_hora + ':00',))  # Agregar segundos
        
        if cursor.rowcount > 0:
            connection.commit()
            
            return jsonify({
                'success': True,
                'message': f'Hora límite actualizada a {nueva_hora}',
                'data': {
                    'hora_limite': nueva_hora
                }
            })
        else:
            return jsonify({
                'success': False,
                'error': 'No se pudo actualizar la hora límite'
            }), 500
            
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error al actualizar hora límite: {str(e)}'
        }), 500
    finally:
        if 'connection' in locals() and connection.is_connected():
            cursor.close()
            connection.close()

@app.route('/api/configuracion/hora-inicio-operacion', methods=['GET'])
def obtener_hora_inicio_operacion_analistas():
    try:
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        cursor = connection.cursor(dictionary=True)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS hora_inicio_operacion_analistas (
                id_hora_inicio_operacion INT AUTO_INCREMENT PRIMARY KEY,
                hora_limite_inicio_operacion TIME NOT NULL DEFAULT '10:00:00'
            )
        """)

        cursor.execute("""
            SELECT id_hora_inicio_operacion, hora_limite_inicio_operacion
            FROM hora_inicio_operacion_analistas
            ORDER BY id_hora_inicio_operacion DESC
            LIMIT 1
        """)
        resultado = cursor.fetchone()

        if not resultado:
            cursor.execute(
                """
                INSERT INTO hora_inicio_operacion_analistas (hora_limite_inicio_operacion)
                VALUES ('10:00:00')
                """
            )
            connection.commit()
            cursor.execute("""
                SELECT id_hora_inicio_operacion, hora_limite_inicio_operacion
                FROM hora_inicio_operacion_analistas
                ORDER BY id_hora_inicio_operacion DESC
                LIMIT 1
            """)
            resultado = cursor.fetchone()

        hora_limite = str(resultado['hora_limite_inicio_operacion'])

        return jsonify({
            'success': True,
            'data': {
                'id': resultado['id_hora_inicio_operacion'],
                'hora_limite': hora_limite,
                'hora_limite_formatted': hora_limite[:5]
            }
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Error al obtener hora límite: {str(e)}'
        }), 500
    finally:
        if 'connection' in locals() and connection.is_connected():
            cursor.close()
            connection.close()

@app.route('/api/configuracion/hora-inicio-operacion', methods=['PUT'])
@login_required()
def actualizar_hora_inicio_operacion_analistas():
    try:
        if session.get('user_role') != 'administrativo':
            return jsonify({
                'success': False,
                'error': 'No tienes permisos para realizar esta acción'
            }), 403

        data = request.get_json()
        nueva_hora = data.get('hora_limite')
        if not nueva_hora:
            return jsonify({'success': False, 'error': 'La hora límite es obligatoria'}), 400

        import re
        if not re.match(r'^([01]?[0-9]|2[0-3]):[0-5][0-9]$', nueva_hora):
            return jsonify({'success': False, 'error': 'Formato de hora inválido. Use HH:MM'}), 400

        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        cursor = connection.cursor(dictionary=True)

        cursor.execute("""
            CREATE TABLE IF NOT EXISTS hora_inicio_operacion_analistas (
                id_hora_inicio_operacion INT AUTO_INCREMENT PRIMARY KEY,
                hora_limite_inicio_operacion TIME NOT NULL DEFAULT '10:00:00'
            )
        """)

        cursor.execute("SELECT id_hora_inicio_operacion FROM hora_inicio_operacion_analistas ORDER BY id_hora_inicio_operacion DESC LIMIT 1")
        fila = cursor.fetchone()

        if fila:
            cursor.execute(
                """
                UPDATE hora_inicio_operacion_analistas
                SET hora_limite_inicio_operacion = %s
                WHERE id_hora_inicio_operacion = %s
                """,
                (nueva_hora + ':00', fila['id_hora_inicio_operacion'])
            )
        else:
            cursor.execute(
                """
                INSERT INTO hora_inicio_operacion_analistas (hora_limite_inicio_operacion)
                VALUES (%s)
                """,
                (nueva_hora + ':00',)
            )

        connection.commit()

        return jsonify({
            'success': True,
            'message': f'Hora límite actualizada a {nueva_hora}',
            'data': {'hora_limite': nueva_hora}
        })
    except Exception as e:
        return jsonify({'success': False, 'error': f'Error al actualizar hora límite: {str(e)}'}), 500
    finally:
        if 'connection' in locals() and connection.is_connected():
            cursor.close()
            connection.close()

# ============================================================================
# APIs DEL SISTEMA DE GESTIÓN DE ESTADOS
# ============================================================================

@app.route('/api/devoluciones/<int:devolucion_id>/estado', methods=['PUT'])
@login_required()
@role_required('logistica')
def actualizar_estado_devolucion(devolucion_id):
    """
    API para actualizar el estado de una devolución específica
    """
    try:
        # Obtener datos de la petición
        data = request.get_json() if request.is_json else request.form
        
        nuevo_estado = data.get('nuevo_estado')
        motivo = data.get('motivo')
        usuario_id = session.get('user_id')
        
        # Validaciones de entrada
        if not nuevo_estado:
            return jsonify({
                'success': False,
                'error': 'El nuevo estado es obligatorio'
            }), 400
        
        if not motivo:
            return jsonify({
                'success': False,
                'error': 'El motivo del cambio es obligatorio'
            }), 400
        
        estados_validos = ['PENDIENTE', 'REGISTRADA', 'EN_REVISION', 'APROBADA', 'RECHAZADA', 'PROCESADA', 'COMPLETADA', 'CANCELADA']
        if nuevo_estado not in estados_validos:
            return jsonify({
                'success': False,
                'error': f'Estado no válido. Estados permitidos: {", ".join(estados_validos)}'
            }), 400
        
        # Conectar a la base de datos
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener estado actual de la devolución
        cursor.execute(
            "SELECT estado FROM devoluciones_dotacion WHERE id = %s",
            (devolucion_id,)
        )
        
        devolucion = cursor.fetchone()
        
        if not devolucion:
            cursor.close()
            connection.close()
            return jsonify({
                'success': False,
                'error': 'Devolución no encontrada'
            }), 404
        
        estado_actual = devolucion['estado']
        
        # Verificar si el estado ya es el mismo
        if estado_actual == nuevo_estado:
            cursor.close()
            connection.close()
            return jsonify({
                'success': False,
                'error': f'La devolución ya se encuentra en estado {nuevo_estado}'
            }), 400
        
        # Validar transición
        validacion = validar_transicion_estado(estado_actual, nuevo_estado, usuario_id)
        
        if not validacion['valida']:
            cursor.close()
            connection.close()
            return jsonify({
                'success': False,
                'error': validacion['mensaje'],
                'rol_usuario': validacion['rol_usuario']
            }), 403
        
        # Actualizar estado en la base de datos
        cursor.execute(
            "UPDATE devoluciones_dotacion SET estado = %s, updated_by = %s WHERE id = %s",
            (nuevo_estado, usuario_id, devolucion_id)
        )
        
        connection.commit()
        
        # Registrar en auditoría
        auditoria_exitosa = registrar_auditoria_estado(
            devolucion_id, estado_actual, nuevo_estado, usuario_id, motivo
        )
        
        # Enviar notificaciones
        notificacion_exitosa = enviar_notificacion_cambio_estado(
            devolucion_id, estado_actual, nuevo_estado
        )
        
        cursor.close()
        connection.close()
        
        response_data = {
            'success': True,
            'mensaje': 'Estado actualizado exitosamente',
            'estado_anterior': estado_actual,
            'estado_nuevo': nuevo_estado,
            'devolucion_id': devolucion_id
        }
        
        # Agregar información adicional sobre procesos auxiliares
        if not auditoria_exitosa:
            response_data['advertencia_auditoria'] = 'No se pudo registrar en auditoría'
        
        if not notificacion_exitosa:
            response_data['advertencia_notificacion'] = 'No se pudieron enviar notificaciones'
        
        return jsonify(response_data)
        
    except Exception as e:
        print(f"Error al actualizar estado de devolución: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500

@app.route('/api/devoluciones/<int:devolucion_id>/transiciones', methods=['GET'])
@login_required()
@role_required('logistica')
def obtener_transiciones_devolucion(devolucion_id):
    """
    API para obtener las transiciones válidas para una devolución específica
    """
    try:
        usuario_id = session.get('user_id')
        
        # Conectar a la base de datos
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Obtener estado actual de la devolución
        cursor.execute(
            "SELECT estado FROM devoluciones_dotacion WHERE id = %s",
            (devolucion_id,)
        )
        
        devolucion = cursor.fetchone()
        
        if not devolucion:
            cursor.close()
            connection.close()
            return jsonify({
                'success': False,
                'error': 'Devolución no encontrada'
            }), 404
        
        estado_actual = devolucion['estado']
        
        # Obtener transiciones válidas
        transiciones_validas = obtener_transiciones_validas(estado_actual, usuario_id)
        
        # Obtener información del rol del usuario
        cursor.execute("""
            SELECT r.nombre_rol as rol_nombre
            FROM recurso_operativo ro
            JOIN roles r ON ro.id_roles = r.id_roles
            WHERE ro.id_codigo_consumidor = %s
        """, (usuario_id,))
        
        usuario_info = cursor.fetchone()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'devolucion_id': devolucion_id,
            'estado_actual': estado_actual,
            'transiciones_validas': transiciones_validas,
            'rol_usuario': usuario_info['rol_nombre'] if usuario_info else 'Sin rol'
        })
        
    except Exception as e:
        print(f"Error al obtener transiciones: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500

@app.route('/api/devoluciones/<int:devolucion_id>/historial', methods=['GET'])
@login_required_api(role='logistica')
def obtener_historial_estados(devolucion_id):
    """
    API para consultar el historial de cambios de estado de una devolución
    """
    try:
        # Conectar a la base de datos
        import mysql.connector
        connection = mysql.connector.connect(
            host='localhost',
            user='root',
            password='732137A031E4b@',
            database='capired'
        )
        
        cursor = connection.cursor(dictionary=True)
        
        # Verificar que la devolución existe
        cursor.execute(
            "SELECT id, estado FROM devoluciones_dotacion WHERE id = %s",
            (devolucion_id,)
        )
        
        devolucion = cursor.fetchone()
        
        if not devolucion:
            cursor.close()
            connection.close()
            return jsonify({
                'success': False,
                'error': 'Devolución no encontrada'
            }), 404
        
        # Obtener historial de auditoría
        cursor.execute("""
            SELECT 
                ae.id,
                ae.estado_anterior,
                ae.estado_nuevo,
                ae.motivo_cambio,
                ae.fecha_cambio,
                ro.nombre as usuario_nombre,
                r.nombre_rol as rol_usuario
            FROM auditoria_estados_devolucion ae
            LEFT JOIN recurso_operativo ro ON ae.usuario_id = ro.recurso_operativo_cedula
            LEFT JOIN roles r ON ro.id_roles = r.id_roles
            WHERE ae.devolucion_id = %s
            ORDER BY ae.fecha_cambio DESC
        """, (devolucion_id,))
        
        historial = cursor.fetchall()
        
        # Formatear fechas para JSON
        for registro in historial:
            if registro['fecha_cambio']:
                registro['fecha_cambio'] = registro['fecha_cambio'].isoformat()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'success': True,
            'devolucion_id': devolucion_id,
            'estado_actual': devolucion['estado'],
            'total_cambios': len(historial),
            'historial': historial
        })
        
    except Exception as e:
        print(f"Error al obtener historial de estados: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500

@app.route('/api/estados/validar-transicion', methods=['POST'])
@login_required()
@role_required('logistica')
def validar_transicion_api():
    """
    API para validar si una transición de estado es permitida
    """
    try:
        data = request.get_json() if request.is_json else request.form
        
        estado_actual = data.get('estado_actual')
        estado_nuevo = data.get('estado_nuevo')
        usuario_id = session.get('user_id')
        
        if not estado_actual or not estado_nuevo:
            return jsonify({
                'success': False,
                'error': 'Estado actual y nuevo estado son obligatorios'
            }), 400
        
        validacion = validar_transicion_estado(estado_actual, estado_nuevo, usuario_id)
        
        return jsonify({
            'success': True,
            'validacion': validacion
        })
        
    except Exception as e:
        print(f"Error al validar transición: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Error interno del servidor: {str(e)}'
        }), 500

# ==================== MÓDULO DE LÍMITES FERRETERO ====================

@app.route('/logistica/limites_ferretero')
@login_required()
@role_required('logistica')
def ver_limites_ferretero():
    """Página principal del módulo de límites ferretero"""
    try:
        connection = get_db_connection()
        if connection is None:
            return render_template('error.html', 
                               mensaje='Error de conexión a la base de datos',
                               error='No se pudo establecer conexión con la base de datos')
                               
        cursor = connection.cursor(dictionary=True)
        
        # Obtener todas las áreas de trabajo disponibles
        cursor.execute("SELECT DISTINCT area_trabajo FROM limites_ferretero ORDER BY area_trabajo")
        areas_trabajo = cursor.fetchall()
        
        # Obtener todos los límites configurados
        cursor.execute("""
            SELECT 
                id_limite,
                area_trabajo,
                material_tipo,
                cantidad_limite,
                periodo_dias,
                unidad_medida,
                descripcion,
                activo,
                fecha_actualizacion,
                usuario_actualizacion
            FROM limites_ferretero 
            ORDER BY area_trabajo, material_tipo
        """)
        limites = cursor.fetchall()
        
        cursor.close()
        connection.close()
        
        return render_template('modulos/logistica/limites_ferretero.html',
                             areas_trabajo=areas_trabajo,
                             limites=limites)
                             
    except Exception as e:
        print(f"Error al cargar límites ferretero: {str(e)}")
        return render_template('error.html',
                           mensaje='Error al cargar los límites de ferretero',
                           error=str(e))

@app.route('/api/limites_ferretero', methods=['GET'])
@login_required()
@role_required('logistica')
def obtener_limites_ferretero():
    """API para obtener límites ferretero con filtros opcionales"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'status': 'error',
                'message': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener parámetros de filtro
        area_trabajo = request.args.get('area_trabajo')
        material_tipo = request.args.get('material_tipo')
        activo = request.args.get('activo', 'true').lower() == 'true'
        
        # Construir consulta con filtros
        query = """
            SELECT 
                id_limite,
                area_trabajo,
                material_tipo,
                cantidad_limite,
                periodo_dias,
                unidad_medida,
                descripcion,
                activo,
                fecha_creacion,
                fecha_actualizacion,
                usuario_creacion,
                usuario_actualizacion
            FROM limites_ferretero 
            WHERE activo = %s
        """
        params = [activo]
        
        if area_trabajo:
            query += " AND area_trabajo = %s"
            params.append(area_trabajo)
            
        if material_tipo:
            query += " AND material_tipo = %s"
            params.append(material_tipo)
            
        query += " ORDER BY area_trabajo, material_tipo"
        
        cursor.execute(query, params)
        limites = cursor.fetchall()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'status': 'success',
            'data': limites,
            'total': len(limites)
        })
        
    except Exception as e:
        print(f"Error al obtener límites: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error al obtener límites: {str(e)}'
        }), 500

@app.route('/api/limites_ferretero/<int:id_limite>', methods=['GET'])
@login_required()
@role_required('logistica')
def obtener_limite_ferretero(id_limite):
    """API para obtener un límite específico"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'status': 'error',
                'message': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                id_limite,
                area_trabajo,
                material_tipo,
                cantidad_limite,
                periodo_dias,
                unidad_medida,
                descripcion,
                activo,
                fecha_creacion,
                fecha_actualizacion,
                usuario_creacion,
                usuario_actualizacion
            FROM limites_ferretero 
            WHERE id_limite = %s
        """, (id_limite,))
        
        limite = cursor.fetchone()
        
        cursor.close()
        connection.close()
        
        if limite:
            return jsonify({
                'status': 'success',
                'data': limite
            })
        else:
            return jsonify({
                'status': 'error',
                'message': 'Límite no encontrado'
            }), 404
            
    except Exception as e:
        print(f"Error al obtener límite: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error al obtener límite: {str(e)}'
        }), 500

@app.route('/api/limites_ferretero', methods=['POST'])
@login_required()
@role_required('logistica')
def crear_limite_ferretero():
    """API para crear un nuevo límite"""
    try:
        data = request.get_json() if request.is_json else request.form
        
        # Validar campos requeridos
        campos_requeridos = ['area_trabajo', 'material_tipo', 'cantidad_limite', 'periodo_dias']
        for campo in campos_requeridos:
            if not data.get(campo):
                return jsonify({
                    'status': 'error',
                    'message': f'El campo {campo} es requerido'
                }), 400
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'status': 'error',
                'message': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor()
        
        # Verificar si ya existe un límite para esta combinación
        cursor.execute("""
            SELECT id_limite FROM limites_ferretero 
            WHERE area_trabajo = %s AND material_tipo = %s
        """, (data['area_trabajo'], data['material_tipo']))
        
        if cursor.fetchone():
            return jsonify({
                'status': 'error',
                'message': 'Ya existe un límite para esta área de trabajo y material'
            }), 400
        
        # Insertar nuevo límite
        cursor.execute("""
            INSERT INTO limites_ferretero 
            (area_trabajo, material_tipo, cantidad_limite, periodo_dias, unidad_medida, descripcion, usuario_creacion)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (
            data['area_trabajo'],
            data['material_tipo'],
            int(data['cantidad_limite']),
            int(data['periodo_dias']),
            data.get('unidad_medida', 'unidades'),
            data.get('descripcion', ''),
            session.get('username', 'sistema')
        ))
        
        id_limite = cursor.lastrowid
        connection.commit()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'status': 'success',
            'message': 'Límite creado exitosamente',
            'id_limite': id_limite
        }), 201
        
    except Exception as e:
        print(f"Error al crear límite: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error al crear límite: {str(e)}'
        }), 500

@app.route('/api/limites_ferretero/<int:id_limite>', methods=['PUT'])
@login_required()
@role_required('logistica')
def actualizar_limite_ferretero(id_limite):
    """API para actualizar un límite existente"""
    try:
        data = request.get_json() if request.is_json else request.form
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'status': 'error',
                'message': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor()
        
        # Verificar que el límite existe
        cursor.execute("SELECT id_limite FROM limites_ferretero WHERE id_limite = %s", (id_limite,))
        if not cursor.fetchone():
            return jsonify({
                'status': 'error',
                'message': 'Límite no encontrado'
            }), 404
        
        # Construir consulta de actualización dinámicamente
        campos_actualizables = ['cantidad_limite', 'periodo_dias', 'unidad_medida', 'descripcion', 'activo']
        campos_update = []
        valores = []
        
        for campo in campos_actualizables:
            if campo in data:
                campos_update.append(f"{campo} = %s")
                valores.append(data[campo])
        
        if not campos_update:
            return jsonify({
                'status': 'error',
                'message': 'No se proporcionaron campos para actualizar'
            }), 400
        
        # Agregar usuario y fecha de actualización
        campos_update.append("usuario_actualizacion = %s")
        valores.append(session.get('username', 'sistema'))
        valores.append(id_limite)
        
        query = f"UPDATE limites_ferretero SET {', '.join(campos_update)} WHERE id_limite = %s"
        cursor.execute(query, valores)
        
        connection.commit()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'status': 'success',
            'message': 'Límite actualizado exitosamente'
        })
        
    except Exception as e:
        print(f"Error al actualizar límite: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error al actualizar límite: {str(e)}'
        }), 500

@app.route('/api/limites_ferretero/<int:id_limite>', methods=['DELETE'])
@login_required()
@role_required('logistica')
def eliminar_limite_ferretero(id_limite):
    """API para eliminar (desactivar) un límite"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'status': 'error',
                'message': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor()
        
        # Verificar que el límite existe
        cursor.execute("SELECT id_limite FROM limites_ferretero WHERE id_limite = %s", (id_limite,))
        if not cursor.fetchone():
            return jsonify({
                'status': 'error',
                'message': 'Límite no encontrado'
            }), 404
        
        # Desactivar el límite en lugar de eliminarlo
        cursor.execute("""
            UPDATE limites_ferretero 
            SET activo = FALSE, usuario_actualizacion = %s 
            WHERE id_limite = %s
        """, (session.get('username', 'sistema'), id_limite))
        
        connection.commit()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'status': 'success',
            'message': 'Límite desactivado exitosamente'
        })
        
    except Exception as e:
        print(f"Error al eliminar límite: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error al eliminar límite: {str(e)}'
        }), 500

@app.route('/api/limites_ferretero/areas', methods=['GET'])
@login_required()
@role_required('logistica')
def obtener_areas_trabajo():
    """API para obtener todas las áreas de trabajo disponibles"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'status': 'error',
                'message': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor(dictionary=True)
        
        cursor.execute("SELECT DISTINCT area_trabajo FROM limites_ferretero ORDER BY area_trabajo")
        areas = cursor.fetchall()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'status': 'success',
            'data': [area['area_trabajo'] for area in areas]
        })
        
    except Exception as e:
        print(f"Error al obtener áreas: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error al obtener áreas: {str(e)}'
        }), 500

@app.route('/api/limites_ferretero/historial/<int:id_limite>', methods=['GET'])
@login_required()
@role_required('logistica')
def obtener_historial_limite(id_limite):
    """API para obtener el historial de cambios de un límite"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'status': 'error',
                'message': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                id_historial,
                area_trabajo,
                material_tipo,
                cantidad_limite_anterior,
                cantidad_limite_nueva,
                periodo_dias_anterior,
                periodo_dias_nuevo,
                accion,
                usuario,
                fecha_cambio,
                observaciones
            FROM historial_limites_ferretero 
            WHERE id_limite = %s 
            ORDER BY fecha_cambio DESC
        """, (id_limite,))
        
        historial = cursor.fetchall()
        
        cursor.close()
        connection.close()
        
        return jsonify({
            'status': 'success',
            'data': historial
        })
        
    except Exception as e:
        print(f"Error al obtener historial: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'Error al obtener historial: {str(e)}'
        }), 500

# ============================================================================
# MÓDULO SSTT (SEGURIDAD Y SALUD EN EL TRABAJO) - RUTAS Y API ENDPOINTS
# ============================================================================

@app.route('/sstt')
@login_required()
def sstt_dashboard():
    """Dashboard principal del módulo SSTT"""
    if session.get('user_role') not in ['sstt', 'administrativo']:
        flash('No tienes permisos para acceder a este módulo', 'error')
        return redirect(url_for('dashboard'))
    return render_template('modulos/sstt/dashboard.html')

@app.route('/sstt/inspecciones')
@login_required()
def sstt_inspecciones():
    """Página de gestión de inspecciones de seguridad"""
    if session.get('user_role') not in ['sstt', 'administrativo']:
        flash('No tienes permisos para acceder a este módulo', 'error')
        return redirect(url_for('dashboard'))
    return render_template('modulos/sstt/inspecciones.html')

@app.route('/sstt/capacitaciones')
@login_required()
def sstt_capacitaciones():
    """Página de gestión de capacitaciones"""
    if session.get('user_role') not in ['sstt', 'administrativo']:
        flash('No tienes permisos para acceder a este módulo', 'error')
        return redirect(url_for('dashboard'))
    return render_template('modulos/sstt/capacitaciones.html')

@app.route('/sstt/incidentes')
@login_required()
def sstt_incidentes():
    """Página de gestión de incidentes laborales"""
    if session.get('user_role') not in ['sstt', 'administrativo']:
        flash('No tienes permisos para acceder a este módulo', 'error')
        return redirect(url_for('dashboard'))
    return render_template('modulos/sstt/incidentes.html')

@app.route('/sstt/epp')
@login_required()
def sstt_epp():
    """Página de control de EPP (Elementos de Protección Personal)"""
    if session.get('user_role') not in ['sstt', 'administrativo']:
        flash('No tienes permisos para acceder a este módulo', 'error')
        return redirect(url_for('dashboard'))
    return render_template('modulos/sstt/epp.html')

# API Endpoints para SSTT

@app.route('/api/sstt/tipos-riesgo', methods=['GET'])
@login_required()
def api_sstt_tipos_riesgo():
    """API para obtener tipos de riesgo"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        cursor.execute("SELECT * FROM sstt_tipos_riesgo ORDER BY nombre_riesgo")
        tipos_riesgo = cursor.fetchall()
        
        return jsonify({'tipos_riesgo': tipos_riesgo})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

@app.route('/api/sstt/inspecciones', methods=['GET', 'POST'])
@login_required()
def api_sstt_inspecciones():
    """API para gestionar inspecciones de seguridad"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        if request.method == 'GET':
            # Obtener inspecciones con filtros opcionales
            fecha_desde = request.args.get('fecha_desde')
            fecha_hasta = request.args.get('fecha_hasta')
            estado = request.args.get('estado')
            
            query = """
                SELECT i.*, tr.nombre_riesgo, ro.nombre as inspector_nombre
                FROM sstt_inspecciones i
                LEFT JOIN sstt_tipos_riesgo tr ON i.tipo_riesgo_id = tr.id
                LEFT JOIN recurso_operativo ro ON i.usuario_creador = ro.id_codigo_consumidor
                WHERE 1=1
            """
            params = []
            
            if fecha_desde:
                query += " AND i.fecha_inspeccion >= %s"
                params.append(fecha_desde)
            if fecha_hasta:
                query += " AND i.fecha_inspeccion <= %s"
                params.append(fecha_hasta)
            if estado:
                query += " AND i.estado = %s"
                params.append(estado)
                
            query += " ORDER BY i.fecha_inspeccion DESC"
            
            cursor.execute(query, params)
            inspecciones = cursor.fetchall()
            
            return jsonify({'inspecciones': inspecciones})
            
        elif request.method == 'POST':
            # Crear nueva inspección
            data = request.get_json()
            
            cursor.execute("""
                INSERT INTO sstt_inspecciones 
                (area_inspeccionada, fecha_inspeccion, tipo_riesgo_id, descripcion, 
                 estado, usuario_creador, fecha_creacion)
                VALUES (%s, %s, %s, %s, %s, %s, NOW())
            """, (
                data['area_inspeccionada'],
                data['fecha_inspeccion'],
                data['tipo_riesgo_id'],
                data['descripcion'],
                data.get('estado', 'pendiente'),
                session.get('user_id')
            ))
            
            connection.commit()
            inspeccion_id = cursor.lastrowid
            
            return jsonify({
                'success': True, 
                'message': 'Inspección creada exitosamente',
                'inspeccion_id': inspeccion_id
            })
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

@app.route('/api/sstt/capacitaciones', methods=['GET', 'POST'])
@login_required()
def api_sstt_capacitaciones():
    """API para gestionar capacitaciones"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        if request.method == 'GET':
            cursor.execute("""
                SELECT c.*, ro.nombre as instructor_nombre
                FROM sstt_capacitaciones c
                LEFT JOIN recurso_operativo ro ON c.usuario_creador = ro.id_codigo_consumidor
                ORDER BY c.fecha_programada DESC
            """)
            capacitaciones = cursor.fetchall()
            
            return jsonify({'capacitaciones': capacitaciones})
            
        elif request.method == 'POST':
            data = request.get_json()
            
            cursor.execute("""
                INSERT INTO sstt_capacitaciones 
                (titulo, descripcion, fecha_programada, duracion_horas, 
                 ubicacion, estado, usuario_creador, fecha_creacion)
                VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
            """, (
                data['titulo'],
                data['descripcion'],
                data['fecha_programada'],
                data['duracion_horas'],
                data['ubicacion'],
                data.get('estado', 'programada'),
                session.get('user_id')
            ))
            
            connection.commit()
            capacitacion_id = cursor.lastrowid
            
            return jsonify({
                'success': True, 
                'message': 'Capacitación creada exitosamente',
                'capacitacion_id': capacitacion_id
            })
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

@app.route('/api/sstt/incidentes', methods=['GET', 'POST'])
@login_required()
def api_sstt_incidentes():
    """API para gestionar incidentes laborales"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        if request.method == 'GET':
            cursor.execute("""
                SELECT i.*, ro.nombre as reporta_nombre
                FROM sstt_incidentes i
                LEFT JOIN recurso_operativo ro ON i.usuario_reporta = ro.id_codigo_consumidor
                ORDER BY i.fecha_incidente DESC
            """)
            incidentes = cursor.fetchall()
            
            return jsonify({'incidentes': incidentes})
            
        elif request.method == 'POST':
            data = request.get_json()
            
            cursor.execute("""
                INSERT INTO sstt_incidentes 
                (fecha_incidente, ubicacion, descripcion, tipo_incidente, 
                 gravedad, usuario_reporta, fecha_reporte)
                VALUES (%s, %s, %s, %s, %s, %s, NOW())
            """, (
                data['fecha_incidente'],
                data['ubicacion'],
                data['descripcion'],
                data['tipo_incidente'],
                data['gravedad'],
                session.get('user_id')
            ))
            
            connection.commit()
            incidente_id = cursor.lastrowid
            
            return jsonify({
                'success': True, 
                'message': 'Incidente reportado exitosamente',
                'incidente_id': incidente_id
            })
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

@app.route('/api/sstt/epp', methods=['GET', 'POST'])
@login_required()
def api_sstt_epp():
    """API para control de EPP (Elementos de Protección Personal)"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        if request.method == 'GET':
            cursor.execute("""
                SELECT e.*, ro.nombre as usuario_nombre
                FROM sstt_epp_control e
                LEFT JOIN recurso_operativo ro ON e.usuario_id = ro.id_codigo_consumidor
                ORDER BY e.fecha_entrega DESC
            """)
            epp_registros = cursor.fetchall()
            
            return jsonify({'epp_registros': epp_registros})
            
        elif request.method == 'POST':
            data = request.get_json()
            
            cursor.execute("""
                INSERT INTO sstt_epp_control 
                (usuario_id, tipo_epp, marca, modelo, fecha_entrega, 
                 fecha_vencimiento, estado, observaciones)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                data['usuario_id'],
                data['tipo_epp'],
                data['marca'],
                data['modelo'],
                data['fecha_entrega'],
                data.get('fecha_vencimiento'),
                data.get('estado', 'activo'),
                data.get('observaciones', '')
            ))
            
            connection.commit()
            epp_id = cursor.lastrowid
            
            return jsonify({
                'success': True, 
                'message': 'Registro de EPP creado exitosamente',
                'epp_id': epp_id
            })
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

@app.route('/api/sstt/dashboard-stats', methods=['GET'])
@login_required()
def api_sstt_dashboard_stats():
    """API para estadísticas del dashboard SSTT"""
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        stats = {}
        
        # Estadísticas de inspecciones
        cursor.execute("SELECT COUNT(*) as total FROM sstt_inspecciones")
        stats['total_inspecciones'] = cursor.fetchone()['total']
        
        cursor.execute("SELECT COUNT(*) as pendientes FROM sstt_inspecciones WHERE estado = 'pendiente'")
        stats['inspecciones_pendientes'] = cursor.fetchone()['pendientes']
        
        # Estadísticas de capacitaciones
        cursor.execute("SELECT COUNT(*) as total FROM sstt_capacitaciones")
        stats['total_capacitaciones'] = cursor.fetchone()['total']
        
        cursor.execute("SELECT COUNT(*) as programadas FROM sstt_capacitaciones WHERE estado = 'programada'")
        stats['capacitaciones_programadas'] = cursor.fetchone()['programadas']
        
        # Estadísticas de incidentes
        cursor.execute("SELECT COUNT(*) as total FROM sstt_incidentes")
        stats['total_incidentes'] = cursor.fetchone()['total']
        
        cursor.execute("""
            SELECT COUNT(*) as recientes 
            FROM sstt_incidentes 
            WHERE fecha_incidente >= DATE_SUB(NOW(), INTERVAL 30 DAY)
        """)
        stats['incidentes_mes'] = cursor.fetchone()['recientes']
        
        # Estadísticas de EPP
        cursor.execute("SELECT COUNT(*) as total FROM sstt_epp_control")
        stats['total_epp'] = cursor.fetchone()['total']
        
        cursor.execute("""
            SELECT COUNT(*) as vencidos 
            FROM sstt_epp_control 
            WHERE fecha_vencimiento <= NOW() AND estado = 'activo'
        """)
        stats['epp_vencidos'] = cursor.fetchone()['vencidos']
        
        return jsonify(stats)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

# ============================================================================
# ENDPOINTS PARA INICIO DE OPERACIÓN - NUEVOS CAMPOS DE ASISTENCIA
# ============================================================================

@app.route('/api/asistencia/actualizar-campo', methods=['POST'])
@login_required_api()
def actualizar_campo_asistencia():
    """Actualizar un campo específico de asistencia para inicio de operación"""
    try:
        data = request.get_json()
        
        # Validar campos requeridos
        required_fields = ['cedula', 'campo', 'valor']
        for field in required_fields:
            if field not in data:
                return jsonify({
                    'success': False, 
                    'message': f'Campo requerido faltante: {field}'
                }), 400
        
        cedula = data['cedula']
        campo = data['campo']
        valor = data['valor']
        fecha_filtro = data.get('fecha')  # Obtener fecha del frontend
        
        # Validar que el campo sea uno de los permitidos
        campos_permitidos = ['hora_inicio', 'estado', 'novedad']
        if campo not in campos_permitidos:
            return jsonify({
                'success': False,
                'message': f'Campo no permitido: {campo}. Campos válidos: {campos_permitidos}'
            }), 400
        
        # Validar valores específicos para el campo 'estado'
        if campo == 'estado':
            estados_validos = ['CUMPLE', 'NO CUMPLE', 'NOVEDAD', 'NO APLICA']
            if valor and valor not in estados_validos:
                return jsonify({
                    'success': False,
                    'message': f'Estado no válido: {valor}. Estados válidos: {estados_validos}'
                }), 400
        
        # Validar formato de hora para 'hora_inicio'
        if campo == 'hora_inicio' and valor:
            try:
                # Validar formato HH:MM
                datetime.strptime(valor, '%H:%M')
            except ValueError:
                return jsonify({
                    'success': False,
                    'message': 'Formato de hora inválido. Use HH:MM (ej: 08:30)'
                }), 400
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'success': False, 
                'message': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Determinar fecha a usar para la búsqueda
        bogota_tz = pytz.timezone('America/Bogota')
        if fecha_filtro:
            # Usar la fecha enviada desde el frontend
            try:
                fecha_busqueda = datetime.strptime(fecha_filtro, '%Y-%m-%d').date()
            except ValueError:
                return jsonify({
                    'success': False,
                    'message': 'Formato de fecha inválido. Use YYYY-MM-DD'
                }), 400
        else:
            # Si no se envía fecha, usar la fecha actual
            fecha_busqueda = datetime.now(bogota_tz).date()
        
        # Buscar registro de asistencia para la fecha especificada y esta cédula
        cursor.execute("""
            SELECT id_asistencia, DATE(fecha_asistencia) as fecha_registro
            FROM asistencia 
            WHERE cedula = %s AND DATE(fecha_asistencia) = %s
            ORDER BY fecha_asistencia DESC
            LIMIT 1
        """, (cedula, fecha_busqueda))
        
        resultado = cursor.fetchone()
        
        if not resultado:
            # No existe registro para la fecha especificada, crear uno nuevo
            try:
                # Obtener información del técnico desde la tabla recurso_operativo
                cursor.execute("SELECT nombre FROM recurso_operativo WHERE recurso_operativo_cedula = %s", (cedula,))
                usuario = cursor.fetchone()
                
                if not usuario:
                    # Si no está en recurso_operativo, usar la cédula como nombre por defecto
                    nombre_tecnico = f"Técnico {cedula}"
                else:
                    nombre_tecnico = usuario['nombre'] or f"Técnico {cedula}"
                
                # Crear nuevo registro de asistencia para la fecha especificada
                fecha_completa = datetime.combine(fecha_busqueda, datetime.now(bogota_tz).time())
                cursor.execute("""
                    INSERT INTO asistencia (cedula, tecnico, carpeta_dia, carpeta, super, fecha_asistencia, estado, novedad)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, (cedula, nombre_tecnico, '', '', '', fecha_completa, None, None))
                
                id_asistencia = cursor.lastrowid
                
                app.logger.info(f"Creado nuevo registro de asistencia para cédula {cedula} con ID {id_asistencia}")
                
            except Exception as e:
                app.logger.error(f"Error creando registro de asistencia: {str(e)}")
                return jsonify({
                    'success': False,
                    'message': f'Error creando registro de asistencia: {str(e)}'
                }), 500
        else:
            id_asistencia = resultado['id_asistencia']
            app.logger.info(f"Registro encontrado - ID: {id_asistencia}, Fecha: {resultado['fecha_registro']}")
        
        # Verificar que el registro aún existe antes de actualizar
        cursor.execute("SELECT id_asistencia, estado, hora_inicio FROM asistencia WHERE id_asistencia = %s", (id_asistencia,))
        verificacion = cursor.fetchone()
        if not verificacion:
            app.logger.error(f"CRÍTICO: El registro {id_asistencia} ya no existe en la base de datos")
            return jsonify({
                'success': False,
                'message': f'El registro {id_asistencia} no existe'
            }), 500
        
        app.logger.info(f"Verificación antes de UPDATE - ID: {id_asistencia}, Estado actual: {verificacion.get('estado')}, Hora actual: {verificacion.get('hora_inicio')}")
        
        # Actualizar el campo específico
        app.logger.info(f"Preparando actualización - ID: {id_asistencia}, Campo: {campo}, Valor: {valor}")
        query = f"UPDATE asistencia SET {campo} = %s WHERE id_asistencia = %s"
        app.logger.info(f"Ejecutando query: {query} con valores: ({valor}, {id_asistencia})")
        cursor.execute(query, (valor if valor else None, id_asistencia))
        
        app.logger.info(f"Filas afectadas: {cursor.rowcount}")
        
        # Verificar si la actualización fue exitosa o si el valor ya era el mismo
        if cursor.rowcount > 0:
            connection.commit()
            app.logger.info(f"Campo {campo} actualizado exitosamente para cédula {cedula}")
            return jsonify({
                'success': True,
                'message': f'Campo {campo} actualizado correctamente para cédula {cedula}'
            })
        else:
            # Verificar si el valor ya era el mismo (no es un error)
            cursor.execute(f"SELECT {campo} FROM asistencia WHERE id_asistencia = %s", (id_asistencia,))
            valor_actual = cursor.fetchone()
            
            # Comparar valores considerando diferentes formatos
            valores_iguales = False
            if valor_actual:
                val_actual = valor_actual[campo]
                if campo == 'hora_inicio':
                    # Para hora_inicio, comparar considerando formatos diferentes
                    # Convertir ambos a formato HH:MM para comparar
                    try:
                        if val_actual:
                            # Convertir 8:30:00 a 08:30 para comparar
                            hora_actual = str(val_actual).split(':')[:2]  # Tomar solo HH:MM
                            hora_actual_fmt = f"{hora_actual[0].zfill(2)}:{hora_actual[1]}"
                            hora_nueva_fmt = str(valor)
                            if ':' not in hora_nueva_fmt and len(hora_nueva_fmt) == 4:
                                # Convertir 0830 a 08:30
                                hora_nueva_fmt = f"{hora_nueva_fmt[:2]}:{hora_nueva_fmt[2:]}"
                            valores_iguales = hora_actual_fmt == hora_nueva_fmt
                            app.logger.info(f"Comparación hora - Actual: {hora_actual_fmt}, Nueva: {hora_nueva_fmt}, Iguales: {valores_iguales}")
                    except Exception as e:
                        app.logger.error(f"Error comparando horas: {e}")
                        valores_iguales = str(val_actual) == str(valor)
                else:
                    # Para otros campos, comparación directa
                    valores_iguales = str(val_actual) == str(valor)
            
            if valores_iguales:
                app.logger.info(f"Campo {campo} ya tenía el valor {valor} para cédula {cedula} - no requiere actualización")
                connection.commit()  # Commit por si acaso
                return jsonify({
                    'success': True,
                    'message': f'Campo {campo} ya tenía el valor correcto para cédula {cedula}'
                })
            else:
                app.logger.error(f"No se pudo actualizar el registro. ID: {id_asistencia}, Campo: {campo}, Valor: {valor}")
                return jsonify({
                    'success': False,
                    'message': 'No se pudo actualizar el registro'
                }), 500
        
    except Exception as e:
        return jsonify({
            'success': False, 
            'message': f'Error: {str(e)}'
        }), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

@app.route('/api/asistencia/obtener-datos', methods=['GET'])
@login_required_api()
def obtener_datos_asistencia():
    """Obtener datos existentes de asistencia para una cédula específica"""
    try:
        cedula = request.args.get('cedula')
        
        if not cedula:
            return jsonify({
                'success': False,
                'message': 'Cédula requerida'
            }), 400
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'success': False, 
                'message': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener fecha actual en zona horaria de Bogotá
        bogota_tz = pytz.timezone('America/Bogota')
        fecha_actual = datetime.now(bogota_tz).date()
        
        # Buscar registro de asistencia para hoy y esta cédula
        cursor.execute("""
            SELECT 
                id_asistencia,
                cedula,
                tecnico,
                hora_inicio,
                estado,
                novedad,
                fecha_asistencia
            FROM asistencia 
            WHERE cedula = %s AND DATE(fecha_asistencia) = %s
            ORDER BY fecha_asistencia DESC
            LIMIT 1
        """, (cedula, fecha_actual))
        
        resultado = cursor.fetchone()
        
        if resultado:
            # Formatear hora_inicio si existe
            if resultado['hora_inicio']:
                # El campo time se devuelve como timedelta desde MySQL
                if hasattr(resultado['hora_inicio'], 'total_seconds'):
                    # Es un timedelta
                    total_seconds = int(resultado['hora_inicio'].total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    resultado['hora_inicio'] = f"{hours:02d}:{minutes:02d}"
                else:
                    # Es un string o datetime.time
                    resultado['hora_inicio'] = str(resultado['hora_inicio'])[:5]  # HH:MM
            
            return jsonify({
                'success': True,
                'data': resultado,
                'exists': True
            })
        else:
            return jsonify({
                'success': True,
                'data': {
                    'cedula': cedula,
                    'hora_inicio': '',
                    'estado': '',
                    'novedad': ''
                },
                'exists': False
            })
        
    except Exception as e:
        return jsonify({
            'success': False, 
            'message': f'Error: {str(e)}'
        }), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

@app.route('/api/asistencia/actualizar-masivo', methods=['POST'])
@login_required_api()
def actualizar_masivo_asistencia():
    """Actualizar múltiples registros de asistencia para inicio de operación"""
    try:
        data = request.get_json()
        
        # Validar que se envíen datos
        if not data.get('asistencias'):
            return jsonify({
                'success': False,
                'message': 'No se enviaron datos de asistencias'
            }), 400
        
        asistencias = data['asistencias']
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'success': False,
                'message': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Obtener fecha actual en zona horaria de Bogotá
        bogota_tz = pytz.timezone('America/Bogota')
        fecha_actual = datetime.now(bogota_tz).date()
        
        actualizaciones_exitosas = 0
        errores = []
        
        for asistencia in asistencias:
            try:
                cedula = asistencia.get('cedula')
                hora_inicio = asistencia.get('hora_inicio')
                estado = asistencia.get('estado')
                novedad = asistencia.get('novedad')
                
                if not cedula:
                    errores.append(f'Cédula faltante en uno de los registros')
                    continue
                
                # Validar estado si se proporciona
                if estado:
                    estados_validos = ['CUMPLE', 'NO CUMPLE', 'NOVEDAD', 'NO APLICA']
                    if estado not in estados_validos:
                        errores.append(f'Estado inválido para cédula {cedula}: {estado}')
                        continue
                
                # Validar formato de hora si se proporciona
                if hora_inicio:
                    try:
                        datetime.strptime(hora_inicio, '%H:%M')
                    except ValueError:
                        errores.append(f'Formato de hora inválido para cédula {cedula}: {hora_inicio}')
                        continue
                
                # Buscar registro de asistencia para hoy y esta cédula
                cursor.execute("""
                    SELECT id_asistencia
                    FROM asistencia 
                    WHERE cedula = %s AND DATE(fecha_asistencia) = %s
                    ORDER BY fecha_asistencia DESC
                    LIMIT 1
                """, (cedula, fecha_actual))
                
                resultado = cursor.fetchone()
                
                if not resultado:
                    errores.append(f'No se encontró registro de asistencia para cédula {cedula}')
                    continue
                
                id_asistencia = resultado[0]
                
                # Actualizar los campos
                cursor.execute("""
                    UPDATE asistencia 
                    SET hora_inicio = %s, estado = %s, novedad = %s
                    WHERE id_asistencia = %s
                """, (
                    hora_inicio if hora_inicio else None,
                    estado if estado else None,
                    novedad if novedad else None,
                    id_asistencia
                ))
                
                if cursor.rowcount > 0:
                    actualizaciones_exitosas += 1
                else:
                    errores.append(f'No se pudo actualizar el registro para cédula {cedula}')
                    
            except Exception as e:
                errores.append(f'Error procesando cédula {cedula}: {str(e)}')
                continue
        
        # Confirmar cambios si hubo actualizaciones exitosas
        if actualizaciones_exitosas > 0:
            connection.commit()
        
        # Preparar respuesta
        response_data = {
            'success': actualizaciones_exitosas > 0,
            'actualizaciones_exitosas': actualizaciones_exitosas,
            'total_procesados': len(asistencias)
        }
        
        if errores:
            response_data['errores'] = errores
            response_data['message'] = f'Se actualizaron {actualizaciones_exitosas} de {len(asistencias)} registros. {len(errores)} errores encontrados.'
        else:
            response_data['message'] = f'Se actualizaron {actualizaciones_exitosas} registros exitosamente'
        
        return jsonify(response_data)
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        }), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

@app.route('/api/asistencia/obtener-campos-inicio', methods=['GET'])
@login_required(role='administrativo')
def obtener_campos_inicio_operacion():
    """Obtener los campos de inicio de operación para una fecha específica"""
    try:
        # Obtener parámetros
        fecha = request.args.get('fecha')
        supervisor = request.args.get('supervisor')
        
        # Si no se proporciona fecha, usar la fecha actual
        if not fecha:
            bogota_tz = pytz.timezone('America/Bogota')
            fecha = datetime.now(bogota_tz).date().strftime('%Y-%m-%d')
        
        try:
            fecha_obj = datetime.strptime(fecha, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({
                'success': False,
                'message': 'Formato de fecha inválido. Use YYYY-MM-DD'
            }), 400
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({
                'success': False,
                'message': 'Error de conexión a la base de datos'
            }), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Construir consulta base
        query = """
            SELECT 
                a.cedula,
                a.tecnico,
                a.carpeta,
                a.super as supervisor,
                a.carpeta_dia,
                a.hora_inicio,
                a.estado,
                a.novedad,
                a.fecha_asistencia
            FROM asistencia a
            WHERE DATE(a.fecha_asistencia) = %s
        """
        
        params = [fecha_obj]
        
        # Agregar filtro por supervisor si se proporciona
        if supervisor:
            query += " AND a.super = %s"
            params.append(supervisor)
        
        query += " ORDER BY a.tecnico"
        
        cursor.execute(query, params)
        registros = cursor.fetchall()
        
        # Formatear los datos para el frontend
        datos_formateados = []
        for registro in registros:
            datos_formateados.append({
                'cedula': registro['cedula'],
                'tecnico': registro['tecnico'],
                'carpeta': registro['carpeta'],
                'supervisor': registro['supervisor'],
                'carpeta_dia': registro['carpeta_dia'],
                'hora_inicio': registro['hora_inicio'].strftime('%H:%M') if registro['hora_inicio'] else '',
                'estado': registro['estado'] or '',
                'novedad': registro['novedad'] or '',
                'fecha_asistencia': registro['fecha_asistencia'].strftime('%Y-%m-%d %H:%M:%S')
            })
        
        return jsonify({
            'success': True,
            'data': datos_formateados,
            'fecha': fecha,
            'total_registros': len(datos_formateados)
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        }), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

@app.route('/api/asistencia/graficas/operacion-recurso', methods=['GET'])
@login_required_api(role='administrativo')
def obtener_datos_grafica_operacion_recurso():
    """Obtener datos temporales para la gráfica de Operación x Recurso Operativo"""
    try:
        # Obtener parámetros
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        supervisor_filtro = request.args.get('supervisor')
        tipo_agrupacion = request.args.get('agrupacion', 'mes')  # 'mes' o 'dia'
        
        # Validar fechas
        if not fecha_inicio or not fecha_fin:
            return jsonify({
                'success': False,
                'message': 'Se requieren fecha_inicio y fecha_fin'
            }), 400
        
        try:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({
                'success': False,
                'message': 'Formato de fecha inválido. Use YYYY-MM-DD'
            }), 400
        
        # Validar rango de fechas
        if fecha_inicio > fecha_fin:
            return jsonify({
                'success': False,
                'message': 'La fecha de inicio no puede ser mayor que la fecha de fin'
            }), 400
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Determinar el formato de agrupación temporal
        if tipo_agrupacion == 'mes':
            formato_fecha = "DATE_FORMAT(a.fecha_asistencia, '%Y-%m')"
            formato_orden = "DATE_FORMAT(a.fecha_asistencia, '%Y-%m')"
        else:  # dia
            formato_fecha = "DATE(a.fecha_asistencia)"
            formato_orden = "DATE(a.fecha_asistencia)"
        
        # Consulta para obtener datos agrupados por tiempo y grupo (sin tabla tipificacion_asistencia)
        query = f"""
            SELECT 
                {formato_fecha} as periodo,
                CASE 
                    WHEN a.carpeta IN ('ARREGLOS HFC', 'MANTENIMIENTO FTTH','DX') THEN 'ARREGLOS'
                    WHEN a.carpeta IN ('BROWNFIELD', 'FTTH INSTALACIONES', 'INSTALACIONES DOBLES') THEN 'INSTALACIONES'
                    WHEN a.carpeta IN ('POSTVENTA', 'POSTVENTA FTTH') THEN 'POSTVENTA'
                    ELSE 'OTROS'
                END as grupo,
                COUNT(DISTINCT a.id_codigo_consumidor) as total_tecnicos
            FROM asistencia a
            WHERE DATE(a.fecha_asistencia) BETWEEN %s AND %s
                AND a.carpeta IS NOT NULL
                AND a.carpeta != ''
                AND a.carpeta IN ('DX','ARREGLOS HFC', 'MANTENIMIENTO FTTH', 'BROWNFIELD', 'FTTH INSTALACIONES', 'INSTALACIONES DOBLES', 'POSTVENTA', 'POSTVENTA FTTH')
        """
        
        params = [fecha_inicio, fecha_fin]
        
        # Agregar filtro por supervisor si se proporciona
        if supervisor_filtro:
            query += " AND a.super = %s"
            params.append(supervisor_filtro)
        
        query += f"""
            GROUP BY {formato_fecha}, CASE 
                WHEN a.carpeta IN ('ARREGLOS HFC', 'MANTENIMIENTO FTTH','DX') THEN 'ARREGLOS'
                WHEN a.carpeta IN ('BROWNFIELD', 'FTTH INSTALACIONES', 'INSTALACIONES DOBLES') THEN 'INSTALACIONES'
                WHEN a.carpeta IN ('POSTVENTA', 'POSTVENTA FTTH') THEN 'POSTVENTA'
                ELSE 'OTROS'
            END
            ORDER BY {formato_orden}, grupo
        """
        
        # Log the SQL query for debugging
        logging.info(f"Executing SQL query for operacion-recurso: {query}")
        logging.info(f"Query parameters: {params}")
        
        try:
            cursor.execute(query, tuple(params))
            resultados = cursor.fetchall()
            logging.info(f"Query executed successfully, returned {len(resultados)} rows")
        except Exception as sql_error:
            logging.error(f"SQL execution error in operacion-recurso endpoint: {str(sql_error)}")
            logging.error(f"Failed query: {query}")
            logging.error(f"Query parameters: {params}")
            logging.error(f"tipo_agrupacion: {tipo_agrupacion}")
            raise sql_error
        
        # Organizar datos para la gráfica
        datos_grafica = {}
        periodos = set()
        grupos = ['ARREGLOS', 'INSTALACIONES', 'POSTVENTA']
        
        # Inicializar estructura de datos
        for grupo in grupos:
            datos_grafica[grupo] = {}
        
        # Procesar resultados
        for resultado in resultados:
            periodo = str(resultado['periodo'])
            grupo = resultado['grupo']
            total = resultado['total_tecnicos']
            
            periodos.add(periodo)
            if grupo in datos_grafica:
                datos_grafica[grupo][periodo] = total
        
        # Convertir a lista ordenada de períodos
        periodos_ordenados = sorted(list(periodos))
        
        # Preparar datos finales para Chart.js
        datasets = []
        colores = {
            'ARREGLOS': '#FF6384',
            'INSTALACIONES': '#36A2EB', 
            'POSTVENTA': '#FFCE56'
        }
        
        for grupo in grupos:
            data = []
            for periodo in periodos_ordenados:
                data.append(datos_grafica[grupo].get(periodo, 0))
            
            datasets.append({
                'label': grupo,
                'data': data,
                'borderColor': colores[grupo],
                'backgroundColor': colores[grupo] + '20',
                'fill': False,
                'tension': 0.1
            })
        
        return jsonify({
            'success': True,
            'datos': datasets,
            'labels': periodos_ordenados,
            'agrupacion': tipo_agrupacion,
            'fecha_inicio': fecha_inicio.strftime('%Y-%m-%d'),
            'fecha_fin': fecha_fin.strftime('%Y-%m-%d'),
            'supervisor_filtro': supervisor_filtro
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

@app.route('/api/asistencia/graficas/carpeta-dia', methods=['GET'])
@login_required_api(role='administrativo')
def obtener_datos_grafica_carpeta_dia():
    """Obtener datos temporales para la gráfica de Operación x Carpeta Día"""
    try:
        # Obtener parámetros
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        supervisor_filtro = request.args.get('supervisor')
        tipo_agrupacion = request.args.get('agrupacion', 'mes')  # 'mes' o 'dia'
        
        # Validar fechas
        if not fecha_inicio or not fecha_fin:
            return jsonify({
                'success': False,
                'message': 'Se requieren fecha_inicio y fecha_fin'
            }), 400
        
        try:
            fecha_inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({
                'success': False,
                'message': 'Formato de fecha inválido. Use YYYY-MM-DD'
            }), 400
        
        # Validar rango de fechas
        if fecha_inicio > fecha_fin:
            return jsonify({
                'success': False,
                'message': 'La fecha de inicio no puede ser mayor que la fecha de fin'
            }), 400
        
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Determinar el formato de agrupación temporal
        if tipo_agrupacion == 'mes':
            formato_fecha = "DATE_FORMAT(a.fecha_asistencia, '%Y-%m')"
            formato_orden = "DATE_FORMAT(a.fecha_asistencia, '%Y-%m')"
        else:  # dia
            formato_fecha = "DATE(a.fecha_asistencia)"
            formato_orden = "DATE(a.fecha_asistencia)"
        
        # Consulta para obtener datos agrupados por tiempo y grupo (basado en carpeta_dia)
        query = f"""
            SELECT 
                {formato_fecha} as periodo,
                CASE 
                    WHEN a.carpeta_dia IN ('ARGHFC', 'MTFTTH') THEN 'ARREGLOS'
                    WHEN a.carpeta_dia IN ('BROW', 'FTTHI', 'HFCI') THEN 'INSTALACIONES'
                    WHEN a.carpeta_dia IN ('POST', 'FTTHPOST') THEN 'POSTVENTA'
                    ELSE 'OTROS'
                END as grupo,
                COUNT(DISTINCT a.id_codigo_consumidor) as total_tecnicos
            FROM asistencia a
            WHERE DATE(a.fecha_asistencia) BETWEEN %s AND %s
                AND a.carpeta_dia IS NOT NULL 
                AND a.carpeta_dia != ''
                AND a.carpeta_dia IN ('ARGHFC', 'MTFTTH', 'BROW', 'FTTHI', 'HFCI', 'POST', 'FTTHPOST')
        """
        
        params = [fecha_inicio, fecha_fin]
        
        # Agregar filtro por supervisor si se proporciona
        if supervisor_filtro:
            query += " AND a.super = %s"
            params.append(supervisor_filtro)
        
        query += f"""
            GROUP BY {formato_fecha}, CASE 
                WHEN a.carpeta_dia IN ('ARGHFC', 'MTFTTH') THEN 'ARREGLOS'
                WHEN a.carpeta_dia IN ('BROW', 'FTTHI', 'HFCI') THEN 'INSTALACIONES'
                WHEN a.carpeta_dia IN ('POST', 'FTTHPOST') THEN 'POSTVENTA'
                ELSE 'OTROS'
            END
            ORDER BY {formato_orden}, grupo
        """
        
        # Log the SQL query for debugging
        logging.info(f"Executing SQL query for carpeta-dia: {query}")
        logging.info(f"Query parameters: {params}")
        
        try:
            cursor.execute(query, tuple(params))
            resultados = cursor.fetchall()
            logging.info(f"Query executed successfully, returned {len(resultados)} rows")
        except Exception as sql_error:
            logging.error(f"SQL execution error in carpeta-dia endpoint: {str(sql_error)}")
            logging.error(f"Failed query: {query}")
            logging.error(f"Query parameters: {params}")
            logging.error(f"tipo_agrupacion: {tipo_agrupacion}")
            raise sql_error
        
        # Organizar datos para la gráfica
        datos_grafica = {}
        periodos = set()
        grupos = ['ARREGLOS', 'INSTALACIONES', 'POSTVENTA']
        
        # Inicializar estructura de datos
        for grupo in grupos:
            datos_grafica[grupo] = {}
        
        # Procesar resultados
        for resultado in resultados:
            periodo = str(resultado['periodo'])
            grupo = resultado['grupo']
            total = resultado['total_tecnicos']
            
            periodos.add(periodo)
            if grupo in datos_grafica:
                datos_grafica[grupo][periodo] = total
        
        # Convertir a lista ordenada de períodos
        periodos_ordenados = sorted(list(periodos))
        
        # Preparar datos finales para Chart.js
        datasets = []
        colores = {
            'ARREGLOS': '#FF6384',
            'INSTALACIONES': '#36A2EB', 
            'POSTVENTA': '#FFCE56'
        }
        
        for grupo in grupos:
            data = []
            for periodo in periodos_ordenados:
                data.append(datos_grafica[grupo].get(periodo, 0))
            
            datasets.append({
                'label': grupo,
                'data': data,
                'borderColor': colores[grupo],
                'backgroundColor': colores[grupo] + '20',
                'fill': False,
                'tension': 0.1
            })
        
        return jsonify({
            'success': True,
            'datos': datasets,
            'labels': periodos_ordenados,
            'agrupacion': tipo_agrupacion,
            'fecha_inicio': fecha_inicio.strftime('%Y-%m-%d'),
            'fecha_fin': fecha_fin.strftime('%Y-%m-%d'),
            'supervisor_filtro': supervisor_filtro
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

# ===== RUTAS VENCIMIENTOS =====

@app.route('/mpa/vencimientos')
def mpa_vencimientos():
    """Módulo de gestión de vencimientos consolidados"""
    # Temporalmente sin autenticación para pruebas
    # if not current_user.has_role('administrativo'):
    #     flash('No tienes permisos para acceder a este módulo.', 'error')
    #     return redirect(url_for('mpa_dashboard'))
    
    return render_template('modulos/mpa/vencimientos.html')

# ===== APIs VENCIMIENTOS =====

# API consolidada para obtener todos los vencimientos
@app.route('/api/mpa/vencimientos', methods=['GET'])
def api_vencimientos_consolidados():
    """API consolidada para obtener vencimientos de SOAT, Técnico Mecánica y Licencias de Conducir"""
    # Temporalmente sin autenticación para pruebas
    # if not current_user.has_role('administrativo'):
    #     return jsonify({'error': 'Sin permisos'}), 403
    
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        vencimientos_consolidados = []
        
        # 1. Obtener vencimientos de SOAT
        query_soat = """
        SELECT 
            s.id_mpa_soat as id,
            s.placa,
            s.fecha_vencimiento,
            s.tecnico_asignado,
            ro.nombre as tecnico_nombre,
            'SOAT' as tipo,
            s.estado
        FROM mpa_soat s
        LEFT JOIN recurso_operativo ro ON s.tecnico_asignado = ro.id_codigo_consumidor
        WHERE s.fecha_vencimiento IS NOT NULL
          AND (
              s.tecnico_asignado IS NULL
              OR TRIM(s.tecnico_asignado) = ''
              OR ro.estado = 'Activo'
          )
        ORDER BY s.fecha_vencimiento ASC
        """
        
        cursor.execute(query_soat)
        soats = cursor.fetchall()
        
        for soat in soats:
            vencimientos_consolidados.append({
                'id': soat['id'],
                'tipo': 'SOAT',
                'placa': soat['placa'],
                'fecha_vencimiento': soat['fecha_vencimiento'].strftime('%Y-%m-%d') if soat['fecha_vencimiento'] else None,
                'tecnico_nombre': soat['tecnico_nombre'] or 'Sin asignar',
                'estado_original': soat['estado']
            })
        
        # 2. Obtener vencimientos de Técnico Mecánica
        query_tm = """
        SELECT 
            tm.id_mpa_tecnico_mecanica as id,
            tm.placa,
            tm.fecha_vencimiento,
            tm.tecnico_asignado,
            ro.nombre as tecnico_nombre,
            'Técnico Mecánica' as tipo,
            tm.estado
        FROM mpa_tecnico_mecanica tm
        LEFT JOIN recurso_operativo ro ON tm.tecnico_asignado = ro.id_codigo_consumidor
        WHERE tm.fecha_vencimiento IS NOT NULL
          AND (
              tm.tecnico_asignado IS NULL
              OR TRIM(tm.tecnico_asignado) = ''
              OR ro.estado = 'Activo'
          )
        ORDER BY tm.fecha_vencimiento ASC
        """
        
        cursor.execute(query_tm)
        tecnicos = cursor.fetchall()
        
        for tm in tecnicos:
            vencimientos_consolidados.append({
                'id': tm['id'],
                'tipo': 'Técnico Mecánica',
                'placa': tm['placa'],
                'fecha_vencimiento': tm['fecha_vencimiento'].strftime('%Y-%m-%d') if tm['fecha_vencimiento'] else None,
                'tecnico_nombre': tm['tecnico_nombre'] or 'Sin asignar',
                'estado_original': tm['estado']
            })
        
        # 3. Obtener vencimientos de Licencias de Conducir
        query_lc = """
        SELECT 
            lc.id_mpa_licencia_conducir as id,
            lc.fecha_vencimiento,
            lc.tecnico,
            ro.nombre as tecnico_nombre,
            'Licencia de Conducir' as tipo
        FROM mpa_licencia_conducir lc
        LEFT JOIN recurso_operativo ro ON lc.tecnico = ro.id_codigo_consumidor
        WHERE lc.fecha_vencimiento IS NOT NULL
          AND (
              lc.tecnico IS NULL
              OR TRIM(lc.tecnico) = ''
              OR ro.estado = 'Activo'
          )
        ORDER BY lc.fecha_vencimiento ASC
        """
        
        cursor.execute(query_lc)
        licencias = cursor.fetchall()
        
        for lc in licencias:
            vencimientos_consolidados.append({
                'id': lc['id'],
                'tipo': 'Licencia de Conducir',
                'placa': None,  # Las licencias no tienen placa
                'fecha_vencimiento': lc['fecha_vencimiento'].strftime('%Y-%m-%d') if lc['fecha_vencimiento'] else None,
                'tecnico_nombre': lc['tecnico_nombre'] or 'Sin asignar',
                'estado_original': None  # Las licencias no tienen campo estado
            })
        
        # 4. Calcular días restantes y estado para todos los vencimientos
        from datetime import datetime
        import pytz
        
        colombia_tz = pytz.timezone('America/Bogota')
        fecha_actual = datetime.now(colombia_tz).date()
        
        for vencimiento in vencimientos_consolidados:
            if vencimiento['fecha_vencimiento']:
                fecha_venc = datetime.strptime(vencimiento['fecha_vencimiento'], '%Y-%m-%d').date()
                dias_restantes = (fecha_venc - fecha_actual).days
                
                vencimiento['dias_restantes'] = dias_restantes
                
                # Determinar estado basado en días restantes
                if dias_restantes < 0:
                    vencimiento['estado'] = 'Vencido'
                elif dias_restantes <= 30:
                    vencimiento['estado'] = 'Próximo a vencer'
                else:
                    vencimiento['estado'] = 'Vigente'
            else:
                vencimiento['dias_restantes'] = None
                vencimiento['estado'] = 'Sin fecha'
        
        # 5. Ordenar por fecha de vencimiento (más próximos primero)
        vencimientos_consolidados.sort(key=lambda x: (
            x['fecha_vencimiento'] if x['fecha_vencimiento'] else '9999-12-31'
        ))
        
        return jsonify({
            'success': True,
            'data': vencimientos_consolidados,
            'total': len(vencimientos_consolidados)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

# API para obtener detalles de un vencimiento específico
@app.route('/api/mpa/vencimiento/<string:tipo>/<int:documento_id>', methods=['GET'])
def api_get_vencimiento_detalle(tipo, documento_id):
    """API para obtener detalles de un vencimiento específico"""
    # Temporalmente sin autenticación para pruebas
    # if not current_user.has_role('administrativo'):
    #     return jsonify({'error': 'Sin permisos'}), 403
    
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'error': 'Error de conexión a la base de datos'}), 500
            
        cursor = connection.cursor(dictionary=True)
        
        # Determinar la tabla y consulta según el tipo
        if tipo.lower() == 'soat':
            query = """
            SELECT 
                s.*,
                ro.nombre as tecnico_nombre,
                ro.recurso_operativo_cedula as tecnico_cedula
            FROM mpa_soat s
            LEFT JOIN recurso_operativo ro ON s.tecnico_asignado = ro.id_codigo_consumidor
            WHERE s.id_mpa_soat = %s
            """
        elif tipo.lower() == 'tecnico_mecanica':
            query = """
            SELECT 
                tm.*,
                ro.nombre as tecnico_nombre,
                ro.recurso_operativo_cedula as tecnico_cedula
            FROM mpa_tecnico_mecanica tm
            LEFT JOIN recurso_operativo ro ON tm.tecnico_asignado = ro.id_codigo_consumidor
            WHERE tm.id_mpa_tecnico_mecanica = %s
            """
        elif tipo.lower() == 'licencia_conducir':
            query = """
            SELECT 
                lc.*,
                ro.nombre as tecnico_nombre,
                ro.recurso_operativo_cedula as tecnico_cedula
            FROM mpa_licencia_conducir lc
            LEFT JOIN recurso_operativo ro ON lc.tecnico = ro.id_codigo_consumidor
            WHERE lc.id_mpa_licencia_conducir = %s
            """
        else:
            return jsonify({'success': False, 'error': 'Tipo de documento no válido'}), 400
        
        cursor.execute(query, (documento_id,))
        documento = cursor.fetchone()
        
        if not documento:
            return jsonify({'success': False, 'error': 'Documento no encontrado'}), 404
        
        # Formatear fechas y calcular días restantes
        from datetime import datetime
        import pytz
        
        colombia_tz = pytz.timezone('America/Bogota')
        fecha_actual = datetime.now(colombia_tz).date()
        
        # Formatear fechas
        for key, value in documento.items():
            if isinstance(value, datetime):
                documento[key] = value.strftime('%Y-%m-%d %H:%M:%S')
            elif hasattr(value, 'strftime'):  # Para objetos date
                documento[key] = value.strftime('%Y-%m-%d')
        
        # Calcular días restantes si hay fecha de vencimiento
        if documento.get('fecha_vencimiento'):
            try:
                fecha_venc = datetime.strptime(documento['fecha_vencimiento'], '%Y-%m-%d').date()
                dias_restantes = (fecha_venc - fecha_actual).days
                documento['dias_restantes'] = dias_restantes
                
                # Determinar estado
                if dias_restantes < 0:
                    documento['estado_calculado'] = 'Vencido'
                elif dias_restantes <= 30:
                    documento['estado_calculado'] = 'Próximo a vencer'
                else:
                    documento['estado_calculado'] = 'Vigente'
            except:
                documento['dias_restantes'] = None
                documento['estado_calculado'] = 'Sin fecha'
        
        return jsonify({
            'success': True,
            'data': documento,
            'tipo': tipo
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

# API de prueba con ruta diferente
@app.route('/api/mpa/test-vencimientos', methods=['GET'])
def api_test_vencimientos():
    """API de prueba para vencimientos"""
    return jsonify({
        'success': True,
        'data': [],
        'message': 'API de prueba funcionando correctamente'
    })

@app.route('/api/mi/vencimientos', methods=['GET'])
@login_required()
def api_mis_vencimientos():
    """Devuelve vencimientos próximos para el usuario actual:
    - SOAT y Tecnomecánica de sus vehículos asignados
    - Licencia de conducir del usuario
    Filtra por ventana de días (param "dias", default 45).
    """
    try:
        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500

        cursor = connection.cursor(dictionary=True)

        dias_anticipacion = int(request.args.get('dias', 45))
        user_id = session.get('user_id') or session.get('id_codigo_consumidor')
        if not user_id:
            return jsonify({'success': False, 'message': 'Usuario no autenticado'}), 401

        vencimientos_items = []

        # SOAT desde mpa_soat por técnico asignado
        query_soat = """
            SELECT 
                s.placa,
                s.fecha_vencimiento,
                DATEDIFF(s.fecha_vencimiento, CURDATE()) AS dias_soat
            FROM mpa_soat s
            WHERE s.tecnico_asignado = %s
              AND s.fecha_vencimiento IS NOT NULL
              AND s.fecha_vencimiento NOT LIKE '0000-00-00%'
              AND s.fecha_vencimiento > '1900-01-01'
              AND s.fecha_vencimiento <= DATE_ADD(CURDATE(), INTERVAL %s DAY)
            ORDER BY s.fecha_vencimiento ASC
        """
        cursor.execute(query_soat, (user_id, dias_anticipacion))
        soats = cursor.fetchall()

        # Tecnomecánica desde mpa_tecnico_mecanica por técnico asignado
        query_tm = """
            SELECT 
                tm.placa,
                tm.fecha_vencimiento,
                DATEDIFF(tm.fecha_vencimiento, CURDATE()) AS dias_tm
            FROM mpa_tecnico_mecanica tm
            WHERE tm.tecnico_asignado = %s
              AND tm.fecha_vencimiento IS NOT NULL
              AND tm.fecha_vencimiento NOT LIKE '0000-00-00%'
              AND tm.fecha_vencimiento > '1900-01-01'
              AND tm.fecha_vencimiento <= DATE_ADD(CURDATE(), INTERVAL %s DAY)
            ORDER BY tm.fecha_vencimiento ASC
        """
        cursor.execute(query_tm, (user_id, dias_anticipacion))
        tecnomecanicas = cursor.fetchall()

        def estado_desde_dias(dias:int) -> str:
            if dias is None:
                return 'desconocido'
            if dias < 0:
                return 'vencido'
            if dias <= 7:
                return 'critico'
            if dias <= 30:
                return 'proximo'
            return 'vigente'

        # Armar items SOAT
        for s in soats:
            dias = s.get('dias_soat')
            fecha = s.get('fecha_vencimiento')
            vencimientos_items.append({
                'tipo': 'SOAT',
                'placa': s.get('placa'),
                'fecha_vencimiento': fecha.strftime('%Y-%m-%d') if fecha else None,
                'dias_restantes': dias,
                'estado': estado_desde_dias(dias)
            })

        # Armar items Tecnomecánica
        for tm in tecnomecanicas:
            dias = tm.get('dias_tm')
            fecha = tm.get('fecha_vencimiento')
            vencimientos_items.append({
                'tipo': 'Tecnomecánica',
                'placa': tm.get('placa'),
                'fecha_vencimiento': fecha.strftime('%Y-%m-%d') if fecha else None,
                'dias_restantes': dias,
                'estado': estado_desde_dias(dias)
            })

        # Licencia de conducir del usuario
        query_lic = """
            SELECT 
                lc.tipo_licencia,
                lc.fecha_vencimiento,
                DATEDIFF(lc.fecha_vencimiento, CURDATE()) as dias_licencia
            FROM mpa_licencia_conducir lc
            WHERE lc.tecnico = %s
              AND lc.fecha_vencimiento IS NOT NULL
              AND lc.fecha_vencimiento NOT IN ('0000-00-00', '1900-01-01')
              AND lc.fecha_vencimiento <= DATE_ADD(CURDATE(), INTERVAL %s DAY)
            ORDER BY lc.fecha_vencimiento ASC
        """
        cursor.execute(query_lic, (user_id, dias_anticipacion))
        licencias = cursor.fetchall()

        for lc in licencias:
            dias = lc.get('dias_licencia')
            vencimientos_items.append({
                'tipo': 'Licencia de Conducir',
                'placa': None,
                'tipo_licencia': lc.get('tipo_licencia'),
                'fecha_vencimiento': lc['fecha_vencimiento'].strftime('%Y-%m-%d') if lc['fecha_vencimiento'] else None,
                'dias_restantes': dias,
                'estado': estado_desde_dias(dias)
            })

        cursor.close()
        connection.close()

        items_filtrados = [i for i in vencimientos_items if i.get('fecha_vencimiento')]

        return jsonify({
            'success': True,
            'total': len(items_filtrados),
            'items': items_filtrados
        })

    except Exception as e:
        try:
            if 'cursor' in locals() and cursor:
                cursor.close()
            if 'connection' in locals() and connection and connection.is_connected():
                connection.close()
        except Exception:
            pass
        print(f"Error en /api/mi/vencimientos: {str(e)}")
        return jsonify({'success': False, 'message': f'Error al obtener vencimientos: {str(e)}'}), 500


@app.route('/api/operativo/vencimientos_tecnicos', methods=['GET'])
@login_required(role=['operativo', 'administrativo'])
def api_vencimientos_tecnicos_operativo():
    """Devuelve documentos vencidos de los técnicos a cargo de un supervisor (SOAT, Técnico Mecánica, Licencia)."""
    try:
        # Usar SIEMPRE el usuario logueado como supervisor responsable de técnicos
        supervisor = session.get('user_name', '')
        dias_param = request.args.get('dias', default='30')
        try:
            dias_ventana = int(dias_param)
        except Exception:
            dias_ventana = 30

        if not supervisor:
            return jsonify({'success': False, 'message': 'Usuario no identificado en sesión'}), 400

        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'}), 500
        cursor = connection.cursor(dictionary=True)

        # SOAT por técnicos del supervisor
        # Mejora: considerar vínculos por técnico_asignado o por placa del vehículo asignado
        cursor.execute(
            """
            SELECT DISTINCT
                s.id_mpa_soat AS id,
                s.fecha_vencimiento,
                ro.id_codigo_consumidor AS tecnico_id,
                ro.nombre AS tecnico_nombre,
                'SOAT' AS tipo
            FROM capired.recurso_operativo ro
            LEFT JOIN capired.mpa_vehiculos mv ON mv.tecnico_asignado = ro.id_codigo_consumidor
            LEFT JOIN capired.mpa_soat s ON (s.tecnico_asignado = ro.id_codigo_consumidor OR (mv.placa IS NOT NULL AND s.placa = mv.placa))
            WHERE TRIM(UPPER(ro.super)) = TRIM(UPPER(%s)) AND ro.estado = 'Activo'
              AND s.fecha_vencimiento IS NOT NULL
              AND s.fecha_vencimiento NOT LIKE '0000-00-00%'
              AND s.fecha_vencimiento > '1900-01-01'
            """,
            (supervisor,)
        )
        soats = cursor.fetchall()

        # Técnico Mecánica por técnicos del supervisor
        cursor.execute(
            """
            SELECT DISTINCT
                tm.id_mpa_tecnico_mecanica AS id,
                tm.fecha_vencimiento,
                ro.id_codigo_consumidor AS tecnico_id,
                ro.nombre AS tecnico_nombre,
                'Técnico Mecánica' AS tipo
            FROM capired.recurso_operativo ro
            LEFT JOIN capired.mpa_vehiculos mv ON mv.tecnico_asignado = ro.id_codigo_consumidor
            LEFT JOIN capired.mpa_tecnico_mecanica tm ON (tm.tecnico_asignado = ro.id_codigo_consumidor OR (mv.placa IS NOT NULL AND tm.placa = mv.placa))
            WHERE TRIM(UPPER(ro.super)) = TRIM(UPPER(%s)) AND ro.estado = 'Activo'
              AND tm.fecha_vencimiento IS NOT NULL
              AND tm.fecha_vencimiento NOT LIKE '0000-00-00%'
              AND tm.fecha_vencimiento > '1900-01-01'
            """,
            (supervisor,)
        )
        tms = cursor.fetchall()

        # Licencias de conducir por técnicos del supervisor
        cursor.execute(
            """
            SELECT 
                lc.id_mpa_licencia_conducir AS id,
                lc.fecha_vencimiento,
                ro.id_codigo_consumidor AS tecnico_id,
                ro.nombre AS tecnico_nombre,
                'Licencia de Conducir' AS tipo
            FROM capired.mpa_licencia_conducir lc
            LEFT JOIN capired.recurso_operativo ro ON lc.tecnico = ro.id_codigo_consumidor
            WHERE TRIM(UPPER(ro.super)) = TRIM(UPPER(%s)) AND ro.estado = 'Activo'
              AND lc.fecha_vencimiento IS NOT NULL
              AND lc.fecha_vencimiento NOT LIKE '0000-00-00%'
              AND lc.fecha_vencimiento > '1900-01-01'
            """,
            (supervisor,)
        )
        lcs = cursor.fetchall()

        # Calcular estado y filtrar vencidos
        from datetime import datetime, date
        import re
        import pytz
        # Asegurar cálculo de hoy en zona horaria de Colombia para evitar desfases
        colombia_tz = pytz.timezone('America/Bogota')
        hoy = datetime.now(colombia_tz).date()

        def validar_fecha(fecha_val):
            if not fecha_val:
                return False, None
            if hasattr(fecha_val, 'date'):
                return True, fecha_val.date()
            s = str(fecha_val)
            if s == '0000-00-00' or not re.match(r'^\d{4}-\d{2}-\d{2}$', s):
                return False, None
            try:
                y, m, d = map(int, s.split('-'))
                return True, date(y, m, d)
            except Exception:
                return False, None

        vencidos = []
        for grupo in (soats, tms, lcs):
            for v in grupo:
                ok, fv = validar_fecha(v['fecha_vencimiento'])
                if not ok:
                    continue
                dias = (fv - hoy).days
                # Consistencia de bloqueo: Vencido solo cuando días restantes < 0
                estado = 'Vencido' if dias < 0 else ('Próximo a vencer' if dias <= dias_ventana else 'Vigente')
                if estado == 'Vencido':
                    vencidos.append({
                        'id': v['id'],
                        'tipo': v['tipo'],
                        'tecnico_id': v['tecnico_id'],
                        'tecnico_nombre': v['tecnico_nombre'],
                        'fecha_vencimiento': fv.strftime('%Y-%m-%d'),
                        'dias_restantes': dias,
                        'estado': estado
                    })

        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

        return jsonify({'success': True, 'data': vencidos, 'supervisor': supervisor, 'total_vencidos': len(vencidos)})

    except mysql.connector.Error as e:
        return jsonify({'success': False, 'message': f'Error de base de datos: {str(e)}'}), 500
    except Exception as e:
        try:
            if 'cursor' in locals() and cursor:
                cursor.close()
            if 'connection' in locals() and connection and connection.is_connected():
                connection.close()
        except Exception:
            pass
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

@app.route('/dev/login', methods=['POST'])
def dev_login():
    try:
        # Solo disponible en modo debug
        if not app.debug:
            return jsonify({'message': 'No disponible en producción'}), 403
        data = request.get_json(force=True) or {}
        name = (data.get('user_name') or '').strip()
        cedula = (data.get('user_cedula') or '').strip()
        role = (data.get('user_role') or 'operativo').strip()
        if not name:
            return jsonify({'message': 'user_name requerido'}), 400
        session['user_name'] = name
        session['user_cedula'] = cedula
        session['user_role'] = role
        return jsonify({'success': True, 'message': 'Sesión de desarrollo creada', 'user_name': name, 'user_role': role})
    except Exception as e:
        return jsonify({'message': 'Error en dev login', 'error': str(e)}), 500

# Submódulo: Novedades asistencia (vista calendario por mes)
@app.route('/administrativo/novedades', methods=['GET'], endpoint='administrativo_novedades')
@login_required
def administrativo_novedades():
    """Renderiza el submódulo de Novedades asistencia"""
    try:
        # Datos iniciales: mes actual
        now = get_bogota_datetime()
        current_year = now.year
        current_month = now.month
        return render_template(
            'modulos/administrativo/novedades_asistencia.html',
            current_year=current_year,
            current_month=f"{current_month:02d}"
        )
    except Exception as e:
        return render_template('error.html', mensaje='Error al cargar Novedades', error=str(e))

# API: matriz mensual de novedades de asistencia
@app.route('/api/asistencia/novedades', methods=['GET'], endpoint='api_asistencia_novedades')
@login_required
def api_asistencia_novedades():
    """Devuelve matriz de asistencia por usuario y día del mes.
    Verde: tiene registro de asistencia.
    Rojo: registro con carpeta_dia=0 o carpeta=0 (ausencia injustificada).
    """
    try:
        year = request.args.get('year', type=int)
        month = request.args.get('month', type=int)

        now = get_bogota_datetime()
        if not year:
            year = now.year
        if not month:
            month = now.month

        # Determinar cantidad de días del mes
        import calendar
        days_in_month = calendar.monthrange(year, month)[1]

        connection = get_db_connection()
        if connection is None:
            return jsonify({'success': False, 'message': 'Error de conexión a la base de datos'})

        cursor = connection.cursor(dictionary=True)

        # Usuarios con asistencia registrada en el mes (indexar por cédula, nombre desde recurso_operativo)
        cursor.execute(
            """
            SELECT DISTINCT a.cedula AS usuario_id,
                            COALESCE(ro.nombre, a.tecnico) AS nombre
            FROM asistencia a
            LEFT JOIN recurso_operativo ro ON ro.recurso_operativo_cedula = a.cedula
            WHERE YEAR(a.fecha_asistencia) = %s AND MONTH(a.fecha_asistencia) = %s
            ORDER BY nombre
            """,
            (year, month)
        )
        usuarios = cursor.fetchall() or []

        # Registros por día para el mes (clave por cédula) usando SOLO carpeta_dia
        # Además, recolectar los valores de carpeta_dia para mostrarlos en tooltips
        cursor.execute(
            """
            SELECT a.cedula AS usuario_id,
                   DATE(a.fecha_asistencia) AS fecha,
                   MAX(CASE WHEN TRIM(COALESCE(a.carpeta_dia, '')) = '0' THEN 1 ELSE 0 END) AS tiene_cero,
                   COUNT(*) AS registros,
                   GROUP_CONCAT(DISTINCT TRIM(COALESCE(a.carpeta_dia, '')) ORDER BY TRIM(COALESCE(a.carpeta_dia, '')) SEPARATOR '|') AS carpetas_raw
            FROM asistencia a
            WHERE YEAR(a.fecha_asistencia) = %s AND MONTH(a.fecha_asistencia) = %s
            GROUP BY a.cedula, DATE(a.fecha_asistencia)
            """,
            (year, month)
        )
        registros = cursor.fetchall() or []

        # Indexar por usuario y día con detalle de carpetas
        por_usuario = {}
        for r in registros:
            uid = r['usuario_id']
            dia = int(str(r['fecha']).split('-')[-1])
            valor = None
            if (r['registros'] or 0) > 0:
                valor = 0 if r['tiene_cero'] == 1 else 1
            raw = (r.get('carpetas_raw') or '')
            # Convertir el GROUP_CONCAT en lista, filtrando vacíos
            carpetas_list = [c for c in raw.split('|') if c]
            por_usuario.setdefault(uid, {})[dia] = {
                'valor': valor,
                'carpetas': carpetas_list if carpetas_list else None
            }

        # Construir matriz
        matriz = []
        for u in usuarios:
            fila = []
            fila_carpetas = []
            for d in range(1, days_in_month + 1):
                item = por_usuario.get(u['usuario_id'], {}).get(d, None)
                valor = item['valor'] if isinstance(item, dict) else item
                carpetas = (item.get('carpetas') if isinstance(item, dict) else None)
                fila.append(valor)
                fila_carpetas.append(carpetas)
            matriz.append({
                'cedula': u['usuario_id'],
                'tecnico': u['nombre'],
                'dias': fila,
                'carpetas': fila_carpetas
            })

        return jsonify({
            'success': True,
            'year': year,
            'month': month,
            'days_in_month': days_in_month,
            'usuarios': matriz
        })
    except mysql.connector.Error as e:
        return jsonify({'success': False, 'message': f'Error de base de datos: {str(e)}'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})
    finally:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'connection' in locals() and connection and connection.is_connected():
            connection.close()

if __name__ == '__main__':
    # Handlers de error para evitar exposición de información y listados
    @app.errorhandler(404)
    def handle_404(e):
        try:
            if request.path.startswith('/api/'):
                return jsonify({'success': False, 'message': 'Recurso no encontrado'}), 404
            return render_template('error.html', code=404, message='Recurso no encontrado'), 404
        except Exception:
            return 'Recurso no encontrado', 404

    @app.errorhandler(500)
    def handle_500(e):
        try:
            if request.path.startswith('/api/'):
                return jsonify({'success': False, 'message': 'Error interno del servidor'}), 500
            return render_template('error.html', code=500, message='Error interno del servidor'), 500
        except Exception:
            return 'Error interno del servidor', 500

    @app.errorhandler(405)
    def handle_405(e):
        try:
            if request.path.startswith('/api/'):
                return jsonify({'success': False, 'message': 'Método no permitido'}), 405
            return render_template('error.html', code=405, message='Método no permitido'), 405
        except Exception:
            return 'Método no permitido', 405

    # Deshabilitar modo debug por defecto para evitar fugas de información
    import os
    debug_mode = os.environ.get('FLASK_DEBUG', '0') == '1'
    app.run(debug=debug_mode, host='0.0.0.0', port=8080)
